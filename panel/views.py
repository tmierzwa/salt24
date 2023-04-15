# -*- coding: utf-8 -*-

import time
import pytz, babel.dates
from datetime import date, datetime, timedelta
from dateutil.relativedelta import relativedelta
from django.utils.timezone import localtime
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, fills, Border, borders, Side, Font
from openpyxl.writer.excel import save_virtual_workbook
from django.utils import timezone
from django import forms
from django.urls import reverse
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.views.generic.list import ListView
from django.views.generic.detail import DetailView
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import get_object_or_404, render
from django.forms.models import modelform_factory
from django.forms import Textarea, TextInput, Form, ChoiceField, DateField
from django.forms.widgets import TimeInput
from django.contrib.auth.decorators import permission_required, user_passes_test, login_required
from django.contrib.admin.widgets import AdminDateWidget
from django.contrib import messages
from django.utils.decorators import method_decorator
from django.db.models import Sum, Q
from django_datatables_view.base_datatable_view import BaseDatatableView


def class_view_decorator(function_decorator):

    def simple_decorator(View):
        View.dispatch = method_decorator(function_decorator)(View.dispatch)
        return View

    return simple_decorator

from salt.forms import duration_string

from camo.models import Aircraft
from panel.models import FBOUser, PDT, Operation, Pilot, PilotAircraft, PilotFlightType, Duty, FlightTypes, Rating
from ato.models import Training_inst, Phase_inst, Exercise_inst, Instructor
from fin.models import FuelTank, Contractor, Voucher, BalanceOperation, LocalFueling, PDTFueling, RemoteFueling
from sms.models import SMSFailure, SMSReport, SMSEvent
from res.models import Reservation

from panel.forms import PilotAircraftForm, PilotTypesForm
from panel.forms import CheckOutFormOpen, CheckOutFormServices, CheckOutFormVouchers, CheckOutFormTrainings
from panel.forms import CheckOutFormRent, CheckOutFormCrew, CheckInForm
from panel.forms import OperationEditForm, OperationOpenForm, OperationCloseForm
from panel.forms import PanelFuelingForm, PDTEditForm

from res.views import reservation_check, reservation_msg


@login_required()
def PanelHome (request):

    this_day = datetime(datetime.today().year, datetime.today().month, datetime.today().day)

    # Jeśli jest pilotem
    if hasattr(request.user.fbouser, 'pilot'):
        # Wybierz te które utworzyłeś lub jesteś właścicielem
        query_ac = Reservation.objects.filter(
            Q(owner=request.user.fbouser.pilot) | Q(participant=request.user.fbouser.pilot) | Q(open_user=request.user.fbouser)). \
            exclude(end_time__lt=this_day).exclude(status='Zrealizowana').order_by('start_time')
    else:
        # Pozostałe przypadki
        query_ac = Reservation.objects.none()

    my_open_pdt = request.user.fbouser.open_pdt()

    context = {}

    # Lokalne menu
    local_menu = []
    if hasattr(request.user.fbouser, 'pilot'):
        local_menu.append({'text': 'Nowa rezerwacja SP', 'path': reverse("res:reservation-create")})
    local_menu.append({'text': 'Kalendarz rezerwacji', 'path': reverse("res:reservation-calendar")})
    context['local_menu1'] = local_menu

    # Nagłówki tabeli
    header_list = []
    header_list.append({'header': 'SP'})
    header_list.append({'header': 'Data\nrezerwacji'})
    header_list.append({'header': 'Termin\nrezerwacji'})
    header_list.append({'header': 'Właściciel\nrezerwacji'})
    header_list.append({'header': 'Uczestnik\nrezerwacji'})
    header_list.append({'header': 'Rodzaj\nlotu'})
    header_list.append({'header': '...'})
    context['header_list1'] = header_list

    next_pdt_res = None

    row_list = []
    for res in query_ac:
        fields = []
        days_length = (res.end_time.date() - res.start_time.date()).days
        fields.append({'name': 'resource', 'value': res.aircraft,
                       'link': reverse('res:reservation-info', args=[res.pk]), 'just': 'center'})
        fields.append({'name': 'date', 'value': babel.dates.format_date(localtime(res.start_time), "EEE dd.MM", locale='pl_PL'),
                       'just': 'center'})
        fields.append({'name': 'time', 'value': "%s - %s" % (localtime(res.start_time).strftime("%H:%M"),
                                                             localtime(res.end_time).strftime("%H:%M") + (
                                                                 ' (+%d)' % days_length if days_length else ''))})
        fields.append({'name': 'owner', 'value': res.owner.__str__()[:20]})
        fields.append({'name': 'participant', 'value': res.participant.__str__()[:20] if res.participant else ''})
        fields.append({'name': 'title', 'value': res.planned_type, 'just': 'center'})

        open_pdt = reservation_check(res, request.user.fbouser)

        if open_pdt and not next_pdt_res:
            next_pdt_res = res

        if open_pdt:
            fields.append({'name': 'change', 'report_link': ('%s?res=%d' % (reverse('panel:pdt-open'), res.pk))})
        else:
            fields.append({'name': 'change', 'no_report_msg': reservation_msg(res, request.user.fbouser)})

        row_list.append({'fields': fields})

    context['row_list1'] = row_list

    # Jeśli jest pilotem
    if hasattr(request.user.fbouser, 'pilot'):
        flight_list = PDT.objects.filter(Q(pic=request.user.fbouser.pilot) | Q(sic=request.user.fbouser.pilot) |
                                         Q(open_user=request.user.fbouser) | Q(close_user=request.user.fbouser)).order_by('-date', '-pk')[:50]
    else:
        flight_list = PDT.objects.filter(Q(open_user=request.user.fbouser) | Q(close_user=request.user.fbouser)).order_by('-date', '-pk')[:50]

    # Lokalne menu
    local_menu = []
    if my_open_pdt:
        local_menu.append({'text': 'Przejdź do otwartego PDT', 'path': reverse("panel:pdt-update", args=[my_open_pdt.pk])})
    else:
        if next_pdt_res:
            path = '%s?res=%d' % (reverse('panel:pdt-open'), next_pdt_res.pk)
        else:
            path = reverse('panel:pdt-open')

            # if request.user.fbouser.mechanic:
            #     path = reverse('panel:pdt-open')
            # else:
            #     path = reverse('res:reservation-calendar')

        local_menu.append({'text': 'Otwórz PDT', 'path': path})

    context['local_menu2'] = local_menu

    # Nagłówki tabeli
    header_list = []
    header_list.append({'header': 'Data'})
    header_list.append({'header': 'SP'})
    header_list.append({'header': 'PDT'})
    header_list.append({'header': 'Rodzaj\nlotu'})
    header_list.append({'header': 'TTH'})
    header_list.append({'header': 'Czas'})
    header_list.append({'header': 'Ldg.'})
    header_list.append({'header': 'Status'})
    header_list.append({'header': '...'})
    context['header_list2'] = header_list

    # Zawartość tabeli
    row_list = []
    for object in flight_list:
        fields = []
        fields.append({'name': 'date', 'value': object.date})
        fields.append({'name': 'aircraft', 'value': object.aircraft})
        fields.append({'name': 'pdt', 'value': '{:0>6d}'.format(object.pdt_ref),
                       'link': reverse('panel:pdt-info', args=[object.pk]), 'just': 'center'})
        fields.append({'name': 'flight_type', 'value': object.flight_type_name()})
        fields.append({'name': 'tth', 'value': '%.1f' % object.tth_sum(), 'just': 'center'})
        fields.append({'name': 'time', 'value': '%02.d:%02.d' % (object.hours_sum()[1], object.hours_sum()[2]), 'just': 'center'})
        fields.append({'name': 'landings', 'value': object.landings_sum(), 'just': 'center'})
        fields.append({'name': 'status', 'value': object.status_name(), 'color': object.status_color()})
        fields.append({'name': 'change', 'report_link': reverse('panel:pdt-info', args=[object.pk])})
        row_list.append({'fields': fields})

    context['row_list2'] = row_list

    trainings_list = Training_inst.objects.none()

    # Jeśli jest pilotem
    if hasattr(request.user.fbouser, 'pilot'):
        if hasattr(request.user.fbouser.pilot, 'instructor'):
            trainings_list = Training_inst.objects.filter(instructor=request.user.fbouser.pilot.instructor, open=True).order_by('-start_date')
        elif hasattr(request.user.fbouser.pilot, 'student'):
            trainings_list = Training_inst.objects.filter(student=request.user.fbouser.pilot.student, open=True).order_by('-start_date')

    # Lokalne menu
    local_menu = []
    context['local_menu3'] = local_menu

    # Nagłówki tabeli
    header_list = []
    header_list.append({'header': 'Kod'})
    if hasattr(request.user.fbouser, 'pilot') and hasattr(request.user.fbouser.pilot, 'instructor'):
        header_list.append({'header': 'Student'})
    else:
        header_list.append({'header': 'Instruktor'})
    header_list.append({'header': 'Data\nrozpoczęcia'})
    header_list.append({'header': 'Godzin\nz instr.'})
    header_list.append({'header': 'Godzin\nsolo'})
    header_list.append({'header': 'Status'})
    header_list.append({'header': '...'})
    context['header_list3'] = header_list

    # Zawartość tabeli
    row_list = []
    for object in trainings_list:
        fields = []
        fields.append({'name': 'code', 'value': object.training.code,
                       'link': reverse('ato:training-inst-details', args=[object.pk]), 'just': 'center'})
        if hasattr(request.user.fbouser, 'pilot') and hasattr(request.user.fbouser.pilot, 'instructor'):
            fields.append({'name': 'student', 'value': object.student, 'link': reverse('ato:student-info', args=[object.student.pk])})
        else:
            fields.append({'name': 'instructor', 'value': object.instructor})
        fields.append({'name': 'date', 'value': object.start_date, 'just': 'center'})
        fields.append({'name': 'time_instr', 'value': duration_string(object.time_instr()) if object.time_instr() else '---', 'just': 'center'})
        fields.append({'name': 'time_solo', 'value': duration_string(object.time_solo()) if object.time_solo() else '---', 'just': 'center'})
        fields.append({'name': 'status', 'value': object.status(), 'just': 'center'})
        fields.append({'name': 'change', 'report_link': reverse('ato:training-inst-details', args=[object.pk])})

        row_list.append({'fields': fields})

    context['row_list3'] = row_list

    return render(request, 'panel/panel_main.html', context)


@class_view_decorator(login_required())
class PilotInfo (DetailView):
    model = Pilot
    template_name = 'panel/details_template.html'

    # jeśli nie superuser to ustaw na zalogowanego
    def get_object(self, queryset=None):
        obj = super(PilotInfo, self).get_object()
        if not self.request.user.is_staff:
            obj = self.request.user.fbouser.pilot
        return obj

    def get_context_data(self, **kwargs):
        context = super(PilotInfo, self).get_context_data(**kwargs)
        context['page_title'] = 'Pilot'
        context['header_text'] = 'Informacje o pilocie %s' % self.object
        context['type'] = 'info'
        context['submenu_template'] = 'panel/pilot_submenu.html'

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Aktualizuj', 'path': reverse('panel:pilot-update', args=[self.object.pk])})
        if self.request.user.is_staff:
            local_menu.append({'text': 'Usuń', 'path': reverse('fbo:pilot-delete', args=[self.object.pk])})
        context['local_menu'] = local_menu

        field_list = []
        field_list.append({'header': 'Nazwisko i imię', 'value': self.object.fbouser, 'bold': True})
        field_list.append({'header': 'Adres e-mail', 'value': self.object.fbouser.email})
        field_list.append({'header': 'Numer telefonu', 'value': self.object.fbouser.telephone})
        field_list.append({'header': 'Licencja', 'value': (self.object.licence or "") + (' - ważna do %s' % self.object.licence_date if self.object.licence_date else '')})
        field_list.append({'header': 'Badania lotnicze', 'value': (self.object.medical or "") + (' - ważne do %s' % self.object.medical_date if self.object.medical_date else '')})
        field_list.append({'header': 'Upoważnienie SALT', 'value': self.object.clearance})
        field_list.append({'header': 'Czas pracy', 'value': ("TAK" if self.object.employee else "NIE")})
        field_list.append({'header': 'Uwagi', 'value': self.object.remarks})
        context['field_list'] = field_list
        return context


@class_view_decorator(login_required())
class PilotRatings (DetailView):
    model = Pilot
    template_name = 'panel/list_template.html'

    # jeśli nie superuser to ustaw na zalogowanego
    def get_object(self, queryset=None):
        obj = super(PilotRatings, self).get_object()
        if not self.request.user.is_staff:
            obj = self.request.user.fbouser.pilot
        return obj

    def get_context_data(self, **kwargs):
        context = super(PilotRatings, self).get_context_data(**kwargs)
        context['page_title'] = 'Uprawnienia'
        context['header_text'] = 'Uprawnienia pilota %s' % self.object
        context['empty_text'] = 'Brak zarejestrowanych uprawnień.'
        context['type'] = 'ratings'
        context['submenu_template'] = 'panel/pilot_submenu.html'

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Dodaj nowe uprawnienie', 'path': reverse("panel:rating-create", args=[self.object.pk])})
        context['local_menu'] = local_menu

        # Nagłówki tabeli
        header_list = []
        header_list.append({'header': 'Uprawnienie'})
        header_list.append({'header': 'Data\nważności'})
        header_list.append({'header': 'Uwagi'})
        header_list.append({'header': '...'})
        context['header_list'] = header_list

        # Zawartość tabeli
        row_list = []
        for object in self.object.rating_set.all():
            fields = []
            fields.append({'name': 'rating', 'value': object.rating, 'link': reverse('panel:rating-details', args=[object.pk]),
                           'color': 'lightcoral' if object.valid_date and object.valid_date <= date.today() else 'auto', 'just': 'center'})
            fields.append({'name': 'valid_date', 'value': object.valid_date if object.valid_date else 'bezterminowo',
                           'color': 'lightcoral' if object.valid_date and object.valid_date <= date.today() else 'auto', 'just': 'center'})
            fields.append({'name': 'remarks', 'value': object.remarks})
            fields.append({'name': 'change', 'edit_link': reverse('panel:rating-update', args=[object.pk]),
                           'delete_link': reverse('panel:rating-delete', args=[object.pk])})
            row_list.append({'fields': fields})

        context['row_list'] = row_list

        return context


@class_view_decorator(login_required())
class PilotAircraftList (DetailView):
    model = Pilot
    template_name = 'panel/list_template.html'

    # jeśli nie superuser to ustaw na zalogowanego
    def get_object(self, queryset=None):
        obj = super(PilotAircraftList, self).get_object()
        if not self.request.user.is_staff:
            obj = self.request.user.fbouser.pilot
        return obj

    def get_context_data(self, **kwargs):
        context = super(PilotAircraftList, self).get_context_data(**kwargs)
        context['page_title'] = 'Pilot'
        context['header_text'] = 'Statki powietrzne dostępne dla %s' % self.object
        context['empty_text'] = 'Brak dostępnych statków powietrznych.'
        context['type'] = 'aircraft'
        context['submenu_template'] = 'panel/pilot_submenu.html'

        pilot_aircraft_list = [relation.aircraft for relation in PilotAircraft.objects.filter(pilot=self.object)]
        pilot_aircraft_list.sort(key=lambda x: x.type)

        # Lokalne menu
        local_menu = []
        if self.request.user.is_staff:
            local_menu.append({'text': 'Zmień uprawnienia', 'path': reverse("panel:pilot-aircraft-auth", args=[self.object.pk])})
        context['local_menu'] = local_menu

        # Nagłówki tabeli
        header_list = []
        header_list.append({'header': 'Znaki SP'})
        header_list.append({'header': 'Typ'})
        context['header_list'] = header_list

        # Zawartość tabeli
        row_list = []
        for object in pilot_aircraft_list:
            fields = []
            fields.append({'name': 'aircraft', 'value': object, 'bold': True, 'just': 'center'})
            fields.append({'name': 'type', 'value': object.type})
            row_list.append({'fields': fields})
        context['row_list'] = row_list

        return context


@user_passes_test(lambda u: u.is_staff)
def PilotAircraftAuth (request, pilot_id):

    pilot = get_object_or_404(Pilot, pk=pilot_id)
    pilot_aircraft = [relation.aircraft for relation in PilotAircraft.objects.filter(pilot=pilot)]
    aircraft_auth = [(aircraft, aircraft in pilot_aircraft) for aircraft in Aircraft.objects.order_by('type')]

    form = PilotAircraftForm(request.POST or None, aircraft_auth=aircraft_auth)

    context = {}
    context['form'] = form
    context['page_title'] = 'Pilot'
    context['header_text'] = 'Statki powietrzne dostępne dla %s' % pilot
    context['type'] = 'aircraft'
    context['submenu_template'] = 'panel/pilot_submenu.html'
    context['pilot'] = pilot

    # jeśli formularz został poprawnie wypełniony
    if form.is_valid():

        # zapisanie zmian
        for relation in PilotAircraft.objects.filter(pilot=pilot):
            relation.delete()
        for aircraft in Aircraft.objects.all():
            if form.cleaned_data.get('%s' % aircraft.pk, None):
                pilot_aircraft = PilotAircraft(pilot=pilot, aircraft=aircraft)
                pilot_aircraft.full_clean()
                pilot_aircraft.save()

        # przejście do kolejnego formularza
        return HttpResponseRedirect(reverse('panel:pilot-aircraft', args=[pilot.pk]))

    return render(request, 'panel/update_template.html', context)


@class_view_decorator(login_required())
class PilotTypesList (DetailView):
    model = Pilot
    template_name = 'panel/list_template.html'

    # jeśli nie superuser to ustaw na zalogowanego
    def get_object(self, queryset=None):
        obj = super(PilotTypesList, self).get_object()
        if not self.request.user.is_staff:
            obj = self.request.user.fbouser.pilot
        return obj

    def get_context_data(self, **kwargs):
        context = super(PilotTypesList, self).get_context_data(**kwargs)
        context['page_title'] = 'Pilot'
        context['header_text'] = 'Rodzaje lotów dostępne dla %s' % self.object
        context['empty_text'] = 'Brak dostępnych rodzajów lotów.'
        context['type'] = 'flight_type'
        context['submenu_template'] = 'panel/pilot_submenu.html'

        pilot_types_list = [relation.flight_type for relation in PilotFlightType.objects.filter(pilot=self.object)]
        pilot_types_list.sort()

        # Lokalne menu
        local_menu = []
        if self.request.user.is_staff:
            local_menu.append({'text': 'Zmień uprawnienia', 'path': reverse("panel:pilot-types-auth", args=[self.object.pk])})
        context['local_menu'] = local_menu

        # Nagłówki tabeli
        header_list = []
        header_list.append({'header': 'Kod'})
        header_list.append({'header': 'Nazwa'})
        context['header_list'] = header_list

        # Zawartość tabeli
        row_list = []
        for object in pilot_types_list:
            if FlightTypes()[object]:
                fields = []
                fields.append({'name': 'code', 'value': object, 'bold': True, 'just': 'center'})
                fields.append({'name': 'name', 'value': FlightTypes()[object]})
                row_list.append({'fields': fields})
        context['row_list'] = row_list

        return context


@user_passes_test(lambda u: u.is_staff)
def PilotTypesAuth (request, pilot_id):

    pilot = get_object_or_404(Pilot, pk=pilot_id)
    pilot_types = [relation.flight_type for relation in PilotFlightType.objects.filter(pilot=pilot)]
    types_auth = [(flight_type, flight_type in pilot_types) for flight_type in ['01','01A','02','02H','03A','03B','03C',
                                                                                '03D','03E','04','05','06','06A','07']]

    form = PilotTypesForm(request.POST or None, types_auth=types_auth)

    context = {}
    context['form'] = form
    context['page_title'] = 'Pilot'
    context['header_text'] = 'Rodzaje lotów dostępne dla %s' % pilot
    context['type'] = 'flight_type'
    context['submenu_template'] = 'panel/pilot_submenu.html'
    context['pilot'] = pilot

    # jeśli formularz został poprawnie wypełniony
    if form.is_valid():

        # zapisanie zmian
        for relation in PilotFlightType.objects.filter(pilot=pilot):
            relation.delete()
        for flight_type in ['01','01A','02','02H','03A','03B','03C','03D','03E','04','05','06','06A','07']:
            if form.cleaned_data.get(flight_type, None):
                pilot_flight_type = PilotFlightType(pilot=pilot, flight_type=flight_type)
                pilot_flight_type.full_clean()
                pilot_flight_type.save()

        # przejście do kolejnego formularza
        return HttpResponseRedirect(reverse('panel:pilot-types', args=[pilot.pk]))

    return render(request, 'panel/update_template.html', context)


@class_view_decorator(login_required())
class PilotFlights (DetailView):
    model = Pilot
    template_name = 'panel/list_template.html'

    # jeśli nie superuser to ustaw na zalogowanego
    def get_object(self, queryset=None):
        obj = super(PilotFlights, self).get_object()
        if not self.request.user.is_staff:
            obj = self.request.user.fbouser.pilot
        return obj

    def get_context_data(self, **kwargs):
        context = super(PilotFlights, self).get_context_data(**kwargs)
        context['page_title'] = 'Naloty'
        context['header_text'] = 'Naloty pilota %s (%s)' % (self.object, self.kwargs['type'].upper())
        context['empty_text'] = 'Brak zarejestrowanych lotów.'
        context['type'] = 'flights_' + self.kwargs['type']
        context['submenu_template'] = 'panel/pilot_submenu.html'
        context['fbouser'] = self.object.fbouser

        if self.kwargs['type'] == 'pic':
            flight_list = PDT.objects.filter(pic=self.object.pk).exclude(status='open').order_by('-date', '-pk')
        else:
            flight_list = PDT.objects.filter(sic=self.object.pk).exclude(status='open').order_by('-date', '-pk')

        # Lokalne menu
        local_menu = []
        open_pdt = self.request.user.fbouser.open_pdt()
        if open_pdt:
            local_menu.append({'text': 'Przejdź do otwartego PDT', 'path': reverse("panel:pdt-info", args=[open_pdt.pk])})
        else:
            local_menu.append({'text': 'Otwórz nowy PDT', 'path': reverse("panel:pdt-open")})
        context['local_menu'] = local_menu

        # Nagłówki tabeli
        header_list = []
        header_list.append({'header': 'Data'})
        header_list.append({'header': 'SP'})
        header_list.append({'header': 'PDT'})
        header_list.append({'header': 'Rodzaj\nlotu'})
        if self.kwargs['type'] == 'pic':
            header_list.append({'header': 'SIC'})
        else:
            header_list.append({'header': 'PIC'})
        header_list.append({'header': 'TTH'})
        header_list.append({'header': 'Czas'})
        header_list.append({'header': 'Ldg.'})
        header_list.append({'header': '...'})
        context['header_list'] = header_list

        # Zawartość tabeli
        row_list = []
        for object in flight_list:
            fields = []
            fields.append({'name': 'date', 'value': object.date})
            fields.append({'name': 'aircraft', 'value': object.aircraft})
            fields.append({'name': 'pdt', 'value': '{:0>6d}'.format(object.pdt_ref), 'link': reverse('panel:pdt-info', args=[object.pk]), 'just': 'center'})
            fields.append({'name': 'flight_type', 'value': object.flight_type_name()})
            if self.kwargs['type'] == 'pic':
                fields.append({'name': 'sic', 'value': object.sic})
            else:
                fields.append({'name': 'pic', 'value': object.pic})
            fields.append({'name': 'tth', 'value': '%.1f' % object.tth_sum(), 'just': 'center'})
            fields.append({'name': 'tth', 'value': '%02.d:%02.d' % (object.hours_sum()[1], object.hours_sum()[2]), 'just': 'center'})
            fields.append({'name': 'landings', 'value': object.landings_sum(), 'just': 'center'})
            fields.append({'name': 'change', 'print_link': reverse('panel:pdt-form', args=[object.pk])})
            row_list.append({'fields': fields})

        context['row_list'] = row_list

        return context


@class_view_decorator(login_required())
class PilotDuties (DetailView):
    model = Pilot
    template_name = 'panel/pilot_duties.html'

    # jeśli nie superuser to ustaw na zalogowanego
    def get_object(self, queryset=None):
        obj = super(PilotDuties, self).get_object()
        if not self.request.user.is_staff:
            obj = self.request.user.fbouser.pilot
        return obj

    def get_context_data(self, **kwargs):
        context = super(PilotDuties, self).get_context_data(**kwargs)
        context['page_title'] = 'Czas pracy'
        context['header_text'] = 'Czas pracy pilota %s' % self.object
        context['empty_text'] = 'Brak pozycji czasu pracy.'
        context['type'] = 'duties'
        context['submenu_template'] = 'panel/pilot_submenu.html'
        context['pilot'] = self.object

        duties_list = self.object.duty_set.order_by('-date', '-pk')

        first_of_year  = date.today().replace(month=1, day=1)
        first_of_month = date.today().replace(day=1)
        beg_four_months = (date.today()-relativedelta(months=4)).replace(day=1)
        end_four_months = beg_four_months+relativedelta(months=4)-relativedelta(days=1)

        d1_duties_list = duties_list.filter(date__gte=date.today()-relativedelta(days=1))
        m1_duties_list = duties_list.filter(date__gte=first_of_month)
        m4_duties_list = duties_list.filter(date__gte=beg_four_months).filter(date__lte=end_four_months)
        y1_duties_list = duties_list.filter(date__gte=first_of_year)

        # d1_fdp_time =  d1_duties_list.aggregate(Sum('fdp_time'))['fdp_time__sum'] or timedelta(seconds=0)
        # m1_fdp_time =  m1_duties_list.aggregate(Sum('fdp_time'))['fdp_time__sum'] or timedelta(seconds=0)
        # m4_fdp_time =  m4_duties_list.aggregate(Sum('fdp_time'))['fdp_time__sum'] or timedelta(seconds=0)
        # y1_fdp_time =  y1_duties_list.aggregate(Sum('fdp_time'))['fdp_time__sum'] or timedelta(seconds=0)

        d1_block_time = d1_duties_list.aggregate(Sum('block_time'))['block_time__sum'] or timedelta(seconds=0)
        m1_block_time = m1_duties_list.aggregate(Sum('block_time'))['block_time__sum'] or timedelta(seconds=0)
        m4_block_time = m4_duties_list.aggregate(Sum('block_time'))['block_time__sum'] or timedelta(seconds=0)
        y1_block_time = y1_duties_list.aggregate(Sum('block_time'))['block_time__sum'] or timedelta(seconds=0)

        d1_landings = d1_duties_list.aggregate(Sum('landings'))['landings__sum'] or timedelta(seconds=0)
        m1_landings = m1_duties_list.aggregate(Sum('landings'))['landings__sum'] or timedelta(seconds=0)
        m4_landings = m4_duties_list.aggregate(Sum('landings'))['landings__sum'] or timedelta(seconds=0)
        y1_landings = y1_duties_list.aggregate(Sum('landings'))['landings__sum'] or timedelta(seconds=0)

        d1_flights_list = d1_duties_list.filter(fdp_time__isnull=False)
        m1_flights_list = m1_duties_list.filter(fdp_time__isnull=False)
        m4_flights_list = m4_duties_list.filter(fdp_time__isnull=False)
        y1_flights_list = y1_duties_list.filter(fdp_time__isnull=False)

        d1_others_list = d1_duties_list.filter(fdp_time__isnull=True)
        m1_others_list = m1_duties_list.filter(fdp_time__isnull=True)
        m4_others_list = m4_duties_list.filter(fdp_time__isnull=True)
        y1_others_list = y1_duties_list.filter(fdp_time__isnull=True)

        d1_duty_time = d1_flights_list.aggregate(Sum('duty_time'))['duty_time__sum'] or timedelta(seconds=0)
        m1_duty_time = m1_flights_list.aggregate(Sum('duty_time'))['duty_time__sum'] or timedelta(seconds=0)
        m4_duty_time = m4_flights_list.aggregate(Sum('duty_time'))['duty_time__sum'] or timedelta(seconds=0)
        y1_duty_time = y1_flights_list.aggregate(Sum('duty_time'))['duty_time__sum'] or timedelta(seconds=0)

        d1_other_time = d1_others_list.aggregate(Sum('duty_time'))['duty_time__sum'] or timedelta(seconds=0)
        m1_other_time = m1_others_list.aggregate(Sum('duty_time'))['duty_time__sum'] or timedelta(seconds=0)
        m4_other_time = m4_others_list.aggregate(Sum('duty_time'))['duty_time__sum'] or timedelta(seconds=0)
        y1_other_time = y1_others_list.aggregate(Sum('duty_time'))['duty_time__sum'] or timedelta(seconds=0)

        context['d1_other_time'] = duration_string(d1_other_time)
        context['m1_other_time'] = duration_string(m1_other_time)
        context['m4_other_time'] = duration_string(m4_other_time)
        context['y1_other_time'] = duration_string(y1_other_time)

        context['d1_duty_time']  = duration_string(d1_duty_time)
        context['m1_duty_time']  = duration_string(m1_duty_time)
        context['m4_duty_time']  = duration_string(m4_duty_time)
        context['y1_duty_time']  = duration_string(y1_duty_time)

        context['d1_block_time'] = duration_string(d1_block_time)
        context['m1_block_time'] = duration_string(m1_block_time)
        context['m4_block_time'] = duration_string(m4_block_time)
        context['y1_block_time'] = duration_string(y1_block_time)

        context['d1_landings'] = d1_landings
        context['m1_landings'] = m1_landings
        context['m4_landings'] = m4_landings
        context['y1_landings'] = y1_landings

        context['d1_sum_time'] = duration_string(d1_other_time + d1_duty_time)
        context['m1_sum_time'] = duration_string(m1_other_time + m1_duty_time)
        context['m4_sum_time'] = duration_string(m4_other_time + m4_duty_time)
        context['y1_sum_time'] = duration_string(y1_other_time + y1_duty_time)

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Dodaj nową pozycję', 'path': reverse("panel:duty-create", args=[self.object.pk])})
        local_menu.append({'text': 'Pokaż grafik', 'path': reverse("res:duty-calendar")})
        local_menu.append({'text': 'Raport wykonanych lotów', 'path': reverse("panel:duty-report", args=[self.object.pk])})
        local_menu.append({'text': 'Raport planowanych lotów', 'path': reverse("panel:plan-report", args=[self.object.pk])})
        context['local_menu'] = local_menu

        # Nagłówki tabeli
        header_list = []
        header_list.append({'header': 'Data'})
        header_list.append({'header': 'Firma'})
        header_list.append({'header': 'Stanowisko'})
        header_list.append({'header': 'Rodzaj\nobowiązków'})
        header_list.append({'header': 'Czas'})
        header_list.append({'header': 'Od\ngodziny'})
        header_list.append({'header': 'Do\ngodziny'})
        # header_list.append({'header': 'Czas\nczynności'})
        header_list.append({'header': 'Czas\nlotu'})
        header_list.append({'header': 'Uwagi'})
        header_list.append({'header': '...'})
        context['header_list'] = header_list

        # Zawartość tabeli
        row_list = []
        for object in duties_list:
            fields = []
            fields.append({'name': 'date', 'value': object.date, 'link': reverse('panel:duty-details', args=[object.pk]), 'just': 'center'})
            fields.append({'name': 'company', 'value': object.comp_str(), 'just': 'center'})
            fields.append({'name': 'role', 'value': object.role})
            fields.append({'name': 'type', 'value': object.duty_type})
            fields.append({'name': 'duty_time', 'value': duration_string(object.duty_time) if object.duty_time else None, 'just': 'center'})
            fields.append({'name': 'duty_from', 'value': object.duty_time_from.strftime('%H:%M') if object.duty_time_from else None, 'just': 'center'})
            fields.append({'name': 'duty_to', 'value': object.duty_time_to.strftime('%H:%M') if object.duty_time_to else None, 'just': 'center'})
            # fields.append({'name': 'fdp_time', 'value': duration_string(object.fdp_time) if object.fdp_time else None, 'just': 'center'})
            fields.append({'name': 'block_time', 'value': duration_string(object.block_time) if object.block_time else None, 'just': 'center'})
            fields.append({'name': 'remarks', 'value': object.remarks})
            fields.append({'name': 'change', 'edit_link': reverse('panel:duty-update', args=[object.pk]),
                                   'delete_link': reverse('panel:duty-delete', args=[object.pk]),
                                   'view_link': reverse('res:duty-cal-day', args=[object.date.year,object.date.month,object.date.day])})
            row_list.append({'fields': fields})

        context['row_list'] = row_list

        return context


@class_view_decorator(login_required())
class PilotUpdate (UpdateView):
    model = Pilot
    template_name = 'panel/update_template.html'

    # jeśli nie superuser to ustaw na zalogowanego
    def get_object(self, queryset=None):
        obj = super(PilotUpdate, self).get_object()
        if not self.request.user.is_staff:
            obj = self.request.user.fbouser.pilot
        return obj

    def get_context_data(self, **kwargs):
        context = super(PilotUpdate, self).get_context_data(**kwargs)
        context['page_title'] = 'Aktualizacja pilota'
        context['header_text'] = 'Aktualizacja danych pilota %s' % self.object
        context['type'] = 'info'
        context['submenu_template'] = 'panel/pilot_submenu.html'
        context['pilot'] = self.object
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(Pilot, exclude=['fbouser'], widgets = {'licence_date': AdminDateWidget(),
                                                                             'medical_date': AdminDateWidget(),
                                                                             'remarks':Textarea(attrs={'rows':2, 'cols':100})})
        if not self.request.user.is_staff:
            form_class.base_fields['employee'].widget.attrs = {'disabled': 'true'}

        return form_class

    def get_success_url(self):
        return reverse('panel:pilot-info', args=(self.object.pk,))


@class_view_decorator(login_required())
class PDTActive_feed (BaseDatatableView):
    model = PDT
    max_display_length = 100
    order_columns = ['aircraft', 'pdt_ref', 'date', 'flight_type', 'pic', 'sic', 'remarks', 'status']

    def get_initial_queryset(self):
        return PDT.objects.order_by('date', 'tth_start', 'pdt_ref', 'open_time').reverse()

    def filter_queryset(self, qs):
        search = self.request.GET.get('search[value]', None)
        if search:
            qs = qs.filter(Q(aircraft__registration__icontains=search)|
                           Q(pdt_ref__icontains=search)|
                           Q(pic__fbouser__first_name__icontains=search)|
                           Q(pic__fbouser__second_name__icontains=search) |
                           Q(sic__fbouser__first_name__icontains=search)|
                           Q(sic__fbouser__second_name__icontains=search) |
                           Q(remarks__icontains=search))
        return qs

    def prepare_results(self, qs):
        json_data = []
        for item in qs:
            json_data.append({
                'aircraft':item.aircraft.registration,
                'pdt_ref':'{:0>6d}'.format(item.pdt_ref),
                'info_link': reverse('panel:pdt-info', args=[item.pk]),
                'date':item.date,
                'flight_type':item.flight_type_name(),
                'pic':item.pic.__str__(),
                'sic':(item.sic.__str__() if item.sic else ''),
                'remarks':item.remarks[:80],
                'status':item.status_name(),
                'color': item.status_color(),
                'tool_link':(reverse('camo:pdt-info', args=[item.pk]) if item.status != 'checked' and self.request.user.has_perm('camo.camo_admin') else None),
                'print_link': (reverse('panel:pdt-form', args=[item.pk]))
            })
        return json_data


@class_view_decorator(login_required())
class PDTActive (ListView):
    model = PDT
    template_name = 'panel/pdt_list_template.html'

    # Posortuj listę PDT po datach
    def get_queryset(self):
        return PDT.objects.order_by('date', 'tth_start', 'pdt_ref', 'open_time').reverse()[:100]

    def get_context_data(self, **kwargs):
        context = super(PDTActive, self).get_context_data(**kwargs)
        context['page_title'] = 'Lista PDT'
        context['header_text'] = 'Lista zarejestrowanych PDT'
        context['empty_text'] = 'Brak zarejestrowanych PDT.'

        # Lokalne menu
        local_menu = []
        context['local_menu'] = local_menu

        # Nagłówki tabeli
        header_list = []
        header_list.append({'header': 'Statek\npowietrzny'})
        header_list.append({'header': 'PDT ref'})
        header_list.append({'header': 'Data'})
        header_list.append({'header': 'Rodzaj lotu'})
        header_list.append({'header': 'PIC'})
        header_list.append({'header': 'SIC'})
        header_list.append({'header': 'Uwagi'})
        header_list.append({'header': 'Status'})
        header_list.append({'header': '...'})
        context['header_list'] = header_list

        return context


@class_view_decorator(login_required())
class PDTInfo (DetailView):
    model = PDT
    template_name = 'panel/pdt_info.html'

    def get_context_data(self, **kwargs):
        context = super(PDTInfo, self).get_context_data(**kwargs)
        context['page_title'] = 'PDT'
        context['type'] = 'info'
        context['submenu_template'] = 'panel/pdt_submenu.html'
        context['pdt'] = self.object

        context['landings_total'] = self.object.prev_landings + self.object.landings_sum()
        context['tth_total'] = self.object.tth_start + self.object.tth_sum()
        context['operations'] = len(self.object.operation_set.all())

        add_empty = 4 - self.object.operation_set.count()
        if self.object.fuel_after:
            add_empty -= 1
        if add_empty < 0:
            add_empty = 0
        context['add_empty'] = [i + (5 - add_empty) for i in range(add_empty)]

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Drukuj', 'path': reverse('panel:pdt-form', args=[self.object.pk])})
        if (self.object.status != 'checked' and self.object.open_user == self.request.user.fbouser) or self.request.user.is_staff:
            local_menu.append({'text': 'Aktualizuj', 'path': reverse('panel:pdt-update', args=[self.object.pk])})
            local_menu.append({'text': 'Usuń', 'path': reverse('panel:pdt-delete', args=[self.object.pk])})
        context['local_menu'] = local_menu

        return context


@class_view_decorator(login_required())
class PDTFuel (DetailView):
    model = PDT
    template_name = 'panel/pdt_fuel.html'

    def get_context_data(self, **kwargs):
        context = super(PDTFuel, self).get_context_data(**kwargs)
        context['page_title'] = 'PDT'
        context['header_text'] = 'Rozliczenie paliwa na PDT'
        context['empty_text'] = 'Brak zarejestrowanych operacji.'
        context['type'] = 'fuel'
        context['submenu_template'] = 'panel/pdt_submenu.html'
        context['pdt'] = self.object

        self.request.session.pop('pdt_id', None)
        oper_list = self.object.operation_set.order_by('operation_no')

        # Lokalne menu
        local_menu = []
        context['local_menu'] = local_menu

        # Nagłówki tabeli
        header_list = []
        header_list.append({'header': 'Etap lotu'})
        header_list.append({'header': 'Uzupełniono'})
        header_list.append({'header': 'Zbiornik'})
        header_list.append({'header': 'Stan do lotu'})
        header_list.append({'header': 'Zużyto'})
        context['header_list'] = header_list

        # Zawartość tabeli
        row_list = []
        refill, used = 0, 0
        for object in oper_list:
            fields = []
            fields.append({'name': 'phase', 'value': '%s - %s' % (object.loc_start, object.loc_end or '...'), 'just': 'right'})
            fields.append({'name': 'refill', 'value': '%.1f L' % object.fuel_refill if object.fuel_refill != 0 else '---', 'just': 'center'})
            refill += object.fuel_refill or 0
            fueltank = None
            if object.fuel_refill != 0:
                fueltank_string = object.fuel_source
                if fueltank_string[:4] == 'SALT':
                    try:
                        fueltank_id = int(fueltank_string[5:])
                    except:
                        fueltank_id = None
                    if fueltank_id:
                        fueltank = FuelTank.objects.get(pk=fueltank_id).__str__()
                elif fueltank_string[:5] == 'other':
                    fueltank = 'Poza SALT'
            fields.append({'name': 'fueltank', 'value': fueltank})
            fields.append({'name': 'volume', 'value': '%.1f L' % object.fuel_available, 'just': 'center'})
            fields.append({'name': 'used', 'value': '%.1f L' % object.fuel_used if object.fuel_used else '---', 'just': 'center'})
            used += object.fuel_used or 0
            row_list.append({'fields': fields})

        if (self.object.fuel_after or 0) > 0:
            fields = []
            fields.append({'name': 'phase', 'value': 'Po zakończeniu', 'just': 'right'})
            fields.append({'name': 'refill', 'value': '%.1f L' % self.object.fuel_after, 'just': 'center'})
            refill += self.object.fuel_after or 0
            fueltank = None
            if self.object.fuel_after != 0:
                fueltank_string = self.object.fuel_after_source
                if fueltank_string[:4] == 'SALT':
                    try:
                        fueltank_id = int(fueltank_string[5:])
                    except:
                        fueltank_id = None
                    if fueltank_id:
                        fueltank = FuelTank.objects.get(pk=fueltank_id).__str__()
                elif fueltank_string[:5] == 'other':
                    fueltank = 'Poza SALT'
            fields.append({'name': 'fueltank', 'value': fueltank})
            fields.append({'name': 'volume', 'value': '%.1f L' % self.object.aircraft.fuel_volume, 'just': 'center'})
            fields.append({'name': 'used', 'value': '---', 'just': 'center'})
            row_list.append({'fields': fields})

        fields = []
        fields.append({'name': 'phase', 'value': 'Razem:', 'bold': True, 'just': 'right', 'color': '#CCCCCC'})
        fields.append({'name': 'refill', 'value': '%.1f L' % refill, 'bold': True, 'just': 'center', 'color': '#CCCCCC'})
        fields.append({'name': 'spce', 'value': '', 'color': '#CCCCCC'})
        fields.append({'name': 'volume', 'value': '%.1f L' % self.object.aircraft.fuel_volume, 'bold': True, 'just': 'center', 'color': '#CCCCCC'})
        fields.append({'name': 'used', 'value': '%.1f L' % used, 'bold': True, 'just': 'center', 'color': '#CCCCCC'})
        row_list.append({'fields': fields})
        context['row_list'] = row_list

        remote_fueling_list = self.object.remotefueling_set.order_by('-pk')
        if remote_fueling_list:
            context['sub_header_text'] = 'Tankowania poza SALT'
            context['sub_header_list'] = ['Miejsce', 'Ilość paliwa', 'Faktura', 'Akcyza', 'Cena paliwa', 'Uwagi', '...']

            row_list = []
            for fueling in remote_fueling_list:
                fields = []
                fields.append({'name': 'location', 'value': fueling.location, 'link': reverse('panel:remote-fueling-details', args=[fueling.pk])})
                fields.append({'name': 'fuel_volume', 'value': '%.1f L' % fueling.fuel_volume, 'just': 'center'})
                fields.append({'name': 'document', 'value': fueling.document or '[brak]'})
                fields.append({'name': 'excise', 'value': fueling.excise})
                fields.append({'name': 'fuel_price', 'value': '%.2d PLN' % fueling.total_price if fueling.total_price else '[brak]', 'just': 'center'})
                fields.append({'name': 'remarks', 'value': fueling.remarks})
                fields.append({'name': 'change', 'edit_link': reverse('panel:remote-fueling-update', args=[fueling.pk]),
                               'delete_link': reverse('panel:remote-fueling-delete', args=[fueling.pk])})
                row_list.append({'fields': fields})
            context['sub_row_list'] = row_list

        return context


@class_view_decorator(login_required())
class PDTAOC (DetailView):
    model = PDT
    template_name = 'panel/details_template.html'

    def get_context_data(self, **kwargs):
        context = super(PDTAOC, self).get_context_data(**kwargs)
        context['page_title'] = 'PDT'
        context['header_text'] = 'Informacje o locie AOC'
        context['type'] = 'aoc'
        context['submenu_template'] = 'panel/pdt_submenu.html'
        context['pdt'] = self.object

        # Lokalne menu
        local_menu = []
        context['local_menu'] = local_menu

        field_list = []
        field_list.append({'header': 'Statek powietrzny', 'value': self.object.aircraft, 'bold': True, 'just': 'center'})
        field_list.append({'header': 'Numer PDT', 'value': '{:0>6d}'.format(self.object.pdt_ref), 'bold': True, 'just': 'center'})
        field_list.append({'header': 'Data PDT', 'value': self.object.date, 'just': 'center'})
        if self.object.contractor:
            field_list.append({'header': 'Kontrahent', 'value': self.object.contractor, 'link': reverse('fin:contr-info', args=[self.object.contractor.pk])})
            aoc_operation = self.object.balanceoperation_set.filter(type='Usługa AOC').last()
            if aoc_operation:
                field_list.append({'header': 'Opłata za usługę', 'value': '%.2f PLN' % -aoc_operation.aoc_amount})
        field_list.append({'header': 'Opis usługi', 'value': self.object.service_remarks})

        context['field_list'] = field_list
        return context


@class_view_decorator(login_required())
class PDTView (DetailView):
    model = PDT
    template_name = 'panel/details_template.html'

    def get_context_data(self, **kwargs):
        context = super(PDTView, self).get_context_data(**kwargs)
        context['page_title'] = 'PDT'
        context['header_text'] = 'Informacje o locie widokowym'
        context['type'] = 'view'
        context['submenu_template'] = 'panel/pdt_submenu.html'
        context['pdt'] = self.object

        # Lokalne menu
        local_menu = []
        context['local_menu'] = local_menu

        field_list = []
        field_list.append({'header': 'Statek powietrzny', 'value': self.object.aircraft, 'bold': True, 'just': 'center'})
        field_list.append({'header': 'Numer PDT', 'value': '{:0>6d}'.format(self.object.pdt_ref), 'bold': True, 'just': 'center'})
        field_list.append({'header': 'Data PDT', 'value': self.object.date, 'just': 'center'})
        if self.object.voucher:
            field_list.append({'header': 'Voucher SALT', 'value': '%s z dnia %s' % (self.object.voucher, self.object.date),
                               'link': reverse('fin:voucher-details', args=[self.object.voucher.pk])})
            field_list.append({'header': 'Opis vouchera', 'value': self.object.voucher.description})
        elif self.object.ext_voucher:
            field_list.append({'header': 'Voucher zewn.', 'value': self.object.ext_voucher})

        context['field_list'] = field_list
        return context


@class_view_decorator(login_required())
class PDTIntro (DetailView):
    model = PDT
    template_name = 'panel/details_template.html'

    def get_context_data(self, **kwargs):
        context = super(PDTIntro, self).get_context_data(**kwargs)
        context['page_title'] = 'PDT'
        context['header_text'] = 'Informacje o locie zapoznawczym'
        context['type'] = 'view'
        context['submenu_template'] = 'panel/pdt_submenu.html'
        context['pdt'] = self.object

        # Lokalne menu
        local_menu = []
        context['local_menu'] = local_menu

        field_list = []
        field_list.append({'header': 'Statek powietrzny', 'value': self.object.aircraft, 'bold': True, 'just': 'center'})
        field_list.append({'header': 'Numer PDT', 'value': '{:0>6d}'.format(self.object.pdt_ref), 'bold': True, 'just': 'center'})
        field_list.append({'header': 'Data PDT', 'value': self.object.date, 'just': 'center'})
        if self.object.voucher:
            field_list.append({'header': 'Voucher SALT', 'value': '%s z dnia %s' % (self.object.voucher, self.object.date),
                               'link': reverse('fin:voucher-details', args=[self.object.voucher.pk])})
            field_list.append({'header': 'Opis vouchera', 'value': self.object.voucher.description})
        elif self.object.ext_voucher:
            field_list.append({'header': 'Voucher zewn.', 'value': self.object.ext_voucher})
        field_list.append({'header': 'Uwagi', 'value': self.object.remarks})

        context['field_list'] = field_list
        return context


@class_view_decorator(login_required())
class PDTSPO (DetailView):
    model = PDT
    template_name = 'panel/details_template.html'

    def get_context_data(self, **kwargs):
        context = super(PDTSPO, self).get_context_data(**kwargs)
        context['page_title'] = 'PDT'
        context['header_text'] = 'Informacje o locie SPO'
        context['type'] = 'spo'
        context['submenu_template'] = 'panel/pdt_submenu.html'
        context['pdt'] = self.object

        # Lokalne menu
        local_menu = []
        context['local_menu'] = local_menu

        field_list = []
        field_list.append({'header': 'Statek powietrzny', 'value': self.object.aircraft, 'bold': True, 'just': 'center'})
        field_list.append({'header': 'Numer PDT', 'value': '{:0>6d}'.format(self.object.pdt_ref), 'bold': True, 'just': 'center'})
        field_list.append({'header': 'Data PDT', 'value': self.object.date, 'just': 'center'})
        if self.object.contractor:
            field_list.append({'header': 'Kontrahent', 'value': self.object.contractor, 'link': reverse('fin:contr-info', args=[self.object.contractor.pk])})
            spo_operation = self.object.balanceoperation_set.filter(type='Usługa SPO').last()
            if spo_operation:
                field_list.append({'header': 'Opłata za usługę', 'value': '%.2f PLN' % -spo_operation.spo_amount})
        field_list.append({'header': 'Opis usługi', 'value': self.object.service_remarks})

        context['field_list'] = field_list
        return context


@class_view_decorator(login_required())
class PDTTraining (DetailView):
    model = PDT
    template_name = 'panel/details_template.html'

    def get_context_data(self, **kwargs):
        context = super(PDTTraining, self).get_context_data(**kwargs)
        context['page_title'] = 'PDT'
        context['header_text'] = 'Informacje o locie szkolnym'
        context['type'] = 'ato'
        context['submenu_template'] = 'panel/pdt_submenu.html'
        context['pdt'] = self.object

        # Lokalne menu
        local_menu = []
        context['local_menu'] = local_menu

        field_list = []
        field_list.append({'header': 'Szkolenie', 'value': self.object.training.training.code,
                           'link': reverse('ato:training-inst-details', args=[self.object.training.pk])})
        field_list.append({'header': 'Student', 'value': self.object.training.student,
                           'link': reverse('ato:student-info', args=[self.object.training.student.pk])})
        field_list.append({'header': 'Instruktor na locie', 'value': self.object.instructor})
        field_list.append({'header': 'Instruktor prowadzący', 'value': self.object.training.instructor})
        if self.object.contractor:
            field_list.append({'header': 'Kontrahent', 'value': self.object.contractor, 'link': reverse('fin:contr-info', args=[self.object.contractor.pk])})
            ato_operation = self.object.balanceoperation_set.filter(type='Szkolenie').last()
            if ato_operation:
                field_list.append({'header': 'Opłata za szkolenie', 'value': '%.2f PLN' % -ato_operation.ato_amount})

        context['field_list'] = field_list
        return context


@class_view_decorator(login_required())
class PDTExam (DetailView):
    model = PDT
    template_name = 'panel/details_template.html'

    def get_context_data(self, **kwargs):
        context = super(PDTExam, self).get_context_data(**kwargs)
        context['page_title'] = 'PDT'
        context['header_text'] = 'Informacje o egzaminie'
        context['type'] = 'exam'
        context['submenu_template'] = 'panel/pdt_submenu.html'
        context['pdt'] = self.object

        # Lokalne menu
        local_menu = []
        context['local_menu'] = local_menu

        field_list = []
        field_list.append({'header': 'Statek powietrzny', 'value': self.object.aircraft, 'bold': True, 'just': 'center'})
        field_list.append({'header': 'Numer PDT', 'value': '{:0>6d}'.format(self.object.pdt_ref), 'bold': True, 'just': 'center'})
        field_list.append({'header': 'Data PDT', 'value': self.object.date, 'just': 'center'})
        field_list.append({'header': 'Szkolenie', 'value': self.object.training.training.code, 'link': reverse('ato:training-inst-details', args=[self.object.training.pk])})
        field_list.append({'header': 'Student', 'value': self.object.training.student, 'link': reverse('ato:student-info', args=[self.object.training.student.pk])})
        field_list.append({'header': 'Egzaminator', 'value': self.object.pic})
        field_list.append({'header': 'Instruktor prowadzący', 'value': self.object.training.instructor})
        if self.object.contractor:
            field_list.append({'header': 'Kontrahent', 'value': self.object.contractor, 'link': reverse('fin:contr-info', args=[self.object.contractor.pk])})
            ato_operation = self.object.balanceoperation_set.filter(type='Egzamin').last()
            if ato_operation:
                field_list.append({'header': 'Opłata za egzamin', 'value': '%.2f PLN' % -ato_operation.ato_amount})

        context['field_list'] = field_list
        return context


@class_view_decorator(login_required())
class PDTRent (DetailView):
    model = PDT
    template_name = 'panel/details_template.html'

    def get_context_data(self, **kwargs):
        context = super(PDTRent, self).get_context_data(**kwargs)
        context['page_title'] = 'PDT'
        context['header_text'] = 'Informacje o wynajmie'
        context['type'] = 'rent'
        context['submenu_template'] = 'panel/pdt_submenu.html'
        context['pdt'] = self.object

        # Lokalne menu
        local_menu = []
        context['local_menu'] = local_menu

        field_list = []
        field_list.append({'header': 'Statek powietrzny', 'value': self.object.aircraft, 'bold': True})
        field_list.append({'header': 'Numer PDT', 'value': '{:0>6d}'.format(self.object.pdt_ref), 'bold': True})
        field_list.append({'header': 'Data PDT', 'value': self.object.date})
        field_list.append({'header': 'Czas lotu TTH', 'value': self.object.tth_sum})
        if self.object.contractor:
            field_list.append({'header': 'Kontrahent', 'value': self.object.contractor, 'link': reverse('fin:contr-info', args=[self.object.contractor.pk])})
            rent_operation = self.object.balanceoperation_set.filter(type='Wynajem').last()
            if rent_operation:
                field_list.append({'header': 'Opłata za wynajem', 'value': '%.2f PLN' % -rent_operation.rent_amount})
            instr_operation = self.object.balanceoperation_set.filter(type='Instruktor').last()
            if instr_operation:
                field_list.append({'header': 'Opłata za instruktora', 'value': '%.2f PLN' % -instr_operation.rent_amount})

        context['field_list'] = field_list
        return context


@class_view_decorator(login_required())
class PDTDuties (DetailView):
    model = PDT
    template_name = 'panel/list_template.html'


@class_view_decorator(login_required())
class PDTForm (DetailView):
    model = PDT
    template_name = 'panel/pdt.html'

    def get_context_data(self, **kwargs):
        context = super(PDTForm, self).get_context_data(**kwargs)

        context['landings_total'] = self.object.prev_landings + self.object.landings_sum()
        context['tth_total'] = self.object.tth_start + self.object.tth_sum()
        context['operations'] = len(self.object.operation_set.all())

        add_empty = 4 - self.object.operation_set.count()
        if self.object.fuel_after:
            add_empty -= 1
        if add_empty < 0:
            add_empty = 0
        context['add_empty'] = [i+(5-add_empty) for i in range(add_empty)]

        return context


def fueling_save(pdt, fueltank_string, location, fuel_volume, remarks):
    # aktualizacja stanu zbiorników
    if fueltank_string[:4] == 'SALT':
        try:
            fueltank_id = int(fueltank_string[5:])
        except:
            fueltank_id = None
        if fueltank_id:
            fueltank = FuelTank.objects.get(pk=fueltank_id)
            fueling = PDTFueling(pdt=pdt, fueltank=fueltank, fuel_volume=fuel_volume)
            fueling.full_clean()
            fueling.save()
    elif fueltank_string[:5] == 'other':
        fueling = RemoteFueling(pdt=pdt, location=location, fuel_volume=fuel_volume)
        fueling.full_clean()
        fueling.save()

        # aktualizacja konta kontrahenta
        if pdt.contractor:
            balance_operation = BalanceOperation(contractor=pdt.contractor, date=pdt.date, type='Zwrot za paliwo',
                                                 pdt=pdt, aoc_amount=0, spo_amount=0, ato_amount=0, rent_amount=0,
                                                 remarks=remarks)
            balance_operation.full_clean()
            balance_operation.save()

    return


def pdt_close (pdt, user, cleaned_data):

    # zapis informacji o tankowaniu
    last_operation = pdt.operation_set.order_by('operation_no').last()
    loc_end = last_operation.loc_end if last_operation else '...'

    if cleaned_data.get('fuel_after') and cleaned_data.get('fuel_after') != 0:
        fueling_save(pdt=pdt, fueltank_string=cleaned_data['fuel_after_source'],
                     location=loc_end, fuel_volume=cleaned_data['fuel_after'],
                     remarks='Tankowanie po locie w %s' % loc_end)

    # Aktualizacja stanu SP
    aircraft = pdt.aircraft
    aircraft.refresh_from_db()

    # Zapisanie informacji o usterce
    if cleaned_data['failure_desc']:

        # Informacja o usterce w SMS
        sms_falure = SMSFailure(aircraft=aircraft, person=user.__str__(),
                                failure_date=pdt.date,
                                description=cleaned_data['failure_desc'], pdt_ref=pdt)
        sms_falure.full_clean()
        sms_falure.save()

        # Aktualizacja statusu SP
        aircraft.status = 'damaged'
        aircraft.full_clean()
        aircraft.save()

    # aktualizacja stanu vouchera
    if pdt.flight_type in ('01A', '03D') and pdt.voucher:
        voucher = pdt.voucher
        voucher.done_date = pdt.date
        voucher.save()

    # aktualizacja konta kontrahenta
    if pdt.contractor:

        rent_price = aircraft.rent_price
        special_price = pdt.contractor.specialprice_set.filter(ac_type=aircraft.type).first()
        if special_price:
            rent_price = min(rent_price, special_price.hour_price)

        # TODO: dodać konfigurację ceny instruktora!
        instructor_price = Decimal(100)
        balance_operation = None

        if pdt.flight_type == '01':
            aoc_amount = -pdt.tth_sum() * rent_price
            if aoc_amount:
                aoc_amount = aoc_amount.quantize(Decimal(10) ** -2)
            else:
                aoc_amount = Decimal(0)
            balance_operation = BalanceOperation(contractor=pdt.contractor, date=pdt.date,
                                                 type='Usługa AOC',
                                                 pdt=pdt, aoc_amount=aoc_amount)
        elif pdt.flight_type in ('02', '02H'):
            spo_amount = -pdt.tth_sum() * rent_price
            if spo_amount:
                spo_amount = spo_amount.quantize(Decimal(10) ** -2)
            else:
                spo_amount = Decimal(0)
            balance_operation = BalanceOperation(contractor=pdt.contractor, date=pdt.date,
                                                 type='Usługa SPO',
                                                 pdt=pdt, spo_amount=spo_amount)

        elif pdt.flight_type in ('03A', '03B', '03C'):
            ato_amount = -pdt.tth_sum() * rent_price
            if ato_amount:
                ato_amount = ato_amount.quantize(Decimal(10) ** -2)
            else:
                ato_amount = Decimal(0)
            balance_operation = BalanceOperation(contractor=pdt.contractor, date=pdt.date,
                                                 type='Szkolenie',
                                                 pdt=pdt, ato_amount=ato_amount)

        elif pdt.flight_type == '03E':
            ato_amount = -pdt.tth_sum() * rent_price
            if ato_amount:
                ato_amount = ato_amount.quantize(Decimal(10) ** -2)
            else:
                ato_amount = Decimal(0)
            balance_operation = BalanceOperation(contractor=pdt.contractor, date=pdt.date,
                                                 type='Egzamin',
                                                 pdt=pdt, ato_amount=ato_amount)

        elif pdt.flight_type == '04':

            # wyliczenie łącznej ceny wynajmu i aktualizacja pakietów
            packages = [soldpackage for soldpackage in
                        pdt.contractor.soldpackage_set.filter(ac_type=aircraft.type).order_by(
                            'date')]
            rent_amount = 0
            tth_left = pdt.tth_sum()
            for package in packages:
                if tth_left > 0:
                    if tth_left <= package.left_hours:
                        package.left_hours -= tth_left
                        tth_left = 0
                    else:
                        tth_left -= package.left_hours
                        package.left_hours = 0
                    package.save()
            if tth_left > 0:
                rent_amount += -tth_left * rent_price

            if rent_amount:
                rent_amount = rent_amount.quantize(Decimal(10) ** -2)

                balance_operation = BalanceOperation(contractor=pdt.contractor, date=pdt.date,
                                                     type='Wynajem',
                                                     pdt=pdt, rent_amount=rent_amount)

        if balance_operation:
            balance_operation.full_clean()
            balance_operation.save()

        if pdt.flight_type == '04' and pdt.instructor:
            rent_amount = -pdt.tth_sum() * instructor_price
            if rent_amount:
                rent_amount.quantize(Decimal(10) ** -2)
            else:
                rent_amount = Decimal(0)
            balance_operation = BalanceOperation(contractor=pdt.contractor, date=pdt.date,
                                                 type='Instruktor',
                                                 pdt=pdt, rent_amount=rent_amount)
            balance_operation.full_clean()
            balance_operation.save()

    # Ustawienie statusu zamkniętego
    pdt.status = 'closed'
    pdt.close_user = user
    pdt.close_time = timezone.now()
    pdt.save()


@login_required()
def PDTOpen(request):

    # return render(request, 'panel/pdt_edit_blocked.html', {})

    if 'res' in request.GET:
        reservation_id = request.GET.get('res')
        try:
            reservation = Reservation.objects.get(pk=reservation_id)
        except:
            reservation = None
    else:
        reservation = None

    # Weryfikacja czy wywołano z modułu CAMO
    camo = ('camo' in request.GET)

    aircraft_id = flight_type = pic = sic = pdt_ref = ms = open_pdt = None
    contractor = request.user.fbouser.contractor

    # Ustawienie wartości inicjalnych na podstawie rezerwacji
    if reservation:
        pdt_date = reservation.start_time.date()
        flight_type = reservation.planned_type
        pic = reservation.owner
        if flight_type == '04':
            contractor = pic.fbouser.contractor
        sic = reservation.participant
        aircraft_id = reservation.aircraft_id
        aircraft = reservation.aircraft
        if aircraft:
            open_pdt = aircraft.pdt_set.filter(status='open').last()
            pdt_ref = aircraft.last_pdt_ref + 1
            ms_report = aircraft.ms_report_set.order_by('done_date').last()
            if ms_report:
                ms = "%s / %s" % (ms_report.next_hours, ms_report.next_date)
    else:
        pdt_date = date.today()

    context = {}

    if not open_pdt:

        # Inicjalizacja formularza
        form = PDTEditForm(request.POST or None, fbouser=request.user.fbouser, camo=camo,
                           initial={'date': pdt_date, 'flight_type': flight_type, 'aircraft': aircraft_id,
                                    'fuel_after': 0, 'pdt_ref': pdt_ref, 'pic': pic, 'sic': sic, 'ms': ms,
                                    'contractor': contractor})

        # wspólne zmienne kontekstowe
        context['form'] = form
        context['camo'] = camo
        context['new_pdt'] = True
        context['open_oper'] = None

        # jeśli formularz został poprawnie wypełniony
        if form.is_valid():

            aircraft = form.cleaned_data['aircraft']
            ms_report = aircraft.ms_report_set.order_by('done_date').last()

            # ustalenie licznika początkowego z ostatniego PDT
            last_pdt = aircraft.pdt_set.filter(pdt_ref=aircraft.last_pdt_ref).last()
            if last_pdt:
                tth_start = last_pdt.tth_end()
            else:
                tth_start = aircraft.tth

            # ustalenie wartości z przeniesienia
            prev_tth = aircraft.hours_count
            prev_landings = aircraft.landings_count
            prev_cycles = aircraft.cycles_count

            pdt = PDT(pdt_ref=form.cleaned_data['pdt_ref'], aircraft=aircraft,
                      flight_type=form.cleaned_data['flight_type'], date=form.cleaned_data['date'],
                      pic=form.cleaned_data['pic'], sic=form.cleaned_data['sic'],
                      instructor=form.cleaned_data['instructor'],
                      voucher=form.cleaned_data['voucher'], ext_voucher=form.cleaned_data['ext_voucher'],
                      training=form.cleaned_data['training'], contractor=form.cleaned_data['contractor'],
                      fuel_after=form.cleaned_data['fuel_after'],
                      fuel_after_source=form.cleaned_data['fuel_after_source'],
                      service_remarks=form.cleaned_data['service_remarks'], remarks=form.cleaned_data['remarks'],
                      failure_desc=form.cleaned_data['failure_desc'],
                      tth_start=tth_start, prev_tth=prev_tth, prev_landings=prev_landings, pre_cycles=prev_cycles,
                      ms_report=ms_report, reservation=reservation, status='open', open_user=request.user.fbouser)

            pdt.full_clean()
            pdt.save()

            # przejście do kolejnego formularza
            if '_new_oper_' in request.POST:
                link = reverse('panel:operation-open', args=[pdt.pk])
                if camo:
                    link += '?camo=1'
                return HttpResponseRedirect(link)
            else:
                if '_close_ptd_' in request.POST or '_close_ptd_next_' in request.POST:
                    pdt_close(pdt, request.user.fbouser, form.cleaned_data)
                if camo:
                    if '_close_ptd_next_' in request.POST:
                        link = reverse('panel:pdt-open') + '?camo=1'
                        return HttpResponseRedirect(link)
                    else:
                        return HttpResponseRedirect(reverse('camo:aircraft-list'))
                else:
                    return HttpResponseRedirect(reverse('panel:pdt-info', args=[pdt.pk]))

        # Wyświetlenie formularza
        return render(request, 'panel/pdt_edit.html', context)

    else:

        context['open_pdt'] = open_pdt

        # Wyświetlenie formularza
        return render(request, 'panel/pdt_error.html', context)


@login_required()
def PDTUpdate(request, pk):

    # return render(request, 'panel/pdt_edit_blocked.html', {})

    # Weryfikacja czy wywołano z modułu CAMO
    camo = ('camo' in request.GET)

    pdt = PDT.objects.get(pk=pk)
    reservation = pdt.reservation
    last_oper = pdt.operation_set.order_by('operation_no').last()

    if pdt.ms_report:
        ms = "%s / %s" % (pdt.ms_report.next_hours, pdt.ms_report.next_date)
    else:
        ms = "Brak MS!"

    # Inicjalizacja formularza
    form = PDTEditForm(request.POST or None, instance=pdt, fbouser=request.user.fbouser, camo=camo, initial={'ms': ms})

    # wspólne zmienne kontekstowe
    context = {}
    context['form'] = form
    context['camo'] = camo
    context['open_oper'] = None
    if last_oper and last_oper.status == 'open':
        context['open_oper'] = last_oper

    # jeśli formularz został poprawnie wypełniony
    if form.is_valid():

        pdt.date = form.cleaned_data['date']
        pdt.pdt_ref = form.cleaned_data['pdt_ref']
        pdt.flight_type = form.cleaned_data['flight_type']
        pdt.pic = form.cleaned_data['pic']
        pdt.sic = form.cleaned_data['sic']
        pdt.instructor = form.cleaned_data['instructor']
        pdt.voucher = form.cleaned_data['voucher']
        pdt.ext_voucher = form.cleaned_data['ext_voucher']
        pdt.training = form.cleaned_data['training']
        pdt.contractor = form.cleaned_data['contractor']
        pdt.service_remarks = form.cleaned_data['service_remarks']
        pdt.remarks = form.cleaned_data['remarks']
        pdt.fuel_after = form.cleaned_data['fuel_after']
        pdt.fuel_after_source = form.cleaned_data['fuel_after_source']
        pdt.failure_desc = form.cleaned_data['failure_desc']
        pdt.reservation = reservation

        pdt.full_clean()
        pdt.save()

        # przejście do kolejnego formularza
        if '_new_oper_' in request.POST:
            link = reverse('panel:operation-open', args=[pdt.pk])
        elif '_close_oper_' in request.POST:
            operation_pk = request.POST['_close_oper_']
            link = reverse('panel:operation-update', args=[operation_pk])
        elif any('_edit_oper_' in s for s in request.POST):
            operation_pk = next((s for s in request.POST if '_edit_oper_' in s), None)[11:-2]
            link = reverse('panel:operation-update', args=[operation_pk])
        elif any('_del_oper_' in s for s in request.POST):
            operation_pk = next((s for s in request.POST if '_del_oper_' in s), None)[10:-2]
            link = reverse('panel:operation-delete', args=[operation_pk])
        else:
            if '_close_ptd_' in request.POST or '_close_ptd_next_' in request.POST:
                pdt_close(pdt, request.user.fbouser, form.cleaned_data)
            if camo:
                if '_close_ptd_next_' in request.POST:
                    link = reverse('panel:pdt-open')
                else:
                    link = reverse('camo:aircraft-list')
            else:
                link = reverse('panel:pdt-info', args=[pdt.pk])

        if camo:
            link += '?camo=1'
        return HttpResponseRedirect(link)

    # Wyświetlenie formularza
    return render(request, 'panel/pdt_edit.html', context)


@class_view_decorator(login_required())
class PDTDelete (DeleteView):
    model = PDT
    template_name = 'panel/delete_template.html'

    def get_context_data(self, **kwargs):
        context = super(PDTDelete, self).get_context_data(**kwargs)
        context['page_title'] = 'Usunięcie PDT'
        context['header_text'] = 'Usunięcie PDT'
        context['type'] = 'info'
        context['submenu_template'] = 'panel/pdt_submenu.html'
        context['pdt'] = self.object
        return context

    def get_success_url(self):
        return reverse('panel:pdt-active')


def operation_close(operation):

    # odświeżenie wartości z bazy
    operation.pdt.refresh_from_db()
    operation.pdt.aircraft.refresh_from_db()

    # zmiana statusu operacji
    operation.status = 'closed'
    operation.full_clean()
    operation.save()


@login_required()
def OperationOpen (request, pdt_id):

    # Weryfikacja czy wywołano z modułu CAMO
    camo = ('camo' in request.GET)

    pdt = PDT.objects.get(pk=pdt_id)
    fuel_available = pdt.aircraft.fuel_volume
    last_operation = pdt.operation_set.order_by('operation_no').last()

    # Ustalenie inicjalnych wartości atrybutów operacji
    if last_operation:
        # Na podstawie poprzedniej operacji na tym PDT
        loc_start = last_operation.loc_end
        tth_start = last_operation.tth_end
        operation_no = last_operation.operation_no + 1
    else:
        # Na podstawie ostatniej operacji ostatniego PDTa
        prev_pdt = pdt.aircraft.pdt_set.exclude(pk=pdt.pk).order_by('pk').last()
        if prev_pdt:
            prev_operation = prev_pdt.operation_set.order_by('operation_no').last()
        else:
            prev_operation = None
        if prev_operation:
            loc_start = prev_operation.loc_end
        else:
            loc_start = 'EPMO'
        tth_start = pdt.tth_start
        operation_no = 1

    # Domyślne źródło dla uzupełnianego paliwa
    if loc_start and loc_start.upper() != 'EPMO':
        fuel_source = 'other'
    else:
        fuel_source = None

    # Odczytanie lisczby pasażerów z rezerwacji
    if pdt.flight_type in ('01', '01A') and pdt.reservation:
        pax = len(list(filter(None, pdt.reservation.pax.split(sep="\r\n"))))
    else:
        pax = None

    form = OperationEditForm(request.POST or None, pdt=pdt, user=request.user, pfi=False,
                            initial={'operation_no': operation_no, 'fuel_available': int(fuel_available),
                                     'loc_start': loc_start, 'tth_start': tth_start, 'fuel_source': fuel_source,
                                     'pax': pax})

    # wspólne zmienne kontekstowe
    context = {}
    context['form'] = form
    context['type'] = 'oper'
    context['submenu_template'] = 'panel/pdt_submenu.html'
    context['header_text'] = 'Nowa operacja na PDT'
    context['pdt'] = pdt
    context['camo'] = camo

    # jeśli formularz został poprawnie wypełniony
    if form.is_valid():

        # zapis informacji o tankowaniu
        if form.cleaned_data.get('fuel_refill', 0) != 0:
            fueling_save(pdt=pdt, fueltank_string=form.cleaned_data['fuel_source'],
                         location=form.cleaned_data['loc_start'],
                         fuel_volume=form.cleaned_data['fuel_refill'],
                         remarks='Operacja numer %d w %s' % (operation_no, form.cleaned_data['loc_start']))

        # jeśli jest to ostatni PDT to aktualizacja paliwa na SP
        aircraft = pdt.aircraft
        aircraft.refresh_from_db()
        if pdt == aircraft.pdt_set.last() and form.cleaned_data.get('fuel_available') and form.cleaned_data.get('fuel_used'):
            aircraft.fuel_volume = max(form.cleaned_data['fuel_available'] - form.cleaned_data['fuel_used'], 0)
            aircraft.full_clean()
            aircraft.save()

        # zapis nowej operacji
        operation = Operation(pdt=pdt, operation_no=operation_no,
                              fuel_refill=form.cleaned_data['fuel_refill'],
                              fuel_source=form.cleaned_data['fuel_source'],
                              fuel_available=form.cleaned_data['fuel_available'],
                              fuel_used=form.cleaned_data['fuel_used'], loc_start=form.cleaned_data['loc_start'],
                              tth_start=form.cleaned_data['tth_start'], time_start=form.cleaned_data['time_start'],
                              loc_end=form.cleaned_data['loc_end'], tth_end=form.cleaned_data['tth_end'],
                              time_end=form.cleaned_data['time_end'], landings=form.cleaned_data['landings'],
                              cycles=form.cleaned_data.get('cycles', 0), status='open')

        operation.full_clean()
        operation.save()

        # zamknięcie operacji
        if '_close_oper_' in request.POST:
            operation_close(operation)

        # przejście do kolejnego formularza
        link = reverse('panel:pdt-update', args=[pdt_id])
        if camo:
            link += '?camo=1'
        return HttpResponseRedirect(link)

    return render(request, 'panel/operation_edit.html', context)


@login_required()
def OperationUpdate(request, pk):

    # Weryfikacja czy wywołano z modułu CAMO
    camo = ('camo' in request.GET)

    operation = Operation.objects.get(pk=pk)

    # Inicjalizacja formularza

    form = OperationEditForm(request.POST or None, pdt=operation.pdt, user=request.user, pfi=False, instance=operation)

    # wspólne zmienne kontekstowe
    context = {}
    context['form'] = form
    context['page_title'] = 'Edycja operacji'
    context['type'] = 'oper'
    context['submenu_template'] = 'panel/pdt_submenu.html'
    context['header_text'] = 'Edycja operacji na PDT'
    context['pdt'] = operation.pdt
    context['camo'] = camo

    # jeśli formularz został poprawnie wypełniony
    if form.is_valid():

        # jeśli jest to ostatni PDT to aktualizacja paliwa na SP
        aircraft = operation.pdt.aircraft
        aircraft.refresh_from_db()
        if operation.pdt == aircraft.pdt_set.last() and form.cleaned_data.get('fuel_available') and form.cleaned_data.get('fuel_used'):
            aircraft.fuel_volume = max(form.cleaned_data['fuel_available'] - form.cleaned_data['fuel_used'], 0)
            aircraft.full_clean()
            aircraft.save()

        # zapis zmian w operacji
        operation.pax = form.cleaned_data['pax']
        operation.bags = form.cleaned_data['bags']
        operation.fuel_refill = form.cleaned_data['fuel_refill']
        operation.fuel_source = form.cleaned_data['fuel_source']
        operation.fuel_available = form.cleaned_data['fuel_available']
        operation.fuel_used = form.cleaned_data['fuel_used']
        operation.oil_refill = form.cleaned_data['oil_refill']
        operation.trans_oil_refill = form.cleaned_data['trans_oil_refill']
        operation.hydr_oil_refill = form.cleaned_data['hydr_oil_refill']
        operation.loc_start = form.cleaned_data['loc_start']
        operation.time_start = form.cleaned_data['time_start']
        operation.tth_start = form.cleaned_data['tth_start']
        operation.loc_end = form.cleaned_data['loc_end']
        operation.time_end = form.cleaned_data['time_end']
        operation.tth_end = form.cleaned_data['tth_end']
        operation.landings = form.cleaned_data['landings']
        operation.cycles = form.cleaned_data['cycles']

        operation.full_clean()
        operation.save()

        # przejście do kolejnego formularza
        if '_close_oper_' in request.POST:
            operation_close(operation)

        link = reverse('panel:pdt-update', args=[operation.pdt.pk])
        if camo:
            link += '?camo=1'
        return HttpResponseRedirect(link)

    # Wyświetlenie formularza
    return render(request, 'panel/operation_edit.html', context)


@class_view_decorator(login_required())
class OperationDelete (DeleteView):

    model = Operation
    template_name = 'panel/delete_template.html'

    def get_context_data(self, **kwargs):
        context = super(OperationDelete, self).get_context_data(**kwargs)
        context['page_title'] = 'Operacja'
        context['header_text'] = 'Usunięcie operacji na PDT'
        context['type'] = 'oper'
        context['submenu_template'] = 'panel/pdt_submenu.html'
        context['pdt'] = self.object.pdt
        return context

    def get_success_url(self):
        return reverse('panel:pdt-update', args=[self.object.pdt.pk])


@class_view_decorator(login_required())
class RemoteFuelingDetails (DetailView):
    model = RemoteFueling
    template_name = 'panel/details_template.html'

    def get_context_data(self, **kwargs):
        context = super(RemoteFuelingDetails, self).get_context_data(**kwargs)
        context['page_title'] = 'Tankowanie'
        context['header_text'] = 'Szczegóły tankowania poza SALT'
        context['type'] = 'fuel'
        context['submenu_template'] = 'panel/pdt_submenu.html'
        context['pdt'] = self.object.pdt

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Aktualizuj', 'path': reverse('panel:remote-fueling-update', args=[self.object.pk])})
        local_menu.append({'text': 'Usuń', 'path': reverse('panel:remote-fueling-delete', args=[self.object.pk])})
        context['local_menu'] = local_menu

        field_list = []
        field_list.append({'header': 'Data tankowania', 'value': self.object.pdt.date})
        field_list.append({'header': 'Miejsce tankowania', 'value': self.object.location})
        field_list.append({'header': 'Ilość paliwa', 'value': '%.1f L' % self.object.fuel_volume})
        field_list.append({'header': 'Faktura', 'value': self.object.document})
        field_list.append({'header': 'Dokument akcyzy', 'value': self.object.excise})
        field_list.append({'header': 'Cena paliwa', 'value': '%.2f PLN' % self.object.total_price if self.object.total_price else None})
        field_list.append({'header': 'Uwagi', 'value': self.object.remarks})

        context['field_list'] = field_list

        return context


@class_view_decorator(login_required())
class RemoteFuelingUpdate (UpdateView):
    model = RemoteFueling
    template_name = 'panel/update_template.html'

    def get_context_data(self, **kwargs):
        context = super(RemoteFuelingUpdate, self).get_context_data(**kwargs)
        context['page_title'] = 'Tankowanie'
        context['header_text'] = 'Modyfikacja tankowania poza SALT'
        context['type'] = 'fuel'
        context['submenu_template'] = 'panel/pdt_submenu.html'
        context['pdt'] = self.object.pdt
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(RemoteFueling, exclude=['pdt', 'fuel_volume'],
                                       widgets={'document':TextInput(attrs={'size':50}),
                                                'excise':TextInput(attrs={'size':50}),
                                                'remarks':Textarea(attrs={'rows':4, 'cols':100})})
        return form_class

    def get_success_url(self):
        return reverse('panel:pdt-fuel', args=[self.object.pdt.pk,])


@class_view_decorator(login_required())
class RemoteFuelingDelete (DeleteView):
    model = RemoteFueling
    template_name = 'panel/delete_template.html'

    def get_context_data(self, **kwargs):
        context = super(RemoteFuelingDelete, self).get_context_data(**kwargs)
        context['page_title'] = 'Tankowanie'
        context['header_text'] = 'Usunięcie tankowania poza SALT'
        context['description'] = 'Na podstawie PDT: %s' % self.object.pdt
        context['type'] = 'fuel'
        context['submenu_template'] = 'panel/pdt_submenu.html'
        context['pdt'] = self.object.pdt
        return context

    def get_success_url(self):
        return reverse('panel:pdt-fuel', args=[self.object.pdt.pk,])


@class_view_decorator(login_required())
class RatingDetails (DetailView):
    model = Rating
    template_name = 'panel/details_template.html'

    def get_context_data(self, **kwargs):
        context = super(RatingDetails, self).get_context_data(**kwargs)
        context['page_title'] = 'Uprawnienia'
        context['header_text'] = 'Informacje o uprawnieniu pilota'
        context['type'] = 'ratings'
        context['submenu_template'] = 'panel/pilot_submenu.html'
        context['pilot'] = self.object.pilot

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Aktualizuj', 'path': reverse('panel:rating-update', args=[self.object.pk])})
        local_menu.append({'text': 'Usuń', 'path': reverse('panel:rating-delete', args=[self.object.pk])})
        context['local_menu'] = local_menu

        field_list = []
        field_list.append({'header': 'Nazwa uprawnienia', 'value': self.object.rating, 'bold': True})
        field_list.append({'header': 'Data ważności', 'value': self.object.valid_date if self.object.valid_date else 'bezterminowo' })
        field_list.append({'header': 'Uwagi', 'value': self.object.remarks})
        context['field_list'] = field_list
        return context


@class_view_decorator(login_required())
class RatingCreate (CreateView):
    model = Rating
    template_name = 'panel/create_template.html'

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(Rating, exclude=('pilot',),
                                       widgets={'valid_date': AdminDateWidget,
                                                'remarks':Textarea(attrs={'rows':4, 'cols':100})})
        return form_class

    def form_valid(self, form):
        form.instance.pilot = Pilot.objects.get(pk=self.kwargs['pilot_id'])
        return super(RatingCreate, self).form_valid(form)

    def get_context_data(self, **kwargs):
        context = super(RatingCreate, self).get_context_data(**kwargs)
        context['page_title'] = 'Uprawnienia'
        context['header_text'] = 'Nowe uprawnienie pilota'
        context['type'] = 'ratings'
        context['submenu_template'] = 'panel/pilot_submenu.html'
        context['pilot'] = Pilot.objects.get(pk=self.kwargs['pilot_id'])
        return context

    def get_success_url(self):
        return reverse('panel:pilot-ratings', args=(self.kwargs['pilot_id'],))


@class_view_decorator(login_required())
class RatingUpdate (UpdateView):
    model = Rating
    template_name = 'panel/update_template.html'

    def get_context_data(self, **kwargs):
        context = super(RatingUpdate, self).get_context_data(**kwargs)
        context['page_title'] = 'Uprawnienia'
        context['header_text'] = 'Modyfikacja uprawnienia pilota'
        context['type'] = 'ratings'
        context['submenu_template'] = 'panel/pilot_submenu.html'
        context['pilot'] = self.object.pilot
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(Rating, exclude=('pilot',),
                                       widgets={'remarks':Textarea(attrs={'rows':4, 'cols':100})})
        return form_class

    def get_success_url(self):
        return reverse('panel:pilot-ratings', args=(self.object.pilot.pk,))


@class_view_decorator(login_required())
class RatingDelete (DeleteView):
    model = Rating
    template_name = 'panel/delete_template.html'

    def get_context_data(self, **kwargs):
        context = super(RatingDelete, self).get_context_data(**kwargs)
        context['page_title'] = 'Uprawnienia'
        context['header_text'] = 'Usunięcie uprawnienia pilota'
        context['type'] = 'ratings'
        context['submenu_template'] = 'panel/pilot_submenu.html'
        context['pilot'] = self.object.pilot
        return context

    def get_success_url(self):
        return reverse('panel:pilot-ratings', args=(self.object.pilot.pk,))


@class_view_decorator(login_required())
class DutyDetails (DetailView):
    model = Duty
    template_name = 'panel/details_template.html'

    def get_context_data(self, **kwargs):
        context = super(DutyDetails, self).get_context_data(**kwargs)
        context['page_title'] = 'Czas pracy'
        context['header_text'] = 'Informacje o pozycji czasu pracy'
        context['type'] = 'duties'
        context['submenu_template'] = 'panel/pilot_submenu.html'
        context['pilot'] = self.object.pilot

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Aktualizuj', 'path': reverse('panel:duty-update', args=[self.object.pk])})
        local_menu.append({'text': 'Usuń', 'path': reverse('panel:duty-delete', args=[self.object.pk])})
        context['local_menu'] = local_menu

        field_list = []
        field_list.append({'header': 'Data', 'value': self.object.date, 'bold': True})
        field_list.append({'header': 'Firma', 'value': self.object.comp_str(), 'bold': True})
        field_list.append({'header': 'Stanowisko', 'value': self.object.role})
        field_list.append({'header': 'Rodzaj obowiązków', 'value': self.object.duty_type})
        field_list.append({'header': 'PDT', 'value': self.object.pdt, 'link': reverse('panel:pdt-info', args=[self.object.pdt.pk]) if self.object.pdt else ''})
        field_list.append({'header': 'Czas łączny', 'value': '%s %s' % (duration_string(self.object.duty_time) if self.object.duty_time else '',
                                                            ('(od %s do %s)' % (self.object.duty_time_from.strftime('%H:%M'), self.object.duty_time_to.strftime('%H:%M'))
                                                            if (self.object.duty_time_from and self.object.duty_time_to) else '')) })
        # field_list.append({'header': 'Czas czynności', 'value': duration_string(self.object.fdp_time) if self.object.fdp_time else None})
        field_list.append({'header': 'Czas lotu', 'value': duration_string(self.object.block_time) if self.object.block_time else None})
        field_list.append({'header': 'Mnożnik czasu', 'value': self.object.time_factor})
        field_list.append({'header': 'Uwagi', 'value': self.object.remarks})
        context['field_list'] = field_list
        return context


@class_view_decorator(login_required())
class DutyCreate (CreateView):
    model = Duty
    template_name = 'panel/duty_edit.html'

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(Duty, exclude=('pilot', 'pdt', 'fdp_time', 'time_factor'),
                                       widgets={'date': AdminDateWidget,
                                                'duty_time_from': TimeInput(format="%H:%M"),
                                                'duty_time_to': TimeInput(format="%H:%M"),
                                                'role': TextInput(attrs={'size': 50}),
                                                'duty_type': TextInput(attrs={'size': 50}),
                                                'remarks':Textarea(attrs={'rows':4, 'cols':100})})

        class create_class(form_class):
            def clean(self):
                cleaned_data = super(create_class, self).clean()
                if cleaned_data['duty_time_from'] and cleaned_data['duty_time_to']:
                    if cleaned_data['duty_time_from'] >= cleaned_data['duty_time_to']:
                        raise forms.ValidationError("Termin rozpoczęcia musi być wcześniejszy niż zakończenia!")
                if cleaned_data['duty_time']:
                    if cleaned_data['duty_time'].seconds > 12*60*60:
                        raise forms.ValidationError("Zbyt długi czas!")
                # if cleaned_data['fdp_time'] and cleaned_data['duty_time']:
                #     if cleaned_data['fdp_time'] > cleaned_data['duty_time']:
                #         raise forms.ValidationError("Czas czynności lotniczych dłuższy niż czas łączny!")
                if cleaned_data['block_time'] and cleaned_data['duty_time']:
                    if cleaned_data['block_time'] > cleaned_data['duty_time']:
                        raise forms.ValidationError("Czas lotu dłuższy niż czas łączny!")
                # if cleaned_data['block_time'] and cleaned_data['fdp_time']:
                #     if cleaned_data['block_time'] > cleaned_data['fdp_time']:
                #         raise forms.ValidationError("Czas lotu dłuższy niż czas czynności lotniczych!")

        create_class.base_fields['date'].initial = date.today()

        if ('start' in self.request.GET) and ('end' in self.request.GET):
            try:
                start = datetime.strptime(self.request.GET['start'], '%Y-%m-%dT%H:%M:%S').time()
                end = datetime.strptime(self.request.GET['end'], '%Y-%m-%dT%H:%M:%S').time()
                today = datetime.strptime(self.request.GET['start'], '%Y-%m-%dT%H:%M:%S').date()
            except:
                start = None
                end = None
                today = None

            if start and end and today:
                create_class.base_fields['date'].initial = today
                create_class.base_fields['duty_time_from'].initial = start
                create_class.base_fields['duty_time_to'].initial = end

        return create_class

    def form_valid(self, form):
        form.instance.pilot = Pilot.objects.get(pk=self.kwargs['pilot_id'])
        return super(DutyCreate, self).form_valid(form)

    def get_context_data(self, **kwargs):
        context = super(DutyCreate, self).get_context_data(**kwargs)
        context['page_title'] = 'Czas pracy'
        context['header_text'] = 'Nowa pozycja czasu pracy'
        context['type'] = 'duties'
        context['submenu_template'] = 'panel/pilot_submenu.html'
        context['pilot'] = Pilot.objects.get(pk=self.kwargs['pilot_id'])
        return context

    def get_success_url(self):
        return reverse('panel:pilot-duties', args=(self.kwargs['pilot_id'],))


@class_view_decorator(login_required())
class DutyUpdate (UpdateView):
    model = Duty
    template_name = 'panel/duty_edit.html'

    def get_context_data(self, **kwargs):
        context = super(DutyUpdate, self).get_context_data(**kwargs)
        context['page_title'] = 'Czas pracy'
        context['header_text'] = 'Modyfikacja pozycji czasu pracy'
        context['type'] = 'duties'
        context['submenu_template'] = 'panel/pilot_submenu.html'
        context['pilot'] = self.object.pilot
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(Duty, exclude=('pilot', 'pdt', 'fdp_time', 'time_factor'),
                                       widgets={'date': AdminDateWidget,
                                                'duty_time_from': TimeInput(format="%H:%M"),
                                                'duty_time_to': TimeInput(format="%H:%M"),
                                                'role': TextInput(attrs={'size': 50}),
                                                'duty_type': TextInput(attrs={'size': 50}),
                                                'remarks':Textarea(attrs={'rows':4, 'cols':100})})
        class update_class(form_class):
            def clean(self):
                cleaned_data = super(update_class, self).clean()
                if cleaned_data['duty_time_from'] and cleaned_data['duty_time_to']:
                    if cleaned_data['duty_time_from'] >= cleaned_data['duty_time_to']:
                        raise forms.ValidationError("Termin rozpoczęcia musi być wcześniejszy niż zakończenia!")
                if cleaned_data['duty_time']:
                    if cleaned_data['duty_time'].seconds > 12*60*60:
                        raise forms.ValidationError("Zbyt długi czas służby!")
                # if cleaned_data['fdp_time'] and cleaned_data['duty_time']:
                #     if cleaned_data['fdp_time'] > cleaned_data['duty_time']:
                #         raise forms.ValidationError("Czas czynności lotniczych dłuższy niż czas służby!")
                if cleaned_data['block_time'] and cleaned_data['duty_time']:
                    if cleaned_data['block_time'] > cleaned_data['duty_time']:
                        raise forms.ValidationError("Czas lotu dłuższy niż czas służby!")
                # if cleaned_data['block_time'] and cleaned_data['fdp_time']:
                #     if cleaned_data['block_time'] > cleaned_data['fdp_time']:
                #         raise forms.ValidationError("Czas lotu dłuższy niż czas czynności lotniczych!")

        return update_class

    def get_success_url(self):
        return reverse('panel:pilot-duties', args=(self.object.pilot.pk,))


@class_view_decorator(login_required())
class DutyDelete (DeleteView):
    model = Duty
    template_name = 'panel/delete_template.html'

    def get_context_data(self, **kwargs):
        context = super(DutyDelete, self).get_context_data(**kwargs)
        context['page_title'] = 'Czas pracy'
        context['header_text'] = 'Usunięcie pozycji czasu pracy'
        context['type'] = 'duties'
        context['submenu_template'] = 'panel/pilot_submenu.html'
        context['pilot'] = self.object.pilot
        return context

    def get_success_url(self):
        return reverse('panel:pilot-duties', args=(self.object.pilot.pk,))


@class_view_decorator(permission_required('ato.instructor'))
class ExercisePass (UpdateView):
    model = Exercise_inst
    template_name = 'panel/update_template.html'

    def get_context_data(self, **kwargs):
        context = super(ExercisePass, self).get_context_data(**kwargs)
        context['page_title'] = 'Zaliczenie'
        context['header_text'] = 'Zaliczenie ćwiczenia %s' % self.object
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(Exercise_inst)
        return form_class

    def get_success_url(self):
        return reverse('panel:my-training-list')


@class_view_decorator(permission_required('ato.instructor'))
class PhasePass (UpdateView):
    model = Phase_inst
    template_name = 'panel/update_template.html'

    def get_context_data(self, **kwargs):
        context = super(PhasePass, self).get_context_data(**kwargs)
        context['page_title'] = 'Zaliczenie'
        context['header_text'] = 'Zaliczenie zadania/fazy %s' % self.object
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(Phase_inst)
        return form_class

    def get_success_url(self):
        return reverse('panel:my-training-list')


@class_view_decorator(permission_required('ato.instructor'))
class TrainingPass (UpdateView):
    model = Training_inst
    template_name = 'panel/update_template.html'

    def get_context_data(self, **kwargs):
        context = super(TrainingPass, self).get_context_data(**kwargs)
        context['page_title'] = 'Zaliczenie'
        context['header_text'] = 'Zaliczenie szkolenia %s' % self.object
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(Training_inst)
        return form_class

    def get_success_url(self):
        return reverse('panel:my-training-list')


@login_required()
def MobilePDTOpen0(request):

    # Skasowanie zmiennych sesyjnych
    request.session.pop('pdt_date', None)
    request.session.pop('pdt_ref', None)
    request.session.pop('aircraft_id', None)
    request.session.pop('flight_type', None)
    request.session.pop('voucher_id', None)
    request.session.pop('ext_voucher', None)
    request.session.pop('training_id', None)
    request.session.pop('instructor_id', None)
    request.session.pop('contractor_id', None)
    request.session.pop('pic_id', None)
    request.session.pop('sic_id', None)
    request.session.pop('service_remarks', None)
    request.session.pop('remarks', None)
    request.session.pop('reservation_id', None)

    return HttpResponseRedirect(reverse('panel:mpdt-open1'))


@login_required()
def MobilePDTOpen1(request):

    if 'res' in request.GET:

        reservation_id = request.GET.get('res')
        try:
            reservation = Reservation.objects.get(pk=reservation_id)
        except:
            reservation = None

        # Skasowanie zmiennych sesyjnych
        request.session.pop('pdt_date', None)
        request.session.pop('pdt_ref', None)
        request.session.pop('voucher_id', None)
        request.session.pop('ext_voucher', None)
        request.session.pop('training_id', None)
        request.session.pop('instructor_id', None)
        request.session.pop('contractor_id', None)
        request.session.pop('service_remarks', None)
        request.session.pop('remarks', None)

        # Ustawienie zmienny sesyjnych na podstawie rezewacji
        if reservation:
            request.session['pdt_date'] = reservation.start_time.date()
            request.session['aircraft_id'] = reservation.aircraft_id
            request.session['flight_type'] = reservation.planned_type
            request.session['pic_id'] = reservation.owner_id
            request.session['sic_id'] = reservation.participant_id
            request.session['reservation_id'] = reservation_id
            aircraft = Aircraft.objects.get(pk=reservation.aircraft_id)
            if aircraft:
                request.session['pdt_ref'] = aircraft.last_pdt_ref + 1
        else:
            request.session.pop('pdt_date', None)
            request.session.pop('aircraft_id', None)
            request.session.pop('flight_type', None)
            request.session.pop('pic_id', None)
            request.session.pop('sic_id', None)
            request.session.pop('reservation_id', None)

    # odczytanie zmiennych sesyjnych
    pdt_date = request.session.get('pdt_date', date.today())
    aircraft_id = request.session.get('aircraft_id', None)
    flight_type = request.session.get('flight_type', None)
    pdt_ref = request.session.get('pdt_ref', None)
    remarks = request.session.get('remarks', None)

    form = CheckOutFormOpen(request.POST or None, pilot=request.user.fbouser.pilot, is_mobile=request.device['is_mobile'],
                            initial={'pdt_date':pdt_date, 'flight_type':flight_type, 'aircraft':aircraft_id, 'pdt_ref':pdt_ref, 'remarks':remarks})

    # wspólne zmienne kontekstowe
    context = {}
    context['form'] = form

    # jeśli formularz został poprawnie wypełniony
    if form.is_valid():

        # zapisanie zmiennych sesyjnych
        request.session['flight_type'] = form.cleaned_data['flight_type']
        request.session['aircraft_id'] = form.cleaned_data['aircraft']
        request.session['remarks'] = form.cleaned_data['remarks']
        request.session['pdt_ref'] = form.cleaned_data['pdt_ref']
        request.session['pdt_date'] = form.cleaned_data['pdt_date']

        # przejście do kolejnego formularza
        if form.cleaned_data['flight_type'] in ('05', '06', '06A','07'):
            return HttpResponseRedirect(reverse('panel:mpdt-open3'))
        else:
            return HttpResponseRedirect(reverse('panel:mpdt-open2'))

    return render(request, 'panel/mpdt_open1.html', context)


@login_required()
def MobilePDTOpen2 (request):

    # sprawdzenie zmiennych sesyjnych
    if not ('flight_type' in request.session and 'aircraft_id' in request.session):
        return HttpResponseRedirect(reverse('panel:mpdt-open1'))

    # odczytanie zmiennych sesyjnych
    flight_type = request.session.get('flight_type')
    aircraft_id = request.session.get('aircraft_id')
    voucher_id = request.session.get('voucher_id', None)
    ext_voucher = request.session.get('ext_voucher', None)
    training_id = request.session.get('training_id', None)
    instructor_id = request.session.get('instructor_id', None)
    contractor_id = request.session.get('contractor_id', None)
    service_remarks = request.session.get('service_remarks', None)

    aircraft = Aircraft.objects.get(pk=aircraft_id) if aircraft_id else None

    # ustawienie kontrahenta na podstawie użytkownika
    if not contractor_id:
        if request.user.fbouser.contractor:
            contractor_id = request.user.fbouser.contractor.pk
        else:
            contractor_id = None

    contractor = Contractor.objects.get(pk=contractor_id) if contractor_id else None

    # wybór właściwego formularza dla typu lotu
    if flight_type in ('01', '02', '02H'):
        form = CheckOutFormServices(request.POST or None, flight_type=flight_type, is_mobile=request.device['is_mobile'],
                                    initial={'contractors': contractor_id, 'service_remarks': service_remarks})
    elif flight_type in ('01A','03D'):
        form = CheckOutFormVouchers(request.POST or None, is_mobile=request.device['is_mobile'],
                                    initial={'vouchers': voucher_id, 'ext_voucher': ext_voucher})
    elif flight_type in ('03A', '03B', '03C', '03E'):
        form = CheckOutFormTrainings(request.POST or None, is_mobile=request.device['is_mobile'], user=request.user.fbouser,
                                     initial={'trainings': training_id})
    elif flight_type == '04':
        form = CheckOutFormRent(request.POST or None, is_mobile=request.device['is_mobile'],
                                initial={'instructor': instructor_id})
    else:
        form = None

    # wspólne zmienne kontekstowe
    context = {}
    context['form'] = form
    context['flight_type'] = flight_type
    context['flight_type_name'] = FlightTypes()[flight_type]
    context['aircraft'] = aircraft

    # zmienne kontekstowe zależne od formularza
    if flight_type == '04':
        context['normal_price'] = '%.2f PLN' % aircraft.rent_price
        if contractor:
            special_price = contractor.specialprice_set.filter(ac_type=aircraft.type).first()
            context['special_price'] = '%.2f PLN' % special_price.hour_price if special_price else None
            packages = []
            for soldpackage in contractor.soldpackage_set.filter(ac_type=aircraft.type).exclude(left_hours=0).order_by('date'):
                label = '<b>%.1f h w cenie %.2f PLN</b> (pakiet %s)' % (soldpackage.left_hours, soldpackage.hour_price, soldpackage.package_id)
                packages.append (label)
            context['packages'] = packages

    # jeśli formularz został poprawnie wypełniony
    if form.is_valid():

        # zapisanie zmiennych sesyjnych
        if flight_type in ('01', '02', '02H'):
            request.session['contractor_id'] = form.cleaned_data['contractors']
            request.session['service_remarks'] = form.cleaned_data['service_remarks']
        elif flight_type in ('01A','03D'):
            request.session['voucher_id'] = form.cleaned_data.get('vouchers', '')
            request.session['ext_voucher'] = form.cleaned_data.get('ext_voucher', None)
        elif flight_type in ('03A', '03B', '03C', '03E'):
            request.session['training_id'] = form.cleaned_data['trainings']
            request.session['instructor_id'] = Training_inst.objects.get(pk=form.cleaned_data['trainings']).instructor_id
        elif flight_type == '04':
            request.session['instructor_id'] = form.cleaned_data['instructor']

        # przejście do kolejnego formularza
        return HttpResponseRedirect(reverse('panel:mpdt-open3'))

    return render(request, 'panel/mpdt_open2.html', context)


@login_required()
def MobilePDTOpen3 (request):

    # sprawdzenie zmiennych sesyjnych
    if not ('flight_type' in request.session and 'aircraft_id' in request.session):
        return HttpResponseRedirect(reverse('panel:mpdt-open1'))

    # odczytanie zmiennych sesyjnych
    flight_type = request.session.get('flight_type')
    aircraft_id = request.session.get('aircraft_id')
    training_id = request.session.get('training_id', None)
    instructor_id = request.session.get('instructor_id', None)

    aircraft = Aircraft.objects.get(pk = aircraft_id)
    training = Training_inst.objects.get(pk=training_id) if training_id else None
    instructor = Instructor.objects.get(pk=instructor_id) if instructor_id else None

    # ustawienie domyślnego pic
    if 'pic_id' in request.session:
        pic_id = request.session['pic_id']
    else:
        if flight_type in ['01','01A','02','02H','03D', '03E','04','05','06','06A','07']:
            if flight_type == '03E' and training:
                pic_id = training.student.pilot.pk
            else:
                pic_id = request.user.fbouser.pilot.pk if request.user.fbouser.pilot else None
        elif flight_type in ['03A', '03B', '03C']:
            pic_id = training.instructor.pilot.pk
            if not 'instructor_id' in request.session:
                request.session['instructor_id'] = training.instructor.pk
        else:
            pic_id = None

    # ustawienie domyślnego sic
    if 'sic_id' in request.session:
        sic_id = request.session['sic_id']
    else:
        if flight_type in ['04'] :
            sic_id = instructor.pilot.pk if instructor else None
        elif flight_type in ['03A', '03B', '03C']:
            sic_id = training.student.pilot.pk
        else:
            sic_id = None

    if flight_type in ['03A', '03B', '03C']:
        student = training.student if training else None
    else:
        student = None

    form = CheckOutFormCrew(request.POST or None, student=student, is_mobile=request.device['is_mobile'],
                            initial={'pic': pic_id, 'sic': sic_id})

    # wspólne zmienne kontekstowe
    context = {}
    context['flight_type'] = flight_type
    context['flight_type_name'] = FlightTypes()[flight_type]
    context['aircraft'] = aircraft
    context['form'] = form

    # jeśli formularz został poprawnie wypełniony
    if form.is_valid():

        # zapisanie zmiennych sesyjnych
        request.session['pic_id'] = form.cleaned_data['pic']
        request.session['sic_id'] = form.cleaned_data['sic']

        contractor = Contractor.objects.get(pk=request.session.pop('contractor_id')) \
                     if 'contractor_id' in request.session and request.session['contractor_id'] != '' else None
        service_remarks = request.session.pop('service_remarks', None)
        voucher = Voucher.objects.get(pk=request.session.pop('voucher_id')) \
                  if 'voucher_id' in request.session and request.session['voucher_id'] != '' else None
        ext_voucher = request.session.pop('ext_voucher', None)
        training = Training_inst.objects.get(pk=request.session.pop('training_id')) \
                   if 'training_id' in request.session and request.session['training_id'] != '' else None
        instructor = Instructor.objects.get(pk=request.session.pop('instructor_id')).pilot \
                     if 'instructor_id' in request.session and request.session['instructor_id'] != '' else None
        pic = Pilot.objects.get(pk=request.session.pop('pic_id')) \
              if 'pic_id' in request.session else None
        sic = Pilot.objects.get(pk=request.session.pop('sic_id')) \
              if 'sic_id' in request.session and request.session['sic_id'] != '' else None
        request.session.pop('sic', None)
        remarks = request.session.pop('remarks', None)

        pdt_ref = request.session.pop('pdt_ref', None)
        if not pdt_ref or pdt_ref == '':
            pdt_ref = aircraft.last_pdt_ref + 1

        if not contractor:
            if flight_type in ('03A', '03B', '03C','03E'):
                contractor = training.student.pilot.fbouser.contractor
            elif flight_type == '04':
                contractor = request.user.fbouser.contractor

        pdt_date = request.session.pop('pdt_date', date.today())

        # Utworzenie nowego otwartego PDT
        ms = aircraft.ms_report_set.order_by('done_date').last()
        try:
            res = Reservation.objects.get(pk=request.session.pop('reservation_id', None))
        except:
            res = None

        # ustalenie licznika początkowego z ostatniego PDT
        last_pdt = aircraft.pdt_set.get(pdt_ref=aircraft.last_pdt_ref)
        if last_pdt:
            tth_start = last_pdt.tth_end()
        else:
            tth_start = aircraft.tth

        pdt = PDT(pdt_ref=pdt_ref, aircraft=aircraft, flight_type=flight_type, date=pdt_date,
                  pic=pic, sic=sic, instructor=instructor, voucher=voucher, ext_voucher=ext_voucher,
                  training=training, contractor=contractor, service_remarks=service_remarks, remarks=remarks,
                  tth_start=tth_start, prev_tth=aircraft.hours_count, prev_landings=aircraft.landings_count,
                  pre_cycles=aircraft.cycles_count, ms_report=ms, reservation=res,
                  status='open', open_user=request.user.fbouser)
        pdt.full_clean()
        pdt.save()

        # Skasowanie zmiennych sesyjnych
        request.session.pop('pdt_date', None)
        request.session.pop('pdt_ref', None)
        request.session.pop('aircraft_id', None)
        request.session.pop('flight_type', None)
        request.session.pop('voucher_id', None)
        request.session.pop('ext_voucher', None)
        request.session.pop('training_id', None)
        request.session.pop('instructor_id', None)
        request.session.pop('contractor_id', None)
        request.session.pop('pic_id', None)
        request.session.pop('sic_id', None)
        request.session.pop('service_remarks', None)
        request.session.pop('remarks', None)

        # przejście do kolejnego formularza
        return HttpResponseRedirect(reverse('panel:moperation-open', args=[pdt.pk]))

    return render(request, 'panel/mpdt_open3.html', context)


@login_required()
def MobileOperationOpen (request, pdt_id):

    pdt = get_object_or_404(PDT, pk=pdt_id)
    fuel_available = pdt.aircraft.fuel_volume
    last_operation = pdt.operation_set.order_by('operation_no').last()

    # Ustalenie inicjalnych wartości atrybutów operacji
    if last_operation:
        # Na podstawie poprzedniej operacji na tym PDT
        loc_start = last_operation.loc_end
        tth_start = last_operation.tth_end
        operation_no = last_operation.operation_no + 1
    else:
        # Na podstawie ostatniej operacji ostatniego PDTa
        prev_pdt = pdt.aircraft.pdt_set.exclude(pk=pdt.pk).order_by('pk').last()
        if prev_pdt:
            prev_operation = prev_pdt.operation_set.order_by('operation_no').last()
        else:
            prev_operation = None
        if prev_operation:
            loc_start = prev_operation.loc_end
        else:
            loc_start = 'EPMO'
        tth_start = pdt.tth_start
        operation_no = 1

    # Domyślne źródło dla uzupełnianego paliwa
    if loc_start.upper() != 'EPMO':
        fuel_source = 'other'
    else:
        fuel_source = None

    # Odczytanie lisczby pasażerów z rezerwacji
    if pdt.flight_type in ('01', '01A') and pdt.reservation:
        pax = len(list(filter(None, pdt.reservation.pax.split(sep="\r\n"))))
    else:
        pax = None

    form = OperationOpenForm(request.POST or None, pdt=pdt, is_mobile=request.device['is_mobile'], user=request.user,
                             initial={'fuel_available': int(fuel_available), 'loc_start': loc_start, 'fuel_source': fuel_source})

    # wspólne zmienne kontekstowe
    context = {}
    context['form'] = form
    context['pdt'] = pdt
    context['fuel_volume'] = fuel_available

    # jeśli formularz został poprawnie wypełniony
    if form.is_valid():

        # zapis informacji o tankowaniu
        if form.cleaned_data.get('fuel_refill', 0) != 0:
            fueling_save(pdt=pdt, fueltank_string=form.cleaned_data['fuel_source'],
                         location=form.cleaned_data['loc_start'],
                         fuel_volume=form.cleaned_data['fuel_refill'],
                         remarks='Operacja numer %d w %s' % (operation_no, form.cleaned_data['loc_start']))

        # jeśli jest to ostatni PDT to aktualizacja paliwa na SP
        aircraft = pdt.aircraft
        if pdt == aircraft.pdt_set.last():
            aircraft.fuel_volume = form.cleaned_data['fuel_available']
            aircraft.full_clean()
            aircraft.save()

        # zapis nowej operacji
        operation = Operation(pdt=pdt, operation_no=operation_no, pax=form.cleaned_data['pax'],
                              bags=form.cleaned_data['bags'],
                              fuel_refill=form.cleaned_data['fuel_refill'],
                              fuel_source=form.cleaned_data['fuel_source'],
                              fuel_available=form.cleaned_data['fuel_available'],
                              oil_refill=form.cleaned_data['oil_refill'],
                              trans_oil_refill=form.cleaned_data['trans_oil_refill'],
                              hydr_oil_refill=form.cleaned_data['hydr_oil_refill'],
                              loc_start=form.cleaned_data['loc_start'], tth_start=tth_start, status='open')
        operation.full_clean()
        operation.save()

        # przejście do kolejnego formularza
        return HttpResponseRedirect(reverse('dispatcher'))

    return render(request, 'panel/moperation_open.html', context)


@login_required()
def MobileOperationClose (request, operation_id):

    operation = get_object_or_404(Operation, pk=operation_id)

    form = OperationCloseForm(request.POST or None, operation=operation, is_mobile=request.device['is_mobile'],
                              initial={'loc_start': operation.loc_start, 'tth_start': operation.tth_start,
                                       'fuel_used': 0, 'landings': 1, 'cycles': 0})

    # wspólne zmienne kontekstowe
    context = {}
    context['form'] = form
    context['operation'] = operation
    context['fuel_consumption'] = operation.pdt.aircraft.fuel_consumption

    # jeśli formularz został poprawnie wypełniony
    if form.is_valid():

        # aktualizacja paliwa na SP
        aircraft = operation.pdt.aircraft
        aircraft.fuel_volume = max(aircraft.fuel_volume - form.cleaned_data['fuel_used'], 0)
        aircraft.full_clean()
        aircraft.save()

        operation.fuel_used=form.cleaned_data['fuel_used']
        operation.loc_start=form.cleaned_data['loc_start']
        operation.loc_end=form.cleaned_data['loc_end']
        operation.tth_start=form.cleaned_data['tth_start']
        operation.tth_end=form.cleaned_data['tth_end']
        operation.time_start=form.cleaned_data['time_start']
        operation.time_end=form.cleaned_data['time_end']
        operation.landings=form.cleaned_data['landings']
        operation.cycles=form.cleaned_data.get('cycles', 0)
        operation.status='closed'
        operation.full_clean()
        operation.save()

        # przejście do kolejnego formularza
        return HttpResponseRedirect(reverse('dispatcher'))

    return render(request, 'panel/moperation_close.html', context)


@login_required()
def MobilePDTClose (request, pdt_id):

    pdt = PDT.objects.get(pk = pdt_id)
    reservation = pdt.reservation
    last_operation = pdt.operation_set.order_by('operation_no').last()

    if last_operation and last_operation.loc_end != 'EPMO':
        fuel_source = 'other'
    else:
        fuel_source = None

    form = CheckInForm(request.POST or None, pdt=pdt, is_mobile=request.device['is_mobile'],
                       initial={'fuel_after_source': fuel_source, 'fuel_after': 0, 'remarks': pdt.remarks})

    # wspólne zmienne kontekstowe
    context = {}
    context['pdt'] = pdt
    context['form'] = form

    # jeśli formularz został poprawnie wypełniony
    if form.is_valid():

        pdt.fuel_after = form.cleaned_data['fuel_after']
        pdt.fuel_after_source = form.cleaned_data['fuel_after_source']
        pdt.failure_desc = form.cleaned_data['failure_desc']
        pdt.remarks = form.cleaned_data['remarks']
        pdt.reservation = reservation

        pdt.full_clean()
        pdt.save()

        # zamknięcie PDT
        pdt_close(pdt, request.user.fbouser, form.cleaned_data)

        # przejście do kolejnego formularza
        return HttpResponseRedirect(reverse('dispatcher'))

    return render(request, 'panel/mpdt_close.html', context)


@login_required()
def PanelFuelingCreate (request):

    form = PanelFuelingForm(request.POST or None, is_mobile=request.device['is_mobile'], user=request.user.fbouser)

    # wspólne zmienne kontekstowe
    context = {}
    context['form'] = form
    context['page_title'] = 'Nowe tankowanie'
    context['header_text'] = 'Nowe tankowanie bez PDT'
    context['type'] = 'info'
    context['submenu_template'] = 'panel/user_submenu.html'
    context['fbouser'] = request.user.fbouser

    # jeśli formularz został poprawnie wypełniony
    if form.is_valid():

        # nowa operacja tankowania
        fueling = LocalFueling(date=form.cleaned_data['date'], person=form.cleaned_data['person'],
                               aircraft=form.cleaned_data['aircraft'], fueltank=form.cleaned_data['fueltank'],
                               fuel_volume=form.cleaned_data['fuel_volume'], remarks=form.cleaned_data['remarks'])
        fueling.full_clean()
        fueling.save()

        # przejście do kolejnego formularza
        if request.device['is_mobile']:
            return HttpResponseRedirect(reverse('dispatcher'))
        else:
            return HttpResponseRedirect(reverse('fbo:user-info', args=[request.user.fbouser.pk]))

    if request.device['is_mobile']:
        return render(request, 'panel/mfueling.html', context)
    else:
        return render(request, 'panel/create_template.html', context)


@login_required()
def PanelSMSReportCreate (request):

    form_class = modelform_factory(SMSReport, fields=['report_date', 'person', 'privacy', 'description'],
                                   widgets={'report_date':AdminDateWidget, 'description':Textarea(attrs={'rows':10, 'cols':100})})

    form_class.base_fields['report_date'].initial = date.today()
    form_class.base_fields['person'].initial = request.user.fbouser.__str__()

    form = form_class(request.POST or None)

    # wspólne zmienne kontekstowe
    context = {}
    context['form'] = form
    context['page_title'] = 'Nowy raport SMS'
    context['header_text'] = 'Zgłoszenie dobrowolnego raportu SMS'
    context['type'] = 'reports'
    context['submenu_template'] = 'panel/user_submenu.html'
    context['fbouser'] = request.user.fbouser
    context['logged'] = request.user.fbouser

    # jeśli formularz został poprawnie wypełniony
    if form.is_valid():

        # nowy raport SMS
        smsreport = SMSReport(report_date=form.cleaned_data['report_date'], person=form.cleaned_data['person'],
                              privacy=form.cleaned_data['privacy'], description=form.cleaned_data['description'],
                              reported_by=request.user.fbouser)
        smsreport.full_clean()
        smsreport.save()

        # przejście do kolejnego formularza
        return HttpResponseRedirect(reverse('panel:smsreport-list', args=[request.user.fbouser.pk]))

    return render(request, 'panel/create_template.html', context)


@class_view_decorator(login_required())
class PanelSMSReportList (DetailView):
    model = FBOUser
    template_name = 'panel/list_template.html'

    # jeśli nie superuser to ustaw na zalogowanego
    def get_object(self, queryset=None):
        obj = super(PanelSMSReportList, self).get_object()
        if not self.request.user.is_staff:
            obj = self.request.user.fbouser
        return obj

    def get_context_data(self, **kwargs):
        context = super(PanelSMSReportList, self).get_context_data(**kwargs)
        context['page_title'] = 'Lista raportów'
        context['header_text'] = 'Lista dobrowolnych raportów SMS'
        context['empty_text'] = 'Brak dobrowolnych raportów SMS.'
        context['type'] = 'reports'
        context['submenu_template'] = 'panel/user_submenu.html'
        context['fbouser'] = self.object
        context['logged'] = self.request.user.fbouser

        report_list = SMSReport.objects.filter(reported_by=self.object.pk).order_by('-report_date', '-pk')

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Rejestracja raportu SMS', 'path': reverse("panel:smsreport-create")})
        context['local_menu'] = local_menu

        # Nagłówki tabeli
        header_list = []
        header_list.append({'header': 'Data\nzgłoszenia'})
        header_list.append({'header': 'Poufność'})
        header_list.append({'header': 'Treść\nzgłoszenia'})
        header_list.append({'header': 'Uwagi'})
        header_list.append({'header': '...'})
        context['header_list'] = header_list

        row_list = []
        for object in report_list:
            fields = []
            fields.append({'name': 'report_date', 'value': object.report_date, 'link': reverse('panel:smsreport-details', args=[object.pk]), 'just': 'center'})
            fields.append({'name': 'privacy', 'value': 'TAK' if object.privacy else 'NIE', 'just': 'center'})
            fields.append({'name': 'description', 'value': object.description})
            fields.append({'name': 'remarks', 'value': object.remarks})
            fields.append({'name': 'change', 'edit_link': reverse('panel:smsreport-update', args=[object.pk])})
            row_list.append({'fields': fields})
        context['row_list'] = row_list
        return context


@class_view_decorator(login_required())
class PanelSMSReportUpdate (UpdateView):
    model = SMSReport
    template_name = 'panel/update_template.html'

    def get_context_data(self, **kwargs):
        context = super(PanelSMSReportUpdate, self).get_context_data(**kwargs)
        context['page_title'] = 'Aktualizacja raportu'
        context['header_text'] = 'Aktualizacja dobrowolonego raportu SMS'
        context['type'] = 'reports'
        context['submenu_template'] = 'panel/user_submenu.html'
        context['fbouser'] = self.object.reported_by
        context['logged'] = self.request.user.fbouser

        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(SMSReport, fields=['privacy', 'description', 'remarks'],
                                       widgets={'description':Textarea(attrs={'rows':10, 'cols':100, 'readonly': True}),
                                                'remarks':Textarea(attrs={'rows':4, 'cols':100})})
        return form_class

    def get_success_url(self):
        return reverse('panel:smsreport-list', args=(self.object.reported_by.pk,))


@class_view_decorator(login_required())
class PanelSMSReportDetails (DetailView):
    model = SMSReport
    template_name = 'panel/details_template.html'

    def get_context_data(self, **kwargs):
        context = super(PanelSMSReportDetails, self).get_context_data(**kwargs)
        context['page_title'] = 'Raport SMS'
        context['header_text'] = 'Szczegóły dobrowolnego raportu SMS'
        context['type'] = 'reports'
        context['submenu_template'] = 'panel/user_submenu.html'
        context['fbouser'] = self.object.reported_by
        context['logged'] = self.request.user.fbouser

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Aktualizuj', 'path': reverse('panel:smsreport-update', args=[self.object.pk])})
        context['local_menu'] = local_menu

        field_list = []
        field_list.append({'header': 'Data zgłoszenia', 'value': self.object.report_date, 'bold': True})
        field_list.append({'header': 'Osoba zgłaszająca', 'value': self.object.person, 'bold': True})
        field_list.append({'header': 'Poufność', 'value': 'TAK' if self.object.privacy else 'NIE'})
        field_list.append({'header': 'Treść zgłoszenia', 'value': self.object.description})
        field_list.append({'header': 'Zagrożenie w rejestrze', 'value': self.object.smshazard})
        field_list.append({'header': 'Wnioski i zalecenia', 'value': self.object.findings})
        field_list.append({'header': 'Uwagi', 'value': self.object.remarks})

        context['field_list'] = field_list
        return context


@login_required()
def PanelSMSEventCreate (request):

    form_class = modelform_factory(SMSEvent,
                                   fields=['event_date', 'aircraft', 'pic', 'oper_type', 'event_type', 'description', 'pkbwl_ref', 'pkbwl_date'],
                                   widgets={'event_date':AdminDateWidget, 'pkbwl_date':AdminDateWidget, 'description':Textarea(attrs={'rows':10, 'cols':100})})

    aircraft_choices = [('', '---------')]
    aircraft_choices += [(aircraft.registration, '%s / %s' % (aircraft.registration, aircraft.type)) for aircraft in Aircraft.objects.all()]
    form_class.base_fields['aircraft'] = forms.ChoiceField(choices=aircraft_choices, label='Statek powietrzny')

    form_class.base_fields['event_date'].initial = date.today()
    form_class.base_fields['pic'].initial = request.user.fbouser.__str__()

    form = form_class(request.POST or None)

    # wspólne zmienne kontekstowe
    context = {}
    context['form'] = form
    context['page_title'] = 'Nowe zdarzenie SMS'
    context['header_text'] = 'Zgłoszenie zdarzenia lotniczego'
    context['type'] = 'events'
    context['submenu_template'] = 'panel/user_submenu.html'
    context['fbouser'] = request.user.fbouser
    context['logged'] = request.user.fbouser

    # jeśli formularz został poprawnie wypełniony
    if form.is_valid():

        # nowe zdarzenie lotnicze
        smsevent = SMSEvent(event_date=form.cleaned_data['event_date'], aircraft=form.cleaned_data['aircraft'],
                            pic=form.cleaned_data['pic'], oper_type=form.cleaned_data['oper_type'],
                            event_type=form.cleaned_data['event_type'], description=form.cleaned_data['description'],
                            pkbwl_ref=form.cleaned_data['pkbwl_ref'], pkbwl_date=form.cleaned_data['pkbwl_date'],
                            reported_by=request.user.fbouser)
        smsevent.full_clean()
        smsevent.save()

        # przejście do kolejnego formularza
        return HttpResponseRedirect(reverse('panel:smsevent-list', args=[request.user.fbouser.pk]))

    return render(request, 'panel/create_template.html', context)


@class_view_decorator(login_required())
class PanelSMSEventList (DetailView):
    model = FBOUser
    template_name = 'panel/list_template.html'

    # jeśli nie superuser to ustaw na zalogowanego
    def get_object(self, queryset=None):
        obj = super(PanelSMSEventList, self).get_object()
        if not self.request.user.is_staff:
            obj = self.request.user.fbouser
        return obj

    def get_context_data(self, **kwargs):
        context = super(PanelSMSEventList, self).get_context_data(**kwargs)
        context['page_title'] = 'Lista zdarzeń'
        context['header_text'] = 'Lista zgłoszonych zdarzeń lotniczych'
        context['empty_text'] = 'Brak zgłoszonych zdarzeń lotniczych.'
        context['type'] = 'events'
        context['submenu_template'] = 'panel/user_submenu.html'
        context['fbouser'] = self.object
        context['logged'] = self.request.user.fbouser

        event_list = SMSEvent.objects.filter(reported_by=self.object.pk).order_by('-event_date', '-pk')

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Rejestracja zdarzenia lotniczego', 'path': reverse("panel:smsevent-create")})
        context['local_menu'] = local_menu

        # Nagłówki tabeli
        header_list = []
        header_list.append({'header': 'Statek\npowietrzny'})
        header_list.append({'header': 'Pilot\ndowódca'})
        header_list.append({'header': 'Data\nzdarzenia'})
        header_list.append({'header': 'Kwalifikacja\nzdarzenia'})
        header_list.append({'header': 'Opis\nzdarzenia'})
        header_list.append({'header': 'Uwagi'})
        header_list.append({'header': '...'})
        context['header_list'] = header_list

        row_list = []
        for object in event_list:
            fields = []
            fields.append({'name': 'aircraft', 'value': object.aircraft, 'link': reverse('panel:smsevent-details', args=[object.pk]), 'just': 'center'})
            fields.append({'name': 'pic', 'value': object.pic})
            fields.append({'name': 'event_date', 'value': object.event_date, 'just': 'center'})
            fields.append({'name': 'event_type', 'value': object.event_type, 'just': 'center'})
            fields.append({'name': 'description', 'value': object.description})
            fields.append({'name': 'remarks', 'value': object.remarks})
            fields.append({'name': 'change', 'edit_link': reverse('panel:smsevent-update', args=[object.pk])})
            row_list.append({'fields': fields})
        context['row_list'] = row_list
        return context


@class_view_decorator(login_required())
class PanelSMSEventUpdate (UpdateView):
    model = SMSEvent
    template_name = 'panel/update_template.html'

    def get_context_data(self, **kwargs):
        context = super(PanelSMSEventUpdate, self).get_context_data(**kwargs)
        context['page_title'] = 'Aktualizacja zdarzenia'
        context['header_text'] = 'Aktualizacja zgłoszonego zdarzenia lotniczego'
        context['type'] = 'events'
        context['submenu_template'] = 'panel/user_submenu.html'
        context['fbouser'] = self.object.reported_by
        context['logged'] = self.request.user.fbouser

        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(SMSEvent, fields=['aircraft', 'pic', 'description', 'remarks'],
                                   widgets={'aircraft':TextInput(attrs={'readonly': True}),
                                            'pic':TextInput(attrs={'readonly': True}),
                                            'description':Textarea(attrs={'rows':10, 'cols':100, 'readonly': True}),
                                            'remarks':Textarea(attrs={'rows':4, 'cols':100})})
        return form_class

    def get_success_url(self):
        return reverse('panel:smsevent-list', args=(self.object.reported_by.pk,))


@class_view_decorator(login_required())
class PanelSMSEventDetails (DetailView):
    model = SMSEvent
    template_name = 'panel/details_template.html'

    def get_context_data(self, **kwargs):
        context = super(PanelSMSEventDetails, self).get_context_data(**kwargs)
        context['page_title'] = 'Zdarzenie lotnicze'
        context['header_text'] = 'Szczegóły zgłoszonego zdarzenia lotniczego'
        context['type'] = 'events'
        context['submenu_template'] = 'panel/user_submenu.html'
        context['fbouser'] = self.object.reported_by
        context['logged'] = self.request.user.fbouser

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Aktualizuj', 'path': reverse('panel:smsevent-update', args=[self.object.pk])})
        context['local_menu'] = local_menu

        field_list = []
        field_list.append({'header': 'Statek powietrzny', 'value': self.object.aircraft, 'bold': True})
        field_list.append({'header': 'Pilot dowódca', 'value': self.object.pic, 'bold': True})
        field_list.append({'header': 'Data zdarzenia', 'value': self.object.event_date, 'bold': True})
        field_list.append({'header': 'Rodzaj operacji', 'value': self.object.oper_type})
        field_list.append({'header': 'Kwalifikacja zdarzenia', 'value': self.object.event_type})
        field_list.append({'header': 'Opis zdarzenia', 'value': self.object.description})
        field_list.append({'header': 'Numer ewidencyjny PKBWL', 'value': self.object.pkbwl_ref})
        field_list.append({'header': 'Data przyjęcia w PKBWL', 'value': self.object.pkbwl_date})
        field_list.append({'header': 'Badający zdarzenie', 'value': self.object.examiner})
        field_list.append({'header': 'Przebieg badania wewnętrznego', 'value': self.object.scrutiny})
        field_list.append({'header': 'Zagrożenie w rejestrze', 'value': self.object.smshazard})
        field_list.append({'header': 'Wnioski i zalecenia', 'value': self.object.findings})
        field_list.append({'header': 'Zamknięcie / raport końcowy', 'value': self.object.closure})
        field_list.append({'header': 'Uwagi', 'value': self.object.remarks})
        context['field_list'] = field_list
        return context


@login_required()
def Report_PKBWL (request):

    # wspólne zmienne kontekstowe
    context = {}

    return render(request, 'panel/pkbwl.html', context)


@login_required()
def PilotDutyReport (request, pilot_id):

    class DutyReportForm (Form):
        duty_ym = ChoiceField(choices=[(ym.replace("-",""), ym) for ym in sorted(set([str(duty.date)[:7] for duty in Duty.objects.filter(pilot=pilot_id)]), reverse=True)],
                                       label='Miesiąc raportu')
        report_type = ChoiceField(choices=[(0, 'Wszystkie loty'), (1, 'Tylko AOC')], label='Zakres raportu')

    MyForm = DutyReportForm
    form = MyForm(request.POST or None)

    context = {}
    context['title'] = 'Raport wykonanych lotów dla %s' % Pilot.objects.get(pk=pilot_id)
    context['form'] = form

    # jeśli formularz został poprawnie wypełniony
    if form.is_valid():

        # przejście do kolejnego formularza
        return HttpResponseRedirect(reverse('panel:duty-export', args=[pilot_id, form.cleaned_data['duty_ym'], form.cleaned_data['report_type']]))

    return render(request, 'panel/file_export.html', context)


@login_required()
def PilotDutyExport (request, pilot_id, ym, type):

    # Nowy arkusz
    wb = Workbook()
    wb.active.title = 'Ewidencja czasu pracy'

    # Definicje stylów
    border = Border(top=Side(border_style=borders.BORDER_THIN, color='000000'), bottom=Side(border_style=borders.BORDER_THIN, color='000000'),
                    left=Side(border_style=borders.BORDER_THIN, color='000000'), right=Side(border_style=borders.BORDER_THIN, color='000000'))
    title_align = Alignment(horizontal='center', vertical='center', wrap_text = True)
    table_align = Alignment(vertical='center', wrap_text=True)
    gray_fill = PatternFill(fill_type=fills.FILL_SOLID, fgColor=('DDDDDD'))

    ws = wb.active

    # Informacje nagłówkowe
    ws['A2'] = "EWIDENCJA MIESIĘCZNA CZASU PRACY I WYPOCZYNKU"
    ws.merge_cells('A2:H2')
    ws['A3'] = "CZŁONKA ZAŁOGI STATKU POWIETRZNEGO"
    ws.merge_cells('A3:H3')
    ws['A4'] = "%s %s" % (Pilot.objects.get(pk=pilot_id), ym[:4]+'-'+ym[4:])
    ws.merge_cells('A4:H4')

    ws['A6'] = "Data"
    ws['B6'] = "Czas służby"
    ws['C6'] = "Czas czynności lotniczych"
    ws['D6'] = "Liczba lądowań"
    ws['E6'] = "Czas lotu"
    ws['F6'] = "Czas wypoczynku i dni wolnych"
    ws['G6'] = "Nr zlecenia na lot"
    ws['H6'] = "Charakter pracy"

    # Format nagłówków
    ws['A2'].alignment = title_align
    ws['A2'].font = Font(size=16, bold=True)
    ws['A3'].alignment = title_align
    ws['A3'].font = Font(size=16, bold=True)
    ws['A4'].alignment = title_align
    ws['A4'].font = Font(size=16, bold=True)

    for col in ['A', 'B', 'C', 'D', 'E', 'F','G','H']:
        ws['%s6' % col].alignment = title_align
        ws['%s6' % col].border = border
        ws['%s6' % col].font = Font(size=12, bold=True)
        ws['%s6' % col].fill = gray_fill

    # Zawartość zakładek
    row = 7
    last_duty_time_to = None
    m1_duty_time = m1_fdp_time = m1_block_time = timedelta(seconds=0)
    m1_landings = 0

    for (first_day, last_day) in [(1,7), (8,14), (15,21), (22,28), (29,31)]:

        d7_duty_time = d7_fdp_time = d7_block_time = timedelta(seconds=0)
        d7_landings = 0
        for duty in Duty.objects.filter(pilot=pilot_id).order_by('date', 'duty_time_from'):

            # Czasy wypoczynku
            this_duty_time_from = datetime.combine(duty.date, duty.duty_time_from) if duty.duty_time_from else None
            rest_time = (this_duty_time_from - last_duty_time_to) \
                        if this_duty_time_from and last_duty_time_to and this_duty_time_from > last_duty_time_to else None

            # Wypełnienie wierszy
            if str(duty.date)[:7].replace("-","") == ym and duty.date.day >= first_day and duty.date.day <= last_day:

                if type == '0' or (type == '1' and duty.duty_type[:2] == '01'):
                    ws['A%d' % row] = duty.date
                    ws['B%d' % row] = duration_string(duty.duty_time)
                    ws['C%d' % row] = duration_string(duty.fdp_time)
                    ws['D%d' % row] = duty.landings
                    ws['E%d' % row] = duration_string(duty.block_time)
                    if rest_time:
                        ws['F%d' % row] = '%d d. %d h. %d m.' % (rest_time.days, rest_time.seconds//3600, (rest_time.seconds//60)%60)
                    ws['G%d' % row] = ''
                    ws['H%d' % row] = duty.duty_type
                    d7_duty_time += duty.duty_time or timedelta(seconds=0)
                    d7_fdp_time += duty.fdp_time or timedelta(seconds=0)
                    d7_block_time += duty.block_time or timedelta(seconds=0)
                    d7_landings += duty.landings or 0
                    row += 1

                last_duty_time_to = datetime.combine(duty.date, duty.duty_time_to) if duty.duty_time_to else None

        ws['A%d' % row] = 'SUMA 7 dni'
        ws['B%d' % row] = duration_string(d7_duty_time)
        ws['C%d' % row] = duration_string(d7_fdp_time)
        ws['D%d' % row] = d7_landings
        ws['E%d' % row] = duration_string(d7_block_time)
        for col in ['A', 'B', 'C', 'D', 'E', 'F','G','H']:
            ws['%s%d' % (col,row)].fill = gray_fill
            ws['%s%d' % (col, row)].font = Font(bold=True)
        m1_duty_time += d7_duty_time
        m1_fdp_time += d7_fdp_time
        m1_block_time += d7_block_time
        m1_landings += d7_landings
        row += 1

    ws['A%d' % row] = 'SUMA miesiąc'
    ws['B%d' % row] = duration_string(m1_duty_time)
    ws['C%d' % row] = duration_string(m1_fdp_time)
    ws['D%d' % row] = m1_landings
    ws['E%d' % row] = duration_string(m1_block_time)
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws['%s%d' % (col, row)].fill = gray_fill
        ws['%s%d' % (col, row)].font = Font(bold=True)
    row += 1

    # Ustawienie szerokości kolumn
    ws.column_dimensions['A'].width = '15'
    ws.column_dimensions['B'].width = '15'
    ws.column_dimensions['C'].width = '15'
    ws.column_dimensions['D'].width = '15'
    ws.column_dimensions['E'].width = '10'
    ws.column_dimensions['F'].width = '18'
    ws.column_dimensions['G'].width = '12'
    ws.column_dimensions['H'].width = '25'

    # Formatowanie komórek tabeli
    max_row = row
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        for row in range(7, max_row):
            ws['%s%d' % (col, row)].border = border
            ws['%s%d' % (col, row)].alignment = table_align
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        for row in range(7, max_row):
            ws['%s%d' % (col, row)].alignment = title_align

    # Zwróć arkusz jako obiekt HTTP Response
    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = ('attachment; filename=duty_report_%s.xlsx' % date.today()).replace('-', '')

    return response


@login_required()
def PilotPlanReport (request, pilot_id):

    class PlanReportForm (Form):
        plan_from = DateField(initial=date.today(), label='Data początkowa')
        plan_to = DateField(initial=date.today()+timedelta(days=30), label='Data końcowa')
        report_type = ChoiceField(choices=[(0, 'Wszystkie loty'), (1, 'Tylko AOC')], label='Zakres raportu')

    MyForm = PlanReportForm
    form = MyForm(request.POST or None)

    context = {}
    context['title'] = 'Raport planowanych lotów dla %s' % Pilot.objects.get(pk=pilot_id)
    context['form'] = form

    # jeśli formularz został poprawnie wypełniony
    if form.is_valid():

        # przejście do kolejnego formularza
        return HttpResponseRedirect(reverse('panel:plan-export',
                                            args=[pilot_id, str(form.cleaned_data['plan_from']).replace('-',''),
                                                  str(form.cleaned_data['plan_to']).replace('-',''),
                                                  form.cleaned_data['report_type']]))

    return render(request, 'panel/file_export.html', context)


@login_required()
def PilotPlanExport (request, pilot_id, date_from, date_to, type):

    try:
        start = datetime(year=int(date_from[0:4]), month=int(date_from[4:6]), day=int(date_from[6:8]))
        end = datetime(year=int(date_to[0:4]), month=int(date_to[4:6]), day=int(date_to[6:8]))
        pilot = Pilot.objects.get(pk=pilot_id)
    except:
        pilot = start = end = None

    if start and end and pilot:
        # Nowy arkusz
        wb = Workbook()
        wb.active.title = 'Planowany czas pracy'

        # Definicje stylów
        border = Border(top=Side(border_style=borders.BORDER_THIN, color='000000'), bottom=Side(border_style=borders.BORDER_THIN, color='000000'),
                        left=Side(border_style=borders.BORDER_THIN, color='000000'), right=Side(border_style=borders.BORDER_THIN, color='000000'))
        title_align = Alignment(horizontal='center', vertical='center', wrap_text = True)
        table_align = Alignment(vertical='center', wrap_text=True)
        gray_fill = PatternFill(fill_type=fills.FILL_SOLID, fgColor=('DDDDDD'))

        ws = wb.active

        # Informacje nagłówkowe
        ws['A2'] = "EWIDENCJA PLANOWANYCH LOTÓW"
        ws.merge_cells('A2:E2')
        ws['A3'] = "CZŁONKA ZAŁOGI STATKU POWIETRZNEGO"
        ws.merge_cells('A3:E3')
        ws['A4'] = "%s  %s" % (Pilot.objects.get(pk=pilot_id), str(start)[:10]+' - '+str(end)[:10])
        ws.merge_cells('A4:E4')

        ws['A6'] = "Termin startu"
        ws['B6'] = "Termin lądowania"
        ws['C6'] = "Planowany czas lotu"
        ws['D6'] = "Planowany rodzaj lotu"
        ws['E6'] = "Planowana rola"

        # Format nagłówków
        ws['A2'].alignment = title_align
        ws['A2'].font = Font(size=16, bold=True)
        ws['A3'].alignment = title_align
        ws['A3'].font = Font(size=16, bold=True)
        ws['A4'].alignment = title_align
        ws['A4'].font = Font(size=16, bold=True)

        for col in ['A', 'B', 'C', 'D', 'E']:
            ws['%s6' % col].alignment = title_align
            ws['%s6' % col].border = border
            ws['%s6' % col].font = Font(size=12, bold=True)
            ws['%s6' % col].fill = gray_fill

        # Zawartość zakładek
        row = 7
        sum_duty_time = sum_block_time = timedelta(seconds=0)

        for res in Reservation.objects.filter(Q(owner=pilot) | Q(participant=pilot), start_time__gte = start, start_time__lte = end).order_by('start_time'):

            # Wypełnienie wierszy
            if type == '0' or (type == '1' and res.planned_type[:2] == '01'):
                ws['A%d' % row] = res.start_time.strftime('%Y-%m-%d  %H:%M')
                ws['B%d' % row] = res.end_time.strftime('%Y-%m-%d  %H:%M')
                ws['C%d' % row] = duration_string(res.planned_time)
                ws['D%d' % row] = FlightTypes()[res.planned_type]
                ws['E%d' % row] = 'PIC' if res.owner == pilot else 'SIC'

                sum_block_time += res.planned_time or timedelta(seconds=0)
                sum_duty_time += (res.end_time - res.start_time) or timedelta(seconds=0)
                row += 1

        # Formatowanie komórek tabeli
        max_row = row
        for col in ['A', 'B', 'C', 'D', 'E']:
            for row in range(7, max_row):
                ws['%s%d' % (col, row)].border = border
                ws['%s%d' % (col, row)].alignment = table_align
        for col in ['A', 'B', 'C', 'D', 'E']:
            for row in range(7, max_row):
                ws['%s%d' % (col, row)].alignment = title_align

        # Podsumowania
        row += 2
        ws['A%d' % row] = 'Suma czasu służby:'
        ws['B%d' % row] = duration_string(sum_duty_time)
        for col in ['A', 'B']:
            ws['%s%d' % (col,row)].fill = gray_fill
            ws['%s%d' % (col, row)].font = Font(size=13, bold=True)
        ws['B%d' % row].alignment = title_align
        row += 1
        ws['A%d' % row] = 'Suma czasu lotu:'
        ws['B%d' % row] = duration_string(sum_block_time)
        for col in ['A', 'B']:
            ws['%s%d' % (col,row)].fill = gray_fill
            ws['%s%d' % (col, row)].font = Font(size=13, bold=True)
        ws['B%d' % row].alignment = title_align

        # Ustawienie szerokości kolumn
        ws.column_dimensions['A'].width = '23'
        ws.column_dimensions['B'].width = '23'
        ws.column_dimensions['C'].width = '23'
        ws.column_dimensions['D'].width = '25'
        ws.column_dimensions['E'].width = '20'

        # Zwróć arkusz jako obiekt HTTP Response
        response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = ('attachment; filename=duty_plan_%s.xlsx' % date.today()).replace('-', '')

        return response

    else:
        return HttpResponseRedirect(reverse('panel:plan-report', args=[pilot_id]))
