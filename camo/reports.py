from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, fills, Border, borders, Side, Font
from openpyxl.writer.excel import save_virtual_workbook

from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, TableStyle, Table, KeepTogether
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet

from django.contrib.auth.decorators import permission_required
from django.http import HttpResponse

from camo.models import Aircraft


@permission_required('camo.camo_reader')
def ReportARC (request, aircraft_id):


    def cut_str(str, num):
        if str:
            if num > 3:
                return (str[:num-3] + '...') if len(str) > num else str
            else:
                return str
        else:
            return ''

    # Szablon strony raportu w PDF
    def ARC_Page(canvas, doc):

        canvas.saveState()
        header_style = ParagraphStyle('header', parent=styles['Normal'], fontName='Helvetica')
        footer_style = ParagraphStyle('footer', parent=styles['Normal'], fontName='Helvetica', fontSize=8)

        # Header
        header = Paragraph('<b>%s</b> - Raport ARC z dnia %s' % (aircraft.registration, date.today()), header_style)
        w, h = header.wrap(doc.width, doc.topMargin)
        header.drawOn(canvas, doc.leftMargin, doc.height+doc.bottomMargin+10)
        canvas.line(doc.leftMargin, doc.height+doc.bottomMargin+5, doc.width+doc.leftMargin, doc.height+doc.bottomMargin+5)

        # Footer
        footer = Paragraph('Strona %d' % doc.page, footer_style)
        w, h = footer.wrap(doc.width, doc.bottomMargin)
        footer.drawOn(canvas, doc.leftMargin, doc.bottomMargin-h-10)
        canvas.line(doc.leftMargin, doc.bottomMargin-5, doc.width+doc.leftMargin, doc.bottomMargin-5)

        canvas.restoreState()

    # Zawartość tabeli z obsługami (nr. 1)
    def OTH_Table(aircraft):
        data = [['POT ref.', 'Paczka zadań', 'Limit\nFH', 'Limit\nmsc.']]
        if aircraft.use_landings:
            data[0].append('Limit\nldg.')
        if aircraft.use_cycles:
            data[0].append('Limit\ncycl.')
        data[0] += ['Wykon.\ndata', 'Wykon.\nTTH']
        if aircraft.use_landings:
            data[0].append('Wykon.\nldg.')
        if aircraft.use_cycles:
            data[0].append('Wykon.\ncycl.')
        data[0] += ['Wykon.\nCRS', 'Nast.\ndata', 'Nast.\nTTH']
        if aircraft.use_landings:
            data[0].append('Nast\nldg.')
        if aircraft.use_cycles:
            data[0].append('Nast\ncycl.')
        data[0] += ['Pozost.\nFH', 'Pozost.\ndni']
        if aircraft.use_landings:
            data[0].append('Pozost.\nldg.')
        if aircraft.use_cycles:
            data[0].append('Pozost.\ncycl.')

        for group in oth_list:
            row = [group.POT_ref, cut_str(group.name, 50), group.due_hours, group.due_months]
            if aircraft.use_landings:
                row.append(group.due_landings)
            if aircraft.use_cycles:
                row.append(group.due_cycles)
            row += [group.done_date, group.done_hours]
            if aircraft.use_landings:
                row.append(group.done_landings)
            if aircraft.use_cycles:
                row.append(group.done_cycles)
            row += [group.done_crs, group.next_date(), group.next_hours()]
            if aircraft.use_landings:
                row.append(group.next_landings())
            if aircraft.use_cycles:
                row.append(group.next_cycles())
            row += [group.left_hours(), group.left_days()]
            if aircraft.use_landings:
                row.append(group.left_landings())
            if aircraft.use_cycles:
                row.append(group.left_cycles())

            data.append(row)

        return data

    # Zawartość tabeli z LLP (nr. 2)
    def LLP_Table(aircraft):
        data = [['POT ref.', 'Nazwa', 'P/N', 'S/N', 'Limit\nFH', 'Limit\nmsc.']]
        if aircraft.use_landings:
            data[0].append('Limit\nldg.')
        if aircraft.use_cycles:
            data[0].append('Limit\ncycl.')
        data[0] += ['Instal.\ndata', 'Instal.\nTTH', 'Instal.\nCRS', 'Data\nprodukcji',
                    'TTH\npoczątkowe', 'Data\nżywotności']
        data[0] += ['Pozost.\nFH', 'Pozost.\ndni']
        if aircraft.use_landings:
            data[0].append('Pozost.\nldg.')
        if aircraft.use_cycles:
            data[0].append('Pozost.\ncycl.')

        for group in llp_list:
            row = [group.POT_ref, cut_str(group.part.name, 45), group.part.part_no, group.part.serial_no, group.due_hours, group.due_months]
            if aircraft.use_landings:
                row.append(group.due_landings)
            if aircraft.use_cycles:
                row.append(group.due_cycles)
            row += [group.part.current_assignment().from_date,
                    group.part.current_assignment().from_hours, group.part.current_assignment().crs,
                    group.part.production_date, group.part.current_assignment().install_tth(), group.next_date()]
            row += [group.left_hours(), group.left_days()]
            if aircraft.use_landings:
                row.append(group.left_landings())
            if aircraft.use_cycles:
                row.append(group.left_cycles())

            data.append(row)

        return data

    # Zawartość tabeli z remontami (nr. 3)
    def OVH_Table(aircraft):
        data = [['POT ref.', 'Nazwa', 'P/N', 'S/N', 'Limit\nFH', 'Limit\nmsc.']]
        if aircraft.use_landings:
            data[0].append('Limit\nldg.')
        if aircraft.use_cycles:
            data[0].append('Limit\ncycl.')
        data[0] += ['Instal.\ndata', 'Instal.\nTTH', 'Instal.\nCRS', 'Data\nprodukcji',
                    'TTH\npoczątkowe', 'Data\nżywotności']
        data[0] += ['Pozost.\nFH', 'Pozost.\ndni']
        if aircraft.use_landings:
            data[0].append('Pozost.\nldg.')
        if aircraft.use_cycles:
            data[0].append('Pozost.\ncycl.')

        for group in ovh_list:
            row = [group.POT_ref, cut_str(group.part.name, 45), group.part.part_no, group.part.serial_no, group.due_hours, group.due_months]
            if aircraft.use_landings:
                row.append(group.due_landings)
            if aircraft.use_cycles:
                row.append(group.due_cycles)
            row += [group.part.current_assignment().from_date,
                    group.part.current_assignment().from_hours, group.part.current_assignment().crs,
                    group.part.production_date, group.part.current_assignment().install_tth(), group.next_date()]
            row += [group.left_hours(), group.left_days()]
            if aircraft.use_landings:
                row.append(group.left_landings())
            if aircraft.use_cycles:
                row.append(group.left_cycles())

            data.append(row)

        return data

    # Zawartość tabeli z ADSB (nr. 4)
    def ADSB_Table(aircraft):

        story = []
        for part in part_list:

            data = [['Typ', 'Numer', 'Treść', 'Organ', 'Data', 'Limit\nTTH', 'Limit\nmsc.']]
            if aircraft.use_landings:
                data[0].append('Limit\nldg.')
            if aircraft.use_cycles:
                data[0].append('Limit\ncycl.')
            data[0] += ['Cykl', 'Wykon.\ndnia', 'Wykon.\nTTH']
            if aircraft.use_landings:
                data[0].append('Wykon.\nldg.')
            if aircraft.use_cycles:
                data[0].append('Wykon.\ncycl.')
            data[0] += ['Wykon\nCRS', 'Pozost.\nFH', 'Pozost.\ndni']
            if aircraft.use_landings:
                data[0].append('Pozost.\nldg.')
            if aircraft.use_cycles:
                data[0].append('Pozost.\ncycl.')

            for group in part[2]:
                row = [group.type.upper(), group.adsb_no, cut_str(group.name, 40), group.adsb_agency, group.adsb_date, group.due_hours, group.due_months]
                if aircraft.use_landings:
                    row.append(group.due_landings)
                if aircraft.use_cycles:
                    row.append(group.due_cycles)
                row += ['TAK' if group.cyclic else 'NIE', group.done_date, group.done_hours]
                if aircraft.use_landings:
                    row.append(group.done_landings)
                if aircraft.use_cycles:
                    row.append(group.done_cycles)
                row += [group.done_crs, group.left_hours(), group.left_days()]
                if aircraft.use_landings:
                    row.append(group.left_landings())
                if aircraft.use_cycles:
                    row.append(group.left_cycles())

                data.append(row)

            if part[2]:
                story += [Paragraph('%s' % part[0], style=part_style)]
                story += [Table(data, repeatRows=1, style=ADSBTableStyle, hAlign='LEFT')]

        return story

    # Zawartość tabeli z modyfikacjami (nr. 5)
    def MOD_Table(aircraft):
        data = [['Opis', 'Wykon.\ndata', 'Wykon.\nTTH']]
        if aircraft.use_landings:
            data[0].append('Wykon.\nldg.')
        if aircraft.use_cycles:
            data[0].append('Wykon.\ncycl.')
        data[0] += ['ASO', 'CRS']

        for mod in mod_list:
            row = [cut_str(mod.description, 60), mod.done_date, mod.done_hours]
            if aircraft.use_landings:
                row.append(mod.done_landings)
            if aircraft.use_cycles:
                row.append(mod.done_cycles)
            row += [mod.aso, mod.done_crs]

            data.append(row)

        return data

    # Zawartość tabeli z raportami WB (nr. 6)
    def WB_Table(aircraft):
        data = [['Opis', 'Doc ref.', 'Zmiana\nmasy', 'Całk.\nzm. masy', 'Zmiana\n% MTOW', 'Masa\npustego',
                 'Lon.\nC.G.', 'Lat.\nC.G.','Lon.\nmoment', 'Lat.\nmoment', 'Wykon.\ndata', 'Wykon.\nTTH']]
        if aircraft.use_landings:
            data[0].append('Wykon.\nldg.')
        if aircraft.use_cycles:
            data[0].append('Wykon.\ncycl.')
        data[0] += ['ASO', 'Dokument']

        for wb in wb_list:
            row = [cut_str(wb['description'], 50), wb['doc_ref'], wb['mass_change'], wb['tot_mass_change'], wb['pct_mass_change'], wb['empty_weight'],
                   wb['lon_cg'], wb['lat_cg'], wb['lon_moment'], wb['lat_moment'], wb['done_date'], wb['done_hours']]
            if aircraft.use_landings:
                row.append(wb['done_landings'])
            if aircraft.use_cycles:
                row.append(wb['done_cycles'])
            row += [wb['aso'], wb['done_doc']]

            data.append(row)

        return data

    aircraft = Aircraft.objects.get(pk=aircraft_id)

    # Wybierz grupy czynności do raportu
    groups = []
    for part in aircraft.components():
        for group in part.pot_group_set.all():
            groups.append(group)

    # Posortuj po ilości pozostałych godzin/dni
    # groups = sorted(groups, key=lambda group: group.leftx())

    oth_list = [group for group in groups if group.type=='oth']
    llp_list = [group for group in groups if group.type=='llp']
    ovh_list = [group for group in groups if group.type=='ovh']

    # Utwórz listę części
    assignments = aircraft.assignment_set.filter(current=True).order_by('ata__chapter')
    part_list = []
    ind = 0
    for assignment in assignments:
        if not assignment.super_ass:
            ind = ind + 1
            part_list.append([assignment.part, "%d" % ind])
            sub = 0
            for child in assignments:
                if child.super_ass == assignment:
                    sub = sub + 1
                    part_list.append([child.part, '%d.%02d' % (ind, sub)])

    # Przpisz do części dyrektywy i biuletyny
    for part in part_list:
        part.append([group for group in groups if group.part==part[0] and group.type in ('ad', 'sb')])

    mod_list = aircraft.modification_set.order_by('-done_date')

    wb_list = aircraft.wb_report_set.order_by('done_date').values()
    mc = 0
    for wb in wb_list:
        mc += wb['mass_change']
        wb['tot_mass_change'] = mc
        if aircraft.mtow != 0:
            wb['pct_mass_change'] = '%.2f' % (100 * (mc / aircraft.mtow))
        else:
            wb['pct_mass_change'] = None
        if wb['lon_cg']:
            wb['lon_moment'] = '%.2f' % (wb['empty_weight']*wb['lon_cg'])
        else:
            wb['lon_moment'] = None
        if wb['lat_cg']:
            wb['lat_moment'] = '%.2f' % (wb['empty_weight']*wb['lat_cg'])
        else:
            wb['lat_moment'] = None

    # Przygotowanie obiektu odpowiedzi jako załącznika
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="%s ARC %s.pdf"' % (aircraft, date.today())

    # Utworzenie obiektu PDF
    doc = SimpleDocTemplate(response,
                            rightMargin=inch/4,
                            leftMargin=inch/4,
                            topMargin=inch/2,
                            bottomMargin=inch/2,
                            pagesize=A4)

    # Dodanie czcionki z polskimi znakami
    pdfmetrics.registerFont(TTFont('Verdana', 'Verdana.ttf'))

    # Utworzenie styli
    styles = getSampleStyleSheet()
    heading_style = ParagraphStyle('myHeading', parent=styles['Heading2'], fontName='Verdana', fontSize=11)
    info_style = ParagraphStyle('myInfo', parent=styles['Normal'], fontName='Verdana', fontSize=8)
    part_style = ParagraphStyle('myHeading', parent=styles['Heading2'], fontName='Verdana', fontSize=7, leading=7)

    # Style dla tabel
    OTHTableStyle = TableStyle([('FONT',(0,0),(-1,-1), 'Verdana', 5),
                                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                                ('ALIGN', (2,0), (-1,-1), 'CENTER'),
                                ('VALIGN', (0,0), (-1,0), 'MIDDLE'),
                                ('LEFTPADDING', (0,0), (-1,-1), 4),
                                ('RIGHTPADDING', (0,0), (-1,-1), 4),
                                ('GRID', (0,0), (-1,-1), 0.25, colors.black),
                                ('BOX', (0,0), (-1,-1), 0.25, colors.black),
                                ('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
                                ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
                                ])
    LLPTableStyle = TableStyle([('FONT',(0,0),(-1,-1), 'Verdana', 5),
                                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                                ('ALIGN', (5,0), (-1,-1), 'CENTER'),
                                ('VALIGN', (0,0), (-1,0), 'MIDDLE'),
                                ('LEFTPADDING', (0,0), (-1,-1), 4),
                                ('RIGHTPADDING', (0,0), (-1,-1), 4),
                                ('GRID', (0,0), (-1,-1), 0.25, colors.black),
                                ('BOX', (0,0), (-1,-1), 0.25, colors.black),
                                ('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
                                ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
                                ])
    OVHTableStyle = TableStyle([('FONT',(0,0),(-1,-1), 'Verdana', 5),
                                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                                ('ALIGN', (5,0), (-1,-1), 'CENTER'),
                                ('VALIGN', (0,0), (-1,0), 'MIDDLE'),
                                ('LEFTPADDING', (0,0), (-1,-1), 4),
                                ('RIGHTPADDING', (0,0), (-1,-1), 4),
                                ('GRID', (0,0), (-1,-1), 0.25, colors.black),
                                ('BOX', (0,0), (-1,-1), 0.25, colors.black),
                                ('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
                                ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
                                ])
    ADSBTableStyle = TableStyle([('FONT',(0,0),(-1,-1), 'Verdana', 5),
                                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                                ('ALIGN', (4,0), (-1,-1), 'CENTER'),
                                ('VALIGN', (0,0), (-1,0), 'MIDDLE'),
                                ('LEFTPADDING', (0,0), (-1,-1), 4),
                                ('RIGHTPADDING', (0,0), (-1,-1), 4),
                                ('GRID', (0,0), (-1,-1), 0.25, colors.black),
                                ('BOX', (0,0), (-1,-1), 0.25, colors.black),
                                ('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
                                ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
                                ])
    MODTableStyle = TableStyle([('FONT',(0,0),(-1,-1), 'Verdana', 5),
                                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                                ('ALIGN', (1,0), (-3,-1), 'CENTER'),
                                ('VALIGN', (0,0), (-1,0), 'MIDDLE'),
                                ('LEFTPADDING', (0,0), (-1,-1), 4),
                                ('RIGHTPADDING', (0,0), (-1,-1), 4),
                                ('GRID', (0,0), (-1,-1), 0.25, colors.black),
                                ('BOX', (0,0), (-1,-1), 0.25, colors.black),
                                ('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
                                ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
                                ])
    WBTableStyle = TableStyle([('FONT',(0,0),(-1,-1), 'Verdana', 5),
                                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                                ('ALIGN', (2,0), (-3,-1), 'CENTER'),
                                ('VALIGN', (0,0), (-1,0), 'MIDDLE'),
                                ('LEFTPADDING', (0,0), (-1,-1), 4),
                                ('RIGHTPADDING', (0,0), (-1,-1), 4),
                                ('GRID', (0,0), (-1,-1), 0.25, colors.black),
                                ('BOX', (0,0), (-1,-1), 0.25, colors.black),
                                ('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
                                ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
                                ])

    # Utworzenie zawartości dokumentu
    story =  [Paragraph('1. Status zgodności z programem obsługi technicznej', style=heading_style)]
    if oth_list:
        story += [Table(OTH_Table(aircraft), repeatRows=1, style=OTHTableStyle, hAlign='LEFT')]
    else:
        story += [Paragraph('Brak zdefiniowanych czynności obsługowych.', style=info_style)]
    story += [Paragraph('2. Status podzespołów o ograniczonej żywotności', style=heading_style)]
    if llp_list:
        story += [Table(LLP_Table(aircraft), repeatRows=1, style=LLPTableStyle, hAlign='LEFT')]
    else:
        story += [Paragraph('Brak zdefiniowanych czynności obsługowych.', style=info_style)]
    story += [Paragraph('3. Status podzespołów o ograniczonej żywotności podlegających remontowi', style=heading_style)]
    if ovh_list:
        story += [Table(OVH_Table(aircraft), repeatRows=1, style=OVHTableStyle, hAlign='LEFT')]
    else:
        story += [Paragraph('Brak zdefiniowanych czynności obsługowych.', style=info_style)]
    story += [Paragraph('4. Status dyrektyw i biuletynów', style=heading_style)]
    story += ADSB_Table(aircraft)
    story += [Paragraph('5. Status wykonania modyfikacji i napraw', style=heading_style)]
    if mod_list:
        story += [Table(MOD_Table(aircraft), repeatRows=1, style=MODTableStyle, hAlign='LEFT')]
    else:
        story += [Paragraph('Brak wykonanych modyfikacji lub napraw.', style=info_style)]
    story += [Paragraph('6. Raporty z ważenia i wyważenia', style=heading_style)]
    if wb_list:
        story += [Table(WB_Table(aircraft), repeatRows=1, style=WBTableStyle, hAlign='LEFT')]
    else:
        story += [Paragraph('Brak raportów z ważenia i wyważenia.', style=info_style)]

    # Wygenerowanie dokumentu PDF
    doc.build(story, onFirstPage=ARC_Page, onLaterPages=ARC_Page)

    return response


@permission_required('camo.camo_reader')
def ReportParts (request, aircraft_id):

    aircraft = Aircraft.objects.get(pk=aircraft_id)

    # Nowy arkusz
    wb = Workbook()
    ws = wb.active
    ws.title = aircraft.registration

    # Definicje stylów
    border = Border(top=Side(border_style=borders.BORDER_THIN, color='000000'), bottom=Side(border_style=borders.BORDER_THIN, color='000000'),
                    left=Side(border_style=borders.BORDER_THIN, color='000000'), right=Side(border_style=borders.BORDER_THIN, color='000000'))
    title_align = Alignment(horizontal='center', vertical='center')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text = True)
    title_fill = PatternFill(fill_type=fills.FILL_SOLID, fgColor=('99CCFF'))

    # Informacje nagłówkowe
    ws['A1'] = "Ind."
    ws['B1'] = "Nazwa"
    ws['C1'] = "Producent"
    ws['D1'] = "Serial no."
    ws['E1'] = "Part no."
    ws['F1'] = "Form 1"
    ws['G1'] = "Data produkcji"
    ws['H1'] = "Licznik TTH"

    # Format nagłówków
    for col in ['A', 'B', 'C', 'D', 'E', 'F','G','H']:
        ws['%s1' % col].fill = title_fill
        ws['%s1' % col].alignment = title_align
        ws['%s1' % col].border = border

    # Utwórz listę aktualnych przypisań
    assignments = aircraft.assignment_set.filter(current=True).order_by('ata__chapter')
    assignment_list = []
    ind = 0
    for assignment in assignments:
        if not assignment.super_ass:
            ind = ind + 1
            assignment_list.append((assignment.part, "%d" % ind, False))
            sub = 0
            for child in assignments:
                if child.super_ass == assignment:
                    sub = sub + 1
                    assignment_list.append((child.part, '%d.%02d' % (ind, sub), True))

    # Zawartość zakładek
    row = 2
    for part, index, child in assignment_list:

        # Wypełnienie wiersza jedynie jeśli istnieje ostatnia wersja
        ws['A%d' % row] = index
        if not child:
            ws['B%d' % row] = part.name
        else:
            ws['B%d' % row] = '└─ %s' % part.name
        ws['C%d' % row] = part.maker
        ws['D%d' % row] = part.serial_no
        ws['E%d' % row] = part.part_no
        ws['F%d' % row] = part.f1
        ws['G%d' % row] = part.production_date
        ws['H%d' % row] = part.hours_count

        row += 1

    # Ustawienie szerokości kolumn
    ws.column_dimensions['A'].width = '6'
    ws.column_dimensions['B'].width = '40'
    ws.column_dimensions['C'].width = '20'
    ws.column_dimensions['D'].width = '25'
    ws.column_dimensions['E'].width = '25'
    ws.column_dimensions['F'].width = '25'
    ws.column_dimensions['G'].width = '15'
    ws.column_dimensions['H'].width = '12'

    # Formatowanie komórek tabeli
    max_row = row
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        for row in range(2, max_row):
            ws['%s%d' % (col, row)].border = border

    for col in ['A', 'G']:
        for row in range(2, max_row):
            ws['%s%d' % (col, row)].alignment = center_align

    # Zwróć arkusz jako obiekt HTTP Response
    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = ('attachment; filename=%s_parts.xlsx' % aircraft).replace('-', '')

    return response