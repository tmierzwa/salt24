# -*- coding: utf-8 -*-

import string, random, re, calendar
import logging
from django.conf import settings
from decimal import Decimal
from datetime import datetime, date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, fills, Border, borders, Side, Font
from openpyxl.writer.excel import save_virtual_workbook
from django.core.urlresolvers import reverse
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.views.generic.list import ListView
from django.views.generic.detail import DetailView
from django.forms.models import modelform_factory
from django.forms import Form, DecimalField, Textarea, TextInput, ChoiceField, DateField
from django.contrib.admin.widgets import AdminDateWidget
from django.core.exceptions import ValidationError
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render
from django.contrib.auth.decorators import permission_required, login_required
from django_datatables_view.base_datatable_view import BaseDatatableView
from django.db.models import Value, Q
from django.db.models.functions import Concat

from django.utils.decorators import method_decorator
def class_view_decorator(function_decorator):

    def simple_decorator(View):
        View.dispatch = method_decorator(function_decorator)(View.dispatch)
        return View

    return simple_decorator

from fin.forms import PackageBuyForm
from fin.models import FuelTank, FuelDelivery, FuelTransfer, FuelCorrection, LocalFueling, PDTFueling, RemoteFueling
from fin.models import Voucher, RentPackage, Contractor, BalanceOperation, SoldPackage, SpecialPrice
from camo.models import Aircraft
from panel.models import PDT, Operation

logger = logging.getLogger(__name__)

@class_view_decorator(login_required())
class PDT_feed (BaseDatatableView):

    model = PDT
    max_display_length = 100
    order_columns = ['aircraft', 'pdt_ref', 'date', 'flight_type', 'pic', 'sic', 'hours', 'tth', 'landings']

    def get_initial_queryset(self):
        return PDT.objects.order_by('date', 'tth_start', 'pdt_ref', 'open_time').reverse()

    def filter_queryset(self, qs):
        search = self.request.GET.get('search[value]', None)
        if search:
            qs = qs.filter(Q(aircraft__registration__icontains=search)|
                           Q(pdt_ref__icontains=search)|
                           Q(pic__fbouser__first_name__icontains=search)|
                           Q(pic__fbouser__second_name__icontains=search) |
                           Q(sic__fbouser__first_name__icontains=search)|
                           Q(sic__fbouser__second_name__icontains=search))

        search = self.request.GET.get('columns[1][search][value]', None)
        if search:
            try:
                search_min = int(search.split("^")[0])
            except:
                search_min = None

            try:
                search_max = int(search.split("^")[1])
            except:
                search_max = None

            if (search_min and search_max):
                if (search_min <= search_max):
                    qs = qs.filter(Q(pdt_ref__gte=search_min) & Q(pdt_ref__lte=search_max))
            elif (search_min and not search_max):
                qs = qs.filter(pdt_ref__gte=search_min)
            elif (search_max):
                qs = qs.filter(pdt_ref__lte=search_max)

        search = self.request.GET.get('columns[0][search][value]', None)
        if search:
            qs = qs.filter(aircraft__registration__icontains=search)

        search = self.request.GET.get('columns[1][search][value]', None)
        if search:
            try:
                search_min = int(search.split("^")[0])
            except:
                search_min = None

            try:
                search_max = int(search.split("^")[1])
            except:
                search_max = None

            if (search_min and search_max):
                if (search_min<=search_max):
                    qs = qs.filter(Q(pdt_ref__gte=search_min) & Q(pdt_ref__lte=search_max))
            elif (search_min and not search_max):
                qs = qs.filter(pdt_ref__gte=search_min)
            elif (search_max):
                qs = qs.filter(pdt_ref__lte=search_max)

        search = self.request.GET.get('columns[2][search][value]', None)
        if search:
            pattern_year = re.compile("^\d{4}$")
            pattern_month = re.compile("^\d{4}\-\d{2}$")

            try:
                search_min = search.split("^")[0]
                if pattern_year.match(search_min):
                    search_min += "-01-01"
                elif pattern_month.match(search_min):
                    search_min += "-01"
                search_min = datetime.strptime(search_min, "%Y-%m-%d").date()
            except:
                search_min = None

            try:
                search_max = search.split("^")[1]
                if pattern_year.match(search_max):
                    search_max += "-12-31"
                elif pattern_month.match(search_max):
                    last_day = calendar.monthrange(int(search_max[0:4]), int(search_max[5:7]))
                    search_max += "-%02d" % last_day[1]
                search_max = datetime.strptime(search_max, "%Y-%m-%d").date()
            except:
                search_max = None

            if (search_min and search_max):
                if (search_min <= search_max):
                    qs = qs.filter(Q(date__gte=search_min) & Q(date__lte=search_max))
            elif (search_min and not search_max):
                qs = qs.filter(date__gte=search_min)
            elif (search_max):
                qs = qs.filter(date__lte=search_max)

        search = self.request.GET.get('columns[3][search][value]', None)
        if search:
            qs = qs.filter(flight_type__icontains=search)

        search = self.request.GET.get('columns[4][search][value]', None)
        if search:
            qs = qs.annotate(full_name=Concat('pic__fbouser__user__last_name', Value(' '), 'pic__fbouser__user__first_name')).\
                    filter(full_name__icontains=search)

        search = self.request.GET.get('columns[5][search][value]', None)
        if search:
            qs = qs.annotate(full_name=Concat('sic__fbouser__user__last_name', Value(' '), 'sic__fbouser__user__first_name')).\
                    filter(full_name__icontains=search)
        return qs

    def prepare_results(self, qs):
        json_data = []
        for item in qs:
            hours = item.hours_sum()
            tth = item.tth_sum()
            json_data.append({
                'aircraft':item.aircraft.registration,
                'pdt_ref':'{:0>6d}'.format(item.pdt_ref),
                'info_link': reverse('panel:pdt-info', args=[item.pk]),
                'date':item.date,
                'flight_type':item.flight_type_name(),
                'pic':item.pic.__str__(),
                'sic':(item.sic.__str__() if item.sic else ''),
                'hours':'%02d:%02d' % (hours[1], hours[2]),
                'tth':'%.2f' % tth
            })

        return json_data

    def prepare_summary(self, qs):
        json_data = []
        sum_minutes = sum_tth = 0
        for item in qs:
            sum_minutes += item.hours_sum()[0]
            sum_tth += item.tth_sum()
        json_data.append({
             'sum_minutes': sum_minutes,
             'sum_tth': sum_tth
        })
        return json_data

    def get_context_data(self, *args, **kwargs):
        try:
            self.initialize(*args, **kwargs)

            qs = self.get_initial_queryset()

            # number of records before filtering
            total_records = qs.count()

            qs = self.filter_queryset(qs)

            # number of records after filtering
            total_display_records = qs.count()

            # summary after filtering
            summary = self.prepare_summary(qs)

            qs = self.ordering(qs)
            qs = self.paging(qs)

            # prepare output data
            if self.pre_camel_case_notation:
                aaData = self.prepare_results(qs)

                ret = {'sEcho': int(self._querydict.get('sEcho', 0)),
                       'iTotalRecords': total_records,
                       'iTotalDisplayRecords': total_display_records,
                       'aaSummary': summary,
                       'aaData': aaData
                       }
            else:
                data = self.prepare_results(qs)

                ret = {'draw': int(self._querydict.get('draw', 0)),
                       'recordsTotal': total_records,
                       'recordsFiltered': total_display_records,
                       'summary': summary,
                       'data': data
                       }
        except Exception as e:
            logger.exception(str(e))

            if settings.DEBUG:
                import sys
                from django.views.debug import ExceptionReporter
                reporter = ExceptionReporter(None, *sys.exc_info())
                text = "\n" + reporter.get_traceback_text()
            else:
                text = "\nAn error occured while processing an AJAX request."

            if self.pre_camel_case_notation:
                ret = {'sError': text,
                       'text': text,
                       'aaData': [],
                       'sEcho': int(self._querydict.get('sEcho', 0)),
                       'iTotalRecords': 0,
                       'iTotalDisplayRecords': 0, }
            else:
                ret = {'error': text,
                       'data': [],
                       'recordsTotal': 0,
                       'recordsFiltered': 0,
                       'draw': int(self._querydict.get('draw', 0))}
        return ret



@class_view_decorator(permission_required('fin.fin_reader'))
class PDTList (ListView):
    model = PDT
    template_name = 'fin/search_template.html'

    # Posortuj listę PDT po datach
    def get_queryset(self):
        return PDT.objects.order_by('date', 'tth_start', 'pdt_ref', 'open_time').reverse()[:100]

    def get_context_data(self, **kwargs):
        context = super(PDTList, self).get_context_data(**kwargs)
        context['page_title'] = 'Lista PDT'
        context['header_text'] = 'Lista PDT'
        context['empty_text'] = 'Brak zarejestrowanych PDT.'

        # Lokalne menu
        local_menu = []
        context['local_menu'] = local_menu

        # Nagłówki tabeli
        header_list = []
        header_list.append({'header': 'Statek\npowietrzny'})
        header_list.append({'header': 'PDT ref'})
        header_list.append({'header': 'Data'})
        header_list.append({'header': 'Rodzaj lotu'})
        header_list.append({'header': 'PIC'})
        header_list.append({'header': 'SIC'})
        header_list.append({'header': 'Czas'})
        header_list.append({'header': 'TTH'})
        context['header_list'] = header_list

        return context


def fuel_name(fuel_type):
    fuels ={'AVGAS': 'Avgas 100LL', 'MOGAS': 'Benzyna samochodowa', 'JETA1': 'Paliwo JET A-1'}
    if fuel_type in fuels:
        return fuels[fuel_type]
    else:
        return 'Nieznane'


@class_view_decorator(permission_required('fin.fin_reader'))
class FuelTanktList (ListView):
    model = FuelTank
    template_name = 'fin/list_template.html'

    def get_context_data(self, **kwargs):
        context = super(FuelTanktList, self).get_context_data(**kwargs)
        context['page_title'] = 'Magazyn paliwa'
        context['header_text'] = 'Lista zbiorników paliwa'
        context['empty_text'] = 'Brak zbiorników paliwa.'

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Nowy zbiornik paliwa', 'path': reverse("fin:tank-create")})
        local_menu.append({'text': 'Raport paliwa', 'path': reverse("fin:fuel-report")})
        context['local_menu'] = local_menu

        # Nagłówki tabeli
        header_list = []
        header_list.append({'header': 'Symbol\nzbiornika'})
        header_list.append({'header': 'Nazwa\nzbiornika'})
        header_list.append({'header': 'Rodzaj\npaliwa'})
        header_list.append({'header': 'Ilość\npaliwa'})
        header_list.append({'header': 'Szacowana\nwartość'})
        header_list.append({'header': 'Uwagi'})
        header_list.append({'header': '...'})
        context['header_list'] = header_list

        row_list = []
        for object in self.object_list:
            fields = []
            fields.append({'name': 'tank_ref', 'value': object.tank_ref, 'link': reverse('fin:tank-info', args=[object.pk]), 'just': 'center'})
            fields.append({'name': 'name', 'value': object.name})
            fields.append({'name': 'fuel_type', 'value': fuel_name(object.fuel_type)})
            fields.append({'name': 'fuel_volume', 'value': '%.1f L' % object.fuel_volume, 'just': 'center'})
            fields.append({'name': 'fuel_value', 'value': '%.2f PLN' % object.fuel_value(), 'just': 'center'})
            fields.append({'name': 'remarks', 'value': object.remarks})
            fields.append({'name': 'change', 'edit_link': reverse('fin:tank-update', args=[object.pk]),
                           'delete_link': (reverse('fin:tank-delete', args=[object.pk]) if object.fuel_volume == 0 else None)})
            row_list.append({'fields': fields})
        context['row_list'] = row_list
        return context


@class_view_decorator(permission_required('fin.fin_reader'))
class FuelTankInfo (DetailView):
    model = FuelTank
    template_name = 'fin/details_template.html'

    def get_context_data(self, **kwargs):
        context = super(FuelTankInfo, self).get_context_data(**kwargs)
        context['page_title'] = 'Zbiornik'
        context['header_text'] = 'Szczegóły zbiornika %s' % self.object.name
        context['type'] = 'info'
        context['submenu_template'] = 'fin/tank_submenu.html'
        context['tank'] = self.object

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Aktualizuj', 'path': reverse('fin:tank-update', args=[self.object.pk])})
        if self.object.fuel_volume == 0:
            local_menu.append({'text': 'Usuń', 'path': reverse('fin:tank-delete', args=[self.object.pk])})
        context['local_menu'] = local_menu

        field_list = []
        field_list.append({'header': 'Symbol zbiornika', 'value': self.object.tank_ref, 'bold': True})
        field_list.append({'header': 'Nazwa zbiornika', 'value': self.object.name})
        field_list.append({'header': 'Rodzaj paliwa', 'value': fuel_name(self.object.fuel_type)})
        field_list.append({'header': 'Ilość paliwa', 'value': '%.1f L' % self.object.fuel_volume})
        field_list.append({'header': 'Szacowana wartość', 'value': '%.2f PLN' % self.object.fuel_value()})
        field_list.append({'header': 'Uwagi', 'value': self.object.remarks})

        context['field_list'] = field_list
        return context


@class_view_decorator(permission_required('fin.fin_reader'))
class FuelTankDeliveries (DetailView):
    model = FuelTank
    template_name = 'fin/list_template.html'

    def get_context_data(self, **kwargs):
        context = super(FuelTankDeliveries, self).get_context_data(**kwargs)
        context['page_title'] = 'Dostawy paliwa'
        context['header_text'] = 'Dostawy paliwa do %s' % self.object.name
        context['empty_text'] = 'Brak dostaw paliwa.'
        context['type'] = 'deliveries'
        context['submenu_template'] = 'fin/tank_submenu.html'
        context['tank'] = self.object

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Nowa dostawa paliwa', 'path': reverse("fin:delivery-create", args=[self.object.pk])})
        context['local_menu'] = local_menu

        # Nagłówki tabeli
        header_list = []
        header_list.append({'header': 'Data\ndostawy'})
        header_list.append({'header': 'Dostawca'})
        header_list.append({'header': 'Dokument\ndostawy'})
        header_list.append({'header': 'Ilość\npaliwa'})
        header_list.append({'header': 'Cena\nlitra'})
        header_list.append({'header': 'Wartość\npaliwa'})
        header_list.append({'header': 'Uwagi'})
        header_list.append({'header': '...'})
        context['header_list'] = header_list

        row_list = []
        for object in self.object.fueldelivery_set.order_by('-date', '-pk'):
            fields = []
            fields.append({'name': 'date', 'value': object.date, 'link': reverse('fin:delivery-details', args=[object.pk]), 'just': 'center'})
            fields.append({'name': 'provider', 'value': object.provider})
            fields.append({'name': 'document', 'value': object.document})
            fields.append({'name': 'fuel_volume', 'value': '%.1f L' % object.fuel_volume, 'just': 'center'})
            fields.append({'name': 'liter_price', 'value': '%.2f PLN' % object.liter_price, 'just': 'center'})
            fields.append({'name': 'fuel_value', 'value': '%.2f PLN' % (object.fuel_volume * object.liter_price), 'just': 'center'})
            fields.append({'name': 'remarks', 'value': object.remarks})
            fields.append({'name': 'change', 'edit_link': reverse('fin:delivery-update', args=[object.pk]),
                           'delete_link': reverse('fin:delivery-delete', args=[object.pk])})
            row_list.append({'fields': fields})

        context['row_list'] = row_list
        return context


@class_view_decorator(permission_required('fin.fin_reader'))
class FuelTankTransfers (DetailView):
    model = FuelTank
    template_name = 'fin/list_template.html'

    def get_context_data(self, **kwargs):
        context = super(FuelTankTransfers, self).get_context_data(**kwargs)
        context['page_title'] = 'Wydania paliwa'
        context['header_text'] = 'Wydania wewnętrzne '
        if self.kwargs['type'] == 'in':
            context['header_text'] += 'do %s' % self.object.name
            context['type'] = 'intrans'
        else:
            context['header_text'] += 'z %s' % self.object.name
            context['type'] = 'outtrans'
        context['empty_text'] = 'Brak wydań paliwa.'
        context['submenu_template'] = 'fin/tank_submenu.html'
        context['tank'] = self.object

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Nowe wydanie', 'path': reverse("fin:transfer-create", args=[self.kwargs['type'], self.object.pk])})
        context['local_menu'] = local_menu

        # Nagłówki tabeli
        header_list = []
        header_list.append({'header': 'Data\nwydania'})
        header_list.append({'header': 'Dokument\nwydania'})
        if self.kwargs['type'] == 'in':
            header_list.append({'header': 'Zbiornik\nźródłowy'})
        else:
            header_list.append({'header': 'Zbiornik\ndocelowy'})
        header_list.append({'header': 'Ilość\npaliwa'})
        header_list.append({'header': 'Średnia\ncena litra'})
        header_list.append({'header': 'Uwagi'})
        header_list.append({'header': '...'})
        context['header_list'] = header_list

        row_list = []
        if self.kwargs['type'] == 'in':
            transfers = self.object.in_transfers.order_by('-date', '-pk')
        else:
            transfers = self.object.out_transfers.order_by('-date', '-pk')

        for object in transfers:
            fields = []
            fields.append({'name': 'date', 'value': object.date, 'link': reverse('fin:transfer-details', args=[self.kwargs['type'], object.pk]), 'just': 'center'})
            fields.append({'name': 'document', 'value': object.document})
            if self.kwargs['type'] == 'in':
                fields.append({'name': 'fueltank_from', 'value': object.fueltank_from})
            else:
                fields.append({'name': 'fueltank_to', 'value': object.fueltank_to})
            fields.append({'name': 'fuel_volume', 'value': '%.1f L' % object.fuel_volume, 'just': 'center'})
            fields.append({'name': 'liter_price', 'value': '%.2f PLN' % object.liter_price, 'just': 'center'})
            fields.append({'name': 'remarks', 'value': object.remarks})
            fields.append({'name': 'change', 'edit_link': reverse('fin:transfer-update', args=[self.kwargs['type'],object.pk]),
                           'delete_link': reverse('fin:transfer-delete', args=[self.kwargs['type'], object.pk])})
            row_list.append({'fields': fields})

        context['row_list'] = row_list
        return context


@class_view_decorator(permission_required('fin.fin_reader'))
class FuelTankCorrections (DetailView):
    model = FuelTank
    template_name = 'fin/list_template.html'

    def get_context_data(self, **kwargs):
        context = super(FuelTankCorrections, self).get_context_data(**kwargs)
        context['page_title'] = 'Protokoły różnic'
        context['header_text'] = 'Protokoły różnic dla %s' % self.object.name
        context['empty_text'] = 'Brak protokołów różnic.'
        context['type'] = 'corrections'
        context['submenu_template'] = 'fin/tank_submenu.html'
        context['tank'] = self.object

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Nowy protokół różnic', 'path': reverse("fin:correction-create", args=[self.object.pk])})
        context['local_menu'] = local_menu

        # Nagłówki tabeli
        header_list = []
        header_list.append({'header': 'Data\nprotokołu'})
        header_list.append({'header': 'Dokument'})
        header_list.append({'header': 'Wielkość\nkorekty'})
        header_list.append({'header': 'Uwagi'})
        header_list.append({'header': '...'})
        context['header_list'] = header_list

        row_list = []
        for object in self.object.fuelcorrection_set.order_by('-date', '-pk'):
            fields = []
            fields.append({'name': 'date', 'value': object.date, 'link': reverse('fin:correction-details', args=[object.pk]), 'just': 'center'})
            fields.append({'name': 'document', 'value': object.document})
            fields.append({'name': 'fuel_volume', 'value': '%.1f L' % object.fuel_volume, 'just': 'center'})
            fields.append({'name': 'remarks', 'value': object.remarks})
            fields.append({'name': 'change', 'edit_link': reverse('fin:correction-update', args=[object.pk]),
                           'delete_link': reverse('fin:correction-delete', args=[object.pk])})
            row_list.append({'fields': fields})

        context['row_list'] = row_list
        return context


@class_view_decorator(permission_required('fin.fin_reader'))
class FuelTankFuelings (DetailView):
    model = FuelTank
    template_name = 'fin/list_template.html'

    def get_context_data(self, **kwargs):
        context = super(FuelTankFuelings, self).get_context_data(**kwargs)
        context['page_title'] = 'Tankowania'
        context['header_text'] = 'Tankowania bez PDT z %s' % self.object.name
        context['empty_text'] = 'Brak tankowań.'
        context['type'] = 'fuelings'
        context['submenu_template'] = 'fin/tank_submenu.html'
        context['tank'] = self.object

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Nowe tankowanie', 'path': reverse("fin:fueling-create", args=[self.object.pk])})
        context['local_menu'] = local_menu

        # Nagłówki tabeli
        header_list = []
        header_list.append({'header': 'Data\ntankowania'})
        header_list.append({'header': 'Statek\npowietrzny'})
        header_list.append({'header': 'Osoba\ntankująca'})
        header_list.append({'header': 'Ilość\npaliwa'})
        header_list.append({'header': 'Uwagi'})
        header_list.append({'header': '...'})
        context['header_list'] = header_list

        row_list = []
        for object in self.object.localfueling_set.order_by('-date', '-pk'):
            fields = []
            fields.append({'name': 'date', 'value': object.date, 'link': reverse('fin:fueling-details', args=[object.pk]), 'just': 'center'})
            fields.append({'name': 'aircraft', 'value': object.aircraft, 'just': 'center'})
            fields.append({'name': 'person', 'value': object.person})
            fields.append({'name': 'fuel_volume', 'value': '%.1f L' % object.fuel_volume, 'just': 'center'})
            fields.append({'name': 'remarks', 'value': object.remarks})
            fields.append({'name': 'change', 'edit_link': reverse('fin:fueling-update', args=[object.pk]),
                           'delete_link': reverse('fin:fueling-delete', args=[object.pk])})
            row_list.append({'fields': fields})

        context['row_list'] = row_list
        return context


@class_view_decorator(permission_required('fin.fin_reader'))
class FuelTankPDTFuelings (DetailView):
    model = FuelTank
    template_name = 'fin/list_template.html'

    def get_context_data(self, **kwargs):
        context = super(FuelTankPDTFuelings, self).get_context_data(**kwargs)
        context['page_title'] = 'Tankowania PDT'
        context['header_text'] = 'Tankowania z %s' % self.object.name
        context['empty_text'] = 'Brak zarejestrowanych tankowań.'
        context['type'] = 'pdtfuelings'
        context['submenu_template'] = 'fin/tank_submenu.html'
        context['tank'] = self.object

        # Lokalne menu
        local_menu = []
        context['local_menu'] = local_menu

        # Nagłówki tabeli
        header_list = []
        header_list.append({'header': 'Data\ntankowania'})
        header_list.append({'header': 'Statek\npowietrzny'})
        header_list.append({'header': 'PDT'})
        header_list.append({'header': 'Ilość\npaliwa'})
        header_list.append({'header': '...'})
        context['header_list'] = header_list

        row_list = []
        for object in self.object.pdtfueling_set.order_by('-pk'):
            fields = []
            fields.append({'name': 'date', 'value': object.pdt.date, 'link': reverse('fin:pdtfueling-details', args=[object.pk]), 'just': 'center'})
            fields.append({'name': 'aircraft', 'value': object.pdt.aircraft, 'just': 'center'})
            fields.append({'name': 'pdt', 'value': object.pdt, 'link': reverse('panel:pdt-fuel', args=[object.pdt.pk]), 'just': 'center'})
            fields.append({'name': 'fuel_volume', 'value': '%.1f L' % object.fuel_volume, 'just': 'center'})
            fields.append({'name': 'change', 'edit_link': reverse('fin:pdtfueling-update', args=[object.pk]),
                           'delete_link': reverse('fin:pdtfueling-delete', args=[object.pk])})
            row_list.append({'fields': fields})

        context['row_list'] = row_list
        return context


@class_view_decorator(permission_required('fin.fin_admin'))
class FuelTankCreate (CreateView):
    model = FuelTank
    template_name = 'fin/create_template.html'

    def get_context_data(self, **kwargs):
        context = super(FuelTankCreate, self).get_context_data(**kwargs)
        context['page_title'] = 'Utworzenie zbiornika'
        context['header_text'] = 'Utworzenie zbiornika paliwa'
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(FuelTank, exclude=['fuel_volume'],
                                       widgets={'name':TextInput(attrs={'size':50}),
                                                'remarks':Textarea(attrs={'rows':4, 'cols':100})})
        return form_class

    def get_success_url(self):
        return reverse('fin:tank-list')


@class_view_decorator(permission_required('fin.fin_admin'))
class FuelTankUpdate (UpdateView):
    model = FuelTank
    template_name = 'fin/update_template.html'

    def get_context_data(self, **kwargs):
        context = super(FuelTankUpdate, self).get_context_data(**kwargs)
        context['page_title'] = 'Modyfikacja zbiornika'
        context['header_text'] = 'Modyfikacja zbiornika paliwa'
        context['type'] = 'info'
        context['submenu_template'] = 'fin/tank_submenu.html'
        context['tank'] = self.object

        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(FuelTank, exclude=['fuel_volume'],
                                       widgets={'name':TextInput(attrs={'size':50}),
                                                'remarks':Textarea(attrs={'rows':4, 'cols':100})})
        return form_class

    def get_success_url(self):
        return reverse('fin:tank-info', args=[self.object.pk,])


@class_view_decorator(permission_required('fin.fin_admin'))
class FuelTankDelete (DeleteView):
    model = FuelTank
    template_name = 'fin/delete_template.html'

    def get_context_data(self, **kwargs):
        context = super(FuelTankDelete, self).get_context_data(**kwargs)
        context['page_title'] = 'Usunięcie zbiornika'
        context['header_text'] = 'Usunięcie zbiornika %s' % self.object.name
        context['description'] = 'Aktualna ilość paliwa wynosi %.1f L' % self.object.fuel_volume
        context['type'] = 'info'
        context['submenu_template'] = 'fin/tank_submenu.html'
        context['tank'] = self.object
        return context

    def get_success_url(self):
        return reverse('fin:tank-list')


@class_view_decorator(permission_required('fin.fin_reader'))
class DeliveryDetails (DetailView):
    model = FuelDelivery
    template_name = 'fin/details_template.html'

    def get_context_data(self, **kwargs):
        context = super(DeliveryDetails, self).get_context_data(**kwargs)
        context['page_title'] = 'Dostawa paliwa'
        context['header_text'] = 'Szczegóły dostawy do %s' % self.object.fueltank.name
        context['type'] = 'deliveries'
        context['submenu_template'] = 'fin/tank_submenu.html'
        context['tank'] = self.object.fueltank

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Aktualizuj', 'path': reverse('fin:delivery-update', args=[self.object.pk])})
        local_menu.append({'text': 'Usuń', 'path': reverse('fin:delivery-delete', args=[self.object.pk])})
        context['local_menu'] = local_menu

        field_list = []
        field_list.append({'header': 'Data dostawy', 'value': self.object.date, 'bold': True})
        field_list.append({'header': 'Dostawca', 'value': self.object.provider})
        field_list.append({'header': 'Dokument dostawy', 'value': self.object.document})
        field_list.append({'header': 'Ilość paliwa', 'value': '%.1f L' % self.object.fuel_volume})
        field_list.append({'header': 'Cena litra', 'value': '%.2f PLN' % self.object.liter_price})
        field_list.append({'header': 'Uwagi', 'value': self.object.remarks})

        context['field_list'] = field_list

        return context


@class_view_decorator(permission_required('fin.fin_admin'))
class DeliveryCreate (CreateView):
    model = FuelDelivery
    template_name = 'fin/create_template.html'

    def get_context_data(self, **kwargs):
        fueltank = FuelTank.objects.get(pk=self.kwargs['fueltank_id'])
        context = super(DeliveryCreate, self).get_context_data(**kwargs)
        context['page_title'] = 'Dostawa paliwa'
        context['header_text'] = 'Nowa dostawa do %s' % fueltank.name
        context['type'] = 'deliveries'
        context['submenu_template'] = 'fin/tank_submenu.html'
        context['tank'] = fueltank
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(FuelDelivery, exclude=['fueltank'],
                                       widgets={'date':AdminDateWidget,
                                                'provider':TextInput(attrs={'size':50}),
                                                'remarks':Textarea(attrs={'rows':4, 'cols':100})})

        form_class.base_fields['date'].initial = date.today()
        return form_class

    def form_valid(self, form):
        form.instance.fueltank = FuelTank.objects.get(pk=self.kwargs['fueltank_id'])
        return super(DeliveryCreate, self).form_valid(form)

    def get_success_url(self):
        return reverse('fin:tank-deliveries', args=[self.kwargs['fueltank_id']])


@class_view_decorator(permission_required('fin.fin_admin'))
class DeliveryUpdate (UpdateView):
    model = FuelDelivery
    template_name = 'fin/update_template.html'

    def get_context_data(self, **kwargs):
        context = super(DeliveryUpdate, self).get_context_data(**kwargs)
        context['page_title'] = 'Modyfikacja dostawy'
        context['header_text'] = 'Modyfikacja dostawy do %s' % self.object.fueltank.name
        context['type'] = 'deliveries'
        context['submenu_template'] = 'fin/tank_submenu.html'
        context['tank'] = self.object.fueltank
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(FuelDelivery, exclude=['fueltank'],
                                       widgets={'date':AdminDateWidget,
                                                'provider':TextInput(attrs={'size':50}),
                                                'remarks':Textarea(attrs={'rows':4, 'cols':100})})
        return form_class

    def get_success_url(self):
        return reverse('fin:tank-deliveries', args=[self.object.fueltank.pk,])


@class_view_decorator(permission_required('fin.fin_admin'))
class DeliveryDelete (DeleteView):
    model = FuelDelivery
    template_name = 'fin/delete_template.html'

    def get_context_data(self, **kwargs):
        context = super(DeliveryDelete, self).get_context_data(**kwargs)
        context['page_title'] = 'Usunięcie dostawy'
        context['header_text'] = 'Usunięcie dostawy do %s' % self.object.fueltank.name
        context['description'] = 'Dokument dostawy: %s' % self.object.document
        context['type'] = 'deliveries'
        context['submenu_template'] = 'fin/tank_submenu.html'
        context['tank'] = self.object.fueltank
        return context

    def get_success_url(self):
        return reverse('fin:tank-deliveries', args=[self.object.fueltank.pk,])


@class_view_decorator(permission_required('fin.fin_reader'))
class TransferDetails (DetailView):
    model = FuelTransfer
    template_name = 'fin/details_template.html'

    def get_context_data(self, **kwargs):
        context = super(TransferDetails, self).get_context_data(**kwargs)
        context['page_title'] = 'Wydanie paliwa'
        context['header_text'] = 'Szczegóły wydania wewnętrznego '
        if self.kwargs['type'] == 'in':
            context['header_text'] += 'do %s' % self.object.fueltank_to.name
            context['type'] = 'intranfers'
            context['tank'] = self.object.fueltank_to
        else:
            context['header_text'] += 'z %s' % self.object.fueltank_from.name
            context['type'] = 'outtranfers'
            context['tank'] = self.object.fueltank_from
        context['submenu_template'] = 'fin/tank_submenu.html'

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Aktualizuj', 'path': reverse('fin:transfer-update', args=[self.kwargs['type'], self.object.pk])})
        local_menu.append({'text': 'Usuń', 'path': reverse('fin:transfer-delete', args=[self.kwargs['type'], self.object.pk])})
        context['local_menu'] = local_menu

        field_list = []
        field_list.append({'header': 'Data wydania', 'value': self.object.date, 'bold': True})
        if self.kwargs['type'] == 'in':
            field_list.append({'header': 'Zbiornik źródłowy', 'value': self.object.fueltank_from})
        else:
            field_list.append({'header': 'Zbiornik docelowy', 'value': self.object.fueltank_to})
        field_list.append({'header': 'Dokument wydania', 'value': self.object.document})
        field_list.append({'header': 'Ilość paliwa', 'value': '%.1f L' % self.object.fuel_volume})
        field_list.append({'header': 'Średnia cena litra', 'value': '%.2f PLN' % self.object.liter_price})
        field_list.append({'header': 'Uwagi', 'value': self.object.remarks})

        context['field_list'] = field_list

        return context


@class_view_decorator(permission_required('fin.fin_admin'))
class TransferCreate (CreateView):
    model = FuelTransfer
    template_name = 'fin/create_template.html'

    def get_context_data(self, **kwargs):
        fueltank = FuelTank.objects.get(pk=self.kwargs['fueltank_id'])
        context = super(TransferCreate, self).get_context_data(**kwargs)
        context['page_title'] = 'Wydanie paliwa'
        context['header_text'] = 'Nowa wydanie wewnętrzne '
        if self.kwargs['type'] == 'in':
            context['header_text'] += 'do %s' % fueltank.name
            context['type'] = 'intrans'
        else:
            context['header_text'] += 'z %s' % fueltank.name
            context['type'] = 'outtrans'
        context['submenu_template'] = 'fin/tank_submenu.html'
        context['tank'] = fueltank
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(FuelTransfer, exclude=[('fueltank_to' if self.kwargs['type'] == 'in' else 'fueltank_from'), 'liter_price'],
                                       widgets={'date':AdminDateWidget,
                                                'remarks':Textarea(attrs={'rows':4, 'cols':100})})

        fueltank = FuelTank.objects.get(pk = self.kwargs['fueltank_id'])
        second_tank = 'fueltank_from' if self.kwargs['type'] == 'in' else 'fueltank_to'

        form_class.base_fields[second_tank].choices = [(fueltank.pk, fueltank.name)
                                                       for fueltank in FuelTank.objects.filter(fuel_type=fueltank.fuel_type).exclude(pk=fueltank.pk)]
        if form_class.base_fields[second_tank].choices == []:
            form_class.base_fields[second_tank].choices = [(None, '---------')]

        form_class.base_fields['date'].initial = date.today()

        return form_class

    def form_valid(self, form):
        if self.kwargs['type'] == 'in':
            form.instance.fueltank_to = FuelTank.objects.get(pk=self.kwargs['fueltank_id'])
        else:
            form.instance.fueltank_from = FuelTank.objects.get(pk=self.kwargs['fueltank_id'])
        return super(TransferCreate, self).form_valid(form)

    def get_success_url(self):
        return reverse('fin:tank-transfers', args=[self.kwargs['type'], self.kwargs['fueltank_id']])


@class_view_decorator(permission_required('fin.fin_admin'))
class TransferUpdate (UpdateView):
    model = FuelTransfer
    template_name = 'fin/update_template.html'

    def get_context_data(self, **kwargs):
        context = super(TransferUpdate, self).get_context_data(**kwargs)
        context['page_title'] = 'Modyfikacja wydania'
        context['header_text'] = 'Modyfikacja wydania ze zbiornika %s do zbiornika %s' % (self.object.fueltank_from.tank_ref, self.object.fueltank_to.tank_ref)
        if self.kwargs['type'] == 'in':
            context['type'] = 'intrans'
            context['tank'] = self.object.fueltank_to
        else:
            context['type'] = 'outtrans'
            context['tank'] = self.object.fueltank_from
        context['submenu_template'] = 'fin/tank_submenu.html'
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(FuelTransfer, exclude=['fueltank_from', 'fueltank_to'],
                                       widgets={'date':AdminDateWidget,
                                                'remarks':Textarea(attrs={'rows':4, 'cols':100})})

        return form_class

    def get_success_url(self):
        if self.kwargs['type'] == 'in':
            return reverse('fin:tank-transfers', args=[self.kwargs['type'], self.object.fueltank_to.pk,])
        else:
            return reverse('fin:tank-transfers', args=[self.kwargs['type'], self.object.fueltank_from.pk,])


@class_view_decorator(permission_required('fin.fin_admin'))
class TransferDelete (DeleteView):
    model = FuelTransfer
    template_name = 'fin/delete_template.html'

    def get_context_data(self, **kwargs):
        context = super(TransferDelete, self).get_context_data(**kwargs)
        context['page_title'] = 'Usunięcie wydania'
        context['header_text'] = 'Usunięcie wydania'
        context['description'] = 'Dokument: %s' % self.object.document
        if self.kwargs['type'] == 'in':
            context['header_text'] += 'do %s' % self.object.fueltank_to.name
            context['type'] = 'intrans'
            context['tank'] = self.object.fueltank_to
        else:
            context['header_text'] += 'z %s' % self.object.fueltank_from.name
            context['type'] = 'outtrans'
            context['tank'] = self.object.fueltank_from
        context['submenu_template'] = 'fin/tank_submenu.html'
        return context

    def get_success_url(self):
        if self.kwargs['type'] == 'in':
            return reverse('fin:tank-transfers', args=[self.kwargs['type'], self.object.fueltank_to.pk,])
        else:
            return reverse('fin:tank-transfers', args=[self.kwargs['type'], self.object.fueltank_from.pk,])


@class_view_decorator(permission_required('fin.fin_reader'))
class CorrectionDetails (DetailView):
    model = FuelCorrection
    template_name = 'fin/details_template.html'

    def get_context_data(self, **kwargs):
        context = super(CorrectionDetails, self).get_context_data(**kwargs)
        context['page_title'] = 'Protokół różnic'
        context['header_text'] = 'Szczegóły protokołu różnic dla %s' % self.object.fueltank.name
        context['type'] = 'corrections'
        context['submenu_template'] = 'fin/tank_submenu.html'
        context['tank'] = self.object.fueltank

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Aktualizuj', 'path': reverse('fin:correction-update', args=[self.object.pk])})
        local_menu.append({'text': 'Usuń', 'path': reverse('fin:correction-delete', args=[self.object.pk])})
        context['local_menu'] = local_menu

        field_list = []
        field_list.append({'header': 'Data protokołu', 'value': self.object.date, 'bold': True})
        field_list.append({'header': 'Dokument', 'value': self.object.document})
        field_list.append({'header': 'Wielkość korekty', 'value': '%.1f L' % self.object.fuel_volume})
        field_list.append({'header': 'Uwagi', 'value': self.object.remarks})

        context['field_list'] = field_list

        return context


@class_view_decorator(permission_required('fin.fin_admin'))
class CorrectionCreate (CreateView):
    model = FuelCorrection
    template_name = 'fin/create_template.html'

    def get_context_data(self, **kwargs):
        fueltank = FuelTank.objects.get(pk=self.kwargs['fueltank_id'])
        context = super(CorrectionCreate, self).get_context_data(**kwargs)
        context['page_title'] = 'Nowy protokół'
        context['header_text'] = 'Nowy protokół różnic dla %s' % fueltank.name
        context['type'] = 'corrections'
        context['submenu_template'] = 'fin/tank_submenu.html'
        context['tank'] = fueltank
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(FuelCorrection, exclude=['fueltank'],
                                       widgets={'date':AdminDateWidget,
                                                'remarks':Textarea(attrs={'rows':4, 'cols':100})})

        form_class.base_fields['date'].initial = date.today()
        return form_class

    def form_valid(self, form):
        form.instance.fueltank = FuelTank.objects.get(pk=self.kwargs['fueltank_id'])
        return super(CorrectionCreate, self).form_valid(form)

    def get_success_url(self):
        return reverse('fin:tank-corrections', args=[self.kwargs['fueltank_id']])


@class_view_decorator(permission_required('fin.fin_admin'))
class CorrectionUpdate (UpdateView):
    model = FuelCorrection
    template_name = 'fin/update_template.html'

    def get_context_data(self, **kwargs):
        context = super(CorrectionUpdate, self).get_context_data(**kwargs)
        context['page_title'] = 'Modyfikacja protokołu'
        context['header_text'] = 'Modyfikacja protokołu różnic dla %s' % self.object.fueltank.name
        context['type'] = 'corrections'
        context['submenu_template'] = 'fin/tank_submenu.html'
        context['tank'] = self.object.fueltank
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(FuelCorrection, exclude=['fueltank'],
                                       widgets={'date':AdminDateWidget,
                                                'remarks':Textarea(attrs={'rows':4, 'cols':100})})
        return form_class

    def get_success_url(self):
        return reverse('fin:tank-corrections', args=[self.object.fueltank.pk,])


@class_view_decorator(permission_required('fin.fin_admin'))
class CorrectionDelete (DeleteView):
    model = FuelCorrection
    template_name = 'fin/delete_template.html'

    def get_context_data(self, **kwargs):
        context = super(CorrectionDelete, self).get_context_data(**kwargs)
        context['page_title'] = 'Usunięcie protokołu'
        context['header_text'] = 'Usunięcie protokołu różnic dla %s' % self.object.fueltank.name
        context['description'] = 'Dokument: %s' % self.object.document
        context['type'] = 'corrections'
        context['submenu_template'] = 'fin/tank_submenu.html'
        context['tank'] = self.object.fueltank
        return context

    def get_success_url(self):
        return reverse('fin:tank-corrections', args=[self.object.fueltank.pk,])


@class_view_decorator(permission_required('fin.fin_reader'))
class FuelingDetails (DetailView):
    model = LocalFueling
    template_name = 'fin/details_template.html'

    def get_context_data(self, **kwargs):
        context = super(FuelingDetails, self).get_context_data(**kwargs)
        context['page_title'] = 'Tankowanie'
        context['header_text'] = 'Szczegóły tankowania bez PDT z %s' % self.object.fueltank.name
        context['type'] = 'fuelings'
        context['submenu_template'] = 'fin/tank_submenu.html'
        context['tank'] = self.object.fueltank

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Aktualizuj', 'path': reverse('fin:fueling-update', args=[self.object.pk])})
        local_menu.append({'text': 'Usuń', 'path': reverse('fin:fueling-delete', args=[self.object.pk])})
        context['local_menu'] = local_menu

        field_list = []
        field_list.append({'header': 'Data tankowania', 'value': self.object.date, 'bold': True})
        field_list.append({'header': 'Statek powietrzny', 'value': self.object.aircraft})
        field_list.append({'header': 'Osoba tankująca', 'value': self.object.person})
        field_list.append({'header': 'Ilość paliwa', 'value': '%.1f L' % self.object.fuel_volume})
        field_list.append({'header': 'Uwagi', 'value': self.object.remarks})

        context['field_list'] = field_list

        return context


@class_view_decorator(permission_required('fin.fin_admin'))
class FuelingCreate (CreateView):
    model = LocalFueling
    template_name = 'fin/create_template.html'

    def get_context_data(self, **kwargs):
        fueltank = FuelTank.objects.get(pk=self.kwargs['fueltank_id'])
        context = super(FuelingCreate, self).get_context_data(**kwargs)
        context['page_title'] = 'Nowe tankowanie'
        context['header_text'] = 'Nowe tankowanie bez PDT z %s' % fueltank.name
        context['type'] = 'fuelings'
        context['submenu_template'] = 'fin/tank_submenu.html'
        context['tank'] = fueltank
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(LocalFueling, exclude=['fueltank'],
                                       widgets={'date':AdminDateWidget,
                                                'remarks':Textarea(attrs={'rows':4, 'cols':100})})

        fueltank = FuelTank.objects.get(pk = self.kwargs['fueltank_id'])

        form_class.base_fields['aircraft'].choices = [(None, '---------')] + \
                                                     [(aircraft.pk, aircraft.registration) for aircraft in Aircraft.objects.filter(fuel_type=fueltank.fuel_type)]

        form_class.base_fields['date'].initial = date.today()

        return form_class

    def form_valid(self, form):
        form.instance.fueltank = FuelTank.objects.get(pk=self.kwargs['fueltank_id'])
        return super(FuelingCreate, self).form_valid(form)

    def get_success_url(self):
        return reverse('fin:tank-fuelings', args=[self.kwargs['fueltank_id']])


@class_view_decorator(permission_required('fin.fin_admin'))
class FuelingUpdate (UpdateView):
    model = LocalFueling
    template_name = 'fin/update_template.html'

    def get_context_data(self, **kwargs):
        context = super(FuelingUpdate, self).get_context_data(**kwargs)
        context['page_title'] = 'Modyfikacja tankowania'
        context['header_text'] = 'Modyfikacja tankowania bez PDT z %s' % self.object.fueltank.name
        context['type'] = 'fuelings'
        context['submenu_template'] = 'fin/tank_submenu.html'
        context['tank'] = self.object.fueltank
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(LocalFueling, exclude=['fueltank'],
                                       widgets={'date':AdminDateWidget,
                                                'remarks':Textarea(attrs={'rows':4, 'cols':100})})
        return form_class

    def get_success_url(self):
        return reverse('fin:tank-fuelings', args=[self.object.fueltank.pk,])


@class_view_decorator(permission_required('fin.fin_admin'))
class FuelingDelete (DeleteView):
    model = LocalFueling
    template_name = 'fin/delete_template.html'

    def get_context_data(self, **kwargs):
        context = super(FuelingDelete, self).get_context_data(**kwargs)
        context['page_title'] = 'Usunięcie tankowania'
        context['header_text'] = 'Usunięcie tankowania bez PDT z %s' % self.object.fueltank.name
        context['description'] = 'Osoba tankująca: %s' % self.object.person
        context['type'] = 'fuelings'
        context['submenu_template'] = 'fin/tank_submenu.html'
        context['tank'] = self.object.fueltank
        return context

    def get_success_url(self):
        return reverse('fin:tank-fuelings', args=[self.object.fueltank.pk,])


@class_view_decorator(permission_required('fin.fin_reader'))
class PDTFuelingDetails (DetailView):
    model = PDTFueling
    template_name = 'fin/details_template.html'

    def get_context_data(self, **kwargs):
        context = super(PDTFuelingDetails, self).get_context_data(**kwargs)
        context['page_title'] = 'Tankowanie'
        context['header_text'] = 'Szczegóły tankowania z %s' % self.object.fueltank.name
        context['type'] = 'pdtfuelings'
        context['submenu_template'] = 'fin/tank_submenu.html'
        context['tank'] = self.object.fueltank

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Aktualizuj', 'path': reverse('fin:pdtfueling-update', args=[self.object.pk])})
        local_menu.append({'text': 'Usuń', 'path': reverse('fin:pdtfueling-delete', args=[self.object.pk])})
        context['local_menu'] = local_menu

        field_list = []
        field_list.append({'header': 'Data tankowania', 'value': self.object.pdt.date, 'bold': True})
        field_list.append({'header': 'Statek powietrzny', 'value': self.object.pdt.aircraft})
        field_list.append({'header': 'Na podstawie PDT', 'value': self.object.pdt, 'link': reverse('panel:pdt-info', args=[self.object.pdt.pk])})
        field_list.append({'header': 'Ilość paliwa', 'value': '%.1f L' % self.object.fuel_volume})

        context['field_list'] = field_list

        return context


@class_view_decorator(permission_required('fin.fin_admin'))
class PDTFuelingUpdate (UpdateView):
    model = PDTFueling
    template_name = 'fin/update_template.html'

    def get_context_data(self, **kwargs):
        context = super(PDTFuelingUpdate, self).get_context_data(**kwargs)
        context['page_title'] = 'Modyfikacja tankowania'
        context['header_text'] = 'Modyfikacja tankowania z %s' % self.object.fueltank.name
        context['type'] = 'pdtfuelings'
        context['submenu_template'] = 'fin/tank_submenu.html'
        context['tank'] = self.object.fueltank
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(PDTFueling, exclude=['fueltank', 'pdt'])
        form_class.base_fields['fuel_volume'] = DecimalField(max_digits=4, decimal_places=1, min_value = 0,
                                                max_value = self.object.pdt.aircraft.fuel_capacity,
                                                label='Objętość paliwa (L)')
        return form_class

    def get_success_url(self):
        return reverse('fin:tank-pdt-fuelings', args=[self.object.fueltank.pk,])


@class_view_decorator(permission_required('fin.fin_admin'))
class PDTFuelingDelete (DeleteView):
    model = PDTFueling
    template_name = 'fin/delete_template.html'

    def get_context_data(self, **kwargs):
        context = super(PDTFuelingDelete, self).get_context_data(**kwargs)
        context['page_title'] = 'Usunięcie tankowania'
        context['header_text'] = 'Usunięcie tankowania z %s' % self.object.fueltank.name
        context['description'] = 'Na podstawie PDT: %s' % self.object.pdt
        context['type'] = 'pdtfuelings'
        context['submenu_template'] = 'fin/tank_submenu.html'
        context['tank'] = self.object.fueltank
        return context

    def get_success_url(self):
        return reverse('fin:tank-pdt-fuelings', args=[self.object.fueltank.pk,])


@class_view_decorator(permission_required('fin.fin_reader'))
class RemoteFuelingDetails (DetailView):
    model = RemoteFueling
    template_name = 'fin/details_template.html'

    def get_context_data(self, **kwargs):
        context = super(RemoteFuelingDetails, self).get_context_data(**kwargs)
        context['page_title'] = 'Tankowanie'
        context['header_text'] = 'Szczegóły tankowania poza SALT'
        context['type'] = 'fuelings'
        context['submenu_template'] = 'fin/contr_submenu.html'
        context['contractor'] = self.object.pdt.contractor

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Aktualizuj', 'path': reverse('fin:remote-fueling-update', args=[self.object.pk])})
        local_menu.append({'text': 'Usuń', 'path': reverse('fin:remote-fueling-delete', args=[self.object.pk])})
        context['local_menu'] = local_menu

        field_list = []
        field_list.append({'header': 'Data tankowania', 'value': self.object.pdt.date})
        field_list.append({'header': 'Miejsce tankowania', 'value': self.object.location})
        field_list.append({'header': 'Ilość paliwa', 'value': '%.1f L' % self.object.fuel_volume})
        field_list.append({'header': 'Faktura', 'value': self.object.document})
        field_list.append({'header': 'Dokument akcyzy', 'value': self.object.excise})
        field_list.append({'header': 'Cena paliwa', 'value': '%.2f PLN' % self.object.total_price if self.object.total_price else None})
        field_list.append({'header': 'Uwagi', 'value': self.object.remarks})

        context['field_list'] = field_list

        return context


@class_view_decorator(permission_required('fin.fin_admin'))
class RemoteFuelingUpdate (UpdateView):
    model = RemoteFueling
    template_name = 'fin/update_template.html'

    def get_context_data(self, **kwargs):
        context = super(RemoteFuelingUpdate, self).get_context_data(**kwargs)
        context['page_title'] = 'Tankowanie'
        context['header_text'] = 'Modyfikacja tankowania poza SALT'
        context['type'] = 'fuelings'
        context['submenu_template'] = 'fin/contr_submenu.html'
        context['contractor'] = self.object.pdt.contractor
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(RemoteFueling, exclude=['pdt', 'fuel_volume'],
                                       widgets={'document':TextInput(attrs={'size':50}),
                                                'excise':TextInput(attrs={'size':50}),
                                                'remarks':Textarea(attrs={'rows':4, 'cols':100})})
        return form_class

    def get_success_url(self):
        return reverse('fin:contr-fuelings', args=[self.object.pdt.contractor.pk,])


@class_view_decorator(permission_required('fin.fin_admin'))
class RemoteFuelingDelete (DeleteView):
    model = RemoteFueling
    template_name = 'fin/delete_template.html'

    def get_context_data(self, **kwargs):
        context = super(RemoteFuelingDelete, self).get_context_data(**kwargs)
        context['page_title'] = 'Tankowanie'
        context['header_text'] = 'Usunięcie tankowania poza SALT'
        context['description'] = 'Na podstawie PDT: %s' % self.object.pdt
        context['type'] = 'fuelings'
        context['submenu_template'] = 'fin/contr_submenu.html'
        context['contractor'] = self.object.pdt.contractor
        return context

    def get_success_url(self):
        return reverse('fin:contr-fuelings', args=[self.object.pdt.contractor.pk,])


@class_view_decorator(permission_required('fin.fin_reader'))
class VoucherList (ListView):
    model = Voucher
    template_name = 'fin/list_template.html'

    def get_queryset(self):
        return Voucher.objects.filter(done_date__isnull=True)

    def get_context_data(self, **kwargs):
        context = super(VoucherList, self).get_context_data(**kwargs)
        context['page_title'] = 'Vouchery'
        context['header_text'] = 'Lista aktualnych voucherów'
        context['empty_text'] = 'Brak zarejestrowanych voucherów.'

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Zarejestruj nowy voucher', 'path': reverse("fin:voucher-create")})
        local_menu.append({'text': 'Vouchery zrealizowane', 'path': reverse("fin:voucher-done")})
        context['local_menu'] = local_menu

        # Nagłówki tabeli
        header_list = []
        header_list.append({'header': 'Numer\nvouchera'})
        header_list.append({'header': 'Kod\nvouchera'})
        header_list.append({'header': 'Data\nsprzedaży'})
        header_list.append({'header': 'Liczba\nosób'})
        header_list.append({'header': 'Czas\ntrwania'})
        header_list.append({'header': 'Opis\nvouchera'})
        header_list.append({'header': 'Opłacony'})
        header_list.append({'header': 'Data\nważności'})
        header_list.append({'header': 'Uwagi'})
        header_list.append({'header': '...'})
        context['header_list'] = header_list

        # Zawartość tabeli
        row_list = []
        for object in self.object_list:
            fields = []
            fields.append({'name': 'id', 'value': object, 'link': reverse('fin:voucher-details', args=[object.pk]), 'just': 'center'})
            fields.append({'name': 'voucher_code', 'value': object.voucher_code})
            fields.append({'name': 'issue_date', 'value': object.issue_date, 'just': 'center'})
            fields.append({'name': 'persons', 'value': object.persons, 'just': 'center'})
            fields.append({'name': 'time', 'value': '%d min' % object.time, 'just': 'center'})
            fields.append({'name': 'description', 'value': object.description})
            fields.append({'name': 'paid', 'value': ('TAK' if object.paid else 'NIE'), 'just': 'center'})
            fields.append({'name': 'valid_date', 'value': object.valid_date, 'just': 'center'})
            fields.append({'name': 'remarks', 'value': object.remarks})
            fields.append({'name': 'change', 'edit_link': reverse('fin:voucher-update', args=[object.pk]),
                           'delete_link': (reverse('fin:voucher-delete', args=[object.pk]) if not object.done_date else None)})
            row_list.append({'fields': fields})
        context['row_list'] = row_list
        return context


@class_view_decorator(permission_required('fin.fin_reader'))
class VoucherDoneList (ListView):
    model = Voucher
    template_name = 'fin/list_template.html'

    def get_queryset(self):
        return Voucher.objects.filter(done_date__isnull=False)

    def get_context_data(self, **kwargs):
        context = super(VoucherDoneList, self).get_context_data(**kwargs)
        context['page_title'] = 'Vouchery'
        context['header_text'] = 'Lista voucherów zrealizowanych'
        context['empty_text'] = 'Brak zrealizowanych voucherów.'

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Vouchery aktualne', 'path': reverse("fin:voucher-list")})
        context['local_menu'] = local_menu

        # Nagłówki tabeli
        header_list = []
        header_list.append({'header': 'Numer\nvouchera'})
        header_list.append({'header': 'Data\nsprzedaży'})
        header_list.append({'header': 'Liczba\nosób'})
        header_list.append({'header': 'Czas\ntrwania'})
        header_list.append({'header': 'Opis\nvouchera'})
        header_list.append({'header': 'Data\nrealizacji'})
        header_list.append({'header': 'Uwagi'})
        context['header_list'] = header_list

        # Zawartość tabeli
        row_list = []
        for object in self.object_list:
            fields = []
            fields.append({'name': 'id', 'value': object, 'link': reverse('fin:voucher-details', args=[object.pk]), 'just': 'center'})
            fields.append({'name': 'issue_date', 'value': object.issue_date, 'just': 'center'})
            fields.append({'name': 'person', 'value': object.persons, 'just': 'center'})
            fields.append({'name': 'person', 'value': '%d min' % object.time, 'just': 'center'})
            fields.append({'name': 'description', 'value': object.description})
            fields.append({'name': 'done_date', 'value': object.done_date, 'just': 'center'})
            fields.append({'name': 'remarks', 'value': object.remarks})
            row_list.append({'fields': fields})
        context['row_list'] = row_list
        return context


@class_view_decorator(permission_required('fin.fin_reader'))
class VoucherDetails (DetailView):
    model = Voucher
    template_name = 'fin/details_template.html'

    def get_context_data(self, **kwargs):
        context = super(VoucherDetails, self).get_context_data(**kwargs)
        context['page_title'] = self.object
        context['header_text'] = 'Informacje na temat vouchera'

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Aktualizuj', 'path': reverse('fin:voucher-update', args=[self.object.pk])})
        if not self.object.done_date:
            local_menu.append({'text': 'Usuń', 'path': reverse('fin:voucher-delete', args=[self.object.pk])})
        context['local_menu'] = local_menu

        payments = {'cash': 'Gotówka', 'transfer': 'Przelew bankowy', 'epay': 'Płatność elektroniczna'}

        field_list = []
        field_list.append({'header': 'Numer vouchera', 'value': self.object, 'bold': True})
        field_list.append({'header': 'Kod vouchera', 'value': self.object.voucher_code, 'bold': True})
        field_list.append({'header': 'Data sprzedaży', 'value': self.object.issue_date})
        field_list.append({'header': 'Liczba osób', 'value': self.object.persons})
        field_list.append({'header': 'Czas trwania', 'value': '%d min' % self.object.time})
        field_list.append({'header': 'Opis vouchera', 'value': self.object.description})
        field_list.append({'header': 'Data ważności', 'value': self.object.valid_date})
        field_list.append({'header': 'Forma płatności', 'value': payments[self.object.payment]})
        field_list.append({'header': 'Opłacony', 'value': ('TAK' if self.object.paid else 'NIE')})
        if self.object.done_date:
            field_list.append({'header': 'Data realizacji', 'value': self.object.done_date})
        field_list.append({'header': 'Uwagi', 'value': self.object.remarks})

        context['field_list'] = field_list

        return context


@class_view_decorator(permission_required('fin.fin_admin'))
class VoucherCreate (CreateView):
    model = Voucher
    template_name = 'fin/create_template.html'

    def get_context_data(self, **kwargs):
        context = super(VoucherCreate, self).get_context_data(**kwargs)
        context['page_title'] = 'Nowy voucher'
        context['header_text'] = 'Rejestracja nowego vouchera'
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(Voucher, exclude=['done_date'], widgets={'description':TextInput(attrs={'size':103}),
                                                                                'issue_date':AdminDateWidget(),
                                                                                'valid_date':AdminDateWidget(),
                                                                                'remarks':Textarea(attrs={'rows':4, 'cols':100})})
        form_class.base_fields['issue_date'].initial = date.today()
        form_class.base_fields['valid_date'].initial = date.today() + timedelta(days=365)
        form_class.base_fields['persons'].initial = 1
        form_class.base_fields['voucher_code'].initial = ''.join(random.choice(string.ascii_uppercase + string.digits) for _ in range(10))
        return form_class

    def get_success_url(self):
        return reverse('fin:voucher-list')


@class_view_decorator(permission_required('fin.fin_admin'))
class VoucherUpdate (UpdateView):
    model = Voucher
    template_name = 'fin/update_template.html'

    def get_context_data(self, **kwargs):
        context = super(VoucherUpdate, self).get_context_data(**kwargs)
        context['page_title'] = 'Zmiana vouchera'
        context['header_text'] = 'Zmiana danych vouchera'
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(Voucher, fields='__all__', widgets={'description':TextInput(attrs={'size':103}),
                                                                           'issue_date':AdminDateWidget(),
                                                                           'valid_date':AdminDateWidget(),
                                                                           'done_date':AdminDateWidget(),
                                                                           'remarks':Textarea(attrs={'rows':4, 'cols':100})})
        return form_class

    def get_success_url(self):
        return reverse('fin:voucher-list')


@class_view_decorator(permission_required('fin.fin_admin'))
class VoucherDelete (DeleteView):
    model = Voucher
    template_name = 'fin/delete_template.html'

    def get_context_data(self, **kwargs):
        context = super(VoucherDelete, self).get_context_data(**kwargs)
        context['page_title'] = 'Usunięcie vouchera'
        context['header_text'] = 'Usunięcie vouchera'
        context['description'] = self.object.description
        return context

    def get_success_url(self):
        return reverse('fin:voucher-list')


@class_view_decorator(permission_required('fin.fin_reader'))
class RentPackageList (ListView):
    model = RentPackage
    template_name = 'fin/list_template.html'

    def get_context_data(self, **kwargs):
        context = super(RentPackageList, self).get_context_data(**kwargs)
        context['page_title'] = 'Pakiety wynajmu'
        context['header_text'] = 'Lista pakietów wynajmu'
        context['empty_text'] = 'Brak pakietów wynajmu.'

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Nowy pakiet wynajmu', 'path': reverse("fin:package-create")})
        context['local_menu'] = local_menu

        # Nagłówki tabeli
        header_list = []
        header_list.append({'header': 'Symbol\npakietu'})
        header_list.append({'header': 'Nazwa\npakietu'})
        header_list.append({'header': 'Typ\nSP'})
        header_list.append({'header': 'Liczba\ngodzin'})
        header_list.append({'header': 'Cena\ngodziny'})
        header_list.append({'header': 'Uwagi'})
        header_list.append({'header': '...'})
        context['header_list'] = header_list

        row_list = []
        for object in self.object_list:
            fields = []
            fields.append({'name': 'package_id', 'value': object, 'link': reverse('fin:package-details', args=[object.pk])})
            fields.append({'name': 'name', 'value': object.name})
            fields.append({'name': 'ac_type', 'value': object.ac_type})
            fields.append({'name': 'hours', 'value': object.hours, 'just': 'center'})
            fields.append({'name': 'hour_price', 'value': '%.2f PLN' % object.hour_price, 'just': 'center'})
            fields.append({'name': 'remarks', 'value': object.remarks})
            fields.append({'name': 'change', 'edit_link': reverse('fin:package-update', args=[object.pk]),
                           'delete_link': reverse('fin:package-delete', args=[object.pk])})
            row_list.append({'fields': fields})
        context['row_list'] = row_list
        return context


@class_view_decorator(permission_required('fin.fin_reader'))
class RentPackageDetails (DetailView):
    model = RentPackage
    template_name = 'fin/details_template.html'

    def get_context_data(self, **kwargs):
        context = super(RentPackageDetails, self).get_context_data(**kwargs)
        context['page_title'] = 'Pakiety wynajmu'
        context['header_text'] = 'Szczegóły pakietu wynajmu %s' % self.object

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Aktualizuj', 'path': reverse('fin:package-update', args=[self.object.pk])})
        local_menu.append({'text': 'Usuń', 'path': reverse('fin:package-delete', args=[self.object.pk])})
        context['local_menu'] = local_menu

        field_list = []
        field_list.append({'header': 'Symbol pakietu', 'value': self.object, 'bold': True})
        field_list.append({'header': 'Nazwa pakietu', 'value': self.object.name})
        field_list.append({'header': 'Typ SP', 'value': self.object.ac_type})
        field_list.append({'header': 'Liczba godzin', 'value': self.object.hours})
        field_list.append({'header': 'Cena godziny', 'value': '%.2f PLN' % self.object.hour_price})
        field_list.append({'header': 'Uwagi', 'value': self.object.remarks})

        context['field_list'] = field_list

        return context


@class_view_decorator(permission_required('fin.fin_admin'))
class RentPackageCreate (CreateView):
    model = RentPackage
    template_name = 'fin/create_template.html'

    def get_context_data(self, **kwargs):
        context = super(RentPackageCreate, self).get_context_data(**kwargs)
        context['page_title'] = 'Nowy pakiet'
        context['header_text'] = 'Utworzenie pakietu wynajmu'
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(RentPackage, fields='__all__',
                                       widgets={'name': TextInput(attrs={'size':50}),
                                                'remarks':Textarea(attrs={'rows':4, 'cols':100})})

        form_class.base_fields['ac_type'] = ChoiceField([(type['type'], type['type'])
                                                         for type in Aircraft.objects.order_by().values('type').distinct()], label='Typ SP')

        return form_class

    def get_success_url(self):
        return reverse('fin:package-list')


@class_view_decorator(permission_required('fin.fin_admin'))
class RentPackageUpdate (UpdateView):
    model = RentPackage
    template_name = 'fin/update_template.html'

    def get_context_data(self, **kwargs):
        context = super(RentPackageUpdate, self).get_context_data(**kwargs)
        context['page_title'] = 'Modyfikacja pakietu'
        context['header_text'] = 'Modyfikacja pakietu wynajmu %s' % self.object
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(RentPackage, fields='__all__',
                                       widgets={'name': TextInput(attrs={'size':50}),
                                                'remarks':Textarea(attrs={'rows':4, 'cols':100})})

        form_class.base_fields['ac_type'] = ChoiceField([(type['type'], type['type'])
                                                         for type in Aircraft.objects.order_by().values('type').distinct()], label='Typ SP')
        return form_class

    def get_success_url(self):
        return reverse('fin:package-list')


@class_view_decorator(permission_required('fin.fin_admin'))
class RentPackageDelete (DeleteView):
    model = RentPackage
    template_name = 'fin/delete_template.html'

    def get_context_data(self, **kwargs):
        context = super(RentPackageDelete, self).get_context_data(**kwargs)
        context['page_title'] = 'Usunięcie pakietu'
        context['header_text'] = 'Usunięcie pakietu wynajmu %s' % self.object
        context['description'] = '%d godzin na %s ' % (self.object.hours, self.object.ac_type)
        return context

    def get_success_url(self):
        return reverse('fin:package-list')


@class_view_decorator(permission_required('fin.fin_reader'))
class ContactorList (ListView):
    model = Contractor
    template_name = 'fin/list_template.html'

    def get_context_data(self, **kwargs):
        context = super(ContactorList, self).get_context_data(**kwargs)
        context['page_title'] = 'Kontrahenci'
        context['header_text'] = 'Lista kontrahentów'
        context['empty_text'] = 'Brak kontrahentów.'

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Nowy kontrahent', 'path': reverse("fin:contr-create")})
        context['local_menu'] = local_menu

        # Nagłówki tabeli
        header_list = []
        header_list.append({'header': 'Id FK'})
        header_list.append({'header': 'Nazwa\nkontrahenta'})
        header_list.append({'header': 'Aktywny'})
        header_list.append({'header': 'Firma'})
        header_list.append({'header': 'Saldo\nszkoleń'})
        header_list.append({'header': 'Saldo\nwynajmu'})
        header_list.append({'header': 'Saldo\nAOC'})
        header_list.append({'header': 'Saldo\nusług'})
        header_list.append({'header': 'Zgoda\nna debet'})
        header_list.append({'header': 'Uwagi'})
        header_list.append({'header': '...'})
        context['header_list'] = header_list

        row_list = []
        for object in self.object_list:
            fields = []
            fields.append({'name': 'contractor_id', 'value': object.contractor_id, 'link': reverse('fin:contr-info', args=[object.pk]), 'just': 'center'})
            fields.append({'name': 'name', 'value': object.name})
            fields.append({'name': 'active', 'value': ('TAK' if object.active else 'NIE'), 'just': 'center'})
            fields.append({'name': 'company', 'value': ('TAK' if object.company else 'NIE'), 'just': 'center'})
            fields.append({'name': 'ato_balance', 'value': '%.2f PLN' % object.ato_balance, 'color': ('auto' if object.ato_balance >= 0 else 'lightcoral'), 'just': 'center'})
            fields.append({'name': 'rent_balance', 'value': '%.2f PLN' % object.rent_balance, 'color': ('auto' if object.rent_balance >= 0 else 'lightcoral'), 'just': 'center'})
            fields.append({'name': 'aoc_balance', 'value': '%.2f PLN' % object.aoc_balance, 'color': ('auto' if object.aoc_balance >= 0 else 'lightcoral'), 'just': 'center'})
            fields.append({'name': 'spo_balance', 'value': '%.2f PLN' % object.spo_balance, 'color': ('auto' if object.spo_balance >= 0 else 'lightcoral'), 'just': 'center'})
            fields.append({'name': 'debet_allowed', 'value': ('TAK' if object.debet_allowed else 'NIE'), 'just': 'center'})
            fields.append({'name': 'remarks', 'value': object.remarks})
            fields.append({'name': 'change', 'edit_link': reverse('fin:contr-update', args=[object.pk]),
                           'delete_link': reverse('fin:contr-delete', args=[object.pk])})
            row_list.append({'fields': fields})
        context['row_list'] = row_list
        return context


@class_view_decorator(permission_required('fin.fin_reader'))
class ContractorInfo (DetailView):
    model = Contractor
    template_name = 'fin/details_template.html'

    def get_context_data(self, **kwargs):
        context = super(ContractorInfo, self).get_context_data(**kwargs)
        context['page_title'] = 'Kontrahenci'
        context['header_text'] = 'Informacje o kontrahencie %s' % self.object.name
        context['type'] = 'info'
        context['submenu_template'] = 'fin/contr_submenu.html'
        context['contractor'] = self.object

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Aktualizuj', 'path': reverse('fin:contr-update', args=[self.object.pk])})
        local_menu.append({'text': 'Usuń', 'path': reverse('fin:contr-delete', args=[self.object.pk])})
        context['local_menu'] = local_menu

        field_list = []
        field_list.append({'header': 'Nazwa kontrahenta', 'value': self.object.name, 'bold': True})
        field_list.append({'header': 'Identyfikator FK', 'value': self.object.contractor_id})
        field_list.append({'header': 'Adres kontrahenta', 'value': self.object.address1})
        if self.object.address2 != '':
            field_list.append({'header': '', 'value': self.object.address2})
        field_list.append({'header': 'Firma', 'value': ('TAK' if self.object.company else 'NIE')})
        if not self.object.company:
            field_list.append({'header': 'PESEL', 'value': self.object.pesel})
        field_list.append({'header': 'NIP', 'value': self.object.nip})
        if self.object.company:
            field_list.append({'header': 'REGON', 'value': self.object.regon})
        field_list.append({'header': 'Zgoda na debet', 'value': ('TAK' if self.object.debet_allowed else 'NIE')})
        field_list.append({'header': 'Aktywny', 'value': ('TAK' if self.object.active else 'NIE')})
        field_list.append({'header': 'Uwagi', 'value': self.object.remarks})

        context['field_list'] = field_list

        return context


@class_view_decorator(permission_required('fin.fin_reader'))
class ContractorOperations (DetailView):
    model = Contractor
    template_name = 'fin/contr_operations.html'

    def get_context_data(self, **kwargs):
        context = super(ContractorOperations, self).get_context_data(**kwargs)
        context['page_title'] = 'Operacje'
        context['header_text'] = 'Operacje na rachunku %s' % self.object.name
        context['empty_text'] = 'Brak operacji na rachunku.'
        context['type'] = 'operations'
        context['submenu_template'] = 'fin/contr_submenu.html'
        context['contractor'] = self.object

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Nowa operacja', 'path': reverse("fin:operation-create", args=[self.object.pk])})
        context['local_menu'] = local_menu

        # Nagłówki tabeli
        header_list = []
        header_list.append({'header': 'Data\noperacji'})
        header_list.append({'header': 'Rodzaj\noperacji'})
        header_list.append({'header': 'Dokument\noperacji'})
        header_list.append({'header': 'Kwota\nszkolenia'})
        header_list.append({'header': 'Kwota\nwynajem'})
        header_list.append({'header': 'Kwota\nAOC'})
        header_list.append({'header': 'Kwota\nSPO'})
        header_list.append({'header': 'Uwagi'})
        header_list.append({'header': '...'})
        context['header_list'] = header_list

        row_list = []
        for object in self.object.balanceoperation_set.order_by('-date', '-pk'):
            fields = []
            fields.append({'name': 'date', 'value': object.date, 'link': reverse('fin:operation-details', args=[object.pk]), 'just': 'center'})
            fields.append({'name': 'type', 'value': object.type, 'just': 'center'})
            fields.append({'name': 'document', 'value': object.pdt or object.document,
                           'link': reverse('panel:pdt-info', args=[object.pdt.pk]) if object.pdt else ''})
            fields.append({'name': 'ato_amount', 'value': '%.2f PLN' % object.ato_amount, 'just': 'center'})
            fields.append({'name': 'rent_amount', 'value': '%.2f PLN' % object.rent_amount, 'just': 'center'})
            fields.append({'name': 'aoc_amount', 'value': '%.2f PLN' % object.aoc_amount, 'just': 'center'})
            fields.append({'name': 'spo_amount', 'value': '%.2f PLN' % object.spo_amount, 'just': 'center'})
            fields.append({'name': 'remarks', 'value': object.remarks})
            fields.append({'name': 'change', 'edit_link': reverse('fin:operation-update', args=[object.pk]),
                           'delete_link': reverse('fin:operation-delete', args=[object.pk])})
            row_list.append({'fields': fields})

        context['row_list'] = row_list
        return context


@class_view_decorator(permission_required('fin.fin_reader'))
class ContractorPackages (DetailView):
    model = Contractor
    template_name = 'fin/contr_packages.html'

    def get_context_data(self, **kwargs):
        context = super(ContractorPackages, self).get_context_data(**kwargs)
        context['page_title'] = 'Pakiety'
        context['header_text'] = 'Pakiety na rachunku %s' % self.object.name
        context['empty_text'] = 'Brak pakietów na rachunku.'
        context['type'] = 'packages'
        context['submenu_template'] = 'fin/contr_submenu.html'
        context['contractor'] = self.object

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Zakup pakietu', 'path': reverse("fin:package-buy", args=[self.object.pk])})
        context['local_menu'] = local_menu

        # Nagłówki tabeli
        header_list = []
        header_list.append({'header': 'Data\nzakupu'})
        header_list.append({'header': 'Symbol\npakietu'})
        header_list.append({'header': 'Nazwa\npakietu'})
        header_list.append({'header': 'Typ\nSP'})
        header_list.append({'header': 'Liczba\ngodzin'})
        header_list.append({'header': 'Pozostało\ngodzin'})
        header_list.append({'header': 'Cena\ngodziny'})
        header_list.append({'header': 'Uwagi'})
        header_list.append({'header': '...'})
        context['header_list'] = header_list

        row_list = []
        for object in self.object.soldpackage_set.order_by('-date', '-pk'):
            fields = []
            fields.append({'name': 'date', 'value': object.date})
            fields.append({'name': 'package_id', 'value': object, 'just': 'center'})
            fields.append({'name': 'name', 'value': object.name})
            fields.append({'name': 'ac_type', 'value': object.ac_type})
            fields.append({'name': 'hours', 'value': object.hours, 'just': 'center'})
            fields.append({'name': 'left_hours', 'value': object.left_hours, 'just': 'center', 'bold': True})
            fields.append({'name': 'hour_price', 'value': '%.2f PLN' % object.hour_price, 'just': 'center'})
            fields.append({'name': 'remarks', 'value': object.remarks})
            fields.append({'name': 'change', 'delete_link': reverse('fin:package-sell', args=[object.pk])})
            row_list.append({'fields': fields})
        context['row_list'] = row_list

        context['sub_header_text'] = 'Ceny specjalne dla %s' % self.object.name
        context['sub_create_path'] = reverse('fin:price-create', args=[self.object.pk])
        context['sub_create_text'] = 'Nowa cena specjalna'
        context['sub_empty_text'] = 'Brak cen specjalnych.'
        context['sub_header_list'] = ['Typ SP', 'Cena\ngodziny', 'Uwagi', '...']

        row_list = []
        for price in self.object.specialprice_set.all():
            fields = []
            fields.append({'name': 'ac_type', 'value': price.ac_type, 'just': 'center' })
            fields.append({'name': 'hour_price', 'value': '%.2f PLN' % price.hour_price, 'just': 'center', 'bold': True})
            fields.append({'name': 'remarks', 'value': price.remarks})
            fields.append({'name': 'change', 'edit_link': reverse('fin:price-update', args=[price.pk]),
                           'delete_link': reverse('fin:price-delete', args=[price.pk])})
            row_list.append({'fields': fields})
        context['sub_row_list'] = row_list

        return context


@class_view_decorator(login_required())
class ContractorFlights (DetailView):
    model = Contractor
    template_name = 'fin/list_template.html'

    # jeśli nie superuser to ustaw na zalogowanego
    def get_object(self, queryset=None):
        obj = super(ContractorFlights, self).get_object()
        if not self.request.user.is_staff:
            obj = self.request.user.fbouser.contractor
        return obj

    def get_context_data(self, **kwargs):
        context = super(ContractorFlights, self).get_context_data(**kwargs)

        types = {}
        types['ato'] = ('szkoleniowe', ['03A', '03B'])
        types['rent'] = ('w wynajmie', ['03C', '04'])
        types['aoc'] = ('AOC', ['01'])
        types['spo'] = ('SPO', ['02'])

        type = self.kwargs['type']

        context['page_title'] = 'Loty kontrahenta'
        context['header_text'] = 'Loty %s kontrahenta %s' % (types[type][0], self.object)
        context['empty_text'] = 'Brak zarejestrowanych lotów.'
        context['type'] = 'flights'
        context['submenu_template'] = 'fin/contr_submenu.html'
        context['contractor'] = self.object

        flight_list = PDT.objects.filter(contractor=self.object.pk, flight_type__in=types[type][1]).exclude(status='open').order_by('-date', '-pk')

        # Lokalne menu
        local_menu = []
        context['local_menu'] = local_menu

        # Nagłówki tabeli
        header_list = []
        header_list.append({'header': 'Data'})
        header_list.append({'header': 'SP'})
        header_list.append({'header': 'PDT'})
        header_list.append({'header': 'PIC'})
        header_list.append({'header': 'SIC'})
        header_list.append({'header': 'TTH'})
        header_list.append({'header': 'Ldg.'})
        header_list.append({'header': 'Uwagi'})
        context['header_list'] = header_list

        # Zawartość tabeli
        row_list = []
        for object in flight_list:
            fields = []
            fields.append({'name': 'date', 'value': object.date})
            fields.append({'name': 'aircraft', 'value': object.aircraft})
            fields.append({'name': 'pdt', 'value': '{:0>6d}'.format(object.pdt_ref), 'link': reverse('panel:pdt-info', args=[object.pk]), 'just': 'center'})
            fields.append({'name': 'pic', 'value': object.pic})
            fields.append({'name': 'sic', 'value': object.sic})
            fields.append({'name': 'tth', 'value': '%.1f' % (object.tth_end() - object.tth_start), 'just': 'center'})
            fields.append({'name': 'landings', 'value': object.landings_sum(), 'just': 'center'})
            fields.append({'name': 'remarks', 'value': object.remarks})
            row_list.append({'fields': fields})
        context['row_list'] = row_list

        return context


@class_view_decorator(login_required())
class ContractorFuelings (DetailView):
    model = Contractor
    template_name = 'fin/list_template.html'

    # jeśli nie superuser to ustaw na zalogowanego
    def get_object(self, queryset=None):
        obj = super(ContractorFuelings, self).get_object()
        if not self.request.user.is_staff:
            obj = self.request.user.fbouser.contractor
        return obj

    def get_context_data(self, **kwargs):
        context = super(ContractorFuelings, self).get_context_data(**kwargs)
        context['page_title'] = 'Tankowania kontrahenta'
        context['header_text'] = 'Tankowania kontrahenta poza SALT'
        context['empty_text'] = 'Brak zarejestrowanych tankowań.'
        context['type'] = 'fuelings'
        context['submenu_template'] = 'fin/contr_submenu.html'
        context['contractor'] = self.object


        # Lokalne menu
        local_menu = []
        context['local_menu'] = local_menu

        remote_fueling_list = RemoteFueling.objects.filter(pdt__contractor=self.object.pk).order_by('-pk')

        if remote_fueling_list:
            context['header_text'] = 'Tankowania poza SALT'

            # Nagłówki tabeli
            header_list = []
            header_list.append({'header': 'Data'})
            header_list.append({'header': 'SP'})
            header_list.append({'header': 'Miejsce'})
            header_list.append({'header': 'PDT'})
            header_list.append({'header': 'Ilość paliwa'})
            header_list.append({'header': 'Faktura'})
            header_list.append({'header': 'Akcyza'})
            header_list.append({'header': 'Cena paliwa'})
            header_list.append({'header': 'Uwagi'})
            header_list.append({'header': '...'})
            context['header_list'] = header_list

            row_list = []
            for fueling in remote_fueling_list:
                fields = []
                fields.append({'name': 'date', 'value': fueling.pdt.date, 'link': reverse('fin:remote-fueling-details', args=[fueling.pk])})
                fields.append({'name': 'aircraft', 'value': fueling.pdt.aircraft, 'just': 'center'})
                fields.append({'name': 'location', 'value': fueling.location})
                fields.append({'name': 'pdt', 'value': fueling.pdt, 'link': reverse('panel:pdt-fuel', args=[fueling.pdt.pk])})
                fields.append({'name': 'fuel_volume', 'value': '%.1f L' % fueling.fuel_volume, 'just': 'center'})
                fields.append({'name': 'document', 'value': fueling.document or '[brak]'})
                fields.append({'name': 'excise', 'value': fueling.excise})
                fields.append({'name': 'fuel_price', 'value': '%.2d PLN' % fueling.total_price if fueling.total_price else '[brak]', 'just': 'center'})
                fields.append({'name': 'remarks', 'value': fueling.remarks})
                fields.append({'name': 'change', 'edit_link': reverse('fin:remote-fueling-update', args=[fueling.pk]),
                               'delete_link': reverse('fin:remote-fueling-delete', args=[fueling.pk])})
                row_list.append({'fields': fields})
            context['row_list'] = row_list

        return context


@class_view_decorator(permission_required('fin.fin_admin'))
class ContractorCreate (CreateView):
    model = Contractor
    template_name = 'fin/create_template.html'

    def get_context_data(self, **kwargs):
        context = super(ContractorCreate, self).get_context_data(**kwargs)
        context['page_title'] = 'Nowy kontrahent'
        context['header_text'] = 'Utworzenie nowego kontrahenta'
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(Contractor, fields='__all__',
                                       widgets={'name': TextInput(attrs={'size':50}),
                                                'address1': TextInput(attrs={'size':50}),
                                                'address2': TextInput(attrs={'size':50}),
                                                'remarks':Textarea(attrs={'rows':2, 'cols':100})})
        return form_class

    def get_success_url(self):
        return reverse('fin:contr-list')


@class_view_decorator(permission_required('fin.fin_admin'))
class ContractorUpdate (UpdateView):
    model = Contractor
    template_name = 'fin/update_template.html'


    def get_context_data(self, **kwargs):
        context = super(ContractorUpdate, self).get_context_data(**kwargs)
        context['page_title'] = 'Modyfikacja kontrahenta'
        context['header_text'] = 'Modyfikacja kontrahenta %s' % self.object.name
        context['type'] = 'info'
        context['submenu_template'] = 'fin/contr_submenu.html'
        context['contractor'] = self.object
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(Contractor, exclude=['ato_balance', 'rent_balance', 'aoc_balance', 'spo_balance'],
                                       widgets={'name': TextInput(attrs={'size':50}),
                                                'address1': TextInput(attrs={'size':50}),
                                                'address2': TextInput(attrs={'size':50}),
                                                'remarks':Textarea(attrs={'rows':2, 'cols':100})})
        return form_class

    def get_success_url(self):
        return reverse('fin:contr-info', args=[self.object.pk,])


@class_view_decorator(permission_required('fin.fin_admin'))
class ContractorDelete (DeleteView):
    model = Contractor
    template_name = 'fin/delete_template.html'

    def get_context_data(self, **kwargs):
        context = super(ContractorDelete, self).get_context_data(**kwargs)
        context['page_title'] = 'Usunięcie kontrahenta'
        context['header_text'] = 'Usunięcie kontrahenta %s' % self.object.name
        context['type'] = 'info'
        context['submenu_template'] = 'fin/contr_submenu.html'
        context['contractor'] = self.object
        return context

    def get_success_url(self):
        return reverse('fin:contr-list')


@class_view_decorator(permission_required('fin.fin_reader'))
class BalanceOperationDetails (DetailView):
    model = BalanceOperation
    template_name = 'fin/details_template.html'

    def get_context_data(self, **kwargs):
        context = super(BalanceOperationDetails, self).get_context_data(**kwargs)
        context['page_title'] = 'Operacja'
        context['header_text'] = 'Szczegóły operacji na rachunku %s' % self.object.contractor.name
        context['type'] = 'operations'
        context['submenu_template'] = 'fin/contr_submenu.html'
        context['contractor'] = self.object.contractor

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Aktualizuj', 'path': reverse('fin:operation-update', args=[self.object.pk])})
        local_menu.append({'text': 'Usuń', 'path': reverse('fin:operation-delete', args=[self.object.pk])})
        context['local_menu'] = local_menu

        field_list = []
        field_list.append({'header': 'Data operacji', 'value': self.object.date, 'bold': True})
        field_list.append({'header': 'Typ operacji', 'value': self.object.type})
        field_list.append({'header': 'Dokument operacji', 'value': self.object.document})
        field_list.append({'header': 'Kwota operacji (szkolenia)', 'value': '%.2f PLN' % self.object.ato_amount})
        field_list.append({'header': 'Kwota operacji (wynajem)', 'value': '%.2f PLN' % self.object.rent_amount})
        field_list.append({'header': 'Kwota operacji (AOC)', 'value': '%.2f PLN' % self.object.aoc_amount})
        field_list.append({'header': 'Kwota operacji (SPO)', 'value': '%.2f PLN' % self.object.spo_amount})
        field_list.append({'header': 'Uwagi', 'value': self.object.remarks})

        context['field_list'] = field_list

        return context


@class_view_decorator(permission_required('fin.fin_admin'))
class BalanceOperationCreate (CreateView):
    model = BalanceOperation
    template_name = 'fin/create_template.html'

    def get_context_data(self, **kwargs):
        contractor = Contractor.objects.get(pk=self.kwargs['contractor_id'])
        context = super(BalanceOperationCreate, self).get_context_data(**kwargs)
        context['page_title'] = 'Nowa operacja'
        context['header_text'] = 'Nowa operacja na rachunku %s' % contractor.name
        context['type'] = 'operations'
        context['submenu_template'] = 'fin/contr_submenu.html'
        context['contractor'] = contractor
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(BalanceOperation, exclude=['contractor', 'pdt'],
                                       widgets={'date':AdminDateWidget,
                                                'remarks':Textarea(attrs={'rows':4, 'cols':100})})

        form_class.base_fields['date'].initial = date.today()
        return form_class

    def form_valid(self, form):
        form.instance.contractor = Contractor.objects.get(pk=self.kwargs['contractor_id'])
        return super(BalanceOperationCreate, self).form_valid(form)

    def get_success_url(self):
        return reverse('fin:contr-operations', args=[self.kwargs['contractor_id']])


@class_view_decorator(permission_required('fin.fin_admin'))
class BalanceOperationUpdate (UpdateView):
    model = BalanceOperation
    template_name = 'fin/update_template.html'

    def get_context_data(self, **kwargs):
        context = super(BalanceOperationUpdate, self).get_context_data(**kwargs)
        context['page_title'] = 'Modyfikacja operacji'
        context['header_text'] = 'Modyfikacja operacji na rachunku %s' % self.object.contractor.name
        context['type'] = 'operations'
        context['submenu_template'] = 'fin/contr_submenu.html'
        context['contractor'] = self.object.contractor
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(BalanceOperation, exclude=['contractor'],
                                       widgets={'date':AdminDateWidget,
                                                'remarks':Textarea(attrs={'rows':4, 'cols':100})})
        return form_class

    def get_success_url(self):
        return reverse('fin:contr-operations', args=[self.object.contractor.pk,])


@class_view_decorator(permission_required('fin.fin_admin'))
class BalanceOperationDelete (DeleteView):
    model = BalanceOperation
    template_name = 'fin/delete_template.html'

    def get_context_data(self, **kwargs):
        context = super(BalanceOperationDelete, self).get_context_data(**kwargs)
        context['page_title'] = 'Usunięcie operacji'
        context['header_text'] = 'Usunięcie operacji na rachunku %s' % self.object.contractor.name
        context['description'] = 'Dokument operacji: %s' % (self.object.pdt or self.object.document)
        context['type'] = 'operations'
        context['submenu_template'] = 'fin/contr_submenu.html'
        context['contractor'] = self.object.contractor
        return context

    def get_success_url(self):
        return reverse('fin:contr-operations', args=[self.object.contractor.pk,])


@class_view_decorator(permission_required('fin.fin_reader'))
class RentPriceList (ListView):
    model = Aircraft
    template_name = 'fin/list_template.html'

    def get_queryset(self):
        return Aircraft.objects.order_by('type')

    def get_context_data(self, **kwargs):
        context = super(RentPriceList, self).get_context_data(**kwargs)
        context['page_title'] = 'Ceny wynajmu'
        context['header_text'] = 'Ceny wynajmu SP'
        context['empty_text'] = 'Brak statków powietrznych.'

        # Lokalne menu
        local_menu = []
        context['local_menu'] = local_menu

        # Nagłówki tabeli
        header_list = []
        header_list.append({'header': 'Rejestracja\nSP'})
        header_list.append({'header': 'Typ SP'})
        header_list.append({'header': 'Cena godziny\nwynajmu'})
        header_list.append({'header': '...'})
        context['header_list'] = header_list

        # Zawartość tabeli
        row_list = []
        for object in self.object_list:
            fields = []
            fields.append({'name': 'registration', 'value': object, 'just': 'center', 'bold': True})
            fields.append({'name': 'type', 'value': object.type})
            fields.append({'name': 'hour_price', 'value': '%.2f PLN' % object.rent_price, 'just': 'right', 'link': reverse('fin:rentprice-update', args=[object.pk])})
            fields.append({'name': 'change', 'edit_link': reverse('fin:rentprice-update', args=[object.pk])})
            row_list.append({'fields': fields})
        context['row_list'] = row_list
        return context


@class_view_decorator(permission_required('fin.fin_admin'))
class RentPriceUpdate (UpdateView):
    model = Aircraft
    template_name = 'fin/update_template.html'

    def get_context_data(self, **kwargs):
        context = super(RentPriceUpdate, self).get_context_data(**kwargs)
        context['page_title'] = 'Zmiana ceny'
        context['header_text'] = 'Zmiana ceny wynajmu dla %s' % self.object
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(Aircraft, fields=['rent_price'])
        return form_class

    def get_success_url(self):
        return reverse('fin:rentprice-list')


@class_view_decorator(permission_required('fin.fin_admin'))
class PriceCreate (CreateView):
    model = SpecialPrice
    template_name = 'fin/create_template.html'

    def get_context_data(self, **kwargs):
        contractor = Contractor.objects.get(pk=self.kwargs['contractor_id'])
        context = super(PriceCreate, self).get_context_data(**kwargs)
        context['page_title'] = 'Nowa cena'
        context['header_text'] = 'Nowa cena specjalna dla %s' % contractor.name
        context['type'] = 'packages'
        context['submenu_template'] = 'fin/contr_submenu.html'
        context['contractor'] = contractor
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(SpecialPrice, exclude=['contractor'],
                                       widgets={'remarks':Textarea(attrs={'rows':2, 'cols':100})})

        # Subklasa z kontrolą unikalności ac_type
        class create_class(form_class):

            # Metoda init przejmuje dodatkowy argument 'contractor'
            def __init__(self, **kwargs):
                super(create_class, self).__init__(**kwargs)
                if 'contractor' in kwargs:
                    self.contractor = kwargs.pop('contractor')

            def clean_ac_type(self, **kwargs):
                cleaned_data = super(create_class, self).clean()
                # Jeśli ten kontrahent ma już ten cenę na ten ac_type to zgłoś błąd
                if self.contractor.specialprice_set.filter(ac_type = cleaned_data['ac_type']):
                    raise ValidationError(('Na ten typ jest już cena specjalna!'), code='duplicated_price')
                return cleaned_data['ac_type']

        create_class.contractor = Contractor.objects.get(pk=self.kwargs['contractor_id'])
        return create_class

    def form_valid(self, form):
        form.instance.contractor = Contractor.objects.get(pk=self.kwargs['contractor_id'])
        return super(PriceCreate, self).form_valid(form)

    def get_success_url(self):
        return reverse('fin:contr-packages', args=[self.kwargs['contractor_id']])


@class_view_decorator(permission_required('fin.fin_admin'))
class PriceUpdate (UpdateView):
    model = SpecialPrice
    template_name = 'fin/update_template.html'

    def get_context_data(self, **kwargs):
        context = super(PriceUpdate, self).get_context_data(**kwargs)
        context['page_title'] = 'Modyfikacja ceny'
        context['header_text'] = 'Modyfikacja ceny specjalnej dla %s' % self.object.contractor.name
        context['type'] = 'packages'
        context['submenu_template'] = 'fin/contr_submenu.html'
        context['contractor'] = self.object.contractor
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(SpecialPrice, exclude=['contractor'],
                                       widgets={'remarks':Textarea(attrs={'rows':2, 'cols':100})})

        # Subklasa z kontrolą unikalności ac_type
        class update_class(form_class):

            # Metoda init przejmuje dodatkowe argumenty
            def __init__(self, **kwargs):
                super(update_class, self).__init__(**kwargs)
                if 'contractor' in kwargs:
                    self.contractor = kwargs.pop('contractor')
                if 'pk' in kwargs:
                    self.contractor = kwargs.pop('pk')

            def clean_ac_type(self, **kwargs):
                cleaned_data = super(update_class, self).clean()
                # Jeśli ten kontrahent ma już ten cenę na ten ac_type to zgłoś błąd
                if self.contractor.specialprice_set.exclude(pk = self.pk).filter(ac_type = cleaned_data['ac_type']):
                    raise ValidationError(('Na ten typ jest już cena specjalna!'), code='duplicated_price')
                return cleaned_data['ac_type']

        update_class.contractor = self.object.contractor
        update_class.pk = self.object.pk
        return update_class

    def get_success_url(self):
        return reverse('fin:contr-packages', args=[self.object.contractor.pk,])


@class_view_decorator(permission_required('fin.fin_admin'))
class PriceDelete (DeleteView):
    model = SpecialPrice
    template_name = 'fin/delete_template.html'

    def get_context_data(self, **kwargs):
        context = super(PriceDelete, self).get_context_data(**kwargs)
        context['page_title'] = 'Usunięcie ceny'
        context['header_text'] = 'Usunięcie ceny specjalnej dla %s' % self.object.contractor.name
        context['type'] = 'packages'
        context['submenu_template'] = 'fin/contr_submenu.html'
        context['contractor'] = self.object.contractor
        return context

    def get_success_url(self):
        return reverse('fin:contr-packages', args=[self.object.contractor.pk,])


@permission_required('fin.fin_admin')
def PackageBuy (request, contractor_id):

    context = {}
    contractor = Contractor.objects.get(pk=contractor_id)
    context['page_title'] = 'Zakup pakietu'
    context['header_text'] = 'Zakup pakietu godzin dla %s' % contractor.name
    context['type'] = 'packages'
    context['submenu_template'] = 'fin/contr_submenu.html'
    context['contractor'] = contractor
    form = PackageBuyForm(request.POST or None, contractor=contractor)
    context['form'] = form

    if form.is_valid():
        # Utworznenie nowego zakupionego pakietu #
        package = RentPackage.objects.get(pk = form.cleaned_data['packages'])
        sold_package = SoldPackage(contractor = contractor,
                                   date = form.cleaned_data['date'],
                                   package_id = package.package_id,
                                   name = package.name,
                                   ac_type = package.ac_type,
                                   hours = package.hours,
                                   left_hours = package.hours,
                                   hour_price = package.hour_price,
                                   remarks = form.cleaned_data['remarks'])
        sold_package.full_clean()
        sold_package.save()

        # Zmniejszenie salda rachunku
        balance_operation = BalanceOperation(contractor=contractor, date=form.cleaned_data['date'], type='Zakup pakietu',
                                             aoc_amount=0, spo_amount=0, ato_amount=0, rent_amount=(-package.hours * package.hour_price).quantize(Decimal(10) ** -2),
                                             remarks=package.name)
        balance_operation.full_clean()
        balance_operation.save()

        return HttpResponseRedirect(reverse('fin:contr-packages', args=[contractor_id,]))
    return render(request, 'fin/package_buy.html', context)


@class_view_decorator(permission_required('fin.fin_admin'))
class PackageSell (DeleteView):
    model = SoldPackage
    template_name = 'fin/delete_template.html'

    def get_context_data(self, **kwargs):
        context = super(PackageSell, self).get_context_data(**kwargs)
        context['page_title'] = 'Usunięcie pakietu'
        context['header_text'] = 'Rezygnacja z pakietu dla %s' % self.object.contractor.name
        context['description'] = 'Niewykorzystane środki wrócą na rachunek wynajmu.'
        context['type'] = 'packages'
        context['submenu_template'] = 'fin/contr_submenu.html'
        context['contractor'] = self.object.contractor
        return context

    def delete(self, request, *args, **kwargs):
        object = super(PackageSell, self).get_object()

        # Zwrot niewykorzystanych środków na rachunek
        balance_operation = BalanceOperation(contractor=object.contractor, date=date.today(),
                                             type='Zwrot za pakiet', aoc_amount=0, spo_amount=0, ato_amount=0,
                                             rent_amount=(object.left_hours * object.hour_price).quantize(Decimal(10) ** -2),
                                             remarks='%d godzin na %s' % (object.left_hours, object.ac_type))
        balance_operation.full_clean()
        balance_operation.save()

        return super(PackageSell, self).delete(request, *args, **kwargs)

    def get_success_url(self):
        return reverse('fin:contr-packages', args=[self.object.contractor.pk,])


@permission_required('fin.fin_reader')
def FuelReport (request):

    class FuelReportForm (Form):

        initial_year = (date.today().year if date.today().month != 1 else date.today().year - 1)
        initial_month = (date.today().month - 1 if date.today().month != 1 else 12)
        initial_start = date(year=initial_year, month=initial_month, day=1)
        next_month = initial_start.replace(day=28) + timedelta(days=4)
        initial_end = next_month - timedelta(days=next_month.day)

        report_type = ChoiceField(choices=[(0, 'Raport pełny'), (1, 'Tylko akcyzowe'), (2, 'Tylko bezakcyzowe')],
                                  label='Rodzaj raportu')
        report_year = ChoiceField(choices=[(year, year) for year in range(2016, date.today().year + 1)],
                                  initial=initial_year, label='Rok raportu')
        report_month = ChoiceField(choices=[(month, month) for month in range(1, 13)],
                                   initial=initial_month, label='Miesiąc raportu')
        report_start = DateField(initial=initial_start, label='Data początkowa')
        report_end = DateField(initial=initial_end, label='Data końcowa')

    MyForm = FuelReportForm
    form = MyForm(request.POST or None)

    context = {}
    context['title'] = 'Generacja raportu paliwa'
    context['form'] = form

    # jeśli formularz został poprawnie wypełniony
    if form.is_valid():

        # przejście do kolejnego formularza
        return HttpResponseRedirect(reverse('fin:fuel-export', args=[form.cleaned_data['report_type'],
                                                                     form.cleaned_data['report_start'], form.cleaned_data['report_end']]))

    return render(request, 'fin/file_export.html', context)


@permission_required('fin.fin_reader')
def FuelExport (request, type, report_start, report_end):

    # Nowy arkusz
    wb = Workbook()

    # Zakładki odpowiadające statkom powiertrznym
    i = 0
    for aircraft in Aircraft.objects.order_by('registration'):
        if i == 0:
            wb.active.title = aircraft.registration
        else:
            wb.create_sheet(aircraft.registration)
        i += 1

    # Definicje stylów
    border = Border(top=Side(border_style=borders.BORDER_THIN, color='000000'), bottom=Side(border_style=borders.BORDER_THIN, color='000000'),
                    left=Side(border_style=borders.BORDER_THIN, color='000000'), right=Side(border_style=borders.BORDER_THIN, color='000000'))
    title_align = Alignment(horizontal='center', vertical='center', wrap_text = True)
    table_align = Alignment(vertical='center', wrap_text=True)
    green_fill = PatternFill(fill_type=fills.FILL_SOLID, fgColor=('90EE90'))
    yellow_fill = PatternFill(fill_type=fills.FILL_SOLID, fgColor=('FFFF00'))
    red_fill = PatternFill(fill_type=fills.FILL_SOLID, fgColor=('FF7F50'))

    # Wypełnienie poszczególnych zakładek
    for ws in wb.worksheets:
        # Informacje nagłówkowe

        ws['A1'] = "Od"
        ws['B1'] = "Do"
        ws['C1'] = "Paliwo uzup."
        ws['D1'] = "Stan do lotu"
        ws['E1'] = "Paliwo zużyte"
        ws['F1'] = "Olej uzup."
        ws['G1'] = "Off-block"
        ws['H1'] = "On-block"
        ws['I1'] = "Załoga"
        ws['J1'] = "Instruktor"
        ws['K1'] = "Data"
        ws['L1'] = "Czas lotu"
        ws['M1'] = "Liczba lądowań"
        ws['N1'] = "Licznik"
        ws['O1'] = "Rodzaj lotu"

        # Format nagłówków
        for col in ['A', 'B', 'C', 'D', 'E', 'F','G','H','I','J','K','L','M','N','O']:
            ws['%s1' % col].alignment = title_align
            ws['%s1' % col].border = border
            ws['%s1' % col].font = Font(size=12, bold=True)

        # Wybór zakresu danych do raportu
        query = Operation.objects.filter(pdt__aircraft__registration = ws.title, pdt__date__range=(report_start, report_end), status='closed')
        if type == '0':
            operations = query
        elif type == '1':
            operations = query.filter(pdt__flight_type='04')
        elif type == '2':
            operations = query.exclude(pdt__flight_type='04')
        else:
            operations = query.none()

        # Zawartość zakładek
        row = 2
        for oper in operations.order_by('pdt__date'):

            # Wypełnienie wierszy
            ws['A%d' % row] = oper.loc_start
            ws['B%d' % row] = oper.loc_end
            ws['C%d' % row] = oper.fuel_refill
            ws['D%d' % row] = oper.fuel_available
            ws['E%d' % row] = oper.fuel_used
            ws['F%d' % row] = oper.oil_refill
            ws['G%d' % row] = oper.time_start.strftime('%H:%M')
            ws['H%d' % row] = oper.time_end.strftime('%H:%M')
            ws['I%d' % row] = oper.pdt.pic.fbouser.second_name + (' / %s' % oper.pdt.sic.fbouser.second_name if oper.pdt.sic else '')
            ws['J%d' % row] = (oper.pdt.instructor.fbouser.second_name if oper.pdt.instructor else '')
            ws['K%d' % row] = oper.pdt.date
            ws['L%d' % row] = '{:02d}:{:02d}'.format(oper.hours()[1], oper.hours()[2])
            ws['M%d' % row] = oper.landings
            ws['N%d' % row] = oper.tth_end - oper.tth_start
            ws['O%d' % row] = oper.pdt.flight_type_name()
            row += 1

        # Ustawienie szerokości kolumn
        ws.column_dimensions['A'].width = '10'
        ws.column_dimensions['B'].width = '10'
        ws.column_dimensions['C'].width = '10'
        ws.column_dimensions['D'].width = '10'
        ws.column_dimensions['E'].width = '10'
        ws.column_dimensions['F'].width = '10'
        ws.column_dimensions['G'].width = '10'
        ws.column_dimensions['H'].width = '10'
        ws.column_dimensions['I'].width = '30'
        ws.column_dimensions['J'].width = '14'
        ws.column_dimensions['K'].width = '12'
        ws.column_dimensions['L'].width = '10'
        ws.column_dimensions['M'].width = '10'
        ws.column_dimensions['N'].width = '10'
        ws.column_dimensions['O'].width = '22'

        # Formatowanie komórek tabeli
        max_row = row
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I','J','K','L','M','N','O']:
            for row in range(2, max_row):
                ws['%s%d' % (col, row)].border = border
                ws['%s%d' % (col, row)].alignment = title_align

    # Zwróć arkusz jako obiekt HTTP Response
    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = ('attachment; filename=Rozliczenie_Paliwa_%s_%s.xlsx' % (report_start, report_end))

    return response
