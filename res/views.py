import json, babel.dates
from datetime import datetime, timedelta, date
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, fills, Border, borders, Side, Font
from io import BytesIO
from django import forms
from django.conf import settings
from django.db.models import Q
from django.shortcuts import render
from django.core.mail import EmailMessage
from django.urls import reverse
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.views.generic.list import ListView
from django.views.generic.detail import DetailView
from django.forms.models import modelform_factory
from django.forms.widgets import DateTimeInput
from django.forms import Textarea
from django.contrib.auth.decorators import permission_required
from django.template.response import TemplateResponse, HttpResponse

from django.utils.decorators import method_decorator
def class_view_decorator(function_decorator):

    def simple_decorator(view):
        view.dispatch = method_decorator(function_decorator)(view.dispatch)
        return view

    return simple_decorator

from salt.forms import duration_string
from res.forms import ReservationForm, ReservationFBOForm

from panel.models import FBOUser, FlightTypes, Pilot, PilotAircraft, Operation, Duty
from res.models import ResourceFBO, Reservation, ReservationFBO, Blackout
from camo.models import Aircraft


def reservation_open_limit(reservation):

    # limit czasu, na jaki można otworzyć PDT przed rozpoczęciem rezerwacji
    if reservation.aircraft.helicopter:
        open_limit = 60 * 60 * 3
    else:
        open_limit = 60 * 60

    return open_limit


def reservation_check(reservation, fbouser):

    # sprawdzenie, czy można otworzyć PDT
    open_pdt = (reservation.status in ('Nowa', 'Potwierdzona')) and \
               hasattr(fbouser, 'pilot') and \
               (datetime.now() > reservation.start_time - timedelta(seconds=reservation_open_limit(reservation))) and \
               (datetime.now() < reservation.end_time + timedelta(days=1)) and \
               (reservation.aircraft.status == 'flying') and \
               reservation.aircraft.airworthy() and \
               not fbouser.open_pdt() and \
               not reservation.aircraft.pdt_set.filter(status='open')

    return open_pdt


def reservation_msg(reservation, fbouser):

    # Generacja komunikatu, dlaczego nie można otworzyć PDT
    if not (reservation.status in ('Nowa', 'Potwierdzona')):
        res_msg = 'Rezerwacja została anulowana.'
    elif not (hasattr(fbouser, 'pilot')):
        res_msg = 'Zalogowany użytkownik nie jest pilotem.'
    elif not (datetime.now() > reservation.start_time - timedelta(seconds=reservation_open_limit(reservation))):
        if reservation.aircraft.helicopter:
            res_msg = 'Do rozpoczęcia rezerwacji pozostało ponad 3 godziny.'
        else:
            res_msg = 'Do rozpoczęcia rezerwacji pozostała ponad godzina.'
    elif not (datetime.now() < reservation.end_time + timedelta(days=1)):
        res_msg = 'Od zakończenia rezerwacji minęła ponad doba.'
    elif not (reservation.aircraft.status == 'flying'):
        res_msg = 'Statek powietrzny jest niesprawny.'
    elif not (reservation.aircraft.airworthy()):
        res_msg = 'Statek powietrzny nie ma ważnego MS.'
    elif fbouser.open_pdt():
        res_msg = 'Zalogowany użytkownik ma otwarty inny PDT.'
    elif reservation.aircraft.pdt_set.filter(status='open'):
        res_msg = 'Dla statku powietrznego jest otwarty inny PDT.'
    else:
        res_msg = ''

    return res_msg


@permission_required('res.res_user')
def ReservationList(request):

    this_day = datetime(datetime.today().year, datetime.today().month, datetime.today().day)

    if request.user.has_perm('res.res_admin'):
        # Jeśli jest administratorem rezerwacji
        query_ac = Reservation.objects.exclude(end_time__lt = this_day).exclude(status = 'Zrealizowana').order_by('start_time')
        query_fbo = ReservationFBO.objects.exclude(end_time__lt=this_day).order_by('start_time')
    else:
        # Jeśli jest pilotem
        if hasattr(request.user.fbouser, 'pilot'):
            # Wybierz te, które utworzyłeś lub jesteś właścicielem
            query_ac = Reservation.objects.filter(Q(owner=request.user.fbouser.pilot) | Q(participant=request.user.fbouser.pilot) | Q(open_user=request.user.fbouser)).\
                    exclude(end_time__lt = this_day).exclude(status = 'Zrealizowana').order_by('start_time')
        else:
            # Pozostałe przypadki
            query_ac = Reservation.objects.none()

        # Wybierz te, które utworzyłeś lub jesteś właścicielem
        query_fbo = ReservationFBO.objects.filter(Q(owner=request.user.fbouser) | Q(participant=request.user.fbouser) | Q(open_user=request.user.fbouser)).\
                    exclude(end_time__lt = this_day).order_by('start_time')

    # lista rezerwacji na podstawie query_ac i query_fbo
    object_list = []
    for res in query_ac:
        object_list.append({'res': res, 'pk':res.pk, 'resource':res.aircraft, 'start_time': res.start_time,
                            'end_time': res.end_time, 'owner': res.owner, 'participant': res.participant,
                            'title': FlightTypes()[res.planned_type] if res.planned_type else '',
                            'planned_time': res.planned_time, 'loc_start': res.loc_start, 'loc_stop': res.loc_stop,
                            'loc_end': res.loc_end, 'remarks': res.remarks, 'internal_remarks': res.internal_remarks,
                            'status': res.status, 'fbo': False})
    for res in query_fbo:
        object_list.append({'res': res, 'pk':res.pk, 'resource':res.resource, 'start_time': res.start_time,
                            'end_time': res.end_time, 'owner': res.owner, 'participant': res.participant,
                            'title': res.title, 'planned_time': None, 'loc_start': '', 'loc_stop': '', 'loc_end': '',
                            'remarks': res.remarks, 'internal_remarks': '', 'status': 'Brak', 'fbo': True})

    object_list.sort(key=lambda k: k['start_time'])

    context = {}
    context['page_title'] = 'Rezerwacje'
    context['header_text'] = 'Lista aktualnych rezerwacji'
    context['empty_text'] = 'Brak rezerwacji.'

    # Lokalne menu
    local_menu = []
    if hasattr(request.user.fbouser, 'pilot'):
        local_menu.append({'text': 'Nowa rezerwacja SP', 'path': reverse("res:reservation-create")})
    if request.user.fbouser.infos:
        local_menu.append({'text': 'Nowa rezerwacja inna', 'path': reverse("res:resfbo-create")})
    local_menu.append({'text': 'Kalendarz', 'path': reverse("res:reservation-calendar")})
    context['local_menu'] = local_menu

    # Nagłówki tabeli
    header_list = []
    header_list.append({'header': 'Zasób'})
    header_list.append({'header': 'Data\nrezerwacji'})
    header_list.append({'header': 'Termin\nrezerwacji'})
    header_list.append({'header': 'Właściciel\nrezerwacji'})
    header_list.append({'header': 'Uczestnik\nrezerwacji'})
    header_list.append({'header': 'Tytuł\nrezerwacji'})
    header_list.append({'header': 'Status'})
    header_list.append({'header': 'Uwagi'})
    if request.user.has_perm('res.res_admin') or request.user.fbouser.infos:
        header_list.append({'header': 'Uwagi SALT'})
    header_list.append({'header': '...'})
    context['header_list'] = header_list

    row_list = []
    for object in object_list:
        fields = []
        days_length = (object['end_time'].date() - object['start_time'].date()).days
        fields.append({'name': 'resource', 'value': object['resource'],
                       'link': reverse(('res:resfbo-info' if object['fbo'] else 'res:reservation-info'), args=[object['pk']]), 'just': 'center'})
        fields.append({'name': 'date', 'value': babel.dates.format_date(object['start_time'], "EEE d MMMM", locale='pl_PL')})
        fields.append({'name': 'time', 'value': "%s - %s" % (object['start_time'].strftime("%H:%M"),
                                                             object['end_time'].strftime("%H:%M") + (' (+%d)' % days_length if days_length else ''))})
        fields.append({'name': 'owner', 'value': object['owner']})
        fields.append({'name': 'participant', 'value': object['participant']})
        fields.append({'name': 'title', 'value': object['title']})
        fields.append({'name': 'status', 'value': object['status'], 'just': 'center'})
        fields.append({'name': 'remarks', 'value': object['remarks']})
        if request.user.has_perm('res.res_admin') or request.user.fbouser.infos:
            fields.append({'name': 'remarks', 'value': object['internal_remarks']})

        open_pdt = (not object['fbo']) and reservation_check(object['res'], request.user.fbouser)

        fields.append({'name': 'change', 'view_link': reverse('res:reservation-cal-day', args=[object['start_time'].year,object['start_time'].month,object['start_time'].day]),
                       'report_link': ('%s?res=%d' % (reverse('panel:pdt-open'), object['pk'])) if open_pdt else None,
                       'edit_link': reverse('res:resfbo-update', args=[object['pk']]) if object['fbo']
                                    else (reverse('res:reservation-update', args=[object['pk']]) if object['status'] in ('Nowa', 'Potwierdzona') else None),
                       'delete_link': reverse('res:resfbo-delete', args=[object['pk']]) if object['fbo']
                                      else (reverse('res:reservation-delete', args=[object['pk']]) if object['status'] in ('Nowa', 'Potwierdzona') else None)})

        row_list.append({'fields': fields})

    context['row_list'] = row_list
    template_name = 'res/list_template.html'

    return render(request, template_name, context)


@permission_required('res.res_user')
def ReservationListMobile(request):

    context={}
    this_day = datetime(datetime.today().year, datetime.today().month, datetime.today().day)

    if hasattr(request.user.fbouser, 'pilot'):
        # Jeśli jest pilotem
        # Wybierz te, w których pilot uczestniczy
        query_ac = Reservation.objects.filter(Q(owner=request.user.fbouser.pilot) | Q(participant=request.user.fbouser.pilot)).\
                   exclude(end_time__lt = this_day).exclude(status = 'Zrealizowana').order_by('start_time')
    else:
        # Pozostałe przypadki
        query_ac = Reservation.objects.none()

    query_fbo = ReservationFBO.objects.filter(Q(owner=request.user.fbouser) | Q(participant=request.user.fbouser)).\
                exclude(end_time__lt = this_day).order_by('start_time')

    # lista rezerwacji na podstawie query_ac i query_fbo
    object_list = []
    for res in query_ac:
        object_list.append(
            {'res': res, 'pk': res.pk, 'resource': res.aircraft, 'start_time': res.start_time, 'end_time': res.end_time,
             'owner': res.owner.fbouser, 'participant': res.participant.fbouser if res.participant else None,
             'title': FlightTypes()[res.planned_type] if res.planned_type else '',
             'planned_time': res.planned_time, 'loc_start': res.loc_start, 'loc_stop': res.loc_stop,
             'loc_end': res.loc_end, 'remarks': res.remarks, 'internal_remarks': res.internal_remarks,
             'status': res.status, 'fbo': False})
    for res in query_fbo:
        object_list.append(
            {'res': res, 'pk': res.pk, 'resource': res.resource, 'start_time': res.start_time, 'end_time': res.end_time,
             'owner': res.owner, 'participant': res.participant, 'title': res.title,
             'planned_time': None, 'loc_start': '', 'loc_stop': '', 'loc_end': '', 'remarks': res.remarks,
             'internal_remarks': '', 'status': 'Brak', 'fbo': True})

    object_list.sort(key=lambda k: k['start_time'])

    context['page_title'] = 'Lista rezerwacji'
    context['header_text'] = 'Lista aktualnych rezerwacji'
    context['empty_text'] = 'Brak rezerwacji...'

    # Lokalne menu
    local_menu = {}
    local_menu['return_path'] = reverse("dispatcher")
    if hasattr(request.user.fbouser, 'pilot'):
        local_menu['create_path'] = reverse("res:reservation-create")
    else:
        local_menu['create_path'] = ''
    local_menu['cal_path'] = reverse("res:reservation-calendar")
    context['local_menu'] = local_menu

    # Wiersze tabeli
    row_list = []
    for object in object_list:
        days_length = (object['end_time'].date() - object['start_time'].date()).days
        fields = {}
        fields['line1'] = "<b>%s</b>, %s-%s" % (babel.dates.format_date(object['start_time'], "EEE d MMMM", locale='pl_PL'),
                                           object['start_time'].strftime("%H:%M"),
                                           object['end_time'].strftime("%H:%M") + (' (+%d)' % days_length if days_length else ''))
        fields['line2'] = "<span style='color: darkred'><b>%s</b></span>, %s" % (object['resource'], object['title'])
        fields['line3'] = (object['owner'].__str__() if object['owner'] != request.user.fbouser else '')
        if object['participant']:
            if object['participant'] != request.user.fbouser:
                if fields['line3']:
                    fields['line3'] += '/'
                fields['line3'] += object['participant'].__str__()
        fields['line4'] = object['remarks']

        if not object['fbo']:
            open_pdt = reservation_check(object['res'], request.user.fbouser)
        else:
            open_pdt = None

        if open_pdt:
            fields['pdt_link'] = '%s?res=%d' % (reverse('panel:mpdt-open1'), object['pk'])

        if object['fbo']:
            fields['edit_link'] = reverse('res:resfbo-update', args=[object['pk']])
        else:
            fields['edit_link'] = reverse('res:reservation-update', args=[object['pk']])

        row_list.append({'fields': fields})

    context['row_list'] = row_list

    return render(request, 'res/mres_list.html', context)


@class_view_decorator(permission_required('res.res_user'))
class ReservationInfo (DetailView):
    model = Reservation
    template_name = 'res/details_template.html'

    def get_context_data(self, **kwargs):
        context = super(ReservationInfo, self).get_context_data(**kwargs)
        context['page_title'] = 'Rezerwacja'
        context['header_text'] = 'Szczegóły rezerwacji %s' % self.object.aircraft
        calendar = ('cal' in self.request.GET)

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Pokaż w kalendarzu',
                           'path': reverse('res:reservation-cal-day', args=[self.object.start_time.year,self.object.start_time.month,self.object.start_time.day])})
        if self.object.status in ('Nowa', 'Potwierdzona'):
            if self.request.user.has_perm('res.res_admin') or \
               (hasattr(self.request.user.fbouser, 'pilot') and (self.request.user.fbouser.pilot in (self.object.owner, self.object.participant))) or \
               (self.object.open_user == self.request.user.fbouser):
                local_menu.append({'text': 'Aktualizuj', 'path': '%s%s' % (reverse('res:reservation-update', args=[self.object.pk]), '?cal=1' if calendar else '')})
                local_menu.append({'text': 'Usuń', 'path': '%s%s' % (reverse('res:reservation-delete', args=[self.object.pk]), '?cal=1' if calendar else '')})
        context['local_menu'] = local_menu

        # Sprawdzenie, czy na podstawie rezerwacji użytkownik może otworzyć PDT
        if reservation_check(self.object, self.request.user.fbouser):
            context['pdt_link'] = '%s?res=%d' % (reverse('panel:pdt-open'), self.object.pk)
        else:
            context['pdt_msg'] = reservation_msg(self.object, self.request.user.fbouser)

        field_list = []
        field_list.append({'header': 'Statek powietrzny', 'value': self.object.aircraft, 'bold': True})
        field_list.append({'header': 'Termin rozpoczęcia', 'value': '{:%Y-%m-%d %H:%M}'.format(self.object.start_time)})
        field_list.append({'header': 'Termin zakończenia', 'value': '{:%Y-%m-%d %H:%M}'.format(self.object.end_time)})
        field_list.append({'header': 'Właściciel rezerwacji', 'value': self.object.owner})
        if self.object.participant:
            field_list.append({'header': 'Uczestnik rezerwacji', 'value': self.object.participant})
        if self.object.planned_type:
            field_list.append({'header': 'Planowany rodzaj lotu', 'value': FlightTypes()[self.object.planned_type]})
        field_list.append({'header': 'Planowany czas lotu', 'value': duration_string(self.object.planned_time)})
        field_list.append({'header': 'Lotnisko startu', 'value': self.object.loc_start})
        field_list.append({'header': 'Pierwszy odcinek', 'value': self.object.loc_stop})
        if self.object.loc_end:
            field_list.append({'header': 'Dalsze odcinki', 'value': self.object.loc_end})
        if self.object.pax:
            field_list.append({'header': 'Lista pasażerów', 'value': self.object.pax})
        if self.object.remarks:
            field_list.append({'header': 'Uwagi', 'value': self.object.remarks})
        if (self.request.user.has_perm('res.res_admin') or self.request.user.fbouser.infos) and self.object.internal_remarks:
            field_list.append({'header': 'Uwagi SALT', 'value': self.object.internal_remarks})
        field_list.append({'header': 'Status', 'value': self.object.status})
        field_list.append({'header': 'Otwarta przez', 'value': '%s - %s' % (self.object.open_user, self.object.open_time.strftime('%y/%m/%d %X'))})
        if self.object.change_user:
            field_list.append({'header': 'Zmieniona przez', 'value': '%s - %s' % (self.object.change_user, self.object.change_time.strftime('%y/%m/%d %X'))})

        context['field_list'] = field_list
        return context


@class_view_decorator(permission_required('res.res_user'))
class ReservationCreate (CreateView):
    model = Reservation
    form_class = ReservationForm

    def get_template_names(self):
        if self.request.device['is_mobile']:
            template_name = 'res/mres_modify.html'
        else:
            template_name = 'res/create_template.html'
        return template_name

    def get_context_data(self, **kwargs):
        context = super(ReservationCreate, self).get_context_data(**kwargs)
        context['page_title'] = 'Utworzenie rezerwacji'
        context['header_text'] = 'Utworzenie nowej rezerwacji'
        context['update'] = False
        return context

    def get_form_kwargs(self):
        kwargs = super(ReservationCreate, self).get_form_kwargs()
        kwargs['is_mobile'] = self.request.device['is_mobile']
        kwargs['user'] = self.request.user
        kwargs['res'] = self.request.GET.get('res')
        kwargs['start'] = self.request.GET.get('start')
        kwargs['end'] = self.request.GET.get('end')
        return kwargs

    def form_valid(self, form):
        form.instance.open_user = self.request.user.fbouser
        form.instance.status = 'Nowa'
        return super(ReservationCreate, self).form_valid(form)

    def get_success_url(self):
        if 'res' in self.request.GET:
            # Jeśli wywołanie było z kalendarza to wróć do kalendarza
            return reverse('res:reservation-cal-day', args=[self.object.start_time.year, self.object.start_time.month,
                                                            self.object.start_time.day])
        else:
            # Wróć do listy rezerwacji
            if self.request.device['is_mobile']:
                return reverse('res:mobile-res')
            else:
                return reverse('res:reservation-list')


@class_view_decorator(permission_required('res.res_user'))
class ReservationUpdate (UpdateView):
    model = Reservation
    form_class = ReservationForm

    def get_template_names(self):
        if self.request.device['is_mobile']:
            template_name = 'res/mres_modify.html'
        else:
            template_name = 'res/update_template.html'
        return template_name

    def get_context_data(self, **kwargs):
        context = super(ReservationUpdate, self).get_context_data(**kwargs)
        context['page_title'] = 'Modyfikacja rezerwacji'
        context['header_text'] = 'Modyfikacja istniejącej rezerwacji'
        context['update'] = True
        context['pk'] = self.object.pk
        return context

    def get_form_kwargs(self):
        kwargs = super(ReservationUpdate, self).get_form_kwargs()
        kwargs['is_mobile'] = self.request.device['is_mobile']
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        form.instance.change_user = self.request.user.fbouser
        return super(ReservationUpdate, self).form_valid(form)

    def get_success_url(self):
        if 'cal' in self.request.GET:
            # Jeśli wywołanie było z kalendarza to wróć do kalendarza
            return reverse('res:reservation-cal-day', args=[self.object.start_time.year, self.object.start_time.month,
                                                            self.object.start_time.day])
        else:
            # Wróć do listy rezerwacji
            if self.request.device['is_mobile']:
                return reverse('res:mobile-res')
            else:
                return reverse('res:reservation-list')


@class_view_decorator(permission_required('res.res_user'))
class ReservationDelete (DeleteView):
    model = Reservation

    def get_template_names(self):
        if self.request.device['is_mobile']:
            template_name = 'res/mres_delete.html'
        else:
            template_name = 'res/delete_template.html'
        return template_name

    def get_context_data(self, **kwargs):
        context = super(ReservationDelete, self).get_context_data(**kwargs)
        context['page_title'] = 'Usunięcie rezerwacji'
        context['header_text'] = 'Usunięcie rezerwacji %s' % self.object.aircraft
        context['object_name'] = "%s / %s" % (self.object.aircraft, self.object.planned_type)
        context['description'] = self.object.remarks
        return context

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        self.object.change_user = request.user.fbouser
        self.object.save()
        return super(ReservationDelete, self).delete(request, *args, **kwargs)

    def get_success_url(self):
        if 'cal' in self.request.GET:
            # Jeśli wywołanie było z kalendarza to wróć do kalendarza
            return reverse('res:reservation-cal-day', args=[self.object.start_time.year, self.object.start_time.month,
                                                            self.object.start_time.day])
        else:
            # Wróć do listy rezerwacji
            if self.request.device['is_mobile']:
                return reverse('res:mobile-res')
            else:
                return reverse('res:reservation-list')


@class_view_decorator(permission_required('res.res_user'))
class ReservationFBOInfo (DetailView):
    model = ReservationFBO
    template_name = 'res/details_template.html'

    def get_context_data(self, **kwargs):
        context = super(ReservationFBOInfo, self).get_context_data(**kwargs)
        context['page_title'] = 'Rezerwacja'
        context['header_text'] = 'Szczegóły rezerwacji %s' % self.object.resource
        calendar = ('cal' in self.request.GET)

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Pokaż w kalendarzu',
                           'path': reverse('res:reservation-cal-day', args=[self.object.start_time.year,self.object.start_time.month,self.object.start_time.day])})
        if self.request.user.has_perm('res.res_admin') or \
                (self.request.user.fbouser in (self.object.owner, self.object.participant)) or \
                (self.object.open_user == self.request.user.fbouser):
            local_menu.append({'text': 'Aktualizuj', 'path': '%s%s' % (reverse('res:resfbo-update', args=[self.object.pk]), '?cal=1' if calendar else '')})
            local_menu.append({'text': 'Usuń', 'path': '%s%s' % (reverse('res:resfbo-delete', args=[self.object.pk]), '?cal=1' if calendar else '')})
        context['local_menu'] = local_menu

        field_list = []
        field_list.append({'header': 'Zasób', 'value': self.object.resource, 'bold': True})
        field_list.append({'header': 'Termin rozpoczęcia', 'value': '{:%Y-%m-%d %H:%M}'.format(self.object.start_time)})
        field_list.append({'header': 'Termin zakończenia', 'value': '{:%Y-%m-%d %H:%M}'.format(self.object.end_time)})
        field_list.append({'header': 'Właściciel rezerwacji', 'value': self.object.owner})
        if self.object.participant:
            field_list.append({'header': 'Uczestnik rezerwacji', 'value': self.object.participant})
        field_list.append({'header': 'Tytuł rezerwacji', 'value': self.object.title})
        if self.object.remarks:
            field_list.append({'header': 'Uwagi', 'value': self.object.remarks})
        field_list.append({'header': 'Otwarta przez', 'value': '%s - %s' % (self.object.open_user, self.object.open_time.strftime('%y/%m/%d %X'))})
        if self.object.change_user:
            field_list.append({'header': 'Zmieniona przez', 'value': '%s - %s' % (self.object.change_user, self.object.change_time.strftime('%y/%m/%d %X'))})

        context['field_list'] = field_list
        return context


@class_view_decorator(permission_required('res.res_user'))
class ReservationFBOCreate (CreateView):
    model = ReservationFBO
    form_class = ReservationFBOForm

    def get_template_names(self):
        if self.request.device['is_mobile']:
            template_name = 'res/mresfbo_modify.html'
        else:
            template_name = 'res/create_template.html'
        return template_name

    def get_context_data(self, **kwargs):
        context = super(ReservationFBOCreate, self).get_context_data(**kwargs)
        context['page_title'] = 'Utworzenie rezerwacji'
        context['header_text'] = 'Utworzenie nowej rezerwacji'
        context['update'] = False
        return context

    def get_form_kwargs(self):
        kwargs = super(ReservationFBOCreate, self).get_form_kwargs()
        kwargs['is_mobile'] = self.request.device['is_mobile']
        kwargs['user'] = self.request.user
        kwargs['res'] = self.request.GET.get('res')
        kwargs['start'] = self.request.GET.get('start')
        kwargs['end'] = self.request.GET.get('end')
        return kwargs

    def form_valid(self, form):
        form.instance.open_user = self.request.user.fbouser
        return super(ReservationFBOCreate, self).form_valid(form)

    def get_success_url(self):
        if 'res' in self.request.GET:
            # Jeśli wywołanie było z kalendarza to wróć do kalendarza
            return reverse('res:reservation-cal-day', args=[self.object.start_time.year, self.object.start_time.month,
                                                            self.object.start_time.day])
        else:
            # Wróć do listy rezerwacji
            if self.request.device['is_mobile']:
                return reverse('res:mobile-res')
            else:
                return reverse('res:reservation-list')


@class_view_decorator(permission_required('res.res_user'))
class ReservationFBOUpdate (UpdateView):
    model = ReservationFBO
    form_class = ReservationFBOForm

    def get_template_names(self):
        if self.request.device['is_mobile']:
            template_name = 'res/mresfbo_modify.html'
        else:
            template_name = 'res/update_template.html'
        return template_name

    def get_context_data(self, **kwargs):
        context = super(ReservationFBOUpdate, self).get_context_data(**kwargs)
        context['page_title'] = 'Modyfikacja rezerwacji'
        context['header_text'] = 'Modyfikacja istniejącej rezerwacji'
        context['update'] = True
        context['pk'] = self.object.pk
        return context

    def get_form_kwargs(self):
        kwargs = super(ReservationFBOUpdate, self).get_form_kwargs()
        kwargs['is_mobile'] = self.request.device['is_mobile']
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        form.instance.change_user = self.request.user.fbouser
        return super(ReservationFBOUpdate, self).form_valid(form)

    def get_success_url(self):
        if 'cal' in self.request.GET:
            # Jeśli wywołanie było z kalendarza to wróć do kalendarza
            return reverse('res:reservation-cal-day', args=[self.object.start_time.year, self.object.start_time.month,
                                                            self.object.start_time.day])
        else:
            # Wróć do listy rezerwacji
            if self.request.device['is_mobile']:
                return reverse('res:mobile-res')
            else:
                return reverse('res:reservation-list')


@class_view_decorator(permission_required('res.res_user'))
class ReservationFBODelete (DeleteView):
    model = ReservationFBO

    def get_template_names(self):
        if self.request.device['is_mobile']:
            template_name = 'res/mres_delete.html'
        else:
            template_name = 'res/delete_template.html'

        return template_name

    def get_context_data(self, **kwargs):
        context = super(ReservationFBODelete, self).get_context_data(**kwargs)
        context['page_title'] = 'Usunięcie rezerwacji'
        context['header_text'] = 'Usunięcie rezerwacji %s' % self.object.resource
        context['object_name'] = "%s / %s" % (self.object.resource, self.object.title)
        context['description'] = self.object.remarks
        return context

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        self.object.change_user = request.user.fbouser
        self.object.save()
        return super(ReservationFBODelete, self).delete(request, *args, **kwargs)

    def get_success_url(self):
        if 'cal' in self.request.GET:
            # Jeśli wywołanie było z kalendarza to wróć do kalendarza
            return reverse('res:reservation-cal-day', args=[self.object.start_time.year, self.object.start_time.month,
                                                            self.object.start_time.day])
        else:
            # Wróć do listy rezerwacji
            if self.request.device['is_mobile']:
                return reverse('res:mobile-res')
            else:
                return reverse('res:reservation-list')


@permission_required('res.res_user')
def ReservationCalendar(request, year=None, month=None, day=None):
    context={'mobile': request.device['is_mobile']}
    context['resources'] = ''

    if year and month and day:
        context['DefaultDate'] = '%s-%s-%s' % (year, month.zfill(2), day.zfill(2))

    # Dodaj Dyżur jako pierwszy zasób
    try:
        on_duty = ResourceFBO.objects.get(name='Dyżur')
    except:
        on_duty = None
    if on_duty:
        context['resources'] += "{id:'fbo%s', title:'%s', status:'%s', order:%d, type:'%s', info:'%s'}" \
                                % (on_duty.pk, on_duty.name, '', 0, '', on_duty.info.replace('\r', '') if on_duty.info else '')

    # Dla użytkowników z dostępem do INFOS dodaj INFOS i zasoby FBO
    if request.user.fbouser.infos:
        # Dodaj INFOS własny
        try:
            infos = ResourceFBO.objects.get(name='INFOS')
        except:
            infos = None
        if infos:
            if context['resources'] != '':
                context['resources'] += ','
            context['resources'] += "{id:'%s', title:'%s', status:'%s', order:%d, type:'%s', info:'%s'}" \
                                    % ('infos', 'INFOS', '', 1, '', infos.info.replace('\r', '') if infos.info else '')

        # Dodaj INFOS pozostałych użytkowników
        for fbouser in FBOUser.objects.filter(infos=True).exclude(pk=request.user.fbouser.pk).order_by('second_name'):
            if context['resources'] != '':
                context['resources'] += ','
            context['resources'] += "{id:'usr%s', parentId:'%s', title:'%s', status:'%s', order:%d, type:'%s', info:'%s'}" \
                                    % (fbouser.pk, 'infos', fbouser.second_name, fbouser.first_name, 1, '', '')

        # Dodaj pozostałe zasoby FBO
        for resource in ResourceFBO.objects.exclude(name='INFOS').exclude(name='Dyżur').filter(scheduled=True):
            if context['resources'] != '':
                context['resources'] += ','
            context['resources'] += "{id:'fbo%s', title:'%s', status:'%s', order:%d, type:'%s', info:'%s'}" \
                                    % (resource.pk, resource.name, '', 0 if resource.name == 'Dyżur' else 2, '',
                                       resource.info.replace('\r','') if resource.info else '')

    # Jeśli administrator albo nie pilot
    if request.user.has_perm('res.res_admin') or not hasattr(request.user.fbouser, 'pilot'):
        # Wybierz wszystkie AC
        aircraft_set = [aircraft for aircraft in Aircraft.objects.filter(scheduled = True)]
    else:
        # W pozostałych przypadkach wybierz tylko jego AC
        aircraft_set = [aircraft.aircraft for aircraft in PilotAircraft.objects.filter(pilot=request.user.fbouser.pilot).filter(aircraft__scheduled = True)]

    # Dodaj listę AC
    for aircraft in aircraft_set:
        if not aircraft.status == 'flying':
            status = 'niesprawny'
            order = 5
        elif not aircraft.airworthy():
            status = 'brak MS'
            order = 4
        else:
            status = '%.2f TTH' % ((aircraft.ms_hours or 99999) - aircraft.hours_count)
            order = 3
        if context['resources'] != '':
            context['resources'] += ','
        context['resources'] += "{id:'%s', title:'%s', status:'%s', order:%d, type:'%s', info:'%s'}" \
                                % (aircraft.pk, aircraft, status, order, aircraft.type,
                                   aircraft.info.replace('\r','') if aircraft.info else '')

    if request.device['is_mobile']:
        return render(request, 'res/mres_calendar.html', context)
    else:
        return render(request, 'res/res_calendar.html', context)


@class_view_decorator(permission_required('res.res_admin'))
class BlackoutList (ListView):
    model = Blackout
    template_name = 'res/list_template.html'

    def get_queryset(self):
        this_day = datetime(datetime.today().year, datetime.today().month, datetime.today().day)
        query = Blackout.objects.exclude(end_time__lt = this_day)
        return query

    def get_context_data(self, **kwargs):
        context = super(BlackoutList, self).get_context_data(**kwargs)
        context['page_title'] = 'Blokady'
        context['header_text'] = 'Lista aktualnych blokad'
        context['empty_text'] = 'Brak blokad.'

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Nowa blokada', 'path': reverse("res:blackout-create")})
        local_menu.append({'text': 'Kalendarz', 'path': reverse("res:reservation-calendar")})
        context['local_menu'] = local_menu

        # Nagłówki tabeli
        header_list = []
        header_list.append({'header': 'Statek\npowietrzny'})
        header_list.append({'header': 'Termin\nrozpoczęcia'})
        header_list.append({'header': 'Termin\nzakończenia'})
        header_list.append({'header': 'Uwagi'})
        header_list.append({'header': '...'})
        context['header_list'] = header_list

        row_list = []
        for object in self.object_list:
            fields = []
            fields.append({'name': 'aircraft', 'value': object.aircraft, 'link': reverse('res:blackout-info', args=[object.pk]), 'just': 'center'})
            fields.append({'name': 'start_time', 'value': object.start_time.strftime("%y/%m/%d %H:%M")})
            fields.append({'name': 'end_time', 'value': object.end_time.strftime("%y/%m/%d %H:%M")})
            fields.append({'name': 'remarks', 'value': object.remarks})
            fields.append({'name': 'change', 'view_link': reverse('res:reservation-cal-day', args=[object.start_time.year,object.start_time.month,object.start_time.day]),
                           'edit_link': reverse('res:blackout-update', args=[object.pk]), 'delete_link': reverse('res:blackout-delete', args=[object.pk])})
            row_list.append({'fields': fields})
        context['row_list'] = row_list
        return context


@class_view_decorator(permission_required('res.res_admin'))
class BlackoutInfo (DetailView):
    model = Blackout
    template_name = 'res/details_template.html'

    def get_context_data(self, **kwargs):
        context = super(BlackoutInfo, self).get_context_data(**kwargs)
        context['page_title'] = 'Blokada'
        context['header_text'] = 'Szczegóły blokady %s' % self.object.aircraft

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Pokaż w kalendarzu',
                           'path': reverse('res:reservation-cal-day', args=[self.object.start_time.year,self.object.start_time.month,self.object.start_time.day])})
        local_menu.append({'text': 'Aktualizuj', 'path': reverse('res:blackout-update', args=[self.object.pk])})
        local_menu.append({'text': 'Usuń', 'path': reverse('res:blackout-delete', args=[self.object.pk])})
        context['local_menu'] = local_menu

        field_list = []
        field_list.append({'header': 'Statek powietrzny', 'value': self.object.aircraft, 'bold': True})
        field_list.append({'header': 'Termin rozpoczęcia', 'value': '{:%Y-%m-%d %H:%M}'.format(self.object.start_time)})
        field_list.append({'header': 'Termin zakończenia', 'value': '{:%Y-%m-%d %H:%M}'.format(self.object.end_time)})
        field_list.append({'header': 'Uwagi', 'value': self.object.remarks})
        field_list.append({'header': 'Utworzona przez', 'value': '%s - %s' % (self.object.open_user, self.object.open_time.strftime('%y/%m/%d %X'))})

        context['field_list'] = field_list
        return context


@class_view_decorator(permission_required('res.res_admin'))
class BlackoutCreate (CreateView):
    model = Blackout
    template_name = 'res/create_template.html'

    def get_context_data(self, **kwargs):
        context = super(BlackoutCreate, self).get_context_data(**kwargs)
        context['page_title'] = 'Utworzenie blokady'
        context['header_text'] = 'Utworzenie nowej blokady'
        return context

    def get_form_class(self, **kwargs):
        exclude_list = ['open_time', 'open_user']
        widgets_list = {'start_time': DateTimeInput(format="%d.%m.%Y %H:%M"),
                        'end_time': DateTimeInput(format="%d.%m.%Y %H:%M"),
                        'remarks': Textarea(attrs={'rows': 4, 'cols': 100})}
        form_class = modelform_factory(Blackout, exclude=exclude_list, widgets=widgets_list)

        class create_class(form_class):
            def clean(self):
                cleaned_data = super(create_class, self).clean()
                if cleaned_data['end_time'] <= cleaned_data['start_time']:
                    raise forms.ValidationError("Termin rozpoczęcia musi być wcześniejszy niż zakończenia!")

                start = datetime(cleaned_data['start_time'].year, cleaned_data['start_time'].month, cleaned_data['start_time'].day)
                end = datetime(cleaned_data['end_time'].year, cleaned_data['end_time'].month, cleaned_data['end_time'].day, hour=23, minute=59)
                overlap = False
                for res in Reservation.objects.filter(aircraft=cleaned_data['aircraft']).exclude(start_time__gt=end).exclude(end_time__lt=start):
                    if (res.start_time >= cleaned_data['start_time'] and res.start_time < cleaned_data['end_time']) or \
                       (res.end_time > cleaned_data['start_time'] and res.end_time <= cleaned_data['end_time']) or \
                       (cleaned_data['start_time'] >= res.start_time and cleaned_data['end_time'] <= res.end_time):
                        overlap = True
                        break
                if overlap:
                    raise forms.ValidationError("Blokada koliduje z istniejącymi rezerwacjami!")

        return create_class

    def form_valid(self, form):
        form.instance.open_user = self.request.user.fbouser
        return super(BlackoutCreate, self).form_valid(form)

    def get_success_url(self):
        return reverse('res:blackout-list')


@class_view_decorator(permission_required('res.res_admin'))
class BlackoutUpdate (UpdateView):
    model = Blackout
    template_name = 'res/update_template.html'

    def get_context_data(self, **kwargs):
        context = super(BlackoutUpdate, self).get_context_data(**kwargs)
        context['page_title'] = 'Modyfikacja blokady'
        context['header_text'] = 'Modyfikacja istniejącej blokady'

        return context

    def get_form_class(self, **kwargs):
        exclude_list = ['open_time', 'open_user']
        widgets_list = {'start_time': DateTimeInput(format="%d.%m.%Y %H:%M"),
                        'end_time': DateTimeInput(format="%d.%m.%Y %H:%M"),
                        'remarks': Textarea(attrs={'rows': 4, 'cols': 100})}
        form_class = modelform_factory(Blackout, exclude=exclude_list, widgets=widgets_list)

        class update_class(form_class):
            def clean(self):
                cleaned_data = super(update_class, self).clean()
                if cleaned_data['end_time'] <= cleaned_data['start_time']:
                    raise forms.ValidationError("Termin rozpoczęcia musi być wcześniejszy niż zakończenia!")

                start = datetime(cleaned_data['start_time'].year, cleaned_data['start_time'].month,cleaned_data['start_time'].day)
                end = datetime(cleaned_data['end_time'].year, cleaned_data['end_time'].month,cleaned_data['end_time'].day, hour=23, minute=59)
                overlap = False
                for res in Reservation.objects.filter(aircraft=cleaned_data['aircraft']).exclude(start_time__gt=end).exclude(end_time__lt=start).exclude(pk=self.instance.id):
                    if (cleaned_data['start_time'] <= res.start_time < cleaned_data['end_time']) or \
                        (cleaned_data['start_time'] < res.end_time <= cleaned_data['end_time']) or \
                        (cleaned_data['start_time'] >= res.start_time and cleaned_data['end_time'] <= res.end_time):
                        overlap = True
                        break
                if overlap:
                    raise forms.ValidationError("Blokada koliduje z istniejącymi rezerwacjami!")

        return update_class

    def get_success_url(self):
        return reverse('res:blackout-list')


@class_view_decorator(permission_required('res.res_admin'))
class BlackoutDelete (DeleteView):
    model = Blackout
    template_name = 'res/delete_template.html'

    def get_context_data(self, **kwargs):
        context = super(BlackoutDelete, self).get_context_data(**kwargs)
        context['page_title'] = 'Usunięcie blokady'
        context['header_text'] = 'Usunięcie blokady %s' % self.object.aircraft
        context['description'] = self.object.remarks
        return context

    def get_success_url(self):
        return reverse('res:blackout-list')


@class_view_decorator(permission_required('res.res_admin'))
class ResourceFBOList (ListView):
    model = ResourceFBO
    template_name = 'res/list_template.html'

    def get_context_data(self, **kwargs):
        context = super(ResourceFBOList, self).get_context_data(**kwargs)
        context['page_title'] = 'Zasoby'
        context['header_text'] = 'Lista dostępnych zasobów'
        context['empty_text'] = 'Brak zasobów.'

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Nowy zasób', 'path': reverse("res:resource-create")})
        context['local_menu'] = local_menu

        # Nagłówki tabeli
        header_list = []
        header_list.append({'header': 'Zasób'})
        header_list.append({'header': 'Opis zasobu'})
        header_list.append({'header': 'Podlega\nrezerwacji'})
        header_list.append({'header': 'Informacja dla\nużytkowników'})
        header_list.append({'header': 'Kolor\nwyświetlania'})
        header_list.append({'header': '...'})
        context['header_list'] = header_list

        row_list = []
        for object in self.object_list:
            fields = []
            fields.append({'name': 'resource', 'value': object.name, 'bold': True, 'just': 'center'})
            fields.append({'name': 'description', 'value': object.description})
            fields.append({'name': 'scheduled', 'value': 'TAK' if object.scheduled else 'NIE', 'just': 'center'})
            fields.append({'name': 'info', 'value': object.info})
            fields.append({'name': 'color', 'value': '     ', 'color': object.color})
            fields.append({'name': 'change', 'edit_link': reverse('res:resource-update', args=[object.pk]),
                           'delete_link': reverse('res:resource-delete', args=[object.pk]) if object.name not in ['INFOS', 'Dyżur'] else ''})
            row_list.append({'fields': fields})

        context['row_list'] = row_list
        context['no_paging'] = True

        return context


@class_view_decorator(permission_required('res.res_admin'))
class ResourceFBOCreate (CreateView):
    model = ResourceFBO
    template_name = 'res/create_template.html'

    def get_context_data(self, **kwargs):
        context = super(ResourceFBOCreate, self).get_context_data(**kwargs)
        context['page_title'] = 'Utworzenie zasobu'
        context['header_text'] = 'Utworzenie nowego zasobu'
        return context

    def get_form_class(self, **kwargs):
        widgets_list = {'description': Textarea(attrs={'rows': 4, 'cols': 100})}
        form_class = modelform_factory(ResourceFBO, exclude=['scheduled', 'info', 'color'], widgets=widgets_list)
        form_class.base_fields['name'].error_messages['unique'] = 'Nazwa już istnieje!'
        return form_class

    def get_success_url(self):
        return reverse('res:resource-list')


@class_view_decorator(permission_required('res.res_admin'))
class ResourceFBOUpdate (UpdateView):
    model = ResourceFBO
    template_name = 'res/params_template.html'

    def get_context_data(self, **kwargs):
        context = super(ResourceFBOUpdate, self).get_context_data(**kwargs)
        context['page_title'] = 'Modyfikacja zasobu'
        context['header_text'] = 'Modyfikacja istniejącego zasobu'
        return context

    def get_form_class(self, **kwargs):
        widgets_list = {'description': Textarea(attrs={'rows': 4, 'cols': 100}),
                        'info': Textarea(attrs={'rows': 4, 'cols': 100})}
        form_class = modelform_factory(ResourceFBO, fields=['name', 'description', 'scheduled', 'info', 'color'], widgets=widgets_list)
        form_class.base_fields['name'].error_messages['unique'] = 'Nazwa już istnieje!'
        if self.object.name in ['INFOS', 'Dyżur']:
            form_class.base_fields['name'].disabled = True
            form_class.base_fields['scheduled'].disabled = True
        return form_class

    def get_success_url(self):
        return reverse('res:resource-list')


@class_view_decorator(permission_required('res.res_admin'))
class ResourceFBODelete (DeleteView):
    model = ResourceFBO
    template_name = 'res/delete_template.html'

    def get_context_data(self, **kwargs):
        context = super(ResourceFBODelete, self).get_context_data(**kwargs)
        context['page_title'] = 'Usunięcie zasobu'
        context['header_text'] = 'Usunięcie zasobu %s' % self.object
        context['description'] = self.object.description
        return context

    def get_success_url(self):
        return reverse('res:resource-list')


@class_view_decorator(permission_required('res.res_admin'))
class ResParamsList (ListView):
    model = Aircraft
    template_name = 'res/list_template.html'

    def get_queryset(self):
        return Aircraft.objects.order_by('registration')

    def get_context_data(self, **kwargs):
        context = super(ResParamsList, self).get_context_data(**kwargs)
        context['page_title'] = 'Parametry'
        context['header_text'] = 'Parametry rezerwacji SP'
        context['empty_text'] = 'Brak SP.'

        # Lokalne menu
        local_menu = []
        context['local_menu'] = local_menu

        # Nagłówki tabeli
        header_list = []
        header_list.append({'header': 'Statek\npowietrzny'})
        header_list.append({'header': 'Podlega\nrezerwacji'})
        header_list.append({'header': 'Informacja dla pilotów'})
        header_list.append({'header': 'Kolor\nwyświetlania'})
        header_list.append({'header': '...'})
        context['header_list'] = header_list

        row_list = []
        for object in self.object_list:
            fields = []
            fields.append({'name': 'aircraft', 'value': object, 'bold': True, 'just': 'center'})
            fields.append({'name': 'scheduled', 'value': 'TAK' if object.scheduled else 'NIE', 'just': 'center'})
            fields.append({'name': 'info', 'value': object.info})
            fields.append({'name': 'color', 'value': '     ', 'color': object.color})
            fields.append({'name': 'change', 'edit_link': reverse('res:params-update', args=[object.pk])})
            row_list.append({'fields': fields})
        context['row_list'] = row_list

        context['no_paging'] = True
        return context


@class_view_decorator(permission_required('res.res_admin'))
class ResParamsUpdate (UpdateView):
    model = Aircraft
    template_name = 'res/params_template.html'

    def get_context_data(self, **kwargs):
        context = super(ResParamsUpdate, self).get_context_data(**kwargs)
        context['page_title'] = 'Parametry'
        context['header_text'] = 'Modyfikacja parametrów rezerwacji dla %s' % self.object
        return context

    def get_form_class(self, **kwargs):
        fields_list = ['color', 'info', 'scheduled']
        widgets_list = {'info': Textarea(attrs={'rows': 4, 'cols': 100})}
        form_class = modelform_factory(Aircraft, fields=fields_list, widgets=widgets_list)
        return form_class

    def get_success_url(self):
        return reverse('res:params-list')


@permission_required('res.res_user')
def DutyCalendar(request, year=None, month=None, day=None):
    context = {'mobile': request.device['is_mobile']}
    context['resources'] = ''

    if year and month and day:
        context['DefaultDate'] = '%s-%s-%s' % (year, month.zfill(2), day.zfill(2))

    for pilot in Pilot.objects.filter(employee=True):
        if context['resources'] != '':
            context['resources'] += ','
        context['resources'] += "{id:'%s', title:'%s'}" % (pilot.pk, pilot)

    if request.device['is_mobile']:
        return render(request, 'res/duty_calendar.html', context)
    else:
        return render(request, 'res/duty_calendar.html', context)


# Funkcja pomocnicza do sprawdzenia, czy zapytanie jest Ajax
def is_ajax(request):
    return request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'


# Pobranie listy rezerwacji poprzez Ajax
def reservation_feed(request):
    if is_ajax(request):
        data_list = []
        try:
            start = datetime(year=int(request.GET['start'][0:4]), month=int(request.GET['start'][5:7]), day=int(request.GET['start'][8:10]))
            end = datetime(year=int(request.GET['end'][0:4]), month=int(request.GET['end'][5:7]), day=int(request.GET['end'][8:10]), hour=23, minute=59)
        except:
            start = end = None

        if start and end:
            today = datetime(year=date.today().year, month=date.today().month, day=date.today().day)
            # Dodaj rezerwacje AC z tego zakresu dat
            for res in Reservation.objects.exclude(start_time__gt = end).exclude(end_time__lt = start):
                editable = False
                update = False
                if res.status in ('Nowa', 'Potwierdzona'):
                    if request.user.has_perm('res.res_admin') or \
                       (hasattr(request.user.fbouser, 'pilot') and (request.user.fbouser.pilot in (res.owner, res.participant))) or \
                       (res.open_user == request.user.fbouser):
                        editable = True
                        if request.device['is_mobile']:
                            update = True
                notes = ('<b style="color: lightcoral">' + res.internal_remarks.replace('\r','').replace('\n', '<br>') + '</b>') \
                        if (res.internal_remarks and (request.user.has_perm('res.res_admin') or request.user.fbouser.infos)) else ''
                notes = notes + ((('<br>' if notes else '') + res.remarks.replace('\r','').replace('\n', '<br>')) if res.remarks else '')
                if res.aircraft.scheduled:
                    if not request.device['is_mobile']:
                        url = '%s?cal=1' % (reverse('res:reservation-info', args=[res.pk]))
                    else:
                        if update:
                            url = '%s?cal=1' % (reverse('res:reservation-update', args=[res.pk]))
                        else:
                            url = ''
                    data_list.append({'id': 'ac' + str(res.pk),
                                      'resourceId': str(res.aircraft.pk),
                                      'title': res.owner.fbouser.second_name +
                                               ('/%s' % res.participant.fbouser.second_name if res.participant else ''),
                                      'start': res.start_time.strftime("%Y-%m-%d %H:%M"),
                                      'end': res.end_time.strftime("%Y-%m-%d %H:%M"),
                                      'type': (FlightTypes()[res.planned_type] if res.planned_type else 'nieokreślony'),
                                      'notes': notes,
                                      'color': res.aircraft.color,
                                      'editable': editable,
                                      'url': url,
                                      'owner': res.owner.__str__(),
                                      'participant': res.participant.__str__() if res.participant else ''
                                      })

            # Dodaj blokady AC z tego zakresu dat
            for blackout in Blackout.objects.exclude(start_time__gt=end).exclude(end_time__lt=start):
                if blackout.aircraft.scheduled:
                    data_list.append({'id': 'bl' + str(blackout.pk),
                                      'resourceId': str(blackout.aircraft.pk),
                                      'title': 'Blokada %s' % blackout.aircraft,
                                      'start': blackout.start_time.strftime("%Y-%m-%d %H:%M"),
                                      'end': blackout.end_time.strftime("%Y-%m-%d %H:%M"),
                                      'notes': blackout.remarks,
                                      'color': 'lightcoral',
                                      'editable': False,
                                      'rendering': 'background'
                                      })

            # Dodaj PDT-y z tego zakresu dat
            for oper in Operation.objects.filter(pdt__date__gte=start, pdt__date__lte=end):
                if oper.time_start and oper.time_end and oper.pdt.aircraft.scheduled:
                    data_list.append({'id': 'op' + str(oper.pk),
                                      'resourceId': str(oper.pdt.aircraft.pk),
                                      'title': 'PDT %s' % oper.pdt.pdt_ref,
                                      'start': datetime.combine(oper.pdt.date, oper.time_start).strftime("%Y-%m-%d %H:%M"),
                                      'end': datetime.combine(oper.pdt.date, oper.time_end).strftime("%Y-%m-%d %H:%M"),
                                      'owner': oper.pdt.pic.__str__(),
                                      'participant': oper.pdt.sic.__str__() if oper.pdt.sic else '',
                                      'type': FlightTypes()[oper.pdt.flight_type],
                                      'notes': '%s - %s' % (oper.loc_start, oper.loc_end),
                                      'color': oper.pdt.aircraft.color,
                                      'editable': False,
                                      'url': ('' if request.device['is_mobile'] else reverse('panel:pdt-info', args=[oper.pdt.pk])),
                                      })

            # Dodaj rezerwacje pozostałych zasobów z tego zakresu dat
            for res in ReservationFBO.objects.exclude(start_time__gt=end).exclude(end_time__lt=start):
                if res.start_time and res.end_time and res.resource.scheduled:
                    if res.resource.name == 'INFOS':
                        # Dodaj do INFOS właściciela
                        data_list.append({'id': 'fb1' + str(res.pk),
                                          'resourceId': 'infos' if res.owner == request.user.fbouser else ('usr%s' % str(res.owner.pk)),
                                          'title': res.title,
                                          'start': res.start_time.strftime("%Y-%m-%d %H:%M"),
                                          'end': res.end_time.strftime("%Y-%m-%d %H:%M"),
                                          'owner': res.owner.__str__(),
                                          'participant': res.participant.__str__() if res.participant else '',
                                          'type': res.title,
                                          'notes': res.remarks,
                                          'color': res.resource.color,
                                          'editable': False,
                                          'url': '%s?cal=1' % (reverse('res:resfbo-update', args=[res.pk]) if request.device['is_mobile']
                                                               else reverse('res:resfbo-info', args=[res.pk])),
                                          })
                        # Dodaj do INFOS uczestnika
                        if res.participant:
                            data_list.append({'id': 'fb2' + str(res.pk),
                                              'resourceId': 'infos' if res.participant == request.user.fbouser else ('usr%s' % str(res.participant.pk)),
                                              'title': res.title,
                                              'start': res.start_time.strftime("%Y-%m-%d %H:%M"),
                                              'end': res.end_time.strftime("%Y-%m-%d %H:%M"),
                                              'owner': res.owner.__str__(),
                                              'participant': res.participant.__str__(),
                                              'type': res.title,
                                              'notes': res.remarks,
                                              'color': res.resource.color,
                                              'editable': False,
                                              'url': '%s?cal=1' % (reverse('res:resfbo-update', args=[res.pk]) if request.device['is_mobile']
                                                                   else reverse('res:resfbo-info', args=[res.pk])),
                                              })
                    elif res.resource.name == 'Dyżur':
                        # Dodaj dyżur
                        data_list.append({'id': 'fb1' + str(res.pk),
                                          'resourceId': 'fbo'+str(res.resource.pk),
                                          'title': res.title,
                                          'subtitle': res.start_time.strftime("%H:%M") + " - " + res.end_time.strftime("%H:%M"),
                                          'allDay': True,
                                          'start': res.start_time.strftime("%Y-%m-%d %H:%M"),
                                          'end': res.end_time.strftime("%Y-%m-%d %H:%M"),
                                          'owner': res.owner.__str__(),
                                          'participant': res.participant.__str__() if res.participant else '',
                                          'type': res.title,
                                          'notes': res.remarks,
                                          'color': res.resource.color,
                                          'editable': False,
                                          'url': '%s?cal=1' % (reverse('res:resfbo-update', args=[res.pk]) if request.device['is_mobile']
                                                               else reverse('res:resfbo-info', args=[res.pk])),
                                          })
                    else:
                        # Dodaj rezerwacje dla pozostałych zasobów
                        data_list.append({'id': 'fbo' + str(res.pk),
                                          'resourceId': 'fbo'+str(res.resource.pk),
                                          'title': res.title,
                                          'start': res.start_time.strftime("%Y-%m-%d %H:%M"),
                                          'end': res.end_time.strftime("%Y-%m-%d %H:%M"),
                                          'owner': res.owner.__str__(),
                                          'participant': res.participant.__str__() if res.participant else '',
                                          'type': res.title,
                                          'notes': res.remarks,
                                          'color': res.resource.color,
                                          'editable': False,
                                          'url': '%s?cal=1' % (reverse('res:resfbo-update', args=[res.pk]) if request.device['is_mobile']
                                                               else reverse('res:resfbo-info', args=[res.pk])),
                                          })
        response = json.dumps(data_list)
        return HttpResponse(response)
    else:
        return TemplateResponse(request, 'res/res_calendar.html', {})


# Zapisanie zmian w rezerwacji przez Ajax
def reservation_move(request):
    if is_ajax(request):
        reservation_id = request.POST['reservation'][2:]
        new_resource = request.POST['newResource']
        new_start = request.POST['newStart']
        new_end = request.POST['newEnd']
        try:
            new_duration = datetime.strptime(new_end, '%Y-%m-%dT%H:%M:%S') - datetime.strptime(new_start, '%Y-%m-%dT%H:%M:%S')
            if new_duration.seconds > 30*60:
                new_duration -= timedelta(seconds=30*60)
        except:
            new_duration = timedelta(seconds=0)
        response = None

        if new_resource and new_start and new_end:
            res = Reservation.objects.get(pk=reservation_id)
            if res:
                res.aircraft = Aircraft.objects.get(pk=new_resource)
                res.start_time = datetime.strptime(new_start, '%Y-%m-%dT%H:%M:%S')
                res.end_time = datetime.strptime(new_end, '%Y-%m-%dT%H:%M:%S')
                if res.planned_time > new_duration:
                    res.planned_time = new_duration
                res.change_user = request.user.fbouser
                res.save()
                response = None
            else:
                response = None
        return HttpResponse(response)
    else:
        return TemplateResponse(request, 'res/res_calendar.html', {})


# Pobranie listy pozycji czasu pracy poprzez Ajax
def duty_feed(request):
    if is_ajax(request):
        data_list = []
        try:
            start = datetime(year=int(request.GET['start'][0:4]), month=int(request.GET['start'][5:7]), day=int(request.GET['start'][8:10]))
            end = datetime(year=int(request.GET['end'][0:4]), month=int(request.GET['end'][5:7]), day=int(request.GET['end'][8:10]), hour=23, minute=59)
        except:
            start = end = None

        if start and end:
            # Dodaj rezerwacje AC z wybranego zakresu dat
            for res in Reservation.objects.exclude(start_time__gt = end).exclude(end_time__lt = max(start, datetime.now()-timedelta(seconds=30*60))):
                notes = ('<b style="color: lightcoral">' + res.internal_remarks.replace('\r','').replace('\n', '<br>') + '</b>') \
                        if (res.internal_remarks and request.user.has_perm('res.res_admin')) else ''
                notes = notes + ((('<br>' if notes else '') + res.remarks.replace('\r','').replace('\n', '<br>')) if res.remarks else '')
                if res.owner.employee:
                    # Dodaj rezerwacje dla właściciela
                    data_list.append({'id': 'ro' + str(res.pk),
                                      'resourceId': str(res.owner_id),
                                      'title': res.aircraft.__str__(),
                                      'start': res.start_time.strftime("%Y-%m-%d %H:%M"),
                                      'end': res.end_time.strftime("%Y-%m-%d %H:%M"),
                                      'type': (FlightTypes()[res.planned_type] if res.planned_type else 'nieokreślony'),
                                      'notes': notes,
                                      'color': res.aircraft.color,
                                      'editable': False,
                                      'url': reverse('res:reservation-info', args=[res.pk]),
                                      'owner': res.owner.__str__(),
                                      'participant': res.participant.__str__() if res.participant else ''
                                      })
                if res.participant and res.participant.employee:
                    # Dodaj rezerwacje dla uczestnika
                    data_list.append({'id': 'rp' + str(res.pk),
                                      'resourceId': str(res.participant_id),
                                      'title': res.aircraft.__str__(),
                                      'start': res.start_time.strftime("%Y-%m-%d %H:%M"),
                                      'end': res.end_time.strftime("%Y-%m-%d %H:%M"),
                                      'type': (FlightTypes()[res.planned_type] if res.planned_type else 'nieokreślony'),
                                      'notes': notes,
                                      'color': res.aircraft.color,
                                      'editable': False,
                                      'url': reverse('res:reservation-info', args=[res.pk]),
                                      'owner': res.owner.__str__(),
                                      'participant': res.participant.__str__() if res.participant else ''
                                      })

            # Dodaj pozycje czasu pracy z wybranego zakresu dat
            for duty in Duty.objects.filter(date__gte=start, date__lte=end):
                if duty.pilot.employee:
                    data_list.append({'id': 'dt' + str(duty.pk),
                                      'resourceId': str(duty.pilot_id),
                                      'title': duty.role,
                                      'start': datetime.combine(duty.date, duty.duty_time_from).strftime("%Y-%m-%d %H:%M"),
                                      'end': datetime.combine(duty.date, duty.duty_time_to).strftime("%Y-%m-%d %H:%M"),
                                      'owner': duty.pilot.__str__(),
                                      'participant': '',
                                      'type': '',
                                      'notes': duty.remarks,
                                      'color': duty.pdt.aircraft.color if duty.pdt else '#ccddff',
                                      'editable': False,
                                      'url': reverse('panel:duty-details', args=[duty.pk]),
                                      })
        response = json.dumps(data_list)
        return HttpResponse(response)
    else:
        return TemplateResponse(request, 'res/res_calendar.html', {})


# Generacja raportu planowanych lotów w formacie wymaganym przez EPMO
def PortExport (min_date, max_date, rep_type):

    res = 0

    if min_date and max_date and rep_type:

        # wybór rezerwacji do raportu i adresów wysyłki
        if rep_type == 9:
            flight_type_list = ['01', '01A', '02', '02H', '03A', '03B', '03C', '03D', '03E', '04', '05', '07']
            email_address = settings.EMAIL_ATO
            created_string1 = 'Justyna Luty, tel. 518 111101'
            created_string2 = None
        elif rep_type == 1:
            flight_type_list = ['04', '07']
            email_address = settings.EMAIL_ATO
            created_string1 = 'Justyna Luty, tel. 518 111101'
            created_string2 = None
        elif rep_type == 2:
            flight_type_list = ['01', '01A', '02', '02H', '03A', '03C', '05']
            email_address = settings.EMAIL_INFO
            created_string1 = 'Łukasz Szeskaz, tel. 518 969243'
            created_string2 = 'Robert Miedza, tel. 501 658716'
        elif rep_type == 3:
            flight_type_list = ['03B', '03D', '03E']
            email_address = settings.EMAIL_ATO
            created_string1 = 'Magdalena Grądek, tel. 512 302241'
            created_string2 = None
        else:
            flight_type_list = []
            email_address = []
            created_string1 = None
            created_string2 = None

        # Wybór rezerwacji A-B wychodzących z EPMO
        query_out = Reservation.objects.filter(start_time__date__range=[min_date, max_date]).\
                    filter(planned_type__in=flight_type_list).filter(loc_start__iexact="epmo").\
                    exclude(loc_stop__iexact="epmo")

        # Wybór rezerwacji A-B przychodzących do EPMO (pierwszy odcinek)
        query_in1 = Reservation.objects.filter(end_time__date__range=[min_date, max_date]).\
                    filter(planned_type__in=flight_type_list).filter(loc_stop__iexact="epmo").\
                    exclude(loc_start__iexact="epmo")

        # Wybór rezerwacji A-B przychodzących do EPMO (docelowo)
        query_in2 = Reservation.objects.filter(end_time__date__range=[min_date, max_date]).\
                    filter(planned_type__in=flight_type_list).filter(loc_end__iexact="epmo").\
                    exclude(loc_stop__iexact="epmo")

        # Integracja list (rezerwacja, rodzaj trasy)
        reservations = [(res, 0) for res in query_out] + [(res, 1) for res in query_in1] + [(res, 2) for res in query_in2]
        # Sortowanie połączonej listy
        reservations.sort(key=lambda res: res[0].start_time if res[1] == 0 else res[0].end_time)

        if reservations:
            # Nowy arkusz
            wb = Workbook()
            wb.active.title = 'SALT General Aviation'

            # Definicje stylów
            border = Border(top=Side(border_style=borders.BORDER_THIN, color='000000'), bottom=Side(border_style=borders.BORDER_THIN, color='000000'),
                            left=Side(border_style=borders.BORDER_THIN, color='000000'), right=Side(border_style=borders.BORDER_THIN, color='000000'))
            title_align = Alignment(horizontal='center', vertical='center', wrap_text = True)
            table_align = Alignment(vertical='center', wrap_text=True)
            blue_fill = PatternFill(fill_type=fills.FILL_SOLID, fgColor=('90EE90'))

            ws = wb.active

            # Informacje nagłówkowe
            ws['A1'] = "SALT Aviation Sp. z o.o."
            ws.merge_cells('A1:C1')
            if min_date == max_date:
                ws['F1'] = "PLANOWANE OPERACJE GA W DNIU %s" % min_date
            else:
                ws['F1'] = "PLANOWANE OPERACJE GA W DNIACH %s - %s" % (min_date, max_date)
            ws.merge_cells('F1:O1')
            ws['C1'].alignment = title_align
            ws['C1'].font = Font(size=12, bold=True)
            ws['F1'].alignment = title_align
            ws['F1'].font = Font(size=15, bold=True)

            row = 3
            lp = 0

            for (res, route_type) in reservations:

                lp += 1
                ws['A%d' % row] = lp
                ws.merge_cells('A%d:A%d' % (row, row+2))
                ws['B%d' % row] = "A/C"
                ws.merge_cells('B%d:C%d' % (row, row))
                ws['D%d' % row] = "Znaki"
                ws['E%d' % row] = "Nr Rejsu"
                ws['F%d' % row] = "Operator"
                ws['G%d' % row] = "Załoga"
                ws['H%d' % row] = "Pax ARR"
                ws['I%d' % row] = "Pax DEP"
                ws['J%d' % row] = "Przylot z"
                ws['K%d' % row] = "Odlot do"
                ws['L%d' % row] = "STA"
                ws.merge_cells('L%d:M%d' % (row, row))
                ws['N%d' % row] = "STD"
                ws.merge_cells('N%d:O%d' % (row, row))
                ws['P%d' % row] = "Miejsce postojowe"
                ws['Q%d' % row] = "Uwagi"
                row += 1

                ws['B%d' % row] = "TYP"
                ws['C%d' % row] = res.aircraft.type
                ws['D%d' % row] = res.aircraft.registration
                ws.merge_cells('D%d:D%d' % (row, row + 1))
                ws['E%d' % row] = ""
                ws.merge_cells('E%d:E%d' % (row, row + 1))
                ws['F%d' % row] = ("%s" % res.owner if res.planned_type in ['04', '07'] else "SALT Aviation")
                ws.merge_cells('F%d:F%d' % (row, row + 1))
                ws['G%d' % row] = "%d" % (1 + (1 if res.participant else 0))
                ws.merge_cells('G%d:G%d' % (row, row + 1))
                if route_type != 0:
                    pax = len(list(filter(None, res.pax.split(sep="\r\n"))))
                    ws['H%d' % row] = pax if pax else ""
                else:
                    ws['H%d' % row] = ""
                ws.merge_cells('H%d:H%d' % (row, row + 1))
                if route_type == 0:
                    pax = len(list(filter(None, res.pax.split(sep="\r\n"))))
                    ws['I%d' % row] = pax if pax else ""
                else:
                    ws['I%d' % row] = ""
                ws.merge_cells('I%d:I%d' % (row, row + 1))
                if route_type == 1:
                    ws['J%d' % row] = res.loc_start
                elif route_type == 2:
                    ws['J%d' % row] = res.loc_stop
                else:
                    ws['J%d' % row] = ""
                ws.merge_cells('J%d:J%d' % (row, row + 1))
                if route_type == 0:
                    ws['K%d' % row] = res.loc_stop
                else:
                    ws['K%d' % row] = ""
                ws.merge_cells('K%d:K%d' % (row, row + 1))
                ws['L%d' % row] = "Data"
                ws['M%d' % row] = "%s" % res.end_time.date()
                ws['N%d' % row] = "Data"
                ws['O%d' % row] = "%s" % res.start_time.date()
                ws['P%d' % row] = ""
                ws.merge_cells('P%d:P%d' % (row, row + 1))
                ws['Q%d' % row] = ""
                ws.merge_cells('Q%d:Q%d' % (row, row + 1))
                row += 1

                ws['B%d' % row] = "MTOW"
                ws['C%d' % row] = res.aircraft.mtow
                ws['L%d' % row] = "Czas LT"
                ws['M%d' % row] = ("%s" % res.end_time.time())[:5]
                ws['N%d' % row] = "Czas LT"
                if route_type != 2:
                    ws['O%d' % row] = ("%s" % res.start_time.time())[:5]
                else:
                    ws['O%d' % row] = ""
                row += 1

                ws['A%d' % (row - 3)].alignment = title_align
                ws['A%d' % (row - 3)].border = border
                ws['A%d' % (row - 2)].border = border
                ws['A%d' % (row - 1)].border = border
                ws['A%d' % (row - 3)].font = Font(size=10)
                for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']:
                    ws['%s%d' % (col, row - 3)].alignment = title_align
                    ws['%s%d' % (col, row - 3)].border = border
                    ws['%s%d' % (col, row - 3)].font = Font(size=10, bold=True)
                    ws['%s%d' % (col, row - 3)].fill = blue_fill

                    ws['%s%d' % (col, row - 2)].alignment = title_align
                    ws['%s%d' % (col, row - 2)].border = border
                    ws['%s%d' % (col, row - 2)].font = Font(size=10)

                    ws['%s%d' % (col, row - 1)].alignment = title_align
                    ws['%s%d' % (col, row - 1)].border = border
                    ws['%s%d' % (col, row - 1)].font = Font(size=10)

            row += 2
            ws['B%d' % row] = "Sporządził:"
            ws.merge_cells('B%d:D%d' % (row, row))
            ws['B%d' % row].alignment = title_align
            ws['B%d' % row].font = Font(size=11, bold=True)

            row += 1
            ws['B%d' % row] = created_string1
            ws.merge_cells('B%d:D%d' % (row, row))
            ws['B%d' % row].alignment = title_align

            if created_string2:
                row += 1
                ws['B%d' % row] = created_string2
                ws.merge_cells('B%d:D%d' % (row, row))
                ws['B%d' % row].alignment = title_align

            # Ustawienie szerokości kolumn
            ws.column_dimensions['A'].width = '4'
            ws.column_dimensions['B'].width = '8'
            ws.column_dimensions['C'].width = '15'
            ws.column_dimensions['D'].width = '9'
            ws.column_dimensions['E'].width = '10'
            ws.column_dimensions['F'].width = '18'
            ws.column_dimensions['G'].width = '10'
            ws.column_dimensions['H'].width = '6'
            ws.column_dimensions['I'].width = '6'
            ws.column_dimensions['J'].width = '12'
            ws.column_dimensions['K'].width = '12'
            ws.column_dimensions['L'].width = '8'
            ws.column_dimensions['M'].width = '12'
            ws.column_dimensions['N'].width = '8'
            ws.column_dimensions['O'].width = '12'
            ws.column_dimensions['P'].width = '15'
            ws.column_dimensions['Q'].width = '30'

            for (res, route_type) in reservations:
                # Nowy arkusz dla każdego lotu
                ws = wb.create_sheet("%s " % res.aircraft.registration +
                                     ("%s" % res.start_time.time())[:5].replace(":","") +
                                     (" DEP" if route_type == 0 else " ARR"))

                # Informacje nagłówkowe
                ws['A1'] = "SALT Aviation Sp. z o.o."
                ws.merge_cells('A1:D1')
                ws['A2'] = "WMI / EPMO"
                ws.merge_cells('A2:D2')
                ws['A2'].alignment = title_align
                ws['A2'].font = Font(size=16, bold=True)

                ws['A4'] = "PAX LIST / GEN. DEC."
                ws.merge_cells('A4:D4')
                ws['A4'].alignment = title_align
                ws['A4'].font = Font(size=14, bold=True)

                ws['A5'] = "Nazwa operatora / Operator name"
                ws.merge_cells('A5:B5')
                ws['C5'] = ("%s" % res.owner if res.planned_type in ['04', '07'] else "SALT Aviation")
                ws.merge_cells('C5:D5')

                ws['A6'] = "Znaki samolotu / A/C Reg.// Nr. rejsu / flt. nbr"
                ws.merge_cells('A6:B6')
                ws['C6'] = res.aircraft.registration
                ws.merge_cells('C6:D6')

                ws['A7'] = "Data / Date"
                ws.merge_cells('A7:B7')
                ws['C7'] = res.start_time.date()
                ws.merge_cells('C7:D7')

                ws['A8'] = "Trasa lotu / Flight route"
                ws.merge_cells('A8:B8')
                if route_type in (0, 1):
                    ws['C8'] = "%s - %s" % (res.loc_start or "", res.loc_stop or "")
                else:
                    ws['C8'] = "%s - %s" % (res.loc_stop or "", res.loc_end or "")
                ws.merge_cells('C8:D8')

                ws['A9'] = "Typ lotu / Flight type"
                ws.merge_cells('A9:B9')
                ws['C9'] = "Lot z A do B"
                ws.merge_cells('C9:D9')

                for row in range(5,10):
                    ws['A%d' % row].alignment = table_align
                    ws['C%d' % row].alignment = title_align
                    ws['A%d' % row].font = Font(size=12, bold=True)
                    ws['C%d' % row].font = Font(size=12)

                ws['A10'] = "ZAŁOGA / CREW"
                ws.merge_cells('A10:D10')
                ws['A10'].alignment = title_align
                ws['A10'].font = Font(size=14, bold=True)

                ws['A11'] = "1."
                ws['B11'] = "%s" % res.owner
                ws.merge_cells('B11:D11')
                ws['A12'] = "2."
                ws['B12'] = "%s" % res.participant if res.participant else ""
                ws.merge_cells('B12:D12')

                ws['A13'] = "PASAŻEROWIE /PAX"
                ws.merge_cells('A13:C13')
                ws['D13'] = "„VISITOR” NR (JEŚLI WYMAGANY)"
                ws['A13'].alignment = title_align
                ws['D13'].alignment = title_align
                ws['D13'].font = Font(size=10, bold=True)

                pax_list = list(filter(None, res.pax.split(sep="\r\n")))
                row = 14

                for pax in pax_list:
                    ws['A%d' % row] = "%d." % (row - 13)
                    ws['B%d' % row] = pax
                    ws.merge_cells('B%d:C%d' % (row, row))
                    row += 1

                for new_row in range(row, 24):
                    ws['A%d' % new_row] = "%d." % (new_row-13)
                    ws.merge_cells('B%d:C%d' % (new_row, new_row))

                for row in range(11,24):
                    ws['A%d' % row].alignment = title_align
                    ws['C%d' % row].alignment = title_align
                    ws['A%d' % row].font = Font(size=12)
                    ws['C%d' % row].font = Font(size=12)

                ws['A13'].font = Font(size=14, bold=True)

                for col in ['A', 'B', 'C', 'D']:
                    for row in range(4,24):
                        ws['%s%d' % (col, row)].border = border

                ws['B26'] = "Sporządził:"
                ws['B26'].alignment = title_align
                ws['B26'].font = Font(size=11, bold=True)

                ws['B27'] = created_string1
                ws['B27'].alignment = title_align

                if created_string2:
                    ws['B28'] = created_string2
                    ws['B28'].alignment = title_align

                # Ustawienie szerokości kolumn
                ws.column_dimensions['A'].width = '5'
                ws.column_dimensions['B'].width = '50'
                ws.column_dimensions['C'].width = '20'
                ws.column_dimensions['D'].width = '20'

            # Zapisanie arkusza w pamięci
            output = BytesIO()
            wb.save(output)

            # przygotowanie emaila
            if min_date == max_date:
                date_str = 'w dniu %s' % min_date
            else:
                date_str = 'w dniach %s - %s' % (min_date, max_date)

            email = EmailMessage(
                'Planowane operacje EPMO / %s' % min_date,
                'Planowane operacje typów: %s %s \n\nSALT24.pl' % (', '.join(flight_type_list), date_str),
                settings.EMAIL_FROM,
                email_address,
                ['mierzwik@me.com']
            )
            email.attach('SALT EPMO.xlsx', output.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

            # wysłanie emaila i wyczyszczenie cache
            res = email.send(fail_silently=True)
            output.close()

    return HttpResponse('%s' % res)


# Automatyczne wysłanie maili z zestawieniem rezerwacji
def PortSend():

    PortExport(date.today(), date.today(), 9)


# Automatyczne wysłanie maili z zestawieniem rezerwacji na weekend
def PortSendWeekend():

    PortExport(date.today()+timedelta(days=1), date.today()+timedelta(days=1), 9)
    PortExport(date.today()+timedelta(days=2), date.today()+timedelta(days=2), 9)
