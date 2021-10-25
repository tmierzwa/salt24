from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, fills, Border, borders, Side, Font
from openpyxl.writer.excel import save_virtual_workbook
from django.core.urlresolvers import reverse
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.views.generic.list import ListView
from django.views.generic.detail import DetailView
from django.forms.models import modelform_factory
from django.forms import Textarea, TextInput, Form, ChoiceField
from django.contrib.admin.widgets import AdminDateWidget
from django.contrib.auth.decorators import permission_required
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import get_object_or_404, render
from django.core.exceptions import ValidationError
from django.core.mail import EmailMessage
from django.conf import settings

from django.utils.decorators import method_decorator
def class_view_decorator(function_decorator):

    def simple_decorator(View):
        View.dispatch = method_decorator(function_decorator)(View.dispatch)
        return View

    return simple_decorator

from sms.models import SMSHazard, SMSHazardRev, SMSRisk, SMSEvent, SMSFailure, SMSReport, NCR
from panel.models import FBOUser, PDT


@class_view_decorator(permission_required('sms.sms_reader'))
class HazardList (ListView):
    model = SMSHazard
    template_name = 'sms/list_template.html'

    def get_context_data(self, **kwargs):
        context = super(HazardList, self).get_context_data(**kwargs)
        context['page_title'] = 'Lista zagrożeń'
        context['header_text'] = 'Lista zarejestrowanych zagrożeń'
        context['empty_text'] = 'Brak zarejestrowanych zagrożeń.'

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Nowe zagrożenie', 'path': reverse("sms:hazard-create")})
        local_menu.append({'text': 'Generuj raport', 'path': reverse("sms:sms-hazards-export")})
        context['local_menu'] = local_menu

        # Nagłówki tabeli
        header_list = []
        header_list.append({'header': 'Identyfikator\nzagrożenia'})
        header_list.append({'header': 'Klasa\nzagrożenia'})
        header_list.append({'header': 'Data\nwprowadzenia'})
        header_list.append({'header': 'Numer\nwersji'})
        header_list.append({'header': 'Data\nwersji'})
        header_list.append({'header': 'Obszar\nfirmy'})
        header_list.append({'header': 'Nazwa\nzagrożenia'})
        header_list.append({'header': '...'})
        context['header_list'] = header_list

        row_list = []
        for object in self.object_list:
            last_rev = SMSHazardRev.objects.filter(hazard=object.pk, rev_last=True).first()
            if last_rev:
                fields = []
                fields.append({'name': 'hazard_ref', 'value': object.hazard_ref, 'link': reverse('sms:hazard-details', args=[object.pk]), 'just': 'center'})
                fields.append({'name': 'hazard_type', 'value': object.hazard_type, 'just': 'center'})
                fields.append({'name': 'hazard_date', 'value': object.hazard_date, 'just':'center'})
                fields.append({'name': 'rev_num', 'value': 'Rev. %d' % last_rev.rev_num, 'just':'center'})
                fields.append({'name': 'rev_date', 'value': last_rev.rev_date, 'just':'center'})
                fields.append({'name': 'company_area', 'value': last_rev.company_area, 'just': 'center'})
                fields.append({'name': 'name', 'value': last_rev.name})
                fields.append({'name': 'change', 'edit_link': reverse('sms:hazard-update', args=[object.pk]),
                               'delete_link': reverse('sms:hazard-delete', args=[object.pk])})
                row_list.append({'fields': fields, 'noactive': not object.hazard_active})
        context['row_list'] = row_list
        return context


@class_view_decorator(permission_required('sms.sms_reader'))
class HazardDetails (DetailView):
    model = SMSHazard
    template_name = 'sms/hazard_details.html'

    def get_context_data(self, **kwargs):
        context = super(HazardDetails, self).get_context_data(**kwargs)
        context['page_title'] = 'Zagrożenie'
        context['header_text'] = 'Szczegóły zagrożenia %s' % self.object.hazard_ref

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Historia zmian', 'path': reverse('sms:hazard-history', args=[self.object.pk])})
        local_menu.append({'text': 'Aktualizuj', 'path': reverse('sms:hazard-update', args=[self.object.pk])})
        local_menu.append({'text': 'Usuń', 'path': reverse('sms:hazard-delete', args=[self.object.pk])})
        context['local_menu'] = local_menu

        field_list = []
        field_list.append({'header': 'Identyfikator zagrożenia', 'value': self.object.hazard_ref, 'bold': True})
        field_list.append({'header': 'Klasa zagrożenia', 'value': self.object.hazard_type})
        field_list.append({'header': 'Data wprowadzenia', 'value': self.object.hazard_date})
        field_list.append({'header': 'Aktywne', 'value': 'TAK' if self.object.hazard_active else 'NIE'})

        last_rev = self.object.smshazardrev_set.filter(rev_last=True).first()

        if last_rev:
            field_list.append({'header': 'Numer / data wersji', 'value': 'Rev. %d / %s' % (last_rev.rev_num, last_rev.rev_date)})
            field_list.append({'header': 'Obszar firmy', 'value': last_rev.company_area})
            field_list.append({'header': 'Nazwa zagrożenia', 'value': last_rev.name})
            field_list.append({'header': 'Termin wykonania', 'value': last_rev.due_date})
            field_list.append({'header': 'Odpowiedzialny', 'value': last_rev.responsible})
            field_list.append({'header': 'Wykonanie / kontrola', 'value': last_rev.control})
            field_list.append({'header': 'Uwagi', 'value': last_rev.remarks})

        context['field_list'] = field_list

        context['risk_header_text'] = 'Ryzyka powiązane z zagrożeniem'
        context['risk_create_path'] = reverse('sms:risk-create', args=[self.object.pk])
        context['risk_create_text'] = 'Nowe ryzyko'
        context['risk_empty_text'] = 'Brak powiązanych ryzyk.'
        context['risk_header_list'] = ['Ref.', 'Wersja', 'Szczegółowy opis<br>natury ryzyka', 'Początkowe<br>prawdop.', 'Początkowa<br>dotkliwość',
                                       'Sposoby ograniczania<br>ryzyka', 'Szczątkowe<br>prawdop.', 'Szczątkowa<br>dotkliwość']
        row_list = []
        for risk in self.object.smsrisk_set.filter(rev_last=True).order_by('risk_ref'):
            fields = []
            fields.append({'name': 'risk_ref', 'value': risk.risk_ref, 'link': reverse('sms:risk-details', args=[risk.pk])})
            fields.append({'name': 'rev_num', 'value': 'Rev. %d' % risk.rev_num, 'just':'center'})
            fields.append({'name': 'description', 'value': risk.description})
            fields.append({'name': 'init_probability', 'value': risk.prob_str()[0], 'just': 'center', 'color': risk.risk_color()[0]})
            fields.append({'name': 'init_impact', 'value': risk.impact_str()[0], 'just': 'center', 'color': risk.risk_color()[0]})
            fields.append({'name': 'mitigation', 'value': risk.mitigation})
            fields.append({'name': 'res_probability', 'value': risk.prob_str()[1], 'just': 'center', 'color': risk.risk_color()[1]})
            fields.append({'name': 'res_impact', 'value': risk.impact_str()[1], 'just': 'center', 'color': risk.risk_color()[1]})
            row_list.append({'fields': fields})
        context['risk_row_list'] = row_list

        context['event_header_text'] = 'Zdarzenia powiązane z zagrożeniem'
        context['event_empty_text'] = 'Brak zarejestrowanych zdarzeń.'
        context['event_header_list'] = ['Statek<br>powietrzny', 'Pilot<br>dowódca', 'Data<br>zdarzenia', 'Kwalifikacja<br>zdarzenia',
                                        'Skrócony opis<br>zdarzenia', 'Zamknięcie /<br>Raport końcowy']
        row_list = []
        for event in self.object.smsevent_set.all().order_by('event_date'):
            fields = []
            fields.append({'name': 'aircraft', 'value': event.aircraft, 'link': reverse('sms:smsevent-details', args=[event.pk]), 'just': 'center'})
            fields.append({'name': 'pic', 'value': event.pic})
            fields.append({'name': 'event_date', 'value': event.event_date})
            fields.append({'name': 'event_type', 'value': event.event_type, 'just': 'center'})
            fields.append({'name': 'description', 'value': event.description})
            fields.append({'name': 'closure', 'value': event.closure, 'just': 'center'})
            row_list.append({'fields': fields})
        context['event_row_list'] = row_list

        context['failure_header_text'] = 'Usterki powiązane z zagrożeniem'
        context['failure_empty_text'] = 'Brak zarejestrowanych usterek.'
        context['failure_header_list'] = ['Statek<br>powietrzny', 'Osoba<br>zgłaszająca', 'Data<br>usterki', 'Opis<br>usterki', 'PDT',
                                          'Data<br>usunięcia', 'Sposób<br>usunięcia', 'CRS']
        row_list = []
        for object in self.object.smsfailure_set.all().order_by('failure_date'):
            fields = []
            fields.append({'name': 'aircraft', 'value': object.aircraft, 'link': reverse('sms:failure-details', args=[object.pk]), 'just': 'center'})
            fields.append({'name': 'person', 'value': object.person})
            fields.append({'name': 'failure_date', 'value': object.failure_date, 'just': 'center'})
            fields.append({'name': 'description', 'value': object.description })
            fields.append({'name': 'pdt_ref', 'value': object.pdt_ref })
            fields.append({'name': 'repair_date', 'value': object.repair_date, 'just': 'center'})
            fields.append({'name': 'repair_desc', 'value': object.repair_desc})
            fields.append({'name': 'crs', 'value': object.crs, 'just': 'center'})
            row_list.append({'fields': fields})
        context['failure_row_list'] = row_list

        context['report_header_text'] = 'Raporty powiązane z zagrożeniem'
        context['report_empty_text'] = 'Brak zarejestrowanych raportów.'
        context['report_header_list'] = ['Data<br>zgłoszenia', 'Osoba<br>zgłaszająca', 'Skrócona treść<br>zgłoszenia',
                                         'Wnioski i<br>zalecenia', 'Uwagi']
        row_list = []
        for report in self.object.smsreport_set.all().order_by('report_date'):
            fields = []
            fields.append({'name': 'report_date', 'value': report.report_date, 'link': reverse('sms:smsreport-details', args=[report.pk])})
            fields.append({'name': 'person', 'value': report.person})
            fields.append({'name': 'description', 'value': report.description})
            fields.append({'name': 'findings', 'value': report.findings})
            fields.append({'name': 'remarks', 'value': report.remarks})
            row_list.append({'fields': fields})
        context['report_row_list'] = row_list

        return context


@class_view_decorator(permission_required('sms.sms_reader'))
class HazardHistory (DetailView):
    model = SMSHazard
    template_name = 'sms/hazard_history.html'

    def get_context_data(self, **kwargs):
        context = super(HazardHistory, self).get_context_data(**kwargs)
        context['page_title'] = 'Zagrożenie'
        context['header_text'] = 'Historia zmian zagrożenia %s' % self.object.hazard_ref

        # Lokalne menu
        local_menu = []
        context['local_menu'] = local_menu

        field_list = []
        field_list.append({'header': 'Identyfikator zagrożenia', 'value': self.object.hazard_ref, 'bold': True})
        field_list.append({'header': 'Klasa zagrożenia', 'value': self.object.hazard_type})
        field_list.append({'header': 'Data wprowadzenia', 'value': self.object.hazard_date})
        context['field_list'] = field_list

        context['history_header_text'] = 'Lista zmian zagrożenia'
        context['history_empty_text'] = 'Brak historii wersji.'
        context['history_create_path'] = reverse('sms:hazard-update', args=[self.object.pk])
        context['history_create_text'] = 'Nowa wersja'
        context['history_header_list'] = ['Numer\nwersji', 'Data\nwersji', 'Obszar\nfirmy', 'Nazwa\nzagrożenia', 'Termin\nwykonania',
                                          'Odpowiedzialny', 'Wykonanie /\nkontrola', 'Uwagi']

        revisions = SMSHazardRev.objects.filter(hazard=self.object.pk).order_by('rev_num')

        row_list = []
        for revision in revisions:
            fields = []
            fields.append({'name': 'rev_num', 'value': 'Rev. %d' % revision.rev_num, 'just':'center'})
            fields.append({'name': 'rev_date', 'value': revision.rev_date, 'just':'center'})
            fields.append({'name': 'company_area', 'value': revision.company_area, 'just': 'center'})
            fields.append({'name': 'name', 'value': revision.name})
            fields.append({'name': 'due_date', 'value': revision.due_date, 'just': 'center'})
            fields.append({'name': 'responsible', 'value': revision.responsible})
            fields.append({'name': 'control', 'value': revision.control})
            fields.append({'name': 'remarks', 'value': revision.remarks})
            row_list.append({'fields': fields})

        context['history_row_list'] = row_list

        return context


@class_view_decorator(permission_required('sms.sms_admin'))
class HazardCreate (CreateView):
    model = SMSHazard
    template_name = 'sms/create_template.html'

    def get_context_data(self, **kwargs):
        context = super(HazardCreate, self).get_context_data(**kwargs)
        context['page_title'] = 'Rejestracja zagrożenia'
        context['header_text'] = 'Rejestracja nowego zagrożenia'
        return context

    def get_form_class(self, **kwargs):
        hazard_form_class = modelform_factory(SMSHazard, fields=['hazard_type', 'hazard_ref', 'hazard_date', 'hazard_active'],
                                              widgets={'hazard_date':AdminDateWidget()})

        rev_form_class = modelform_factory(SMSHazardRev, fields=['company_area', 'name', 'due_date', 'responsible', 'control', 'remarks'],
                                           widgets={'name':Textarea(attrs={'rows':2, 'cols':100}),
                                                    'remarks':Textarea(attrs={'rows':4, 'cols':100})})

        for key in rev_form_class.base_fields:
            hazard_form_class.base_fields[key] = rev_form_class.base_fields[key]

        hazard_form_class.base_fields['hazard_date'].initial = date.today()

        return hazard_form_class

    def form_valid(self, form):
        hazard = form.save()
        hazard_rev = SMSHazardRev(hazard=hazard, rev_num=1, rev_date=hazard.hazard_date, rev_last=True,
                                  company_area=form.cleaned_data['company_area'], name=form.cleaned_data['name'],
                                  due_date=form.cleaned_data['due_date'], responsible=form.cleaned_data['responsible'],
                                  control=form.cleaned_data['control'], remarks=form.cleaned_data['remarks'])
        hazard_rev.save()
        return super(HazardCreate, self).form_valid(form)

    def get_success_url(self):
        return reverse('sms:hazard-list')


@class_view_decorator(permission_required('sms.sms_admin'))
class HazardUpdate (UpdateView):
    model = SMSHazard
    template_name = 'sms/update_template.html'

    def get_context_data(self, **kwargs):
        # Kontekst dla template
        context = super(HazardUpdate, self).get_context_data(**kwargs)
        context['page_title'] = 'Aktualizacja zagrożenia'
        context['header_text'] = 'Aktualizacja zagrożenia w rejestrze'

        # Dodatkowy guzik do tworzenia wersji
        context['versions'] = True

        return context

    def get_form_class(self, **kwargs):

        # Standardowy formularz dla SMSHazard
        hazard_form_class = modelform_factory(SMSHazard, fields=['hazard_type', 'hazard_ref', 'hazard_date', 'hazard_active'], widgets={'hazard_date':AdminDateWidget})

        # Standardowy formularz dla SMSHazardRev
        rev_form_class = modelform_factory(SMSHazardRev, fields=['rev_date', 'company_area', 'name', 'due_date', 'responsible', 'control', 'remarks'],
                                           widgets={'rev_date':AdminDateWidget,
                                                    'responsible':TextInput(attrs={'size':103}),
                                                    'control':TextInput(attrs={'size':103}),
                                                    'name':Textarea(attrs={'rows':2, 'cols':100}),
                                                    'remarks':Textarea(attrs={'rows':4, 'cols':100})})

        # Zblendowanie formularzy w jeden
        last_rev = self.object.smshazardrev_set.filter(rev_last=True).first()
        for key in rev_form_class.base_fields:
            hazard_form_class.base_fields[key] = rev_form_class.base_fields[key]
            if last_rev:
                hazard_form_class.base_fields[key].initial = last_rev.__dict__[key]

        return hazard_form_class

    def form_valid(self, form):

        # Wczytaj ostatnią wersję
        last_rev = self.object.smshazardrev_set.filter(rev_last=True).first()

        # Jesli wybrano stworzenie wersji
        if '_revision' in self.request.POST:

            # Ostatnia wersja już nie jest ostatnią
            if last_rev:
                last_rev.rev_last = False
                last_rev.save()

            # Utworzenie nowej wersji
            new_rev = SMSHazardRev(hazard=self.object, rev_num=(last_rev.rev_num + 1 if last_rev else 1), rev_date=form.cleaned_data['rev_date'],
                                   rev_last=True, company_area=form.cleaned_data['company_area'], name=form.cleaned_data['name'],
                                   due_date=form.cleaned_data['due_date'], responsible=form.cleaned_data['responsible'],
                                   control=form.cleaned_data['control'], remarks=form.cleaned_data['remarks'])
            new_rev.save()

        # Jeśli wybrano aktualizację
        else:

            # Jeśli istniała ostatnia wersja
            if last_rev:
                last_rev.rev_date = form.cleaned_data['rev_date']
                last_rev.company_area=form.cleaned_data['company_area']
                last_rev.name=form.cleaned_data['name']
                last_rev.due_date=form.cleaned_data['due_date']
                last_rev.responsible=form.cleaned_data['responsible']
                last_rev.control=form.cleaned_data['control']
                last_rev.remarks=form.cleaned_data['remarks']
                last_rev.save()

            # Jeśli nie istniała ostatnia wersja
            else:
                new_rev = SMSHazardRev(hazard=self.object, rev_num=1, rev_date=form.cleaned_data['rev_date'],
                                       rev_last=True, company_area=form.cleaned_data['company_area'], name=form.cleaned_data['name'],
                                       due_date=form.cleaned_data['due_date'], responsible=form.cleaned_data['responsible'],
                                       control=form.cleaned_data['control'], remarks=form.cleaned_data['remarks'])
                new_rev.save()

        return super(HazardUpdate, self).form_valid(form)

    def get_success_url(self):
        return reverse('sms:hazard-details', args=[self.object.pk])


@class_view_decorator(permission_required('sms.sms_admin'))
class HazardDelete (DeleteView):
    model = SMSHazard
    template_name = 'sms/delete_template.html'

    def get_context_data(self, **kwargs):
        context = super(HazardDelete, self).get_context_data(**kwargs)
        context['page_title'] = 'Usunięcie zagrożenia'
        context['header_text'] = 'Usunięcie zagrożenia z rejestru'
        last_rev = self.object.smshazardrev_set.filter(rev_last=True).first()
        if last_rev:
            context['description'] = last_rev.name
        return context

    def get_success_url(self):
        return reverse('sms:hazard-list')


@class_view_decorator(permission_required('sms.sms_reader'))
class RiskDetails (DetailView):
    model = SMSRisk
    template_name = 'sms/details_template.html'

    def get_context_data(self, **kwargs):
        context = super(RiskDetails, self).get_context_data(**kwargs)
        context['page_title'] = 'Ryzyko'
        context['header_text'] = 'Szczegóły ryzyka %s' % self.object.risk_ref

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Historia zmian', 'path': reverse('sms:risk-history', args=[self.object.pk])})
        local_menu.append({'text': 'Aktualizuj', 'path': reverse('sms:risk-update', args=[self.object.pk])})
        local_menu.append({'text': 'Usuń', 'path': reverse('sms:risk-delete', args=[self.object.pk])})
        context['local_menu'] = local_menu

        field_list = []
        field_list.append({'header': 'Identyfikator ryzyka', 'value': self.object.risk_ref})
        field_list.append({'header': 'Numer / data wersji', 'value': 'Rev. %d / %s' % (self.object.rev_num, self.object.rev_date)})
        field_list.append({'header': 'Szczegółowy opis natury ryzyka', 'value': self.object.description})
        field_list.append({'header': 'Początkowe prawdopodobieństwo', 'value': self.object.prob_str()[0], 'color': self.object.risk_color()[0]})
        field_list.append({'header': 'Początkowa dotkliwość', 'value': self.object.impact_str()[0], 'color': self.object.risk_color()[0]})
        field_list.append({'header': 'Sposoby ograniczania ryzyka', 'value': self.object.mitigation})
        field_list.append({'header': 'Szczątkowe prawdopodobieństwo', 'value': self.object.prob_str()[1], 'color': self.object.risk_color()[1]})
        field_list.append({'header': 'Szczątkowa dotkliwość', 'value': self.object.impact_str()[1], 'color': self.object.risk_color()[1]})
        context['field_list'] = field_list
        return context


@class_view_decorator(permission_required('sms.sms_reader'))
class RiskHistory (DetailView):
    model = SMSRisk
    template_name = 'sms/risk_history.html'

    def get_context_data(self, **kwargs):
        context = super(RiskHistory, self).get_context_data(**kwargs)
        context['page_title'] = 'Ryzyko'
        context['header_text'] = 'Historia zmian ryzyka %s' % self.object.risk_ref
        context['empty_text'] = 'Brak historii wersji.'

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Powrót', 'path': reverse("sms:risk-details", args=[self.object.pk])})
        local_menu.append({'text': 'Nowa wersja', 'path': reverse("sms:risk-update", args=[self.object.pk])})
        context['local_menu'] = local_menu

        # Nagłówki tabeli
        header_list = []
        header_list.append({'header': 'Numer\nwersji'})
        header_list.append({'header': 'Data\nwersji'})
        header_list.append({'header': 'Szczegółowy opis\nnatury ryzyka'})
        header_list.append({'header': 'Początkowe\nprawdop.'})
        header_list.append({'header': 'Początkowa\ndotkliwość'})
        header_list.append({'header': 'Sposoby ograniczania\nryzyka'})
        header_list.append({'header': 'Szczątkowe\nprawdop.'})
        header_list.append({'header': 'Szczątkowa\ndotkliwość'})
        context['header_list'] = header_list

        revisions = SMSRisk.objects.filter(smshazard=self.object.smshazard, risk_ref=self.object.risk_ref).order_by('rev_num')

        row_list = []
        for revision in revisions:
            fields = []
            fields.append({'name': 'rev_num', 'value': 'Rev. %d' % revision.rev_num, 'just':'center'})
            fields.append({'name': 'rev_date', 'value': revision.rev_date, 'just':'center'})
            fields.append({'name': 'description', 'value': revision.description})
            fields.append({'name': 'init_probability', 'value': revision.prob_str()[0], 'just': 'center', 'color': revision.risk_color()[0]})
            fields.append({'name': 'init_impact', 'value': revision.impact_str()[0], 'just': 'center', 'color': revision.risk_color()[0]})
            fields.append({'name': 'mitigation', 'value': revision.mitigation})
            fields.append({'name': 'res_probability', 'value': revision.prob_str()[1], 'just': 'center', 'color': revision.risk_color()[1]})
            fields.append({'name': 'res_impact', 'value': revision.impact_str()[1], 'just': 'center', 'color': revision.risk_color()[1]})
            row_list.append({'fields': fields})

        context['row_list'] = row_list

        return context


@class_view_decorator(permission_required('sms.sms_admin'))
class RiskCreate (CreateView):
    model = SMSRisk
    template_name = 'sms/create_template.html'

    def get_context_data(self, **kwargs):
        context = super(RiskCreate, self).get_context_data(**kwargs)
        hazard = get_object_or_404(SMSHazard, pk=self.kwargs['hazard_id'])
        context['page_title'] = 'Nowe ryzyko'
        context['header_text'] = 'Nowe ryzyko dla zagrożenia %s' % hazard.hazard_ref
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(SMSRisk, exclude={'smshazard', 'rev_num', 'rev_last'},
                                                widgets={'rev_date':AdminDateWidget,
                                                         'description':Textarea(attrs={'rows':4, 'cols':100}),
                                                         'mitigation':Textarea(attrs={'rows':4, 'cols':100})})
        # Subklasa z kontrolą unikalności risk_ref
        class create_class(form_class, **kwargs):
            # Metoda init przejmuje dodatkowy argument 'hazard'
            def __init__(self, **kwargs):
                super(create_class, self).__init__(**kwargs)
                if 'hazard' in kwargs:
                    self.hazard = kwargs.pop('hazard')

            # Metoda czyszcząca risk_ref
            def clean_risk_ref(self, **kwargs):
                cleaned_data = super(create_class, self).clean()
                # Jeśli ten hazard ma już ten risk_ref to zgłoś błąd
                if self.hazard.smsrisk_set.filter(risk_ref = cleaned_data['risk_ref']):
                    raise ValidationError(('Identyfikator musi być unikalny!'), code='wrong_ref')
                return cleaned_data['risk_ref']

        # Przekazanie parametrów nowej klasie
        create_class.base_fields['rev_date'].initial = date.today()
        create_class.hazard = get_object_or_404(SMSHazard, pk=self.kwargs['hazard_id'])

        return create_class

    def form_valid(self, form):
        form.instance.smshazard = get_object_or_404(SMSHazard, pk=self.kwargs['hazard_id'])
        return super(RiskCreate, self).form_valid(form)

    def get_success_url(self):
        return reverse('sms:hazard-details', args=[self.kwargs['hazard_id']])


@class_view_decorator(permission_required('sms.sms_admin'))
class RiskUpdate (UpdateView):
    model = SMSRisk
    template_name = 'sms/update_template.html'

    def get_context_data(self, **kwargs):
        context = super(RiskUpdate, self).get_context_data(**kwargs)
        context['page_title'] = 'Aktualizacja ryzyka'
        context['header_text'] = 'Aktualizacja ryzyka %s' % self.object.risk_ref
        context['versions'] = True
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(SMSRisk, exclude={'smshazard', 'risk_ref', 'rev_num', 'rev_last'},
                                                widgets={'rev_date':AdminDateWidget,
                                                         'description':Textarea(attrs={'rows':4, 'cols':100}),
                                                         'mitigation':Textarea(attrs={'rows':4, 'cols':100})})
        return form_class

    # def get_initial(self):
    #     return {'rev_date': date.today()}

    def form_valid(self, form):

        # Jesli wybrano stworzenie wersji
        if '_revision' in self.request.POST:

            # Utworzenie nowej wersji
            new_rev = SMSRisk(risk_ref = self.object.risk_ref, rev_num = self.object.rev_num + 1, rev_date = form.cleaned_data['rev_date'],
                              rev_last = True, smshazard = self.object.smshazard, description = form.cleaned_data['description'],
                              init_probability = form.cleaned_data['init_probability'], init_impact = form.cleaned_data['init_impact'],
                              mitigation = form.cleaned_data['mitigation'], res_probability	= form.cleaned_data['res_probability'],
                              res_impact = form.cleaned_data['res_impact'])
            new_rev.save()

            # Bieżąca wersja przestaje być ostatnią
            last_rev = SMSRisk.objects.get(pk=self.object.pk)
            last_rev.rev_last = False
            last_rev.save()

            # Powrót bez zapisu bieżącego obiektu
            return HttpResponseRedirect(reverse('sms:hazard-details', args=[self.object.smshazard.pk]))

        return super(RiskUpdate, self).form_valid(form)

    def get_success_url(self):
        return reverse('sms:risk-details', args=(self.object.pk,))


@class_view_decorator(permission_required('sms.sms_admin'))
class RiskDelete (DeleteView):
    model = SMSRisk
    template_name = 'sms/delete_template.html'

    def get_context_data(self, **kwargs):
        context = super(RiskDelete, self).get_context_data(**kwargs)
        context['page_title'] = 'Usunięcie ryzyka'
        context['header_text'] = 'Usunięcie ryzyka %s' % self.object.risk_ref
        context['description'] = self.object.description
        return context

    def delete(self, request, *args, **kwargs):
        risk_ref = SMSRisk.objects.get(pk=kwargs['pk']).risk_ref
        hazard = SMSRisk.objects.get(pk=kwargs['pk']).smshazard
        # Usuń wszystkie z tym risk_ref poza sobą samym
        for risk in hazard.smsrisk_set.filter(risk_ref=risk_ref).exclude(pk=kwargs['pk']):
            risk.delete()
        return super(RiskDelete, self).delete(request, *args, **kwargs)

    def get_success_url(self):
        return reverse('sms:hazard-details', args=[self.object.smshazard.pk])


@class_view_decorator(permission_required('sms.sms_reader'))
class SMSEventList (ListView):
    model = SMSEvent
    template_name = 'sms/list_template.html'

    def get_queryset(self):
        return SMSEvent.objects.order_by('-event_date', '-pk')

    def get_context_data(self, **kwargs):
        context = super(SMSEventList, self).get_context_data(**kwargs)
        context['page_title'] = 'Lista zdarzeń'
        context['header_text'] = 'Lista zarejestrowanych zdarzeń'
        context['create_path'] = reverse('sms:smsevent-create')
        context['empty_text'] = 'Brak zarejestrowanych zdarzeń.'

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Nowe zdarzenie', 'path': reverse("sms:smsevent-create")})
        local_menu.append({'text': 'Generuj raport', 'path': reverse("sms:sms-events-export")})
        context['local_menu'] = local_menu

        # Nagłówki tabeli
        header_list = []
        header_list.append({'header': 'Statek\npowietrzny'})
        header_list.append({'header': 'Pilot\ndowódca'})
        header_list.append({'header': 'Data\nzdarzenia'})
        header_list.append({'header': 'Kwalifikacja\nzdarzenia'})
        header_list.append({'header': 'Opis\nzdarzenia'})
        header_list.append({'header': 'Zagrożenie\nw rejestrze'})
        header_list.append({'header': '...'})
        context['header_list'] = header_list

        row_list = []
        for object in self.object_list:
            fields = []
            fields.append({'name': 'aircraft', 'value': object.aircraft, 'link': reverse('sms:smsevent-details', args=[object.pk]), 'just': 'center'})
            fields.append({'name': 'pic', 'value': object.pic})
            fields.append({'name': 'event_date', 'value': object.event_date, 'just': 'center'})
            fields.append({'name': 'event_type', 'value': object.event_type, 'just': 'center'})
            fields.append({'name': 'description', 'value': object.description})
            if object.smshazard:
                fields.append({'name': 'hazard', 'value': object.smshazard.hazard_ref, 'link': reverse('sms:hazard-details', args=[object.smshazard.pk]), 'just': 'center'})
            else:
                fields.append({'name': 'hazard', 'value': None})
            fields.append({'name': 'change', 'edit_link': reverse('sms:smsevent-update', args=[object.pk]),
                           'delete_link': reverse('sms:smsevent-delete', args=[object.pk])})
            row_list.append({'fields': fields})
        context['row_list'] = row_list
        return context


@class_view_decorator(permission_required('sms.sms_reader'))
class SMSEventDetails (DetailView):
    model = SMSEvent
    template_name = 'sms/details_template.html'

    def get_context_data(self, **kwargs):
        context = super(SMSEventDetails, self).get_context_data(**kwargs)
        context['page_title'] = 'Zdarzenie'
        context['header_text'] = 'Szczegóły zarejestrowanego zdarzenia'

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Aktualizuj', 'path': reverse('sms:smsevent-update', args=[self.object.pk])})
        local_menu.append({'text': 'Usuń', 'path': reverse('sms:smsevent-delete', args=[self.object.pk])})
        context['local_menu'] = local_menu

        field_list = []
        field_list.append({'header': 'Statek powietrzny', 'value': self.object.aircraft, 'bold': True})
        field_list.append({'header': 'Pilot dowódca', 'value': self.object.pic, 'bold': True})
        field_list.append({'header': 'Data zdarzenia', 'value': self.object.event_date, 'bold': True})
        field_list.append({'header': 'Rodzaj operacji', 'value': self.object.oper_type})
        field_list.append({'header': 'Kwalifikacja zdarzenia', 'value': self.object.event_type})
        field_list.append({'header': 'Opis zdarzenia', 'value': self.object.description})
        field_list.append({'header': 'Numer ewidencyjny PKBWL', 'value': self.object.pkbwl_ref})
        field_list.append({'header': 'Data przyjęcia w PKBWL', 'value': self.object.pkbwl_date})
        field_list.append({'header': 'Badający zdarzenie', 'value': self.object.examiner})
        field_list.append({'header': 'Przebieg badania wewnętrznego', 'value': self.object.scrutiny})
        if self.object.smshazard:
            field_list.append({'header': 'Zagrożenie w rejestrze', 'value': self.object.smshazard.hazard_ref,
                               'link':reverse('sms:hazard-details', args=[self.object.smshazard.pk])})
        else:
            field_list.append({'header': 'Zagrożenie w rejestrze', 'value': self.object.smshazard})
        field_list.append({'header': 'Wnioski i zalecenia', 'value': self.object.findings})
        field_list.append({'header': 'Zamknięcie / raport końcowy', 'value': self.object.closure})
        field_list.append({'header': 'Uwagi', 'value': self.object.remarks})
        context['field_list'] = field_list
        return context


@class_view_decorator(permission_required('sms.sms_admin'))
class SMSEventCreate (CreateView):
    model = SMSEvent
    template_name = 'sms/create_template.html'

    def get_context_data(self, **kwargs):
        context = super(SMSEventCreate, self).get_context_data(**kwargs)
        context['page_title'] = 'Rejestracja zdarzenia'
        context['header_text'] = 'Rejestracja nowego zdarzenia'
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(SMSEvent, exclude=['reported_by'],
                                       widgets={'event_date':AdminDateWidget, 'pkbwl_date':AdminDateWidget,
                                                'description':Textarea(attrs={'rows':4, 'cols':100}),
                                                'scrutiny':Textarea(attrs={'rows':4, 'cols':100}),
                                                'findings':Textarea(attrs={'rows':4, 'cols':100}),
                                                'remarks':Textarea(attrs={'rows':4, 'cols':100})})
        form_class.base_fields['event_date'].initial = date.today()
        form_class.base_fields['closure'].initial = '---'
        return form_class

    def get_success_url(self):
        return reverse('sms:smsevent-list')


@class_view_decorator(permission_required('sms.sms_admin'))
class SMSEventUpdate (UpdateView):
    model = SMSEvent
    template_name = 'sms/update_template.html'

    def get_context_data(self, **kwargs):
        context = super(SMSEventUpdate, self).get_context_data(**kwargs)
        context['page_title'] = 'Modyfikacja zdarzenia'
        context['header_text'] = 'Modyfikacja zdarzenia %s / %s' % (self.object.aircraft, self.object.event_date)
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(SMSEvent, exclude=['reported_by'],
                                       widgets={'event_date':AdminDateWidget(), 'pkbwl_date':AdminDateWidget(),
                                                'description':Textarea(attrs={'rows':4, 'cols':100}),
                                                'scrutiny':Textarea(attrs={'rows':4, 'cols':100}),
                                                'findings':Textarea(attrs={'rows':4, 'cols':100}),
                                                'remarks':Textarea(attrs={'rows':4, 'cols':100})})
        return form_class

    def get_success_url(self):
        return reverse('sms:smsevent-details', args=(self.object.pk,))


@class_view_decorator(permission_required('sms.sms_admin'))
class SMSEventDelete (DeleteView):
    model = SMSEvent
    template_name = 'sms/delete_template.html'

    def get_context_data(self, **kwargs):
        context = super(SMSEventDelete, self).get_context_data(**kwargs)
        context['page_title'] = 'Usunięcie zdarzenia'
        context['header_text'] = 'Usunięcie zdarzenia z rejestru'
        context['description'] = self.object.description
        return context

    def get_success_url(self):
        return reverse('sms:smsevent-list')


@class_view_decorator(permission_required('sms.sms_reader'))
class FailureList (ListView):
    model = SMSFailure
    template_name = 'sms/list_template.html'

    def get_queryset(self):
        return SMSFailure.objects.order_by('-failure_date', '-pk')

    def get_context_data(self, **kwargs):
        context = super(FailureList, self).get_context_data(**kwargs)
        context['page_title'] = 'Lista usterek'
        context['header_text'] = 'Lista zarejestrowanych usterek'
        context['empty_text'] = 'Brak zarejestrowanych usterek.'

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Nowa usterka', 'path': reverse("sms:failure-create-free")})
        context['local_menu'] = local_menu

        # Nagłówki tabeli
        header_list = []
        header_list.append({'header': 'Statek\npowietrzny'})
        header_list.append({'header': 'Osoba\nzgłaszająca'})
        header_list.append({'header': 'Data\nzgłoszenia'})
        header_list.append({'header': 'Opis\nusterki'})
        header_list.append({'header': 'PDT'})
        header_list.append({'header': 'Data\nusunięcia'})
        header_list.append({'header': 'CRS'})
        header_list.append({'header': 'Zagrożenie\nw rejestrze'})
        header_list.append({'header': '...'})
        context['header_list'] = header_list

        row_list = []
        for object in self.object_list:
            fields = []
            fields.append({'name': 'aircraft', 'value': object.aircraft, 'link': reverse('sms:failure-details', args=[object.pk]), 'just': 'center'})
            fields.append({'name': 'person', 'value': object.person})
            fields.append({'name': 'failure_date', 'value': object.failure_date, 'just': 'center'})
            fields.append({'name': 'description', 'value': object.description })
            fields.append({'name': 'pdt_ref', 'value': object.pdt_ref })
            fields.append({'name': 'repair_date', 'value': object.repair_date, 'just': 'center'})
            fields.append({'name': 'crs', 'value': object.crs, 'just': 'center'})
            if object.smshazard:
                fields.append({'name': 'hazard', 'value': object.smshazard.hazard_ref, 'link': reverse('sms:hazard-details', args=[object.smshazard.pk]), 'just': 'center'})
            else:
                fields.append({'name': 'hazard', 'value': None})
            fields.append({'name': 'change', 'edit_link': reverse('sms:failure-update', args=[object.pk]),
                           'delete_link': reverse('sms:failure-delete', args=[object.pk])})
            row_list.append({'fields': fields})
        context['row_list'] = row_list
        return context


@class_view_decorator(permission_required('sms.sms_reader'))
class FailureDetails (DetailView):
    model = SMSFailure
    template_name = 'sms/details_template.html'

    def get_context_data(self, **kwargs):
        context = super(FailureDetails, self).get_context_data(**kwargs)
        context['page_title'] = 'Usterka'
        context['header_text'] = 'Szczegóły zarejestrowanej usterki'

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Aktualizuj', 'path': reverse('sms:failure-update', args=[self.object.pk])})
        local_menu.append({'text': 'Usuń', 'path': reverse('sms:failure-delete', args=[self.object.pk])})
        context['local_menu'] = local_menu

        field_list = []
        field_list.append({'header': 'Statek powietrzny', 'value': self.object.aircraft, 'bold': True})
        field_list.append({'header': 'Osoba zgłaszająca', 'value': self.object.person, 'bold': True})
        field_list.append({'header': 'Data zgłoszenia', 'value': self.object.failure_date, 'bold': True})
        field_list.append({'header': 'Opis usterki', 'value': self.object.description})
        field_list.append({'header': 'PDT', 'value': self.object.pdt_ref})
        field_list.append({'header': 'Data usunięcia', 'value': self.object.repair_date})
        field_list.append({'header': 'Sposób usunięcia', 'value': self.object.repair_desc})
        field_list.append({'header': 'CRS', 'value': self.object.crs})
        if self.object.smshazard:
            field_list.append({'header': 'Zagrożenie w rejestrze', 'value': self.object.smshazard.hazard_ref,
                               'link':reverse('sms:hazard-details', args=[self.object.smshazard.pk])})
        else:
            field_list.append({'header': 'Zagrożenie w rejestrze', 'value': self.object.smshazard})
        field_list.append({'header': 'Wnioski i zalecenia', 'value': self.object.findings})
        context['field_list'] = field_list
        return context


@class_view_decorator(permission_required('sms.sms_admin'))
class FailureCreate (CreateView):
    model = SMSFailure
    template_name = 'sms/create_template.html'

    def get_context_data(self, **kwargs):
        context = super(FailureCreate, self).get_context_data(**kwargs)
        context['page_title'] = 'Rejestracja usterki'
        context['header_text'] = 'Rejestracja nowej usterki'
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(SMSFailure, exclude=['pdt_ref'],
                                       widgets={'failure_date':AdminDateWidget(), 'repair_date':AdminDateWidget(),
                                                'description':Textarea(attrs={'rows':4, 'cols':100}),
                                                'repair_desc':Textarea(attrs={'rows':4, 'cols':100}),
                                                'findings':Textarea(attrs={'rows':4, 'cols':100})})
        return form_class

    def get_success_url(self):
        return reverse('sms:failure-list')


@class_view_decorator(permission_required('sms.sms_admin'))
class FailureUpdate (UpdateView):
    model = SMSFailure
    template_name = 'sms/update_template.html'

    def get_context_data(self, **kwargs):
        context = super(FailureUpdate, self).get_context_data(**kwargs)
        context['page_title'] = 'Modyfikacja usterki'
        context['header_text'] = 'Modyfikacja usterki %s / %s' % (self.object.aircraft, self.object.failure_date)
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(SMSFailure, exclude=['pdt_ref'],
                                       widgets={'failure_date':AdminDateWidget(), 'repair_date':AdminDateWidget(),
                                                'description':Textarea(attrs={'rows':4, 'cols':100}),
                                                'repair_desc':Textarea(attrs={'rows':4, 'cols':100}),
                                                'findings':Textarea(attrs={'rows':4, 'cols':100})})
        return form_class

    def get_success_url(self):
        return reverse('sms:failure-details', args=(self.object.pk,))


@class_view_decorator(permission_required('sms.sms_admin'))
class FailureDelete (DeleteView):
    model = SMSFailure
    template_name = 'sms/delete_template.html'

    def get_context_data(self, **kwargs):
        context = super(FailureDelete, self).get_context_data(**kwargs)
        context['page_title'] = 'Usunięcie usterki'
        context['header_text'] = 'Usunięcie usterki z rejestru'
        context['description'] = self.object.description
        return context

    def get_success_url(self):
        return reverse('sms:failure-list')


@class_view_decorator(permission_required('sms.sms_reader'))
class SMSReportList (ListView):
    model = SMSReport
    template_name = 'sms/list_template.html'

    def get_queryset(self):
        return SMSReport.objects.order_by('-report_date', '-pk')

    def get_context_data(self, **kwargs):
        context = super(SMSReportList, self).get_context_data(**kwargs)
        context['page_title'] = 'Lista raportów'
        context['header_text'] = 'Lista dobrowolnych raportów SMS'
        context['empty_text'] = 'Brak dobrowolnych raportów SMS.'

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Rejestracja raportu SMS', 'path': reverse("sms:smsreport-create")})
        local_menu.append({'text': 'Generuj raport', 'path': reverse("sms:sms-reports-export")})
        context['local_menu'] = local_menu

        # Nagłówki tabeli
        header_list = []
        header_list.append({'header': 'Data\nzgłoszenia'})
        header_list.append({'header': 'Osoba\nzgłaszająca'})
        header_list.append({'header': 'Poufność'})
        header_list.append({'header': 'Treść\nzgłoszenia'})
        header_list.append({'header': 'Zagrożenie\nw rejestrze'})
        header_list.append({'header': 'Uwagi'})
        header_list.append({'header': '...'})
        context['header_list'] = header_list

        row_list = []
        for object in self.object_list:
            fields = []
            fields.append({'name': 'report_date', 'value': object.report_date, 'link': reverse('sms:smsreport-details', args=[object.pk]), 'just': 'center'})
            fields.append({'name': 'person', 'value': object.person})
            fields.append({'name': 'privacy', 'value': 'TAK' if object.privacy else 'NIE', 'just': 'center'})
            fields.append({'name': 'description', 'value': object.description})
            if object.smshazard:
                fields.append({'name': 'hazard', 'value': object.smshazard.hazard_ref, 'link': reverse('sms:hazard-details', args=[object.smshazard.pk]), 'just': 'center'})
            else:
                fields.append({'name': 'hazard', 'value': None})
            fields.append({'name': 'remarks', 'value': object.remarks})
            fields.append({'name': 'change', 'edit_link': reverse('sms:smsreport-update', args=[object.pk]),
                           'delete_link': reverse('sms:smsreport-delete', args=[object.pk])})
            row_list.append({'fields': fields})
        context['row_list'] = row_list
        return context


@class_view_decorator(permission_required('sms.sms_reader'))
class SMSReportDetails (DetailView):
    model = SMSReport
    template_name = 'sms/details_template.html'

    def get_context_data(self, **kwargs):
        context = super(SMSReportDetails, self).get_context_data(**kwargs)
        context['page_title'] = 'Raport SMS'
        context['header_text'] = 'Szczegóły dobrowolnego raportu SMS'

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Aktualizuj', 'path': reverse('sms:smsreport-update', args=[self.object.pk])})
        local_menu.append({'text': 'Usuń', 'path': reverse('sms:smsreport-delete', args=[self.object.pk])})
        context['local_menu'] = local_menu

        field_list = []
        field_list.append({'header': 'Data zgłoszenia', 'value': self.object.report_date, 'bold': True})
        field_list.append({'header': 'Osoba zgłaszająca', 'value': self.object.person, 'bold': True})
        field_list.append({'header': 'Poufność', 'value': 'TAK' if self.object.privacy else 'NIE'})
        field_list.append({'header': 'Treść zgłoszenia', 'value': self.object.description})
        if self.object.smshazard:
            field_list.append({'header': 'Zagrożenie w rejestrze', 'value': self.object.smshazard, 'link':reverse('sms:hazard-details', args=[self.object.smshazard.pk])})
        else:
            field_list.append({'header': 'Zagrożenie w rejestrze', 'value': self.object.smshazard})
        field_list.append({'header': 'Wnioski i zalecenia', 'value': self.object.findings})
        field_list.append({'header': 'Uwagi', 'value': self.object.remarks})

        context['field_list'] = field_list
        return context


@class_view_decorator(permission_required('sms.sms_admin'))
class SMSReportCreate (CreateView):
    model = SMSReport
    template_name = 'sms/create_template.html'

    def get_context_data(self, **kwargs):
        context = super(SMSReportCreate, self).get_context_data(**kwargs)
        context['page_title'] = 'Rejestracja raportu'
        context['header_text'] = 'Rejestracja dobrowolnego raportu SMS'
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(SMSReport, exclude=['reported_by'],
                                       widgets={'report_date':AdminDateWidget(),
                                                'description':Textarea(attrs={'rows':4, 'cols':100}),
                                                'findings':Textarea(attrs={'rows':4, 'cols':100}),
                                                'remarks':Textarea(attrs={'rows':4, 'cols':100})})
        form_class.base_fields['report_date'].initial = date.today()
        return form_class

    def get_success_url(self):
        return reverse('sms:smsreport-list')


@class_view_decorator(permission_required('sms.sms_admin'))
class SMSReportUpdate (UpdateView):
    model = SMSReport
    template_name = 'sms/update_template.html'

    def get_context_data(self, **kwargs):
        context = super(SMSReportUpdate, self).get_context_data(**kwargs)
        context['page_title'] = 'Modyfikacja raportu'
        context['header_text'] = 'Modyfikacja dobrowolnego raportu SMS'
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(SMSReport, exclude=['reported_by'],
                                       widgets={'report_date':AdminDateWidget(),
                                                'description':Textarea(attrs={'rows':4, 'cols':100}),
                                                'findings':Textarea(attrs={'rows':4, 'cols':100}),
                                                'remarks':Textarea(attrs={'rows':4, 'cols':100})})
        return form_class

    def get_success_url(self):
        return reverse('sms:smsreport-details', args=[self.object.pk,])


@class_view_decorator(permission_required('sms.sms_admin'))
class SMSReportDelete (DeleteView):
    model = SMSReport
    template_name = 'sms/delete_template.html'

    def get_context_data(self, **kwargs):
        context = super(SMSReportDelete, self).get_context_data(**kwargs)
        context['page_title'] = 'Usunięcie raportu'
        context['header_text'] = 'Usunięcie dobrowolnego raportu SMS'
        context['description'] = self.object.description
        return context

    def get_success_url(self):
        return reverse('sms:smsreport-list')


@class_view_decorator(permission_required('sms.sms_ncr'))
class NCRList (ListView):
    model = NCR
    template_name = 'sms/list_template.html'

    def get_queryset(self):
        return NCR.objects.order_by('-audit_date', '-pk')

    def get_context_data(self, **kwargs):
        context = super(NCRList, self).get_context_data(**kwargs)
        context['page_title'] = 'Lista NCR'
        context['header_text'] = 'Lista niezgodności (NCR)'
        context['empty_text'] = 'Brak zarejestrowanych niezgodności.'

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Rejestracja NCR', 'path': reverse("sms:ncr-create")})
        local_menu.append({'text': 'Generuj raport', 'path': reverse("sms:ncr-report")})
        context['local_menu'] = local_menu

        # Nagłówki tabeli
        header_list = []
        header_list.append({'header': 'Obszar\naudytu'})
        header_list.append({'header': 'Rodzaj\naudytu'})
        header_list.append({'header': 'Numer\naudytu'})
        header_list.append({'header': 'Data\naudytu'})
        header_list.append({'header': 'Num.\nNCR'})
        header_list.append({'header': 'Treść NCR'})
        header_list.append({'header': 'Wyznaczona\ndata\nusunięcia'})
        header_list.append({'header': 'Osoba\nodpow. 1'})
        header_list.append({'header': 'Osoba\nodpow. 2'})
        header_list.append({'header': 'Pozostało\ndni'})
        header_list.append({'header': 'Data\nusunięcia'})
        header_list.append({'header': 'Audyt\nsprawdz.'})
        header_list.append({'header': '...'})
        context['header_list'] = header_list

        row_list = []
        for object in self.object_list:
            fields = []
            fields.append({'name': 'audit_scope', 'value': object.audit_scope, 'just': 'center'})
            fields.append({'name': 'audit_type', 'value': object.audit_type, 'just': 'center'})
            fields.append({'name': 'audit_nbr', 'value': object.audit_nbr, 'link': reverse('sms:ncr-details', args=[object.pk]), 'just': 'center'})
            fields.append({'name': 'audit_date', 'value': object.audit_date, 'just': 'center'})
            fields.append({'name': 'ncr_nbr', 'value': object.ncr_nbr, 'just': 'center'})
            fields.append({'name': 'description', 'value': object.description})
            fields.append({'name': 'due_date', 'value': object.due_date, 'just': 'center'})
            fields.append({'name': 'ncr_user1', 'value': object.ncr_user1})
            fields.append({'name': 'ncr_user2', 'value': object.ncr_user2})
            if object.left_days() != None:
                if object.left_days() >= 20:
                    color = 'lightgreen'
                elif object.left_days() > 0:
                    color = 'yellow'
                else:
                    color = 'lightcoral'
                fields.append({'name': 'left_days', 'value': object.left_days, 'color': color, 'just': 'center'})
            else:
                fields.append({'name': 'left_days', 'value': ''})
            fields.append({'name': 'done_date', 'value': object.done_date, 'just': 'center'})
            fields.append({'name': 'check_date', 'value': object.check_date, 'just': 'center'})
            fields.append({'name': 'change', 'edit_link': reverse('sms:ncr-update', args=[object.pk]),
                           'delete_link': reverse('sms:ncr-delete', args=[object.pk])})
            row_list.append({'fields': fields})
        context['row_list'] = row_list
        return context


@class_view_decorator(permission_required('sms.sms_ncr'))
class NCRDetails (DetailView):
    model = NCR
    template_name = 'sms/details_template.html'

    def get_context_data(self, **kwargs):
        context = super(NCRDetails, self).get_context_data(**kwargs)
        context['page_title'] = 'NCR'
        context['header_text'] = 'Szczegóły niezgodności (NCR)'

        # Lokalne menu
        local_menu = []
        local_menu.append({'text': 'Aktualizuj', 'path': reverse('sms:ncr-update', args=[self.object.pk])})
        local_menu.append({'text': 'Usuń', 'path': reverse('sms:ncr-delete', args=[self.object.pk])})
        context['local_menu'] = local_menu

        field_list = []
        field_list.append({'header': 'Numer audytu', 'value': self.object.audit_nbr, 'bold': True})
        field_list.append({'header': 'Data audytu', 'value': self.object.audit_date, 'bold': True})
        field_list.append({'header': 'Numer NCR', 'value': self.object.ncr_nbr})
        field_list.append({'header': 'Treść NCR', 'value': self.object.description})
        field_list.append({'header': 'Wyznaczona data usunięcia', 'value': self.object.due_date})
        field_list.append({'header': 'Osoba odpowiedzialna 1', 'value': self.object.ncr_user1})
        field_list.append({'header': 'Osoba odpowiedzialna 2', 'value': self.object.ncr_user2})
        field_list.append({'header': 'Pozostało dni', 'value': self.object.left_days})
        field_list.append({'header': 'Faktyczna data usunięcia', 'value': self.object.done_date})
        field_list.append({'header': 'Data audytu sprawdzającego', 'value': self.object.check_date})

        context['field_list'] = field_list
        return context


@class_view_decorator(permission_required('sms.sms_ncr'))
class NCRCreate (CreateView):
    model = NCR
    template_name = 'sms/create_template.html'

    def get_context_data(self, **kwargs):
        context = super(NCRCreate, self).get_context_data(**kwargs)
        context['page_title'] = 'Rejestracja NCR'
        context['header_text'] = 'Rejestracja niezgodności (NCR)'
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(NCR, fields='__all__',
                                       widgets={'audit_date':AdminDateWidget(),
                                                'due_date': AdminDateWidget(),
                                                'done_date': AdminDateWidget(),
                                                'check_date': AdminDateWidget(),
                                                'description':Textarea(attrs={'rows':4, 'cols':100})})
        # Pole Osoba odpowiedzialna
        choices = [('', '---------')]
        for fbouser in FBOUser.objects.filter(ncr_user=True):
            choices += [(fbouser.pk, fbouser.__str__())]

        form_class.base_fields['audit_date'].initial = date.today()
        form_class.base_fields['ncr_user1'].choices = choices
        form_class.base_fields['ncr_user2'].choices = choices

        return form_class

    def get_success_url(self):
        return reverse('sms:ncr-list')


@class_view_decorator(permission_required('sms.sms_ncr'))
class NCRUpdate (UpdateView):
    model = NCR
    template_name = 'sms/update_template.html'

    def get_context_data(self, **kwargs):
        context = super(NCRUpdate, self).get_context_data(**kwargs)
        context['page_title'] = 'Modyfikacja NCR'
        context['header_text'] = 'Modyfikacja niezgodności (NCR)'
        return context

    def get_form_class(self, **kwargs):
        form_class = modelform_factory(NCR, fields='__all__',
                                       widgets={'audit_date':AdminDateWidget(),
                                                'due_date': AdminDateWidget(),
                                                'done_date': AdminDateWidget(),
                                                'check_date': AdminDateWidget(),
                                                'description':Textarea(attrs={'rows':4, 'cols':100})})
        # Pole Osoba odpowiedzialna
        choices = [('', '---------')]
        for fbouser in FBOUser.objects.filter(ncr_user=True):
            choices += [(fbouser.pk, fbouser.__str__())]

        form_class.base_fields['ncr_user1'].choices = choices
        form_class.base_fields['ncr_user2'].choices = choices

        return form_class

    def get_success_url(self):
        return reverse('sms:ncr-details', args=[self.object.pk,])


@class_view_decorator(permission_required('sms.sms_ncr'))
class NCRDelete (DeleteView):
    model = NCR
    template_name = 'sms/delete_template.html'

    def get_context_data(self, **kwargs):
        context = super(NCRDelete, self).get_context_data(**kwargs)
        context['page_title'] = 'Usunięcie NCR'
        context['header_text'] = 'Usunięcie niezgodności (NCR)'
        context['description'] = self.object.description
        return context

    def get_success_url(self):
        return reverse('sms:ncr-list')


@permission_required('sms.sms_reader')
def SMSHazardsExport (request):

    # Nowy arkusz
    wb = Workbook()

    # Zakładki odpowiadające typom zagrożeń
    for i, type in enumerate(SMSHazard.objects.filter(hazard_active = True).values('hazard_type').distinct()):
        if i == 0:
            wb.active.title = type['hazard_type']
        else:
            wb.create_sheet(type['hazard_type'])

    # Definicje stylów
    border = Border(top=Side(border_style=borders.BORDER_THIN, color='000000'), bottom=Side(border_style=borders.BORDER_THIN, color='000000'),
                    left=Side(border_style=borders.BORDER_THIN, color='000000'), right=Side(border_style=borders.BORDER_THIN, color='000000'))
    title_allign = Alignment(horizontal='center', vertical='center', wrap_text = True)
    table_align = Alignment(horizontal='left', vertical='center', wrap_text = True)
    title_fill = PatternFill(fill_type=fills.FILL_SOLID, fgColor=('99CCFF'))
    green_fill = PatternFill(fill_type=fills.FILL_SOLID, fgColor=('90EE90'))
    yellow_fill = PatternFill(fill_type=fills.FILL_SOLID, fgColor=('FFFF00'))
    red_fill = PatternFill(fill_type=fills.FILL_SOLID, fgColor=('FF7F50'))

    fill_colors = {'lightgreen':green_fill, 'yellow':yellow_fill, 'coral':red_fill}

    # Wypełnienie poszczególnych zakładek
    for ws in wb.worksheets:

        # Informacje nagłówkowe
        ws['A1'] = "ID."
        ws.merge_cells('A1:A2')
        ws['B1'] = "REWIZJA"
        ws.merge_cells('B1:B2')
        ws['C1'] = "OBSZAR"
        ws.merge_cells('C1:C2')
        ws['D1'] = "NAZWA ZAGROŻENIA"
        ws.merge_cells('D1:D2')
        ws['E1'] = "OPIS NATURY RYZYKA"
        ws.merge_cells('E1:E2')
        ws['F1'] = "PIERWOTNE"
        ws.merge_cells('F1:G1')
        ws['F2'] = "PRAWDOPO- DOBIEŃSTWO"
        ws['G2'] = "SZKODLIWOŚĆ"
        ws['H1'] = "SPOSÓB OGRANICZANIA RYZYKA"
        ws.merge_cells('H1:H2')
        ws['I1'] = "SZCZĄTKOWE"
        ws.merge_cells('I1:J1')
        ws['I2'] = "PRAWDOPO- DOBIEŃSTWO"
        ws['J2'] = "SZKODLIWOŚĆ"
        ws['K1'] = "TERMIN"
        ws.merge_cells('K1:K2')
        ws['L1'] = "ODPOWIEDZIALNY"
        ws.merge_cells('L1:L2')
        ws['M1'] = "KONTROLA"
        ws.merge_cells('M1:M2')
        ws['N1'] = "UWAGI"
        ws.merge_cells('N1:N2')

        # Format nagłówków
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            for row in [1,2]:
                ws['%s%d' % (col, row)].fill = title_fill
                ws['%s%d' % (col, row)].alignment = title_allign
                ws['%s%d' % (col, row)].border = border

        # Zawartość zakładek
        row = 3
        for hazard in SMSHazard.objects.filter(hazard_type = ws.title, hazard_active=True).order_by('hazard_ref'):

            # Wybór ostatniej wersji zagrożenia
            last_rev = hazard.smshazardrev_set.filter(rev_last=True).first()
            if last_rev:

                # Wypełnienie wiersza jedynie jeśli istnieje ostatnia wersja
                ws['A%d' % row] = '%s\n%s' % (hazard.hazard_ref, hazard.hazard_date)
                ws['B%d' % row] = 'Rev. %d\n%s' % (last_rev.rev_num, last_rev.rev_date)
                ws['C%d' % row] = last_rev.company_area
                ws['D%d' % row] = last_rev.name
                ws['K%d' % row] = last_rev.due_date
                ws['L%d' % row] = last_rev.responsible
                ws['M%d' % row] = last_rev.control
                ws['N%d' % row] = last_rev.remarks

                # Poszczególne ryzyka dla zagrożenia
                risks = hazard.smsrisk_set.filter(rev_last=True)
                if len(risks) > 0:
                    for risk in risks:
                        ws['E%d' % row] = risk.description
                        ws['F%d' % row] = risk.init_probability
                        ws['F%d' % row].fill = fill_colors[risk.risk_color()[0]]
                        ws['G%d' % row] = risk.init_impact
                        ws['G%d' % row].fill = fill_colors[risk.risk_color()[0]]
                        ws['H%d' % row] = risk.mitigation
                        ws['I%d' % row] = risk.res_probability
                        ws['I%d' % row].fill = fill_colors[risk.risk_color()[1]]
                        ws['J%d' % row] = risk.res_impact
                        ws['J%d' % row].fill = fill_colors[risk.risk_color()[1]]
                        row += 1
                else:
                    row += 1

                # Scalenie komórek w pryzpadku wielu ryzyk dla zagrożenia
                if len(risks) > 1:
                    for col in ['A', 'B', 'C', 'D', 'K', 'L', 'M', 'N']:
                        ws.merge_cells('%s%d:%s%d' % (col,row-len(risks),col,row-1))

        # Ustawienie szerokości kolumn
        ws.column_dimensions['D'].width = '60'
        ws.column_dimensions['E'].width = '60'
        ws.column_dimensions['F'].width = '16'
        ws.column_dimensions['G'].width = '16'
        ws.column_dimensions['H'].width = '60'
        ws.column_dimensions['I'].width = '16'
        ws.column_dimensions['J'].width = '16'
        ws.column_dimensions['K'].width = '12'
        ws.column_dimensions['L'].width = '25'
        ws.column_dimensions['M'].width = '50'
        ws.column_dimensions['N'].width = '60'

        # Formatowanie komórek tabeli
        max_row = row
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            for row in range(3, max_row):
                ws['%s%d' % (col, row)].border = border
                ws['%s%d' % (col, row)].alignment = table_align

        for row in range(3, max_row):
            ws['A%d' % row].alignment = title_allign
            ws['B%d' % row].alignment = title_allign
            ws['C%d' % row].alignment = title_allign

        big_font = Font(size=16, bold=True)
        for col in ['F', 'G', 'I', 'J']:
            for row in range(3, max_row):
                ws['%s%d' % (col, row)].alignment = title_allign
                ws['%s%d' % (col, row)].font = big_font

    # Zwróć arkusz jako obiekt HTTP Response
    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = ('attachment; filename=sms_hazards_%s.xlsx' % date.today()).replace('-', '')

    return response


@permission_required('sms.sms_reader')
def SMSEventsExport (request):

    # Nowy arkusz
    wb = Workbook()
    wb.active.title = 'Zdarzenia lotnicze'

    # Definicje stylów
    border = Border(top=Side(border_style=borders.BORDER_THIN, color='000000'), bottom=Side(border_style=borders.BORDER_THIN, color='000000'),
                    left=Side(border_style=borders.BORDER_THIN, color='000000'), right=Side(border_style=borders.BORDER_THIN, color='000000'))
    title_allign = Alignment(horizontal='center', vertical='center', wrap_text = True)
    table_align = Alignment(horizontal='left', vertical='center', wrap_text = True)
    title_fill = PatternFill(fill_type=fills.FILL_SOLID, fgColor=('99CCFF'))

    ws = wb.active

    # Informacje nagłówkowe
    ws['A1'] = "STATEK POWIETRZNY"
    ws['B1'] = "PILOT DOWÓDCA"
    ws['C1'] = "DATA ZDARZENIA"
    ws['D1'] = "RODZAJ OPERACJI"
    ws['E1'] = "KWALIFIKACJA ZDARZENIA"
    ws['F1'] = "OPIS ZDARZENIA"
    ws['G1'] = "NUMER PKBWL"
    ws['H1'] = "DATA PKBWL"
    ws['I1'] = "BADAJĄCY ZDARZENIE"
    ws['J1'] = "BADANIE WEWNĘTRZNE"
    ws['K1'] = "ZAGROŻENIE"
    ws['L1'] = "WNIOSKI I ZALECENIA"
    ws['M1'] = "ZAMKNIĘCIE RAPORT"
    ws['N1'] = "UWAGI"

    # Format nagłówków
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
        ws['%s1' % col].fill = title_fill
        ws['%s1' % col].alignment = title_allign
        ws['%s1' % col].border = border

    # Zawartość zakładek
    row = 2
    for event in SMSEvent.objects.order_by('event_date'):

        # Wypełnienie wiersza jedynie jeśli istnieje ostatnia wersja
        ws['A%d' % row] = event.aircraft
        ws['B%d' % row] = event.pic
        ws['C%d' % row] = event.event_date
        ws['D%d' % row] = event.oper_type
        ws['E%d' % row] = event.event_type
        ws['F%d' % row] = event.description
        ws['G%d' % row] = event.pkbwl_ref
        ws['H%d' % row] = event.pkbwl_date
        ws['I%d' % row] = event.examiner
        ws['J%d' % row] = event.scrutiny
        ws['K%d' % row] = event.smshazard.hazard_ref if event.smshazard else ''
        ws['L%d' % row] = event.findings
        ws['M%d' % row] = event.closure
        ws['N%d' % row] = event.remarks

        row += 1

    # Ustawienie szerokości kolumn
    ws.column_dimensions['A'].width = '14'
    ws.column_dimensions['B'].width = '20'
    ws.column_dimensions['D'].width = '12'
    ws.column_dimensions['E'].width = '15'
    ws.column_dimensions['F'].width = '60'
    ws.column_dimensions['I'].width = '25'
    ws.column_dimensions['J'].width = '60'
    ws.column_dimensions['K'].width = '14'
    ws.column_dimensions['L'].width = '60'
    ws.column_dimensions['M'].width = '15'
    ws.column_dimensions['N'].width = '60'

    # Formatowanie komórek tabeli
    max_row = row
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
        for row in range(2, max_row):
            ws['%s%d' % (col, row)].border = border
            ws['%s%d' % (col, row)].alignment = table_align

    for col in ['A', 'B', 'C', 'D', 'E', 'G', 'H', 'K', 'M']:
        for row in range(2, max_row):
            ws['%c%d' % (col, row)].alignment = title_allign

    # Zwróć arkusz jako obiekt HTTP Response
    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = ('attachment; filename=sms_events_%s.xlsx' % date.today()).replace('-', '')

    return response


@permission_required('sms.sms_reader')
def SMSReportsExport (request):

    # Nowy arkusz
    wb = Workbook()
    wb.active.title = 'Raporty SMS'

    # Definicje stylów
    border = Border(top=Side(border_style=borders.BORDER_THIN, color='000000'), bottom=Side(border_style=borders.BORDER_THIN, color='000000'),
                    left=Side(border_style=borders.BORDER_THIN, color='000000'), right=Side(border_style=borders.BORDER_THIN, color='000000'))
    title_allign = Alignment(horizontal='center', vertical='center', wrap_text = True)
    table_align = Alignment(horizontal='left', vertical='center', wrap_text = True)
    title_fill = PatternFill(fill_type=fills.FILL_SOLID, fgColor=('99CCFF'))

    ws = wb.active

    # Informacje nagłówkowe
    ws['A1'] = "DATA RAPORTU"
    ws['B1'] = "OSOBA ZGŁASZAJĄCA"
    ws['C1'] = "POUFNOŚĆ"
    ws['D1'] = "TREŚĆ ZGŁOSZENIA"
    ws['E1'] = "ZAGROŻENIE"
    ws['F1'] = "WNIOSKI I ZALECENIA"
    ws['G1'] = "UWAGI"

    # Format nagłówków
    for col in ['A', 'B', 'C', 'D', 'E', 'F','G']:
        ws['%s1' % col].fill = title_fill
        ws['%s1' % col].alignment = title_allign
        ws['%s1' % col].border = border

    # Zawartość zakładek
    row = 2
    for report in SMSReport.objects.order_by('report_date'):

        # Wypełnienie wiersza jedynie jeśli istnieje ostatnia wersja
        ws['A%d' % row] = report.report_date
        ws['B%d' % row] = report.person
        ws['C%d' % row] = 'TAK' if report.privacy else 'NIE'
        ws['D%d' % row] = report.description
        ws['E%d' % row] = report.smshazard.hazard_ref if report.smshazard else ''
        ws['F%d' % row] = report.findings
        ws['G%d' % row] = report.remarks

        row += 1

    # Ustawienie szerokości kolumn
    ws.column_dimensions['B'].width = '18'
    ws.column_dimensions['D'].width = '55'
    ws.column_dimensions['E'].width = '14'
    ws.column_dimensions['F'].width = '55'
    ws.column_dimensions['G'].width = '55'

    # Formatowanie komórek tabeli
    max_row = row
    for col in ['A', 'B', 'C', 'D', 'E', 'F','G']:
        for row in range(2, max_row):
            ws['%s%d' % (col, row)].border = border
            ws['%s%d' % (col, row)].alignment = table_align

    for col in ['A', 'B', 'C', 'E']:
        for row in range(2, max_row):
            ws['%c%d' % (col, row)].alignment = title_allign

    # Zwróć arkusz jako obiekt HTTP Response
    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = ('attachment; filename=sms_reports_%s.xlsx' % date.today()).replace('-', '')

    return response


@permission_required('sms.sms_reader')
def SMSFlightsExport (request):

    # Nowy arkusz
    wb = Workbook()
    wb.active.title = 'Statystyka lotów'

    # Definicje stylów
    border = Border(top=Side(border_style=borders.BORDER_THIN, color='000000'), bottom=Side(border_style=borders.BORDER_THIN, color='000000'),
                    left=Side(border_style=borders.BORDER_THIN, color='000000'), right=Side(border_style=borders.BORDER_THIN, color='000000'))
    title_allign = Alignment(horizontal='center', vertical='center', wrap_text = True)
    table_align = Alignment(horizontal='left', vertical='center', wrap_text = True)
    title_fill = PatternFill(fill_type=fills.FILL_SOLID, fgColor=('99CCFF'))

    ws = wb.active

    # Informacje nagłówkowe
    ws['A1'] = 'SP'
    ws['B1'] = 'Data'
    ws['C1'] = 'TTH Start'
    ws['D1'] = 'PDT Ref'
    ws['E1'] = 'Open time'
    ws['F1'] = 'Filight type'
    ws['G1'] = 'PIC'
    ws['H1'] = 'SIC'
    ws['I1'] = 'Suma TTH'
    ws['J1'] = 'Suma czasu'
    ws['K1'] = 'Suma lądowań'

    # Kolejne loty
    row = 2
    for pdt in PDT.objects.order_by('aircraft', 'date', 'tth_start', 'pdt_ref', 'open_time'):
        ws['A%d' % row] = pdt.aircraft.registration
        ws['B%d' % row] = pdt.date
        ws['C%d' % row] = pdt.tth_start
        ws['D%d' % row] = pdt.pdt_ref
        ws['E%d' % row] = pdt.open_time
        ws['F%d' % row] = pdt.flight_type
        ws['G%d' % row] = pdt.pic.__str__()
        ws['H%d' % row] = pdt.sic.__str__()
        ws['I%d' % row] = pdt.tth_sum()
        ws['J%d' % row] = '%02d:%02d' % (pdt.hours_sum()[1], pdt.hours_sum()[2])
        ws['K%d' % row] = pdt.landings_sum()
        row += 1

    # Zwróć arkusz jako obiekt HTTP Response
    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = ('attachment; filename=flights_stats_%s.xlsx' % date.today()).replace('-', '')

    return response


@permission_required('sms.sms_ncr')
def NCRReport (request):

    class NCRReportForm (Form):
        audit_scope = ChoiceField(choices=[(0, 'AOC'), (1,'SPO'), (2, 'CAMO'),
                                           (3, 'ATO')], label='Obszar audytu')
        audit_type = ChoiceField(choices=[(0, 'Wewnętrzny'), (1, 'Zewnętrzny')],
                                           label='Rodzaj audytu')

    MyForm = NCRReportForm
    form = MyForm(request.POST or None)

    context = {}
    context['title'] = 'Generacja raportu NRC'
    context['form'] = form

    # jeśli formularz został poprawnie wypełniony
    if form.is_valid():

        # przejście do kolejnego formularza
        return HttpResponseRedirect(reverse('sms:ncr-export', args=[form.cleaned_data['audit_scope'], form.cleaned_data['audit_type']]))

    return render(request, 'sms/file_export.html', context)


@permission_required('sms.sms_ncr')
def NCRExport (request, scope, type):

    # Nowy arkusz
    wb = Workbook()
    wb.active.title = 'Zestawienie NCR'

    # Definicje stylów
    border = Border(top=Side(border_style=borders.BORDER_THIN, color='000000'), bottom=Side(border_style=borders.BORDER_THIN, color='000000'),
                    left=Side(border_style=borders.BORDER_THIN, color='000000'), right=Side(border_style=borders.BORDER_THIN, color='000000'))
    title_align = Alignment(horizontal='center', vertical='center', wrap_text = True)
    table_align = Alignment(vertical='center', wrap_text=True)
    green_fill = PatternFill(fill_type=fills.FILL_SOLID, fgColor=('90EE90'))
    yellow_fill = PatternFill(fill_type=fills.FILL_SOLID, fgColor=('FFFF00'))
    red_fill = PatternFill(fill_type=fills.FILL_SOLID, fgColor=('FF7F50'))

    fill_colors = {'lightgreen': green_fill, 'yellow': yellow_fill, 'lightcoral': red_fill}

    ws = wb.active

    # Informacje nagłówkowe
    ws['A2'] = "REJESTR NIEZGODNOŚCI Z AUDYTÓW %s" % ['WEWNĘTRZNYCH', 'ZEWNĘTRZNYCH'][int(type)]
    ws.merge_cells('A2:I2')
    ws['A3'] = "%s SALT W 2016 ROKU" % ['AOC', 'SPO', 'CAMO', 'ATO'][int(scope)]
    ws.merge_cells('A3:I3')

    ws['A5'] = "L.p."
    ws['B5'] = "Nr audytu"
    ws['C5'] = "Audyt z dnia"
    ws['D5'] = "Nr NCR"
    ws['E5'] = "Czego dotyczy"
    ws['F5'] = "Wyznaczony termin usunięcia niezgodności"
    ws['G5'] = "Osoba odpowiedzialna 1"
    ws['H5'] = "Osoba odpowiedzialna 2"
    ws['I5'] = "Flaga (dni do końca terminu)"
    ws['J5'] = "Termin usunięcia niezgodności"
    ws['K5'] = "Termin audytu sprawdzającego usunięcie niezgodności"

    # Format nagłówków
    ws['A2'].alignment = title_align
    ws['A2'].font = Font(size=16, bold=True)
    ws['A3'].alignment = title_align
    ws['A3'].font = Font(size=16, bold=True)

    for col in ['A', 'B', 'C', 'D', 'E', 'F','G','H','I','J','K']:
        ws['%s5' % col].alignment = title_align
        ws['%s5' % col].border = border
        ws['%s5' % col].font = Font(size=12, bold=True)

    # Zawartość zakładek
    row = 6
    scope_str = ['AOC', 'SPO', 'CAMO', 'ATO'][int(scope)]
    type_str = ['Wewnętrzny', 'Zewnętrzny'][int(type)]
    for ncr in NCR.objects.filter(audit_scope=scope_str, audit_type=type_str).order_by('audit_date'):

        # Wypełnienie wierszy
        ws['A%d' % row] = row - 5
        ws['B%d' % row] = ncr.audit_nbr
        ws['C%d' % row] = ncr.audit_date
        ws['D%d' % row] = ncr.ncr_nbr
        ws['E%d' % row] = ncr.description
        ws['F%d' % row] = ncr.due_date
        ws['G%d' % row] = ncr.ncr_user1.__str__() if ncr.ncr_user1 else ''
        ws['H%d' % row] = ncr.ncr_user2.__str__() if ncr.ncr_user2 else ''
        ws['I%d' % row] = ncr.left_days()
        ws['J%d' % row] = ncr.done_date
        ws['K%d' % row] = ncr.check_date

        if ncr.left_days() != None:
            if ncr.left_days() >= 20:
                color = fill_colors['lightgreen']
            elif ncr.left_days() > 0:
                color = fill_colors['yellow']
            else:
                color = fill_colors['lightcoral']

            ws['I%d' % row].fill = color

        row += 1

    # Ustawienie szerokości kolumn
    ws.column_dimensions['A'].width = '8'
    ws.column_dimensions['B'].width = '12'
    ws.column_dimensions['C'].width = '20'
    ws.column_dimensions['D'].width = '8'
    ws.column_dimensions['E'].width = '40'
    ws.column_dimensions['F'].width = '19'
    ws.column_dimensions['G'].width = '20'
    ws.column_dimensions['H'].width = '20'
    ws.column_dimensions['I'].width = '12'
    ws.column_dimensions['J'].width = '20'
    ws.column_dimensions['K'].width = '20'

    # Formatowanie komórek tabeli
    max_row = row
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
        for row in range(6, max_row):
            ws['%s%d' % (col, row)].border = border
            ws['%s%d' % (col, row)].alignment = table_align
    for col in ['A', 'B', 'C', 'D', 'F', 'G', 'H', 'I', 'J', 'K']:
        for row in range(6, max_row):
            ws['%s%d' % (col, row)].alignment = title_align

    # Zwróć arkusz jako obiekt HTTP Response
    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = ('attachment; filename=ncr_report_%s.xlsx' % date.today()).replace('-', '')

    return response


# Generacja informacji o zbliżających się terminach NCR
def NCRCheck (type, days_check, days_left, message):

    res = 0

    # Wybór NCR do zaraportowania
    ncr_query = NCR.objects.filter(audit_type=type, done_date__isnull=True)
    ncr_list = [ncr for ncr in ncr_query if ncr.left_days() == days_check]

    for ncr in ncr_list:

        ncr_string = '- audyt %s nr. %s z dnia %s, NCR nr. %s\n' % (ncr.audit_scope, ncr.audit_nbr, ncr.audit_date, ncr. ncr_nbr)
        email_address = settings.EMAIL_NCR

        if ncr.ncr_user1:
            email_address += [ncr.ncr_user1.email]

        if ncr.ncr_user2:
            email_address += [ncr.ncr_user2.email]

        email = EmailMessage(
            'Zbliżający się termin usunięcia niezgodności',
            message % (days_left, ncr_string),
            settings.EMAIL_FROM,
            email_address,
            ['mierzwik@me.com']
        )

        # wysłanie emaila i wyczysczenie cache
        res += email.send(fail_silently=True)

    return HttpResponse('%s' % res)


# Automatyczne wysłanie maili z zestawieniem rezerwacji
def NCRWarnings():

    NCRMessages = [
        'Za %d dni mija termin zaraportowania do CMM zgodnie z obowiązującymi procedurami sposobu usunięcia następujących niezgodności:\n\n'+
        '%s\n'+
        'W razie uzasadnionej konieczności przedłużenia terminu skontaktuj się z CMM, ze wskazaniem przyczyn nie pozwalających na usunięcie NCR w zaplanowanym terminie.',
        'Za %d dni mija termin usunięcia nastęujących niezgodności:\n\n' +
        '%s\n' +
        'Niezgodność poziomu 2 nieusunięta w terminie lub nieusunięta właściwie traktowana jest jako niezgodność poziomu 1!'
    ]
    NCRCheck('Wewnętrzny', 14, 7, NCRMessages[0])
    NCRCheck('Wewnętrzny', 10, 3, NCRMessages[1])
    NCRCheck('Zewnętrzny', 30, 7, NCRMessages[0])
    NCRCheck('Zewnętrzny', 26, 3, NCRMessages[1])

